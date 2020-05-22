#Copyright © 2020 R. A. Gardner

#from ts_classes_c import *
from ts_extra_vars_and_funcs import *
from ts_widgets import *

import tkinter as tk
from tkinter import ttk, filedialog
import csv as csv_module
from collections import defaultdict, deque, Counter
import os
import datetime
import re
from itertools import islice, repeat, chain, cycle
import json
import pickle
import zlib
import lzma
from base64 import b32encode as b32e, b32decode as b32d
import csv as csv_module
from openpyxl import Workbook, load_workbook
import io
import datetime
from tksheet import Sheet
from math import floor
from ast import literal_eval
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from platform import system as get_os
from fastnumbers import isint, isintlike, isfloat, isreal


class export_flattened_popup(tk.Toplevel):
    def __init__(self,C,
                 width=1280,
                 height=800,
                 theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.title("Export flattened sheet - Click the X button or press escape to go back")
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.wb_ = None

        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(0,weight=1)

        self.selector = single_column_selector(self, theme = theme)
        self.selector.enable_me()
        
        self.selector.set_columns([self.C.headers[h].name for h in self.C.hiers])
        if self.C.pc == -1:
            self.selector.set_col(0)
        else:
            self.selector.set_col(self.C.hiers.index(self.C.pc))
        self.selector.grid(row=0,column=0,sticky="nwe",pady=(10,20),padx=10)

        self.rename_id_col_button = x_checkbutton(self,
                                              text="Rename ID column  ",
                                              style="x_button.Std.TButton",
                                              compound="right",
                                              checked=self.C.xlsx_flattened_rename_id_col.get())
        self.rename_id_col_button.grid(row=1,column=0,sticky="new",pady=(10,5),padx=10)

        self.add_index_button = x_checkbutton(self,
                                              text="Add index column  ",
                                              style="x_button.Std.TButton",
                                              compound="right",
                                              checked=self.C.xlsx_flattened_add_index.get())
        self.add_index_button.grid(row=2,column=0,sticky="new",pady=(10,5),padx=10)

        self.remove_details_button = x_checkbutton(self,
                                                      text="Remove other columns  ",
                                                      style="x_button.Std.TButton",
                                                      compound="right",
                                                      checked=self.C.xlsx_flattened_details.get())
        self.remove_details_button.grid(row=3,column=0,sticky="new",pady=(10,5),padx=10)

        self.split_1st_det_col_button = x_checkbutton(self,
                                                      text="Split treeview detail column  ",
                                                      style="x_button.Std.TButton",
                                                      compound="right",
                                                      checked=self.C.xlsx_flattened_split_1st_det_col.get())
        self.split_1st_det_col_button.grid(row=4,column=0,sticky="new",pady=(10,5),padx=10)

        self.only_base_ids_button = x_checkbutton(self,
                                                  text="Base IDs mode  ",
                                                  style="x_button.Std.TButton",
                                                  compound="right",
                                                  checked=self.C.xlsx_flattened_base_ids.get())
        self.only_base_ids_button.grid(row=5,column=0,sticky="new",pady=(10,5),padx=10)

        self.justify_rows_button = x_checkbutton(self,
                                                     text="Justify rows  ",
                                                     style="x_button.Std.TButton",
                                                     compound="right",
                                                     checked=self.C.xlsx_flattened_justify.get())
        self.justify_rows_button.grid(row=6,column=0,sticky="new",pady=5,padx=10)

        self.order_button = x_checkbutton(self,
                                            text="Reverse Order  ",
                                            style="x_button.Std.TButton",
                                            compound="right",
                                           checked=self.C.xlsx_flattened_reverse_order.get())
        self.order_button.grid(row=7,column=0,sticky="new",pady=5,padx=10)

        self.build_button = button(self,text="  Flatten sheet  ",
                                       style="EF.Std.TButton",
                                       command=self.build_flattened)
        self.build_button.grid(row=8,column=0,pady=10,padx=10,sticky="nsew")
        
        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                  outline_thickness = 0)
        self.sheetdisplay.enable_bindings("enable_all")
        self.sheetdisplay.extra_bindings("begin_edit_cell_use_keypress", self.begin_edit)
        self.sheetdisplay.extra_bindings("escape_edit_cell", self.escape_edit)
        self.sheetdisplay.extra_bindings("edit_cell", self.escape_edit)
        self.sheetdisplay.headers(newheaders = 0)
        self.sheetdisplay.grid(row=0,column=1,rowspan=7,sticky="nswe")

        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(2,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(3,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.button_frame.grid(row=8,column=1,sticky="nswe")
        self.save_button = button(self.button_frame,text="Save as",
                                       style="EF.Std.TButton",
                                       command=self.save_as)
        self.save_button.grid(row=0,column=0,padx=5,pady=20,sticky="nswe")
        self.clipboard_json_button = button(self.button_frame,text="Clipboard as json",
                                               style="EF.Std.TButton",
                                               command=self.clipboard_json)
        self.clipboard_json_button.grid(row=0,column=1,padx=5,pady=20,sticky="nswe")
        self.clipboard_indent_button = button(self.button_frame,text="Clipboard (indent separated)",
                                               style="EF.Std.TButton",
                                               command=self.clipboard_indent)
        self.clipboard_indent_button.grid(row=0,column=2,padx=5,pady=20,sticky="nswe")
        self.clipboard_comma_button = button(self.button_frame,text="Clipboard (comma separated)",
                                               style="EF.Std.TButton",
                                               command=self.clipboard_comma)
        self.clipboard_comma_button.grid(row=0,column=3,padx=5,pady=20,sticky="nswe")
        self.status_bar = StatusBar(self, text = "Use the parent column selector to change hierarchy output", theme = theme)
        self.status_bar.grid(row=9,column=0,columnspan=2,sticky="nswe")
        
        self.bind("<Escape>",self.cancel)
        self.build_flattened()
        
        center(self,width,height)
        self.deiconify()
        self.wait_window()

    def escape_edit(self, event = None):
        self.bind("<Escape>",self.cancel)
        
    def begin_edit(self, event = None):
        self.unbind("<Escape>")

    def start_work(self,msg=""):
        self.status_bar.change_text(msg)
        self.disable_widgets()

    def stop_work(self,msg=""):
        self.status_bar.change_text(msg)
        self.enable_widgets()

    def enable_widgets(self):
        self.sheetdisplay.enable_bindings("enable_all")
        self.sheetdisplay.extra_bindings("begin_edit_cell_use_keypress", self.begin_edit)
        self.sheetdisplay.extra_bindings("escape_edit_cell", self.escape_edit)
        self.sheetdisplay.extra_bindings("edit_cell", self.escape_edit)
        self.sheetdisplay.basic_bindings(True)
        self.save_button.config(state="normal")
        self.clipboard_indent_button.config(state="normal")
        self.clipboard_json_button.config(state="normal")
        self.clipboard_comma_button.config(state="normal")
        self.build_button.config(state="normal")
        self.selector.enable_me()

    def disable_widgets(self):
        self.build_button.config(state="disabled")
        self.sheetdisplay.disable_bindings("disable_all")
        self.sheetdisplay.extra_bindings("begin_edit_cell_use_keypress", None)
        self.sheetdisplay.extra_bindings("escape_edit_cell", None)
        self.sheetdisplay.extra_bindings("edit_cell", None)
        self.sheetdisplay.basic_bindings(False)
        self.save_button.config(state="disabled")
        self.clipboard_json_button.config(state="disabled")
        self.clipboard_indent_button.config(state="disabled")
        self.clipboard_comma_button.config(state="disabled")
        self.selector.disable_me()
        self.update()

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except:
            pass
        try:
            self.wb_ = None
        except:
            pass

    def USER_HAS_CLOSED_WINDOW(self,callback=None):
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except:
            pass
        self.destroy()

    def clipboard_json(self):
        self.start_work("Copying to clipboard...")
        self.C.C.clipboard_clear()
        self.C.C.clipboard_append(json.dumps(self.C.dump_full_sheet_to_json(self.sheetdisplay.get_sheet_data()[0],self.sheetdisplay.get_sheet_data()[1:],include_headers=True)))
        self.C.C.update()
        self.stop_work("Sheet successfully copied to clipboard as json!")

    def clipboard_indent(self):
        self.start_work("Copying to clipboard...")
        s = io.StringIO()
        writer = csv_module.writer(s,dialect=csv_module.excel_tab,lineterminator="\n")
        for row in self.sheetdisplay.get_sheet_data():
            writer.writerow(row)
        s = s.getvalue().rstrip()
        self.C.C.clipboard_clear()
        self.C.C.clipboard_append(s)
        self.C.C.update()
        self.stop_work("Sheet successfully copied to clipboard (indent separated)!")

    def clipboard_comma(self):
        self.start_work("Copying to clipboard...")
        s = io.StringIO()
        writer = csv_module.writer(s,dialect=csv_module.excel,lineterminator="\n")
        for row in self.sheetdisplay.get_sheet_data():
            writer.writerow(row)
        s = s.getvalue().rstrip()
        self.C.C.clipboard_clear()
        self.C.C.clipboard_append(s)
        self.C.C.update()
        self.stop_work("Sheet successfully copied to clipboard (comma separated)!")

    def build_flattened(self):
        self.start_work("Flattening sheet...")
        self.sheetdisplay.deselect("all")
        self.sheetdisplay.set_sheet_data(data = self.C.treebuilder.build_flattened(self.C.sheet,
                                                                                     self.sheetdisplay.get_sheet_data(),
                                                                                   self.C.nodes,
                                                                                   self.C.column_index1,
                                                                                     [f"{hdr.name}" for hdr in self.C.headers],
                                                                                     int(self.C.ic),
                                                                                     int(self.C.hiers[self.selector.get_col()]),
                                                                                     list(self.C.hiers),
                                                                                     self.justify_rows_button.get_checked(),
                                                                                     self.order_button.get_checked(),
                                                                                     self.only_base_ids_button.get_checked(),
                                                                                     self.remove_details_button.get_checked(),
                                                                                     self.split_1st_det_col_button.get_checked(),
                                                                                     self.add_index_button.get_checked(),
                                                                                     self.rename_id_col_button.get_checked()),
                                                         verify = False)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.stop_work("Sheet successfully flattened!")
        
    def save_as(self):
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save flattened sheet as",
                                               filetypes=[('Excel file','.xlsx'),('JSON File','.json'),('CSV File','.csv')],
                                               defaultextension=".xlsx",
                                               confirmoverwrite=True)
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv",".xlsx",".json")):
            self.grab_set()
            self.stop_work("Can only save .json/.csv/.xlsx file types")
            return
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook()
                ws = self.wb_.active
                for rn,row in enumerate(self.sheetdisplay.get_sheet_data()):
                    ws.append(row)
                    if not rn % 50:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status_bar.change_text("".join(("Saving...  rows: ",str(rn))))
                ws.freeze_panes = "A2"
                if self.sheetdisplay.get_sheet_data():
                    for i in range(1, len(self.sheetdisplay.get_sheet_data()[0]) + 1):
                        ws.cell(row = 1, column = i).fill = orange_fill
                        ws.cell(row = 1, column = i).border = openpyxl_thin_border
                        ws.column_dimensions[xl_column_string(i)].width = 25
                self.wb_.save(newfile)
                self.try_to_close_wb()
            elif newfile.lower().endswith(".json"):
                with open(newfile,"w",newline="") as fh:
                    fh.write(json.dumps(self.C.dump_full_sheet_to_json(self.sheetdisplay.get_sheet_data()[0],self.sheetdisplay.get_sheet_data()[1:],include_headers=True)))             
            elif newfile.lower().endswith(".csv"):
                with open(newfile,"w",newline="") as fh:
                    writer = csv_module.writer(fh,dialect=csv_module.excel_tab,lineterminator="\n")
                    for rn,row in enumerate(self.sheetdisplay.get_sheet_data()):
                        writer.writerow(row)
                        if not rn % 50:
                            self.update()
                            if self.USER_HAS_QUIT:
                                return
                            self.status_bar.change_text("".join(("Saving...  rows: ",str(rn))))
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work("Error saving file: " + str(error_msg))
            return
        self.stop_work("Success! Flattened sheet saved")
        
    def cancel(self,event=None):
        self.USER_HAS_CLOSED_WINDOW()


class post_import_changes_popup(tk.Toplevel):
    def __init__(self, C, changes, successful, width = 1200, height = 800, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.title("Successful / Unsuccessful Changes - Click the X button or press escape to go back")
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.total_changes = "Total changes: " + str(len(changes))
        
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(0,weight=1)
        
        self.changes = changes
        self.successful = successful
        
        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                    outline_thickness=0,
                                    row_index_width=245)
        self.sheetdisplay.enable_bindings(("single",
                                           "copy",
                                           "drag_select",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.sheetdisplay.headers(newheaders=["Date","User","Type","ID/Name/Number","Old Value","New Value"])
        self.sheetdisplay.row_index(0)
        self.sheetdisplay.data_reference(newdataref=self.changes,reset_col_positions=False,reset_row_positions=False,redraw=False)
        self.sheetdisplay.display_subset_of_columns(indexes=[1,2,3,4,5],enable=True,reset_col_positions=False)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        for i, b in enumerate(reversed(self.successful)):
            if b:
                self.sheetdisplay.highlight_cells(row = i, canvas = "row_index", bg = theme_green_bg(theme), fg = theme_green_fg(theme))
                for c in range(6):
                    self.sheetdisplay.highlight_cells(row = i, column = c, bg = theme_green_bg(theme), fg = theme_green_fg(theme))
            else:
                self.sheetdisplay.highlight_cells(row = i, canvas = "row_index", bg = theme_red_bg(theme), fg = theme_red_fg(theme))
                for c in range(6):
                    self.sheetdisplay.highlight_cells(row = i, column = c, bg = theme_red_bg(theme), fg = theme_red_fg(theme))
        self.sheetdisplay.grid(row=0,column=0,sticky="nswe")
        self.status_bar = StatusBar(self, text = self.total_changes, theme = theme)
        self.status_bar.grid(row=1,column=0,sticky="nswe")
        self.bind("<Escape>",self.cancel)
        center(self,width,height)
        self.deiconify()
        self.wait_window()

    def cancel(self,event=None):
        self.destroy()


class changelog_popup(tk.Toplevel):
    def __init__(self,C,width=999,height=800, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.title("Changelog - Click the X button or press escape to go back")
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.find_results = []
        self.results_number = 0
        self.wb_ = None
        self.total_changes = "Total changes: " + str(len(self.C.changelog)) + " | "
        
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        
        self.find_frame = frame(self, theme = theme)
        self.find_frame.grid(row=0,column=0,columnspan=2,sticky="nswe")
        self.find_icon = tk.PhotoImage(format="gif",data=find_icon)
        self.search_button = button(self.find_frame,
                                    text=" Find:",
                                    command=self.find)
        self.search_button.config(image=self.find_icon,compound="left")
        self.search_button.pack(side="left",fill="x")
        self.find_window = normal_entry(self.find_frame,font=BF, theme = theme)
        self.find_window.bind("<Return>",self.find)
        self.find_window.pack(side="left",fill="x",expand=True)
        self.find_reset_button = button(self.find_frame,text="X",command=self.find_reset)
        self.find_reset_button.pack(side="left",fill="x")
        self.find_results_label = label(self.find_frame,"0/0",BF, theme = theme)
        self.find_results_label.pack(side="left",fill="x")
        self.find_up_button = button(self.find_frame,text="▲",command=self.find_up)
        self.find_up_button.pack(side="left",fill="x")
        self.find_down_button = button(self.find_frame,text="▼",command=self.find_down)
        self.find_down_button.pack(side="left",fill="x")
        self.changelog = self.C.changelog[::-1]
        
        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  row_index_align = "w",
                                  header_font = ("Calibri", 13, "normal"),
                                       outline_thickness=0)
        self.sheetdisplay.enable_bindings(("single",
                                           "copy",
                                           "drag_select",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.sheetdisplay.headers(newheaders=["Date","User","Type","ID/Name/Number","Old Value","New Value"])
        self.sheetdisplay.row_index(0)
        self.sheetdisplay.data_reference(newdataref=self.changelog,reset_col_positions=False,reset_row_positions=True,redraw=False)
        self.sheetdisplay.display_subset_of_columns(indexes=[1,2,3,4,5],enable=True,reset_col_positions=False)
        self.red_bg = theme_red_bg(theme)
        self.green_bg = theme_green_bg(theme)
        self.red_fg = theme_red_fg(theme)
        self.green_fg = theme_green_fg(theme)
        self.sheetdisplay.highlight_cells(column = 4,
                                          canvas = "header",
                                          bg = self.red_bg,
                                          fg = self.red_fg)
        self.sheetdisplay.highlight_cells(column = 5,
                                          canvas = "header",
                                          bg = self.green_bg,
                                          fg = self.green_fg)
        for r in range(len(self.changelog)):
            self.sheetdisplay.highlight_cells(row = r,
                                              column = 4,
                                              bg = self.red_bg,
                                              fg = self.red_fg)
            self.sheetdisplay.highlight_cells(row = r,
                                              column = 5,
                                              bg = self.green_bg,
                                              fg = self.green_fg)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.sheetdisplay.set_width_of_index_to_text()
        self.sheetdisplay.grid(row=1,column=0,sticky="nswe")
        self.status_bar = StatusBar(self, text = self.total_changes, theme = theme)
        self.status_bar.grid(row=2,column=0,sticky="nswe")
        
        self.buttonframe = frame(self, theme = theme)
        self.buttonframe.grid_columnconfigure(0, weight = 1, uniform = "b")
        self.buttonframe.grid_columnconfigure(1, weight = 1, uniform = "b")
        self.buttonframe.grid_columnconfigure(2, weight = 1, uniform = "b")
        
        self.buttonframe.grid(row=3,column=0,sticky="nswe")
        self.save_text_button = button(self.buttonframe,text="Export all",
                                       style="EF.Std.TButton",
                                       command=self.save_as)
        self.save_text_button.grid(row = 0, column = 0, padx = 25, pady = 20, sticky = "nswe")

        self.export_selected_button = button(self.buttonframe,text="Export selected as",
                                               style="EF.Std.TButton",
                                               command=self.save_selected_as)
        self.export_selected_button.grid(row = 0, column = 1, padx = 25, pady = 20, sticky = "nswe")
        
        self.prune_button = button(self.buttonframe,text="Prune from selected",
                                    style="EF.Std.TButton",
                                    command=self.prune)
        self.prune_button.grid(row = 0, column = 2, padx = 25, pady = 20, sticky = "nswe")
        
        self.bind("<Escape>",self.cancel)
        center(self,width,height)
        self.deiconify()
        self.wait_window()

    def prune(self, event = None):
        selectedrows = self.sheetdisplay.get_selected_rows(get_cells_as_rows = True, return_tuple = True)
        if not selectedrows:
            return
        num = len(selectedrows)
        self.start_work(f"Pruning {num} changes...")
        up_to = len(self.C.changelog) - min(selectedrows) - 1
        if self.C.changelog[up_to][2].endswith(("|", "| ")):
            for i, entry in enumerate(islice(self.C.changelog, up_to, None), up_to):
                if not entry[2].endswith(("|", "| ")):
                    up_to = i
                    break
        self.C.snapshot_prune_changelog(up_to)
        self.C.changelog[:up_to + 1] = []
        self.changelog = self.C.changelog[::-1]
        self.sheetdisplay.headers(newheaders=["Date","User","Type","ID/Name/Number","Old Value","New Value"])
        self.sheetdisplay.row_index(newindex=0)
        self.sheetdisplay.data_reference(newdataref=self.changelog,reset_col_positions=False,reset_row_positions=True,redraw=False)
        self.sheetdisplay.display_subset_of_columns(indexes=[1,2,3,4,5],enable=True,reset_col_positions=False)
        self.sheetdisplay.dehighlight_cells(all_ = True)
        for r in range(len(self.changelog)):
            self.sheetdisplay.highlight_cells(row = r, column = 4, bg = self.red_bg, fg = self.red_fg)
            self.sheetdisplay.highlight_cells(row = r, column = 5, bg = self.green_bg, fg = self.green_fg)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.total_changes = "Total changes: " + str(len(self.C.changelog)) + " | "
        self.status_bar.config(text = self.total_changes)
        self.C.C.status_bar.change_text(self.C.set_status_bar())
        self.sheetdisplay.refresh()
        self.stop_work("Success! Changelog pruned")

    def start_work(self,msg=""):
        self.status_bar.change_text(self.total_changes + msg)
        self.disable_widgets()

    def stop_work(self,msg=""):
        self.status_bar.change_text(self.total_changes + msg)
        self.enable_widgets()

    def enable_widgets(self):
        self.sheetdisplay.enable_bindings(("single",
                                           "copy",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.find_window.bind("<Return>",self.find)
        self.find_reset_button.config(state="normal")
        self.find_up_button.config(state="normal")
        self.find_down_button.config(state="normal")
        self.save_text_button.config(state="normal")

    def disable_widgets(self):
        self.sheetdisplay.disable_bindings(("single",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                           "row_width_resize",
                                           "row_select",
                                           "arrowkeys"))
        self.find_window.unbind("<Return>")
        self.find_reset_button.config(state="disabled")
        self.find_up_button.config(state="disabled")
        self.find_down_button.config(state="disabled")
        self.save_text_button.config(state="disabled")
        self.update()

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except:
            pass
        try:
            self.wb_ = None
        except:
            pass

    def USER_HAS_CLOSED_WINDOW(self,callback=None):
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except:
            pass
        self.destroy()       
        
    def save_as(self):
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save changes as",
                                               filetypes=[('CSV File','.csv'),('Excel file','.xlsx'),('JSON File','.json')],
                                               defaultextension=".csv",
                                               confirmoverwrite=True)
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv",".xlsx",".json")):
            self.grab_set()
            self.stop_work("Can only save .csv/.xlsx/.json file types")
            return
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook()
                ws = self.wb_.active
                ws.append(["Date","User","Type","ID/Name/Number","Old Value","New Value"])
                for rn,row in enumerate(self.changelog):
                    ws.append(row)
                    if not rn % 20:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status_bar.change_text("".join((self.total_changes,"Saving...  changes: ",str(rn))))
                ws.freeze_panes = "A2"
                for i in range(1, 7):
                    if i < 5:
                        ws.cell(row = 1, column = i).fill = orange_fill
                    elif i == 5:
                        ws.cell(row = 1, column = i).fill = red_remove_fill
                    else:
                        ws.cell(row = 1, column = i).fill = green_add_fill
                    ws.cell(row = 1, column = i).border = openpyxl_thin_border
                ws.column_dimensions["A"].width = 37
                ws.column_dimensions["B"].width = 20
                ws.column_dimensions["C"].width = 37
                ws.column_dimensions["D"].width = 52
                ws.column_dimensions["E"].width = 60
                ws.column_dimensions["F"].width = 60
                self.wb_.save(newfile)
                self.try_to_close_wb()
            elif newfile.lower().endswith(".csv"):
                with open(newfile,"w",newline="") as fh:
                    writer = csv_module.writer(fh,dialect=csv_module.excel_tab,lineterminator="\n")
                    writer.writerow(["Date","User","Type","ID/Name/Number","Old Value","New Value"])
                    for rn,row in enumerate(self.changelog):
                        writer.writerow(row)
                        if not rn % 20:
                            self.update()
                            if self.USER_HAS_QUIT:
                                return
                            self.status_bar.change_text("".join((self.total_changes,"Saving...  changes: ",str(rn))))
            elif newfile.lower().endswith(".json"):
                with open(newfile,"w",newline="") as fh:
                    fh.write(json.dumps(self.C.dump_full_sheet_to_json(["Date","User","Type","ID/Name/Number","Old Value","New Value"],self.changelog,include_headers=True)))
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work("Error saving file: " + str(error_msg))
            return
        self.stop_work("Success! Changelog saved")

    def save_selected_as(self):
        selectedrows = self.sheetdisplay.get_selected_rows(get_cells_as_rows = True, return_tuple = True)
        if not selectedrows:
            return
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save selected changes as",
                                               filetypes=[('CSV File','.csv'),('Excel file','.xlsx'),('JSON File','.json')],
                                               defaultextension=".csv",
                                               confirmoverwrite=True)
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv",".xlsx",".json")):
            self.grab_set()
            self.stop_work("Can only save .csv/.xlsx/.json file types")
            return
        from_row = min(selectedrows)
        to_row   = max(selectedrows) + 1
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook()
                ws = self.wb_.active
                ws.append(["Date","User","Type","ID/Name/Number","Old Value","New Value"])
                for rn,row in enumerate(islice(self.changelog, from_row, to_row)):
                    ws.append(row)
                    if not rn % 20:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status_bar.change_text("".join((self.total_changes,"Saving...  changes: ",str(rn))))
                ws.freeze_panes = "A2"
                for i in range(1, 7):
                    if i < 5:
                        ws.cell(row = 1, column = i).fill = orange_fill
                    elif i == 5:
                        ws.cell(row = 1, column = i).fill = red_remove_fill
                    else:
                        ws.cell(row = 1, column = i).fill = green_add_fill
                    ws.cell(row = 1, column = i).border = openpyxl_thin_border
                ws.column_dimensions["A"].width = 37
                ws.column_dimensions["B"].width = 20
                ws.column_dimensions["C"].width = 37
                ws.column_dimensions["D"].width = 52
                ws.column_dimensions["E"].width = 60
                ws.column_dimensions["F"].width = 60
                self.wb_.save(newfile)
                self.try_to_close_wb()
            elif newfile.lower().endswith(".csv"):
                with open(newfile,"w",newline="") as fh:
                    writer = csv_module.writer(fh,dialect=csv_module.excel_tab,lineterminator="\n")
                    writer.writerow(["Date","User","Type","ID/Name/Number","Old Value","New Value"])
                    for rn,row in enumerate(islice(self.changelog, from_row, to_row)):
                        writer.writerow(row)
                        if not rn % 20:
                            self.update()
                            if self.USER_HAS_QUIT:
                                return
                            self.status_bar.change_text("".join((self.total_changes,"Saving...  changes: ",str(rn))))
            elif newfile.lower().endswith(".json"):
                with open(newfile,"w",newline="") as fh:
                    fh.write(json.dumps(self.C.dump_full_sheet_to_json(["Date","User","Type","ID/Name/Number","Old Value","New Value"],self.changelog[from_row:to_row],include_headers=True)))
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work("Error saving file: " + str(error_msg))
            return
        self.stop_work("Success! Changelog saved")

    def find(self,event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        x = self.word.lower()
        for rn,row in enumerate(self.changelog):
            for colno,cell in enumerate(row):
                if x in cell.lower():
                    if colno == 0:
                        self.find_results.append((rn,6))
                        break
                    else:
                        self.find_results.append((rn,colno))
        if self.find_results:
            for rn,colno in islice(self.find_results,1,len(self.find_results)):
                if colno == 6:
                    for i in range(1,6):
                        self.sheetdisplay.highlight_cells(row = rn, column = i, bg = "yellow",fg="black")
                else:
                    self.sheetdisplay.highlight_cells(row=rn,column=colno,bg="yellow",fg="black")
            if self.find_results[self.results_number][1] == 6:
                for i in range(1,6):
                    self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="orange",fg="black")
            else:
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange",fg="black")
            self.find_results_label.config(text="1/"+str(len(self.find_results)))
            self.sheetdisplay.see(row=self.find_results[0][0],column=0,keep_xscroll=True)
        self.sheetdisplay.refresh()
            
    def find_up(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        if self.find_results[self.results_number][1] == 6:
            for i in range(1,6):
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="yellow",fg="black")
        else:
            self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="yellow",fg="black")
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        if self.find_results[self.results_number][1] == 6:
            for i in range(1,6):
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="orange",fg="black")
        else:
            self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange",fg="black")
        self.sheetdisplay.see(row=self.find_results[self.results_number][0],column=0,keep_xscroll=True)
        self.sheetdisplay.refresh()
        
    def find_down(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        if self.find_results[self.results_number][1] == 6:
            for i in range(1,6):
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="yellow",fg="black")
        else:
            self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="yellow",fg="black")
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        if self.find_results[self.results_number][1] == 6:
            for i in range(1,6):
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="orange",fg="black")
        else:
            self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange",fg="black")
        self.sheetdisplay.see(row=self.find_results[self.results_number][0],column=0,keep_xscroll=True)
        self.sheetdisplay.refresh()
        
    def find_reset(self,newfind=False):
        self.find_results = []
        self.results_number = 0
        self.sheetdisplay.dehighlight_cells(all_=True,redraw=False)
        if newfind == False:
            self.find_window.delete(0,"end")
        self.find_results_label.config(text="0/0")
        for r in range(len(self.changelog)):
            self.sheetdisplay.highlight_cells(row = r,
                                              column = 4,
                                              bg = self.red_bg,
                                              fg = self.red_fg)
            self.sheetdisplay.highlight_cells(row = r,
                                              column = 5,
                                              bg = self.green_bg,
                                              fg = self.green_fg)
        self.sheetdisplay.refresh()
        
    def cancel(self,event=None):
        self.USER_HAS_CLOSED_WINDOW()


class compare_report_popup(tk.Toplevel):
    def __init__(self,
                 C,
                 width=1200,
                 height=800,
                 theme = "dark"):
        tk.Toplevel.__init__(self,
                             C,
                             width="1",
                             height="1",
                             bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.title("Comparison Report - Click the X button or press escape to go back")
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.find_results = []
        self.results_number = 0
        self.wb_ = None
        report = self.C.report
        self.sheet1name = self.C.sheetname_1
        self.sheet2name = self.C.sheetname_2

        self.open_tab = 1
        
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)

        self.notebook = ttk.Notebook(self)
        self.notebook.grid(row=1,column=0,sticky="nswe")
        
        self.f1 = frame(self, theme = theme)
        self.f1.grid_columnconfigure(0,weight=1)
        self.f1.grid_rowconfigure(1,weight=1)
        self.notebook.add(self.f1, text = "Matching IDs Differences")

        self.f2 = frame(self, theme = theme)
        self.f2.grid_columnconfigure(0,weight=1)
        self.f2.grid_rowconfigure(1,weight=1)
        self.notebook.add(self.f2, text = "Other Differences")
        self.notebook.select(self.f1)
        self.notebook.enable_traversal()
        self.notebook.bind("<<NotebookTabChanged>>", self.tab_change)
        
        self.find_frame = frame(self, theme = theme)
        self.find_frame.grid(row=0,column=0,columnspan=2,sticky="nswe")
        self.find_icon = tk.PhotoImage(format="gif",data=find_icon)
        self.search_button = button(self.find_frame,
                                    text=" Find:",
                                    command=self.find)
        self.search_button.config(image=self.find_icon,compound="left")
        self.search_button.pack(side="left",fill="x")
        self.find_window = normal_entry(self.find_frame,font=BF, theme = theme)
        self.find_window.bind("<Return>",self.find)
        self.find_window.pack(side="left",fill="x",expand=True)
        self.find_reset_button = button(self.find_frame,text="X",command=self.find_reset)
        self.find_reset_button.pack(side="left",fill="x")
        self.find_results_label = label(self.find_frame,"0/0",BF, theme = theme)
        self.find_results_label.pack(side="left",fill="x")
        self.find_up_button = button(self.find_frame,text="▲",command=self.find_up)
        self.find_up_button.pack(side="left",fill="x")
        self.find_down_button = button(self.find_frame,text="▼",command=self.find_down)
        self.find_down_button.pack(side="left",fill="x")
        
        self.sheetdisplay1 = Sheet(self.f1,
                                   theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                       outline_thickness=0)
        self.sheetdisplay1.enable_bindings(("single",
                                            "copy",
                                           "drag_select",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.sheetdisplay1.headers(newheaders=["ID","Difference",self.sheet1name,self.sheet2name])
        self.sheetdisplay1.data_reference(newdataref = report['ids'],
                                         reset_col_positions = False,
                                         reset_row_positions = False,
                                         redraw = False)
            
        self.sheetdisplay1.set_all_cell_sizes_to_text()
        self.sheetdisplay1.grid(row=1,column=0,sticky="nswe")

        self.sheetdisplay2 = Sheet(self.f2,
                                   theme = theme,
                                   outline_thickness=0)
        self.sheetdisplay2.enable_bindings(("single",
                                            "copy",
                                           "drag_select",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                           "row_width_resize",
                                           "row_select",
                                           "arrowkeys"))
        self.sheetdisplay2.data_reference(newdataref = report['info'],
                                          reset_col_positions = False,
                                          reset_row_positions = False,
                                          redraw = False)

        self.sheetdisplay2.set_all_cell_sizes_to_text()
        self.sheetdisplay2.grid(row=1,column=0,sticky="nswe")
        
        self.buttonframe = frame(self, theme = theme)
        self.buttonframe.grid(row=3,column=0,sticky="nswe")
        self.cancel_button = button(self.buttonframe,text="Done",
                                       style="EF.Std.TButton",
                                       command=self.cancel)
        self.cancel_button.pack(side = "right", padx = (20,100), pady = 20)
        self.save_text_button = button(self.buttonframe,text="Save Report",
                                       style="EF.Std.TButton",
                                       command=self.save_report)
        self.save_text_button.pack(side = "right", padx = (50, 30), pady = 20)
        

        self.bind("<Escape>",self.cancel)
        center(self,width,height)
        self.deiconify()
        self.wait_window()

    def tab_change(self, event = None):
        self.find_reset(True)
        self.open_tab = self.notebook.index(self.notebook.select()) + 1

    def start_work(self,msg=""):
        self.disable_widgets()

    def stop_work(self,msg=""):
        self.enable_widgets()

    def enable_widgets(self):
        self.sheetdisplay1.enable_bindings(("single",
                                           "copy",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.sheetdisplay2.enable_bindings(("single",
                                           "copy",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.find_window.bind("<Return>",self.find)
        self.find_reset_button.config(state="normal")
        self.find_up_button.config(state="normal")
        self.find_down_button.config(state="normal")
        self.save_text_button.config(state="normal")

    def disable_widgets(self):
        self.sheetdisplay1.disable_bindings(("single",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                           "row_width_resize",
                                           "row_select",
                                           "arrowkeys"))
        self.sheetdisplay2.disable_bindings(("single",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                           "row_width_resize",
                                           "row_select",
                                           "arrowkeys"))
        self.find_window.unbind("<Return>")
        self.find_reset_button.config(state="disabled")
        self.find_up_button.config(state="disabled")
        self.find_down_button.config(state="disabled")
        self.save_text_button.config(state="disabled")
        self.update()

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except:
            pass
        try:
            self.wb_ = None
        except:
            pass

    def USER_HAS_CLOSED_WINDOW(self,callback=None):
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except:
            pass
        self.destroy()       
        
    def save_report(self):
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save as",
                                               filetypes=[('Excel file','.xlsx')],
                                               defaultextension=".xlsx",
                                               confirmoverwrite=True)
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith(".xlsx"):
            self.grab_set()
            self.stop_work("Can only save .xlsx file type")
            return
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook()
                ws = self.wb_.active
                ws.title = "Matching IDs Differences"
                ws.append(["ID", "Difference", self.sheet1name, self.sheet2name])
                for rn,row in enumerate(self.sheetdisplay1.get_sheet_data()):
                    ws.append(row)
                ws.freeze_panes = "A2"
                for i in range(1, 6):
                    ws.cell(row = 1, column = i).fill = orange_fill
                    ws.cell(row = 1, column = i).border = openpyxl_thin_border

                ws = self.wb_.create_sheet(title = "Other Differences")
                for rn,row in enumerate(self.sheetdisplay2.get_sheet_data()):
                    ws.append(row)
                
                self.wb_.save(newfile)
                self.try_to_close_wb()
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work("Error saving file: " + str(error_msg))
            return
        self.stop_work("Success! Report saved")

    def find(self,event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        x = self.word.lower()
        if self.open_tab == 1:
            target_sheet = self.sheetdisplay1
            find_res = self.find_results
            res_num = self.results_number
        else:
            target_sheet = self.sheetdisplay2
        for rn, row in enumerate(target_sheet.get_sheet_data()):
            for colno, cell in enumerate(row):
                if x in cell.lower():
                    self.find_results.append((rn, colno))
        if self.find_results:
            for rn,colno in islice(self.find_results,1,len(self.find_results)):
                target_sheet.highlight_cells(row=rn,column=colno,bg="yellow")
            target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange")
            self.find_results_label.config(text="1/"+str(len(self.find_results)))
            target_sheet.see(row=self.find_results[0][0],column=0,keep_xscroll=True)
        target_sheet.refresh()
            
    def find_up(self,event=None):
        if self.open_tab == 1:
            target_sheet = self.sheetdisplay1
        else:
            target_sheet = self.sheetdisplay2
        if not self.find_results or len(self.find_results) == 1:
            return
        target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="yellow")
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange")
        target_sheet.see(row=self.find_results[self.results_number][0],column=0,keep_xscroll=True)
        target_sheet.refresh()
        
    def find_down(self,event=None):
        if self.open_tab == 1:
            target_sheet = self.sheetdisplay1
        else:
            target_sheet = self.sheetdisplay2
        if not self.find_results or len(self.find_results) == 1:
            return
        target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="yellow")
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange")
        target_sheet.see(row=self.find_results[self.results_number][0],column=0,keep_xscroll=True)
        target_sheet.refresh()
        
    def find_reset(self,newfind=False):
        try:
            self.find_results = []
            self.results_number = 0
            self.sheetdisplay1.dehighlight_cells(all_=True,redraw=True)
            self.sheetdisplay2.dehighlight_cells(all_=True,redraw=True)
            if newfind == False:
                self.find_window.delete(0,"end")
            self.find_results_label.config(text="0/0")
        except:
            pass
        
    def cancel(self,event=None):
        self.USER_HAS_CLOSED_WINDOW()


class find_and_replace_popup(tk.Toplevel):
    def __init__(self, C, ss_selection, theme = "dark", within = False, pars = False):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Find and replace - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grid_columnconfigure(0, weight = 1)
        
        self.last_found = None
        self.last_replaced = None
        
        self.notebook = ttk.Notebook(self)
        self.notebook.grid(row=0,column=0,sticky="nswe")
        
        self.f1 = frame(self, theme = theme)
        self.f1.grid_columnconfigure(0,weight=1)
        self.f1.grid_columnconfigure(1, weight = 1)
        self.notebook.add(self.f1, text = "Find in all Hierarchies")

        self.f2 = frame(self, theme = theme)
        self.f2.grid_columnconfigure(0,weight=1)
        self.f2.grid_columnconfigure(1, weight = 1)
        self.notebook.add(self.f2, text = "Find in Sheet")
        self.notebook.select(self.f2)
        self.notebook.enable_traversal()

        self.id_or_detail = ez_dropdown(self.f1,EF)
        self.id_or_detail['values'] = ["Find ID","Find detail"]
        self.id_or_detail.set_my_value("Find ID")
        self.id_or_detail.grid(row=0,column=0,sticky="nswe",padx=(20,5),pady=(10,5))
        self.match_option = ez_dropdown(self.f1,EF)
        self.match_option['values'] = ["Non-exact match","Exact match"]
        self.match_option.set_my_value("Non-exact match")
        self.match_option.grid(row=0,column=1,sticky="nswe",padx=(5,20),pady=(10,5))
        self.tv_find_entry = entry_with_scrollbar(self.f1, theme = theme)
        self.tv_find_entry.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=10)
        self.enter_ss_sel_button = button(self.f1,text="Enter current sheet selection",style="EF.Std.TButton",command=self.enter_ss_sel)
        self.enter_ss_sel_button.grid(row=2,column=0,columnspan=2,sticky="nswe",padx=20,pady=5)
        if not ss_selection:
            self.enter_ss_sel_button.config(state="disabled")
        else:
            self.ss_sel = ss_selection

        self.bf = frame(self.f1, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe")
        
        self.tv_confirm_button = button(self.bf,text="Find",style="EF.Std.TButton",command=self.tv_confirm)
        self.tv_confirm_button.grid(row=0,column=0,sticky="nswe",padx = (45, 30), pady = 10)
        self.tv_cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.tv_cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (30, 45), pady = 10)
        
        self.id_or_detail.bind("<<ComboboxSelected>>",lambda event: self.tv_find_entry.place_cursor())
        self.match_option.bind("<<ComboboxSelected>>",lambda event: self.tv_find_entry.place_cursor())

        self.frframe = frame(self.f2, theme = theme)
        self.frframe.grid_columnconfigure(1, weight = 1)
        self.frframe.grid(row = 0, column = 0, columnspan = 2, sticky = "nswe")

        self.find_label = label(self.frframe,text="Find",font=EF, theme = theme, anchor = "w")
        self.find_label.grid(row=0,column=0,sticky="nswe", pady = (0,14), padx=(20,10))
        self.find_display = entry_with_scrollbar(self.frframe, theme = theme)
        self.find_display.grid(row=0,column=1,sticky="nswe",pady=10,padx=(0,20))

        self.rep_label = label(self.frframe,text="Replace with",font=EF, theme = theme, anchor = "w")
        self.rep_label.grid(row=1,column=0,sticky="nswe", pady = (0, 17), padx=(20,10))
        self.rep_display = entry_with_scrollbar(self.frframe, theme = theme)
        self.rep_display.grid(row=1,column=1,sticky="nswe",pady=10,padx=(0,20))

        self.ids_button = x_checkbutton(self.frframe,
                                          text="  Find Parents",
                                          style="wx_button.Std.TButton",
                                        checked = pars,
                                          compound="left")
        self.ids_button.grid(row=2,column=1,padx=(0, 20),pady=5,sticky="we")

        self.details_button = x_checkbutton(self.frframe,
                                            text="  Find Details",
                                            style="wx_button.Std.TButton",
                                            checked = True,
                                            compound="left")
        self.details_button.grid(row=3,column=1,padx=(0, 20),pady=5,sticky="we")

        self.where = x_checkbutton(self.frframe,
                                  text="  Only within selected cells",
                                  style="wx_button.Std.TButton",
                                   checked = bool(within),
                                 compound="left")
        self.where.grid(row=4,column=1,padx=(0, 20),pady=5,sticky="we")

        self.match_button = x_checkbutton(self.frframe,
                                            text="  Exact match",
                                            style="wx_button.Std.TButton",
                                            compound="left")
        self.match_button.grid(row=5,column=1,padx=(0, 20),pady=5,sticky="we")

        self.allcols_button = x_checkbutton(self.frframe,
                                            text="  Show and include hidden columns",
                                            style="wx_button.Std.TButton",
                                            checked = False,
                                            compound="left")
        self.allcols_button.grid(row=6,column=1,padx=(0, 20),pady=(5,10),sticky="we")

        self.bf2 = frame(self.frframe, theme = theme)
        self.bf2.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf2.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf2.grid_columnconfigure(2, weight = 1, uniform = "x")
        self.bf2.grid_columnconfigure(3, weight = 1, uniform = "x")
        self.bf2.grid(row = 7, column = 0, columnspan = 2, sticky = "nswe")

        self.find_button = button(self.bf2,text="Find next",style="EF.Std.TButton",command=self.find_next)
        self.find_button.grid(row=0,column=0,sticky="nswe",padx = (20, 5), pady = (15, 10))
        
        self.replace_button = button(self.bf2,text="Replace next",style="EF.Std.TButton",command=self.replace_next)
        self.replace_button.grid(row = 0, column = 1, sticky = "nswe", padx = 5, pady = (15, 10))
        
        self.replace_all_button = button(self.bf2,text="Replace all",style="EF.Std.TButton",command=self.replace_all)
        self.replace_all_button.grid(row=0,column=2,sticky="nswe",padx = 5, pady = (15, 10))
        
        self.done_button = button(self.bf2,text="Done",style="EF.Std.TButton",command=self.cancel)
        self.done_button.grid(row = 0, column = 3, sticky = "nswe", padx = (5, 20), pady = (15, 10))

        self.status_bar = readonly_entry_with_scrollbar(self, theme = theme)
        self.status_bar.change_text(text = "Please note that case is currently ignored in search results")
        self.status_bar.my_entry.config(relief = "flat", font = ("Calibri", 11))
        self.status_bar.grid(row = 1, column = 0, sticky = "we")
        
        self.tv_find_entry.my_entry.bind("<Return>",self.tv_confirm)
        self.find_display.my_entry.bind("<Return>", self.find_next)
        self.rep_display.my_entry.bind("<Return>", self.find_next)
        self.bind("<Control-g>", self.find_next)
        self.bind("<Control-G>", self.find_next)
        self.bind("<Escape>",self.cancel)
        self.bind("<Control-z>", self.C.undo)
        self.bind("<Control-Z>", self.C.undo)
        self.result = False
        self.find_display.place_cursor()
        self.tv_find_entry.place_cursor()
        center(self,470,440, move_left = True)
        self.deiconify()

    def see_and_set(self, r, c, just_see = False, lf = False, lr = False):
        if not just_see:
            if self.where.get_checked():
                self.C.sheetdisplay.set_currently_selected(r, c)
            else:
                self.C.sheetdisplay.select_cell(row = r, column = c)
        self.C.sheetdisplay.see(row = r, column = c, keep_yscroll = False, keep_xscroll = False,
                                bottom_right_corner = False, check_cell_visibility = True)
        if lf:
            self.last_found = (r, c)
        if lr:
            self.last_replaced = (r, c)
        return True

    def find_next(self, event  = None):
        self.find_display.place_cursor()
        ids = self.ids_button.get_checked()
        dets = self.details_button.get_checked()
        if not ids and not dets:
            self.status_bar.change_text("Select a search option, IDs and Parents and/or Details")
            return
        search = self.find_display.get_my_value().lower()
        match = self.match_button.get_checked()
        allcols = self.allcols_button.get_checked()
        if allcols and not self.C.all_columns_displayed:
            self.C.displayed_columns = list(range(len(self.C.headers)))
            self.C.all_columns_displayed = True
            self.C.ss_hiding_cols_button.set_checked(not self.C.all_columns_displayed)
            self.C.sheetdisplay.display_subset_of_columns(indexes=self.C.displayed_columns,
                                                      enable=not self.C.all_columns_displayed,
                                                      reset_col_positions=False)
            self.C.set_all_col_widths()
            self.C.disable_paste()
        showing = {c: i for i, c in enumerate(self.C.displayed_columns)}
        ind = self.C.indices
        qic = self.C.ic
        where = self.where.get_checked()
        rst, cst = self.C.sheetdisplay.get_currently_selected(True, True)
        if rst is None or cst is None:
            self.C.sheetdisplay.deselect("all")
            self.see_and_set(0, 0)
            rst, cst = 0, 0
        else:
            self.see_and_set(rst, cst, just_see = True)
        found = False
        if where:
            sels = self.C.sheetdisplay.get_selected_cells(get_rows = True, get_columns = True, sort_by_row = True, sort_by_column = True)
            curridx = next(i for i, t in enumerate(sels) if t[0] == rst and t[1] == cst)
            if self.last_found == (rst, cst):
                if curridx == len(sels) - 1:
                    curridx = 0
                else:
                    curridx += 1
            sels = sels[curridx:] + sels[:curridx]
            for r, c in sels:
                c = self.C.displayed_columns[c]
                e = self.C.sheet[r][c]
                if not allcols and c not in showing:
                    continue
                if c == qic:
                    continue
                if ids and c in ind:
                    if match and e.lower() == search:
                        found = self.see_and_set(r, showing[c], lf = True)
                        break
                    elif not match and search in e.lower():
                        found = self.see_and_set(r, showing[c], lf = True)
                        break
                if dets and c not in ind:
                    if match and e.lower() == search:
                        found = self.see_and_set(r, showing[c], lf = True)
                        break
                    elif not match and search in e.lower():
                        found = self.see_and_set(r, showing[c], lf = True)
                        break
        else:
            if self.last_found == (rst, cst):
                if rst == len(self.C.sheet) - 1 and cst == len(self.C.displayed_columns) - 1:
                    rst, cst = 0, 0
                elif cst == len(self.C.displayed_columns) - 1:
                    rst += 1
                    cst = 0
                else:
                    cst += 1
            for c, e in enumerate(islice(self.C.sheet[rst], self.C.displayed_columns[cst], len(self.C.sheet[rst])), self.C.displayed_columns[cst]):
                if not allcols and c not in showing:
                    continue
                if c == qic:
                    continue
                if ids and c in ind:
                    if match and e.lower() == search:
                        found = self.see_and_set(rst, showing[c], lf = True)
                        break
                    elif not match and search in e.lower():
                        found = self.see_and_set(rst, showing[c], lf = True)
                        break
                if dets and c not in ind:
                    if match and e.lower() == search:
                        found = self.see_and_set(rst, showing[c], lf = True)
                        break
                    elif not match and search in e.lower():
                        found = self.see_and_set(rst, showing[c], lf = True)
                        break
            if not found:
                if len(self.C.sheet) - 1 == rst:
                    rns = range(len(self.C.sheet))
                else:
                    rns = tuple(range(rst + 1, len(self.C.sheet))) + tuple(range(0, rst + 1))
                for rn in rns:
                    if found:
                        break
                    for c, e in enumerate(self.C.sheet[rn]):
                        if not allcols and c not in showing:
                            continue
                        if c == qic:
                            continue
                        if ids and c in ind:
                            if match and e.lower() == search:
                                found = self.see_and_set(rn, showing[c], lf = True)
                                break
                            elif not match and search in e.lower():
                                found = self.see_and_set(rn, showing[c], lf = True)
                                break
                        if dets and c not in ind:
                            if match and e.lower() == search:
                                found = self.see_and_set(rn, showing[c], lf = True)
                                break
                            elif not match and search in e.lower():
                                found = self.see_and_set(rn, showing[c], lf = True)
                                break
        if found:
            self.status_bar.change_text(f"Found {self.find_display.get_my_value()} for {self.C.sheet[self.last_found[0]][self.C.ic]} in {self.C.headers[self.C.displayed_columns[self.last_found[1]]].name}")
        else:
            self.status_bar.change_text(f"Could not find {self.find_display.get_my_value()}")

    def replace_next(self, event = None):
        if self.C.showing_all_hierarchies:
            self.status_bar.change_text("Cannot perform action while showing all hierarchies")
            return
        ids = self.ids_button.get_checked()
        dets = self.details_button.get_checked()
        if not ids and not dets:
            self.status_bar.change_text("Select a search option, IDs and Parents and/or Details")
            return
        search = self.find_display.get_my_value().lower()
        newtext = self.rep_display.get_my_value()
        if search == newtext:
            self.status_bar.change_text("Error: Find value is the same as replace value")
            return
        match = self.match_button.get_checked()
        where = self.where.get_checked()
        allcols = self.allcols_button.get_checked()
        if allcols and not self.C.all_columns_displayed:
            self.C.displayed_columns = list(range(len(self.C.headers)))
            self.C.all_columns_displayed = True
            self.C.ss_hiding_cols_button.set_checked(not self.C.all_columns_displayed)
            self.C.sheetdisplay.display_subset_of_columns(indexes=self.C.displayed_columns,
                                                      enable=not self.C.all_columns_displayed,
                                                      reset_col_positions=False)
            self.C.set_all_col_widths()
            self.C.disable_paste()
        showing = {c: i for i, c in enumerate(self.C.displayed_columns)}
        valids = {c: self.C.detail_is_valid_for_col(c, newtext) for c in range(len(self.C.headers))}
        ind = self.C.indices
        qic = self.C.ic
        with_formulas = set(i for i, c in enumerate(self.C.headers) if c.formula)
        rst, cst = self.C.sheetdisplay.get_currently_selected(True, True)
        if rst is None or cst is None:
            self.C.sheetdisplay.deselect("all")
            self.see_and_set(0, 0)
            rst, cst = 0, 0
        else:
            self.see_and_set(rst, cst, just_see = True)
        found = False
        to_replace = None
        if where:
            sels = self.C.sheetdisplay.get_selected_cells(get_rows = True, get_columns = True, sort_by_row = True, sort_by_column = True)
            curridx = next(i for i, t in enumerate(sels) if t[0] == rst and t[1] == cst)
            if self.last_replaced == (rst, cst):
                if curridx == len(sels) - 1:
                    curridx = 0
                else:
                    curridx += 1
            sels = sels[curridx:] + sels[:curridx]
            for rn, c in sels:
                c = self.C.displayed_columns[c]
                e = self.C.sheet[rn][c]
                if c == qic or c in with_formulas:
                    continue
                if not allcols and c not in showing:
                    continue
                if ids and c in ind:
                    elow = e.lower()
                    if match and elow == search and elow != newtext:
                        found = self.see_and_set(rn, showing[c], lf = True, lr = True)
                        to_replace = rn, c
                        break
                    elif not match and search in elow and elow != newtext:
                        found = self.see_and_set(rn, showing[c], lf = True, lr = True)
                        to_replace = rn, c
                        break
                if dets and c not in ind:
                    elow = e.lower()
                    if match and elow == search and elow != newtext and valids[c]:
                        found = self.see_and_set(rn, showing[c], lf = True, lr = True)
                        to_replace = rn, c
                        break
                    elif not match and search in elow and elow != newtext and self.C.detail_is_valid_for_col(c, case_insensitive_replace(search, newtext, e)):
                        found = self.see_and_set(rn, showing[c], lf = True, lr = True)
                        to_replace = rn, c
                        break
        else:
            if self.last_replaced == (rst, cst):
                if rst == len(self.C.sheet) - 1 and cst == len(self.C.displayed_columns) - 1:
                    rst, cst = 0, 0
                elif cst == len(self.C.displayed_columns) - 1:
                    rst += 1
                    cst = 0
                else:
                    cst += 1
            for c, e in enumerate(islice(self.C.sheet[rst], self.C.displayed_columns[cst], len(self.C.sheet[rst])), self.C.displayed_columns[cst]):
                if c == qic or c in with_formulas:
                    continue
                if not allcols and c not in showing:
                    continue
                if ids and c in ind:
                    elow = e.lower()
                    if match and elow == search and elow != newtext:
                        found = self.see_and_set(rst, showing[c], lf = True, lr = True)
                        to_replace = rst, c
                        break
                    elif not match and search in elow and elow != newtext:
                        found = self.see_and_set(rst, showing[c], lf = True, lr = True)
                        to_replace = rst, c
                        break
                if dets and c not in ind:
                    elow = e.lower()
                    if match and elow == search and elow != newtext and valids[c]:
                        found = self.see_and_set(rst, showing[c], lf = True, lr = True)
                        to_replace = rst, c
                        break
                    elif not match and search in elow and elow != newtext and self.C.detail_is_valid_for_col(c, case_insensitive_replace(search, newtext, e)):
                        found = self.see_and_set(rst, showing[c], lf = True, lr = True)
                        to_replace = rst, c
                        break
            if not found:
                if len(self.C.sheet) - 1 == rst:
                    rns = range(len(self.C.sheet))
                else:
                    rns = tuple(range(rst + 1, len(self.C.sheet))) + tuple(range(0, rst + 1))
                for rn in rns:
                    if found:
                        break
                    for c, e in enumerate(self.C.sheet[rn]):
                        if c == qic or c in with_formulas:
                            continue
                        if not allcols and c not in showing:
                            continue
                        if ids and c in ind:
                            elow = e.lower()
                            if match and elow == search and elow != newtext:
                                found = self.see_and_set(rn, showing[c], lf = True, lr = True)
                                to_replace = rn, c
                                break
                            elif not match and search in elow and elow != newtext:
                                found = self.see_and_set(rn, showing[c], lf = True, lr = True)
                                to_replace = rn, c
                                break
                        if dets and c not in ind:
                            elow = e.lower()
                            if match and elow == search and elow != newtext and valids[c]:
                                found = self.see_and_set(rn, showing[c], lf = True, lr = True)
                                to_replace = rn, c
                                break
                            elif not match and search in elow and elow != newtext and self.C.detail_is_valid_for_col(c, case_insensitive_replace(search, newtext, e)):
                                found = self.see_and_set(rn, showing[c], lf = True, lr = True)
                                to_replace = rn, c
                                break
        if found:
            r = to_replace[0]
            c = to_replace[1]
            if not match:
                newtext = case_insensitive_replace(search, newtext, self.C.sheet[r][c])
            if self.C.headers[c].type_ == "ID":
                self.status_bar.change_text(f"Cannot edit ID column, named: {self.C.headers[c].name}")
                return
            elif self.C.headers[c].formula:
                self.status_bar.change_text(f"Cannot edit column with formula, named: {self.C.headers[c].name}")
                return
            if self.C.headers[c].type_ == "Parent":
                self.C.snapshot_paste_id()
                oldparent = f"{self.C.sheet[r][c]}"
                successful = self.C.cut_paste_edit_cell(self.sheet[r][self.C.ic], oldparent, c, newtext)
                if successful:
                    self.status_bar.change_text(f"Replaced {oldparent} with {newtext} for {self.C.sheet[r][self.C.ic]} in {self.C.headers[c].name}")
                    self.C.changelog.append((self.C.get_datetime_changelog(),
                                           self.C.user_name,
                                           "Cut and paste ID + children" if self.C.nodes[self.C.sheet[r][self.C.ic].lower()].cn[c] else "Cut and paste ID",
                                           self.C.sheet[r][self.C.ic],
                                           f"Old parent: {oldparent} old column #{c + 1} named: {self.C.headers[c].name}",
                                           f"New parent: {newtext} new column #{c + 1} named: {self.headers[c].name}"))
                    self.C.refresh_all_formulas_and_formatting(rows = [r])
                    self.C.prnt_tree()
                    self.C.sheetdisplay.refresh()
                    try:
                        self.C.treedisplay.selection_set(self.C.sheet[r][self.C.ic])
                        self.C.see_item(self.sheet[r][self.C.ic])
                    except:
                        pass
                    self.C.disable_paste()
                    self.C.stop_work(self.C.set_status_bar())
                    return
                else:
                    self.C.vs.pop()
                    self.C.vp -= 1
                    self.C.set_undo_label()
                    self.status_bar.change_text(f"Replaced {self.C.sheet[r][c]} with {newtext} for {self.C.sheet[r][self.C.ic]} in {self.C.headers[c].name}")
                    self.C.changelog.append((self.C.get_datetime_changelog(),
                                           self.C.user_name,
                                           f"Edit cell",
                                           f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                           f"{self.C.sheet[r][c]}",
                                           f"{newtext}"))
                    self.C.snapshot_ctrl_x_v_del_key_id_par()
                    self.C.sheet[r][c] = newtext
                    self.C.nodes = {}
                    self.C.disable_paste()
                    self.C.clear_copied_details()
                    self.C.auto_sort_nodes_bool.set(True)
                    self.C.build_tree_start(add_warnings = False)
                    self.C.fix_associate_sort_edit_cells()
                    self.C.rns = {r[self.C.ic].lower(): i for i,r in enumerate(self.C.sheet)}
                    self.C.sheetdisplay.deselect("all")
                    self.C.sheetdisplay.data_reference(newdataref = self.C.sheet, reset_col_positions = False)
                    self.C.sheetdisplay.display_subset_of_columns(indexes = self.C.displayed_columns, enable = not self.C.all_columns_displayed,
                                                                reset_col_positions = False,
                                                                set_col_positions = False)
                    self.C.reset_tagged_ids_dropdown()
                    self.C.reset_tagged_ids_sheet()
                    self.C.prnt_tree()
                    self.C.refresh_all_formulas_and_formatting()
                    self.C.sheetdisplay.refresh()
                    self.C.stop_work(self.C.set_status_bar())
            else:
                self.C.snapshot_ctrl_x_v_del_key()
                self.C.vs[-1]['cells'][(r, c)] = f"{self.C.sheet[r][c]}"
                self.status_bar.change_text(f"Replaced {self.C.sheet[r][c]} with {newtext} for {self.C.sheet[r][self.C.ic]} in {self.C.headers[c].name}")
                self.C.changelog.append((self.C.get_datetime_changelog(),
                                       self.C.user_name,
                                       "Edit cell",
                                       f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                       f"{self.C.sheet[r][c]}",
                                       f"{newtext}"))
                if self.C.headers[c].type_ == "Date Detail":
                    self.C.sheet[r][c] = self.C.convert_date(newtext, self.C.DATE_FORM)
                else:
                    self.C.sheet[r][c] = newtext
                self.C.disable_paste()
                self.C.refresh_all_formulas_and_formatting(rows = [r])
                self.C.refresh_treedisplay_item(self.C.sheet[r][self.C.ic])
                self.C.sheetdisplay.refresh()
                self.C.stop_work(self.C.set_status_bar())
        else:
            self.status_bar.change_text(f"Could not find an appropriate cell containing {self.find_display.get_my_value()} to replace with {self.rep_display.get_my_value()}")

    def replace_all(self, event = None):
        if self.C.showing_all_hierarchies:
            self.status_bar.change_text("Cannot perform action while showing all hierarchies")
            return
        ids = self.ids_button.get_checked()
        dets = self.details_button.get_checked()
        if not ids and not dets:
            self.status_bar.change_text("Select a search option, IDs and Parents and/or Details")
            return
        search = self.find_display.get_my_value().lower()
        newtext = self.rep_display.get_my_value()
        if search == newtext:
            self.status_bar.change_text("Error: Find value is the same as replace value")
            return
        match = self.match_button.get_checked()
        where = self.where.get_checked()
        allcols = self.allcols_button.get_checked()
        if allcols and not self.C.all_columns_displayed:
            self.C.displayed_columns = list(range(len(self.C.headers)))
            self.C.all_columns_displayed = True
            self.C.ss_hiding_cols_button.set_checked(not self.C.all_columns_displayed)
            self.C.sheetdisplay.display_subset_of_columns(indexes=self.C.displayed_columns,
                                                      enable=not self.C.all_columns_displayed,
                                                      reset_col_positions=False)
            self.C.set_all_col_widths()
            self.C.disable_paste()
        showing = {c: i for i, c in enumerate(self.C.displayed_columns)}
        ind = self.C.indices
        qic = self.C.ic
        with_formulas = set(i for i, c in enumerate(self.C.headers) if c.formula)
        valids = {c: self.C.detail_is_valid_for_col(c, newtext) for c in range(len(self.C.headers))}
        cells_changed = 0
        if ids:
            self.C.snapshot_ctrl_x_v_del_key_id_par()
        else:
            self.C.snapshot_ctrl_x_v_del_key()
        refresh_rows = set()
        if where:
            for r, c in self.C.sheetdisplay.get_selected_cells(get_rows = True, get_columns = True, sort_by_row = True, sort_by_column = True):
                c = self.C.displayed_columns[c]
                e = self.C.sheet[r][c]
                if c == qic or c in with_formulas:
                    continue
                if not allcols and c not in showing:
                    continue
                if ids and c in ind:
                    elow = e.lower()
                    if match and elow == search and elow != newtext:
                        self.C.changelog.append((self.C.get_datetime_changelog(increment_unsaved = False),
                                               self.C.user_name,
                                               "Edit cell |",
                                               f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                               f"{e}",
                                               newtext))
                        if self.C.headers[c].type_ == "Date Detail":
                            self.C.sheet[r][c] = self.C.convert_date(newtext, self.C.DATE_FORM)
                        else:
                            self.C.sheet[r][c] = newtext
                        cells_changed += 1
                    elif not match and search in elow and elow != newtext:
                        newtext2 = case_insensitive_replace(search, newtext, e)
                        self.C.changelog.append((self.C.get_datetime_changelog(increment_unsaved = False),
                                               self.C.user_name,
                                               "Edit cell |",
                                               f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                               f"{e}",
                                               newtext2))
                        if self.C.headers[c].type_ == "Date Detail":
                            self.C.sheet[r][c] = self.C.convert_date(newtext2, self.C.DATE_FORM)
                        else:
                            self.C.sheet[r][c] = newtext2
                        cells_changed += 1
                if dets and c not in ind:
                    elow = e.lower()
                    if match and elow == search and elow != newtext and valids[c]:
                        if not ids:
                            self.C.vs[-1]['cells'][(r, c)] = f"{e}"
                            refresh_rows.add(r)
                        self.C.changelog.append((self.C.get_datetime_changelog(increment_unsaved = False),
                                               self.C.user_name,
                                               "Edit cell |",
                                               f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                               f"{e}",
                                               newtext))
                        if self.C.headers[c].type_ == "Date Detail":
                            self.C.sheet[r][c] = self.C.convert_date(newtext, self.C.DATE_FORM)
                        else:
                            self.C.sheet[r][c] = newtext
                        cells_changed += 1
                    elif not match and search in elow and elow != newtext and self.C.detail_is_valid_for_col(c, case_insensitive_replace(search, newtext, e)):
                        if not ids:
                            self.C.vs[-1]['cells'][(r, c)] = f"{e}"
                            refresh_rows.add(r)
                        newtext2 = case_insensitive_replace(search, newtext, e)
                        self.C.changelog.append((self.C.get_datetime_changelog(increment_unsaved = False),
                                               self.C.user_name,
                                               "Edit cell |",
                                               f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                               f"{e}",
                                               newtext2))
                        if self.C.headers[c].type_ == "Date Detail":
                            self.C.sheet[r][c] = self.C.convert_date(newtext2, self.C.DATE_FORM)
                        else:
                            self.C.sheet[r][c] = newtext2
                        cells_changed += 1
        else:
            for r, row in enumerate(self.C.sheet):
                for c, e in enumerate(row):
                    if c == qic or c in with_formulas:
                        continue
                    if not allcols and c not in showing:
                        continue
                    if ids and c in ind:
                        elow = e.lower()
                        if match and elow == search and elow != newtext:
                            self.C.changelog.append((self.C.get_datetime_changelog(increment_unsaved = False),
                                                   self.C.user_name,
                                                   "Edit cell |",
                                                   f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                                   f"{e}",
                                                   newtext))
                            if self.C.headers[c].type_ == "Date Detail":
                                self.C.sheet[r][c] = self.C.convert_date(newtext, self.C.DATE_FORM)
                            else:
                                self.C.sheet[r][c] = newtext
                            cells_changed += 1
                        elif not match and search in elow and elow != newtext:
                            newtext2 = case_insensitive_replace(search, newtext, e)
                            self.C.changelog.append((self.C.get_datetime_changelog(increment_unsaved = False),
                                                   self.C.user_name,
                                                   "Edit cell |",
                                                   f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                                   f"{e}",
                                                   newtext2))
                            if self.C.headers[c].type_ == "Date Detail":
                                self.C.sheet[r][c] = self.C.convert_date(newtext2, self.C.DATE_FORM)
                            else:
                                self.C.sheet[r][c] = newtext2
                            cells_changed += 1
                    if dets and c not in ind:
                        elow = e.lower()
                        if match and elow == search and elow != newtext and valids[c]:
                            if not ids:
                                self.C.vs[-1]['cells'][(r, c)] = f"{e}"
                                refresh_rows.add(r)
                            self.C.changelog.append((self.C.get_datetime_changelog(increment_unsaved = False),
                                                   self.C.user_name,
                                                   "Edit cell |",
                                                   f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                                   f"{e}",
                                                   newtext))
                            if self.C.headers[c].type_ == "Date Detail":
                                self.C.sheet[r][c] = self.C.convert_date(newtext, self.C.DATE_FORM)
                            else:
                                self.C.sheet[r][c] = newtext
                            cells_changed += 1
                        elif not match and search in elow and elow != newtext and self.C.detail_is_valid_for_col(c, case_insensitive_replace(search, newtext, e)):
                            if not ids:
                                self.C.vs[-1]['cells'][(r, c)] = f"{e}"
                                refresh_rows.add(r)
                            newtext2 = case_insensitive_replace(search, newtext, e)
                            self.C.changelog.append((self.C.get_datetime_changelog(increment_unsaved = False),
                                                   self.C.user_name,
                                                   "Edit cell |",
                                                   f"ID: {self.C.sheet[r][self.C.ic]} column #{c + 1} named: {self.C.headers[c].name} with type: {self.C.headers[c].type_}",
                                                   f"{e}",
                                                   newtext2))
                            if self.C.headers[c].type_ == "Date Detail":
                                self.C.sheet[r][c] = self.C.convert_date(newtext2, self.C.DATE_FORM)
                            else:
                                self.C.sheet[r][c] = newtext2
                            cells_changed += 1
        self.C.disable_paste()
        if not cells_changed:
            self.C.vp -= 1
            self.C.set_undo_label()
            self.C.vs.pop()
            self.C.sheetdisplay.refresh()
            self.C.stop_work(self.C.set_status_bar())
            self.status_bar.change_text(f"Could not find an appropriate cell containing {self.find_display.get_my_value()} to replace with {self.rep_display.get_my_value()}")
            return
        if ids:
            self.C.nodes = {}
            self.C.clear_copied_details()
            self.C.auto_sort_nodes_bool.set(True)
            self.C.build_tree_start(add_warnings = False)
            self.C.fix_associate_sort_edit_cells()
            self.C.rns = {row[self.C.ic].lower(): i for i, row in enumerate(self.C.sheet)}
            self.C.sheetdisplay.data_reference(newdataref = self.C.sheet, reset_col_positions = False)
            self.C.sheetdisplay.display_subset_of_columns(indexes = self.C.displayed_columns, enable = not self.C.all_columns_displayed,
                                                        reset_col_positions = False,
                                                        set_col_positions = False)
            self.C.refresh_all_formulas_and_formatting()
            self.C.sheetdisplay.set_all_row_heights()
            self.C.reset_tagged_ids_dropdown()
            self.C.reset_tagged_ids_sheet()
            self.C.prnt_tree()
        else:
            self.C.refresh_all_formulas_and_formatting(rows = refresh_rows)
            for rn in refresh_rows:
                self.C.refresh_treedisplay_item(self.C.sheet[rn][self.C.ic])
        if cells_changed > 1:
            self.C.changelog.append((self.C.get_datetime_changelog(),
                                   self.C.user_name,
                                   f"Edit {cells_changed} cells",
                                   "",
                                   "",
                                   ""))
        else:
            self.C.changelog_singular("Edit cell")
        self.status_bar.change_text(f"Replaced {cells_changed} cells containing {self.find_display.get_my_value()} with {self.rep_display.get_my_value()}")
        self.C.sheetdisplay.refresh()
        self.C.stop_work(self.C.set_status_bar())
        
    def tv_confirm(self,event=None):
        self.option1 = self.id_or_detail.displayed.get()
        self.option2 = self.match_option.displayed.get()
        if self.option1 == "Find ID":
            self.result = "".join(self.tv_find_entry.get_my_value().strip().split()).lower()
        elif self.option1 == "Find detail":
            self.result = " ".join(self.tv_find_entry.get_my_value().strip().split()).lower()
        id_or_detail = self.option1
        text_or_match = self.option2
        qres = self.result
        found = False
        if id_or_detail == "Find ID":
            if text_or_match == "Non-exact match":
                for row in self.C.sheet:
                    if qres in row[self.C.ic].lower():
                        found = True
                        break
            elif text_or_match == "Exact match":
                for row in self.C.sheet:
                    if qres == row[self.C.ic].lower():
                        found = True
                        break
        elif id_or_detail == "Find detail":
            if text_or_match == "Non-exact match":
                for row in self.C.sheet:
                    if found == True:
                        break
                    for c in row:
                        if qres in c.lower():
                            found = True
                            break
            elif text_or_match == "Exact match":
                for row in self.sheet:
                    if found == True:
                        break
                    for c in row:
                        if qres == c.lower():
                            found = True
                            break
        if found:
            self.destroy()
            if self.C.mirror_var == 0:
                self.C.toggle_mirror()
            if not self.C.showing_all_hierarchies:
                self.C.view_all_hiers()
            if id_or_detail == "Find ID":
                if text_or_match == "Non-exact match":
                    self.C.search_for_ID(qres,False)
                elif text_or_match == "Exact match":
                    self.C.search_for_ID(qres,True)
            elif id_or_detail == "Find detail":
                if text_or_match == "Non-exact match":
                    self.C.search_for_detail(qres,False)
                elif text_or_match == "Exact match":
                    self.C.search_for_detail(qres,True)
        else:
            self.status_bar.change_text("No results found   ")
        
    def enter_ss_sel(self,event=None):
        self.tv_find_entry.set_my_value(self.ss_sel)
        
    def cancel(self,event=None):
        self.destroy()


class column_manager_popup(tk.Toplevel):
    def __init__(self,C):
        tk.Toplevel.__init__(self,C,width="1",height="1")
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Column manager - Green column is Treeview Label - Click the X button or press escape to go back")
        self.C = C
        self.config(bg = theme_bg(self.C.C.theme))
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(0,weight=1)
        self.new_frame = None
        self.window_destroyed = False
        self.total = len(self.C.headers)
        self.cut_col = None
        self.col_sel = None
        self.cond_sel = None
        self.actions = set()

        self.ss_rc_popup_menu_cols = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.ss_rc_popup_menu_cols_add_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.ss_rc_popup_menu_cols_add_menu.add_command(label="Add Detail Column",
                                               command=self.popup_add_col,**menu_kwargs)
        self.ss_rc_popup_menu_cols_add_menu.add_separator()
        self.ss_rc_popup_menu_cols_add_menu.add_command(label="Add Hierarchy",
                                               command=self.popup_add_hier_col,**menu_kwargs)
        self.ss_rc_popup_menu_cols.add_cascade(label="Add",
                                               menu=self.ss_rc_popup_menu_cols_add_menu,**menu_kwargs)
        
        self.ss_rc_popup_menu_cols.add_separator()
        self.ss_rc_popup_menu_cols_delete_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.ss_rc_popup_menu_cols_delete_menu.add_command(label="Delete Column",
                                               command=self.popup_del_col,**menu_kwargs)
        self.ss_rc_popup_menu_cols_delete_menu.add_command(label="Delete Formula",
                                               command=self.del_formula,**menu_kwargs)
        self.ss_rc_popup_menu_cols_delete_menu.add_command(label="Delete Validation",
                                               command=self.del_validation,**menu_kwargs)
        self.ss_rc_popup_menu_cols.add_cascade(label="Delete",
                                               menu=self.ss_rc_popup_menu_cols_delete_menu,**menu_kwargs)

        self.ss_rc_popup_menu_cols.add_separator()
        self.ss_rc_popup_menu_cols_edit_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.ss_rc_popup_menu_cols_edit_menu.add_command(label="Edit Formula",
                                               command=self.popup_edit_formula,**menu_kwargs)
        self.ss_rc_popup_menu_cols_edit_menu.add_command(label="Edit Validation",
                                               command=self.edit_validation,**menu_kwargs)
        self.ss_rc_popup_menu_cols_edit_menu.add_command(label="Edit Formatting",
                                               command=self.go_to_formatting_view,**menu_kwargs)
        self.ss_rc_popup_menu_cols.add_cascade(label="Edit",
                                               menu=self.ss_rc_popup_menu_cols_edit_menu,**menu_kwargs)
        
        self.ss_rc_popup_menu_cols.add_separator()
        self.ss_rc_coltypes_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.ss_rc_coltypes_menu.add_command(label="Text Detail",
                                             command=self.change_coltype_text,**menu_kwargs)
        self.ss_rc_coltypes_menu.add_command(label="Numerical Detail",
                                             command=self.change_coltype_numerical,**menu_kwargs)
        self.ss_rc_coltypes_menu.add_command(label="Date Detail",
                                             command=self.change_coltype_date,**menu_kwargs)
        self.ss_rc_popup_menu_cols.add_cascade(label="Change column type",
                                               menu=self.ss_rc_coltypes_menu,**menu_kwargs)

        self.ss_rc_popup_menu_cols.add_separator()
        self.ss_rc_popup_menu_cols.add_command(label="Cut column",
                                               command=self.popup_cut_col,**menu_kwargs)
        self.ss_rc_popup_menu_cols.add_separator()
        self.ss_rc_popup_menu_cols.add_command(label="Paste columns",
                                               command=self.popup_paste_col,
                                               state="disabled",**menu_kwargs)
        self.ss_rc_popup_menu_cols.add_separator()
        self.ss_rc_popup_menu_cols.add_command(label="Set as Treeview label",
                                               command=self.set_tv_label_col,**menu_kwargs)
        self.ss_rc_popup_menu_cols.add_separator()
        self.ss_rc_popup_menu_cols.add_command(label="Rename column",
                                               command=self.popup_rename_col,**menu_kwargs)

        self.ss_rc_popup_menu_end = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.ss_rc_popup_menu_end.add_command(label="Add Detail column",
                                              command=self.popup_add_col,**menu_kwargs)
        self.ss_rc_popup_menu_end.add_separator()
        self.ss_rc_popup_menu_end.add_command(label="Add Hierarchy",
                                              command=self.popup_add_hier_col,**menu_kwargs)
        self.ss_rc_popup_menu_end.add_separator()
        self.ss_rc_popup_menu_end.add_command(label="Paste columns",
                                              command=self.popup_paste_col,
                                              state="disabled",**menu_kwargs)
        
        self.ss_rc_popup_menu_cols_multiple = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.ss_rc_popup_menu_cols_multiple.add_command(label="Cut columns",
                                                           command=self.popup_cut_col,**menu_kwargs)
        self.ss_rc_popup_menu_cols_multiple.add_separator()
        self.ss_rc_popup_menu_cols_multiple.add_command(label="Delete columns",
                                                           command=self.popup_del_col,**menu_kwargs)

        self.cols_view = Sheet(self,
                               theme = self.C.C.theme,
                               row_drag_and_drop_perform = False,
                               header_font = ("Calibri", 13, "normal"),
                                  align="center",
                                  header_align="center",
                                  row_index_align="center",
                                  row_index_width=190,
                                  headers=["Column Name","Type","Formula","Formatting","Validation"])

        self.headers = [[h.name,h.type_,h.formula,f"{len(h.formatting)}","" if h.formula else ", ".join(h.validation)] for h in self.C.headers]

        for i in range(6):
            self.cols_view.highlight_cells(row = self.C.tv_label_col, column = i, bg = "#8cba66", fg = theme_fg(self.C.C.theme))
        self.cols_view.highlight_cells(row = self.C.tv_label_col, canvas = "row_index", bg = "#8cba66", fg = theme_fg(self.C.C.theme))
        
        self.cols_view.data_reference(newdataref=self.headers)
        self.cols_view.basic_bindings(True)
        self.cols_view.enable_bindings(("single",
                                        "drag_select",
                                        "row_drag_and_drop",
                                        "column_width_resize",
                                        "double_click_column_resize",
                                        "row_select",
                                        "arrowkeys"))
        self.cols_view.extra_bindings([("row_index_drag_drop",self.snapshot_drag_cols_col_manager)])
                                          
        self.cols_view.grid(row=0,column=0,sticky="nswe")
        self.cols_view.set_column_widths(column_widths=[340,150,200,100,350])
        
        self.cols_view.bind(get_platform_rc_binding(), self.cols_view_rc)
        self.cols_view.bind("<Double-Button-1>",self.cols_view_double_b1)
        self.cols_view.bind("<Delete>",self.popup_del_col)

        # ==================== CONDITIONAL FORMATTING FRAME ====================
        self.displayed_colors_dct = {"Yellow": "yellow",
                                     "Red, normal": "firebrick1",
                                     "Brown": "#734021",
                                     "Orange": "orange",
                                     "Green, bright": "lawn green",
                                     "Green, dark": "forest green",
                                     "Red, bright": "red",
                                     "Turquoise": "turquoise",
                                     "Purple": "DarkOrchid1",
                                     "Pink": "orchid1",
                                     "Red, soft": "salmon1",
                                     "Blue, bright": "cyan",
                                     'Scale 1 (green)': "#509f56",
                                     'Scale 2': "#64a85b",
                                     'Scale 3': "#78b160",
                                     'Scale 4': "#8cba66",
                                     'Scale 5': "#a0c36c",
                                     'Scale 6': "#b4cc71",
                                     'Scale 7': "#c8d576",
                                     'Scale 8': "#dcde7c",
                                     'Scale 9': "#f0e782",
                                     'Scale 10 (yellow)': "#ffec87",
                                     'Scale 11': "#ffe182",
                                     'Scale 12': "#ffdc7d",
                                     'Scale 13': "#ffd77b",
                                     'Scale 14': "#ffc873",
                                     'Scale 15': "#ffb469",
                                     'Scale 16': "#fea05f",
                                     'Scale 17': "#fc8c55",
                                     'Scale 18': "#fb784b",
                                     'Scale 19': "#fa6441",
                                     'Scale 20 (red)': "#f85037"}
        self.scale_colors = ("#509f56","#64a85b","#78b160","#8cba66","#a0c36c","#b4cc71","#c8d576",
                             "#dcde7c","#f0e782","#ffec87","#ffe182","#ffdc7d","#ffd77b","#ffc873",
                             "#ffb469","#fea05f","#fc8c55","#fb784b","#fa6441","#f85037")
        self.internal_colors = {v: k for k,v in self.displayed_colors_dct.items()}
        ak = lambda key: [int(c) if c.isdigit() else c.lower() for c in re.split("([0-9]+)",key)]
        self.displayed_colors = sorted(self.displayed_colors_dct.keys(),key=ak)
        
        self.formatting_view_FRAME = frame(self, theme = self.C.C.theme)
        self.formatting_view_FRAME.grid_rowconfigure(1,weight=1)
        self.formatting_view_FRAME.grid_columnconfigure(1,weight=1)

        self.formatting_view_rc_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.formatting_view_rc_menu.add_command(label="Add condition",
                                                 command=self.add_condition,**menu_kwargs)
        self.formatting_view_rc_menu.add_separator()
        self.formatting_view_rc_menu.add_command(label="Edit condition",
                                                 command=self.edit_condition,**menu_kwargs)
        self.formatting_view_rc_menu.add_separator()
        self.formatting_view_rc_menu.add_command(label="Del condition",
                                                 command=self.del_condition,**menu_kwargs)
        self.formatting_view_rc_menu.add_separator()
        self.formatting_view_rc_menu.add_command(label="Del existing/add num scale",
                                                 command=lambda: self.add_auto_conditions("num"),**menu_kwargs)
        self.formatting_view_rc_menu.add_separator()
        self.formatting_view_rc_menu.add_command(label="Del existing/add date scale",
                                                 command=lambda: self.add_auto_conditions("date"),
                                                 state="disabled",**menu_kwargs)
        self.col_name_display = readonly_entry_with_scrollbar(self.formatting_view_FRAME, theme = self.C.C.theme)
        self.col_name_display.grid(row=0,column=1,columnspan=2,sticky="nswe",pady=0,padx=0)
        self.formatting_view_GO_BACK = button(self.formatting_view_FRAME,text="⯇\nGO\nBACK\n⯇",style="EF.Std.TButton",command=self.go_to_cols_view)
        self.formatting_view_GO_BACK.grid(row=0,column=0,rowspan=4,sticky="nswe")

        self.formatting_view = Sheet(self.formatting_view_FRAME,
                                     theme = self.C.C.theme,
                                      align="center",
                                     row_drag_and_drop_perform = False,
                                      header_align="center",
                                      row_index_align="center",
                                      header_font = ("Calibri", 13, "normal"),
                                      row_index_width=190,
                                      headers=["Condition","Color"])
        self.formatting_view.basic_bindings(True)
        self.formatting_view.enable_bindings(("single",
                                              "row_drag_and_drop",
                                              "drag_select",
                                              "column_width_resize",
                                              "double_click_column_resize",
                                              "row_select",
                                              "arrowkeys"))
        self.formatting_view.extra_bindings([("row_index_drag_drop",self.formatting_view_drag)])
        self.formatting_view.set_column_widths(column_widths=[500,100])
        self.formatting_view.grid(row=1,column=1,sticky="nswe")
        self.formatting_view.bind(get_platform_rc_binding(), self.formatting_view_rc)
        self.formatting_view.bind("<Double-Button-1>",self.formatting_view_double_b1)
        self.formatting_view.bind("<Delete>",self.del_condition)
        self.cols_view.bind("<Control-Z>",self.undo)
        self.cols_view.bind("<Control-z>",self.undo)
        self.bind("<Escape>",self.USER_HAS_CLOSED_WINDOW)
        center(self,1200,720)
        self.deiconify()
        self.wait_window()

    def set_tv_label_col(self,event=None):
        self.C.tv_label_col = int(self.col_sel)
        self.repopulate()

    def USER_HAS_CLOSED_WINDOW(self,callback=None):
        self.window_destroyed = True
        try:
            self.new_frame.destroy()
        except:
            pass
        self.destroy()

    def go_to_formatting_view(self,event=None):
        self.cols_view.grid_forget()
        self.formatting_view_FRAME.grid(row=0,column=0,sticky="nswe")
        if self.C.headers[self.col_sel].type_ in ("ID", "Parent", "Text Detail"):
            self.formatting_view_rc_menu.entryconfig("Del existing/add date scale",state="disabled")
            self.formatting_view_rc_menu.entryconfig("Del existing/add num scale",state="disabled")
        elif self.C.headers[self.col_sel].type_ == "Numerical Detail":
            self.formatting_view_rc_menu.entryconfig("Del existing/add num scale",state="normal")
            self.formatting_view_rc_menu.entryconfig("Del existing/add date scale",state="disabled")
        elif self.C.headers[self.col_sel].type_ == "Date Detail":
            self.formatting_view_rc_menu.entryconfig("Del existing/add num scale",state="normal")
            self.formatting_view_rc_menu.entryconfig("Del existing/add date scale",state="normal")
        self.col_name_display.set_my_value("".join(("Column #",str(self.col_sel+1)," named: ",self.C.headers[self.col_sel].name," with type: ",
                                           self.C.headers[self.col_sel].type_,"   ")))
        self.populate_formatting_view(col=self.col_sel)

    def go_to_cols_view(self,event=None):
        self.repopulate()
        self.cols_view.grid(row=0,column=0,sticky="nswe")
        self.formatting_view_FRAME.grid_forget()

    def populate_formatting_view(self,event=None,col=0):
        self.formatting_view.deselect("all")
        self.formatting_view.dehighlight_cells(all_=True,redraw=False)
        self.formatting_view.dehighlight_cells(canvas="row_index",all_=True,redraw=False)
        self.formatting_view.data_reference(newdataref=[[cond, self.internal_colors[color]] for cond,color in self.C.headers[col].formatting])
        for i,(cond,color) in enumerate(self.C.headers[col].formatting):
            self.formatting_view.highlight_cells(row = i, column = 1, bg = color, fg = "black")
        self.formatting_view.set_column_widths(column_widths=[650,200])
        self.formatting_view.refresh()

    def enable_formatting_view_treeview(self):
        self.formatting_view_GO_BACK.config(state="normal")
        self.formatting_view.bind(get_platform_rc_binding(), self.formatting_view_rc)
        self.formatting_view.bind("<Double-Button-1>",self.formatting_view_double_b1)
        self.formatting_view.bind("<Delete>",self.del_condition)
        self.formatting_view.basic_bindings(True)
        self.formatting_view.enable_bindings(("single",
                                            "row_drag_and_drop",
                                            "column_width_resize",
                                            "double_click_column_resize",
                                            "row_select",
                                            "arrowkeys"))
        self.formatting_view.extra_bindings([("row_index_drag_drop",self.formatting_view_drag)])
        
    def disable_formatting_view_treeview(self):
        self.formatting_view_GO_BACK.config(state="disabled")
        self.formatting_view.unbind(get_platform_rc_binding())
        self.formatting_view.unbind("<Double-Button-1>")
        self.formatting_view.unbind("<Delete>")
        self.formatting_view.basic_bindings(False)
        self.formatting_view.disable_bindings(("single",
                                            "row_drag_and_drop",
                                            "column_width_resize",
                                            "double_click_column_resize",
                                            "row_select",
                                            "arrowkeys"))
        self.formatting_view.extra_bindings([("row_index_drag_drop",None)])

    def enable_cols_view_treeview(self):
        self.cols_view.bind(get_platform_rc_binding(), self.cols_view_rc)
        self.cols_view.bind("<Double-Button-1>",self.cols_view_double_b1)
        self.cols_view.bind("<Delete>",self.popup_del_col)
        self.cols_view.basic_bindings(True)
        self.cols_view.enable_bindings(("single",
                                        "drag_select",
                                        "row_drag_and_drop",
                                        "column_width_resize",
                                        "double_click_column_resize",
                                        "row_select",
                                        "arrowkeys"))
        self.cols_view.extra_bindings([("row_index_drag_drop",self.snapshot_drag_cols_col_manager)])

    def formatting_view_drag(self, selected_rows, r):
        r = int(r)
        rowsiter = list(selected_rows)
        rowsiter.sort()
        stins = rowsiter[0]
        endins = rowsiter[-1] + 1
        totalrows = len(rowsiter)
        if stins > r:
            self.C.headers[self.col_sel].formatting = (self.C.headers[self.col_sel].formatting[:r] +
                                                          self.C.headers[self.col_sel].formatting[stins:stins + totalrows] +
                                                          self.C.headers[self.col_sel].formatting[r:stins] +
                                                          self.C.headers[self.col_sel].formatting[stins + totalrows:])
        else:
            self.C.headers[self.col_sel].formatting = (self.C.headers[self.col_sel].formatting[:stins] +
                                                          self.C.headers[self.col_sel].formatting[stins + totalrows:r + 1] +
                                                          self.C.headers[self.col_sel].formatting[stins:stins + totalrows] +
                                                          self.C.headers[self.col_sel].formatting[r + 1:])
        self.populate_formatting_view(col = self.col_sel)

    def snapshot_drag_cols_col_manager(self, selected_cols, c):
        self.C.snapshot_drag_cols_col_manager(selected_cols, c)
        self.repopulate()

    def undo(self, event = None):
        if self.C.vs:
            self.C.undo(col_manager = True)
            self.repopulate()

    def repopulate(self):
        self.cols_view.deselect("all")
        self.cols_view.dehighlight_cells(all_=True,redraw=False)
        self.cols_view.dehighlight_cells(canvas="row_index",all_=True,redraw=False)
        self.headers = [[h.name,h.type_,h.formula,f"{len(h.formatting)}","" if h.formula else ", ".join(h.validation)] for h in self.C.headers]
        for i in range(6):
            self.cols_view.highlight_cells(row = self.C.tv_label_col, column = i, bg = "#8cba66", fg = theme_fg(self.C.C.theme))
        self.cols_view.highlight_cells(row = self.C.tv_label_col, canvas = "row_index", bg = "#8cba66", fg = theme_fg(self.C.C.theme))
        self.cols_view.data_reference(newdataref=self.headers)
        self.total = len(self.C.headers)
        self.cols_view.set_column_widths(column_widths=[340,150,200,100,350])
        self.cols_view.refresh()
        self.cols_view.focus_set()

    def disable_cols_view_treeview(self):
        self.cols_view.unbind(get_platform_rc_binding())
        self.cols_view.unbind("<Double-Button-1>")
        self.cols_view.unbind("<Delete>")
        self.cols_view.basic_bindings(False)
        self.cols_view.disable_bindings(("single",
                                        "drag_select",
                                        "row_drag_and_drop",
                                        "column_width_resize",
                                        "double_click_column_resize",
                                        "row_select",
                                        "arrowkeys"))
        self.cols_view.extra_bindings([("row_index_drag_drop",None)])
        
    def change_coltype_text(self,event=None):
        if self.C.headers[self.col_sel].formula:
            self.disable_cols_view_treeview()
            self.new_frame = error_frame(self,"Cannot change column type when column has a formula.", theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_cols_view_treeview()
            return
        if self.C.headers[self.col_sel].cols:
            self.disable_cols_view_treeview()
            self.new_frame = error_frame(self,
                                         "".join(("Cannot change column type when it's involved in a formula. These are the columns which have formulas which involve this column: ",
                                                       " | ".join([str(c + 1) for c in self.C.headers[self.col_sel].cols]),"   ")),
                                         theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_cols_view_treeview()
            return
        self.C.snapshot_col_type_text(int(self.col_sel))
        self.C.headers[self.col_sel].type_ = "Text Detail"
        if isinstance(self.C.check_validation_validity(self.col_sel,",".join(self.C.headers[self.col_sel].validation)),str):
            self.C.headers[self.col_sel].validation = []
        if self.C.check_formula_validity(self.col_sel,str(self.C.headers[self.col_sel].formula)).startswith("Error:"):
            self.C.headers[self.col_sel].formula = ""
        self.C.headers[self.col_sel].formatting = []
        self.repopulate()
        self.actions.add("coltype")
        #self.enable_cols_view_treeview()
        self.cols_view.select_row(str(self.col_sel))
    
    def change_coltype_numerical(self,event=None):
        self.disable_cols_view_treeview()
        if self.C.headers[self.col_sel].formula:
            self.new_frame = error_frame(self,"Cannot change column type when column has a formula.", theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_cols_view_treeview()
            return
        if self.C.headers[self.col_sel].cols:
            self.new_frame = error_frame(self,"".join(("Cannot change column type when it's involved in a formula. These are the columns which have formulas which involve this column: ",
                                                       " | ".join([str(c + 1) for c in self.C.headers[self.col_sel].cols]),"   ")), theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_cols_view_treeview()
            return
        self.new_frame = askconfirm_frame(self,"".join(("WARNING - This will delete any non-numerical cell entry in this column. Change ",self.C.headers[self.col_sel].name," column type to NUMERICAL   ")),
                                          bgcolor="yellow",fgcolor="black", theme = self.C.C.theme)
        self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if self.new_frame.boolean == False:
            self.enable_cols_view_treeview()
            return
        self.C.snapshot_col_type_num_date(int(self.col_sel),"Numerical Detail")
        self.C.headers[self.col_sel].type_ = "Numerical Detail"
        self.C.change_coltype_numerical(self.col_sel)
        validation = self.C.check_validation_validity(self.col_sel,",".join(self.C.headers[self.col_sel].validation))
        if isinstance(validation,str):
            self.C.headers[self.col_sel].validation = []
        else:
            self.C.headers[self.col_sel].validation = validation
        if self.C.check_formula_validity(self.col_sel,str(self.C.headers[self.col_sel].formula)).startswith("Error:"):
            self.C.headers[self.col_sel].formula = ""
        self.C.headers[self.col_sel].formatting = [tup for tup in self.C.headers[self.col_sel].formatting
                                                   if not self.C.check_condition_validity(self.col_sel,tup[0]).startswith("Error:")]
        self.repopulate()
        self.actions.add("coltype")
        self.enable_cols_view_treeview()
        self.cols_view.select_row(str(self.col_sel))
    
    def change_coltype_date(self,event=None):
        self.disable_cols_view_treeview()
        if self.C.headers[self.col_sel].formula:
            self.new_frame = error_frame(self,"Cannot change column type when column has a formula.", theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_cols_view_treeview()
            return
        if self.C.headers[self.col_sel].cols:
            self.new_frame = error_frame(self,"".join(("Cannot change column type when it's involved in a formula. These are the columns which have formulas which involve this column: ",
                                                       " | ".join(["C" + str(c + 1) for c in self.C.headers[self.col_sel].cols]),"   ")), theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_cols_view_treeview()
            return
        self.new_frame = askconfirm_frame(self,"".join(("WARNING - This will delete any non-date format cell entry in this column. Change ",self.C.headers[self.col_sel].name," column type to DATE   ")),
                                          bgcolor="yellow",fgcolor="black", theme = self.C.C.theme)
        self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if self.new_frame.boolean == False:
            self.enable_cols_view_treeview()
            return
        self.C.snapshot_col_type_num_date(int(self.col_sel),"Date Detail")
        self.C.headers[self.col_sel].type_ = "Date Detail"
        self.C.change_coltype_date(self.col_sel,detect_date_form=True)
        if isinstance(self.C.check_validation_validity(self.col_sel,",".join(self.C.headers[self.col_sel].validation)),str):
            self.C.headers[self.col_sel].validation = []
        if self.C.check_formula_validity(self.col_sel,str(self.C.headers[self.col_sel].formula)).startswith("Error:"):
            self.C.headers[self.col_sel].formula = ""
        self.C.headers[self.col_sel].formatting = [tup for tup in self.C.headers[self.col_sel].formatting
                                                   if not self.C.check_condition_validity(self.col_sel,tup[0]).startswith("Error:")]
        self.repopulate()
        self.actions.add("coltype")
        self.enable_cols_view_treeview()
        self.cols_view.select_row(str(self.col_sel))

    def add_col_to_formula(self,event=None):
        column = self.cols_view.identify_row(event, allow_end = False)
        if column is not None:
            column = int(column)
            if self.C.headers[self.stored_selected_col].type_ == "Text Detail":
                self.new_frame.formula_display.my_entry.insert("end",f"c{column + 1}")
                self.new_frame.formula_display.my_entry.focus_set()
            elif column != self.C.ic and column not in set(self.C.hiers):
                self.new_frame.formula_display.my_entry.insert("end",f"c{column + 1}")
                self.new_frame.formula_display.my_entry.focus_set()
        return "break"

    def del_formula(self,event=None):
        self.C.snapshot_edit_formula(int(self.col_sel),"")
        if self.C.headers[self.col_sel].type_ == "Text Detail":
            for idx in {int("".join(re.findall("([0-9]+)",form_col))) - 1 for form_col in re.split('("[^"]*"|\+)',self.C.headers[self.col_sel].formula) if form_col.startswith(("c","C"))}:
                try:
                    self.C.headers[idx].cols.discard(self.col_sel)
                except:
                    pass
        else:
            old_formula_col_indexes = {int("".join(re.findall("([0-9]+)",form_col))) - 1 for form_col in re.findall("([cC][0-9]+)",self.C.headers[self.col_sel].formula)}
            for idx in old_formula_col_indexes:
                try:
                    self.C.headers[idx].cols.discard(self.col_sel)
                except:
                    pass
        self.C.headers[self.col_sel].formula = ""
        self.repopulate()
        self.actions.add("formula")

    def del_validation(self,event=None):
        if [] != self.C.headers[self.col_sel].validation:
            self.C.snapshot_edit_validation(self.col_sel, [])
        self.C.headers[self.col_sel].validation = []
        self.repopulate()
        self.cols_view.select_row(str(self.col_sel))

    def edit_validation(self,event=None):
        if self.C.headers[self.col_sel].formula:
            return
        self.disable_cols_view_treeview()
        self.new_frame = edit_validation_frame(self,
                                               self.C.headers[self.col_sel].type_,
                                               "".join(("Column #",str(self.col_sel+1)," named:  ",self.C.headers[self.col_sel].name,"   ")),
                                               self.C.headers[self.col_sel].validation,
                                               theme = self.C.C.theme)
        self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if self.new_frame.result == False:
            self.enable_cols_view_treeview()
            return
        if self.new_frame.new_validation:
            validation = self.C.check_validation_validity(self.col_sel,self.new_frame.new_validation)
            if isinstance(validation,str):
                self.new_frame = error_frame(self,"".join((" ",validation,"     see 'Help' under the 'File' menu for instructions on validation in Tree Surgeon   ")), theme = self.C.C.theme)
                self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
                self.bind("<Return>",self.new_frame.confirm)
                self.new_frame.wait_window()
                if self.window_destroyed:
                    return
                self.unbind("<Return>")
                self.enable_cols_view_treeview()
                return
            self.new_frame = askconfirm_frame(self,
                                              "".join(("WARNING - This will overwrite any invalid cells in this column. Set validation for ",
                                                       self.C.headers[self.col_sel].name,"  ")),
                                              confirm_text="Continue",
                                              bgcolor="yellow",fgcolor="black", theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            if self.new_frame.boolean == False:
                self.enable_cols_view_treeview()
                return
        else:
            validation = []
        if validation == self.C.headers[self.col_sel].validation:
            self.enable_cols_view_treeview()
            return
        self.C.snapshot_edit_validation(self.col_sel, validation)
        self.C.headers[self.col_sel].validation = validation
        if validation:
            self.C.apply_validation_to_col(self.col_sel)
        self.repopulate()
        self.enable_cols_view_treeview()
        self.cols_view.select_row(str(self.col_sel))

    def popup_edit_formula(self,event=None):
        if self.C.headers[self.col_sel].validation:
            return
        self.disable_cols_view_treeview()
        self.stored_selected_col = int(self.col_sel)
        if self.C.headers[self.col_sel].cols:
            self.new_frame = error_frame(self,
                                         "".join(("Cannot edit column formula when it's involved in a formula. These are the columns which have formulas which involve this column: ",
                                         " | ".join(["C" + str(c + 1) for c in self.C.headers[self.col_sel].cols]),"   ")), theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_cols_view_treeview()
            return
        self.new_frame = edit_formula_frame(self,
                                            "".join(("Column #",str(self.col_sel+1)," named:  ",self.C.headers[self.col_sel].name,"   ")),
                                            self.C.headers[self.col_sel].formula,
                                            self.C.headers[self.col_sel].type_,
                                            self.C.headers[self.col_sel].formula_only_apply_if_cols,
                                            theme = self.C.C.theme)
        self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.cols_view.bind("<1>",self.add_col_to_formula)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        self.cols_view.unbind("<1>")
        if self.new_frame.result == False:
            self.enable_cols_view_treeview()
            return
        if self.new_frame.new_formula:
            formula = self.C.check_formula_validity(self.col_sel,self.new_frame.new_formula)
            if formula.startswith("Error:"):
                self.new_frame = error_frame(self,"".join((" ",formula,"     see 'Help' under the 'File' menu for instructions on column formulas in Tree Surgeon   ")), theme = self.C.C.theme)
                self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
                self.bind("<Return>",self.new_frame.confirm)
                self.new_frame.wait_window()
                if self.window_destroyed:
                    return
                self.unbind("<Return>")
                self.enable_cols_view_treeview()
                return
            only_apply_result = literal_eval(self.new_frame.formula_only_apply_result)
            self.new_frame = askconfirm_frame(self,
                                              "".join(("WARNING - This will overwrite any applicable cells in this column. Set formula for ",
                                                       self.C.headers[self.col_sel].name,"  ")),
                                              confirm_text="Continue",
                                              bgcolor="yellow",fgcolor="black", theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            if self.new_frame.boolean == False:
                self.enable_cols_view_treeview()
                return
        else:
            only_apply_result = literal_eval(self.new_frame.formula_only_apply_result)
            formula = self.new_frame.new_formula
        if formula == self.C.headers[self.col_sel].formula and self.C.headers[self.col_sel].formula_only_apply_if_cols == only_apply_result:
            self.enable_cols_view_treeview()
            return
        self.C.snapshot_edit_formula(int(self.col_sel),formula)
        if formula:
            if self.C.headers[self.col_sel].type_ == "Text Detail":
                old_formula_col_indexes = {int("".join(re.findall("([0-9]+)",form_col))) - 1 for form_col in re.split('("[^"]*"|\+)',self.C.headers[self.col_sel].formula) if form_col.startswith(("c","C"))}
            else:
                old_formula_col_indexes = {int("".join(re.findall("([0-9]+)",form_col))) - 1 for form_col in re.findall("([cC][0-9]+)",self.C.headers[self.col_sel].formula)}
            for idx in old_formula_col_indexes:
                self.C.headers[idx].cols.discard(self.col_sel)
            if self.C.headers[self.col_sel].type_ == "Text Detail":
                new_formula_col_indexes = {int("".join(re.findall("([0-9]+)",form_col))) - 1 for form_col in re.split('("[^"]*"|\+)',formula) if form_col.startswith(("c","C"))}
            else:
                new_formula_col_indexes = {int("".join(re.findall("([0-9]+)",form_col))) - 1 for form_col in re.findall("([cC][0-9]+)",formula)}
            for idx in new_formula_col_indexes:
                self.C.headers[idx].cols.add(int(self.col_sel))
        elif not formula:
            if self.C.headers[self.col_sel].type_ == "Text Detail":
                for idx in {int("".join(re.findall("([0-9]+)",form_col))) - 1 for form_col in re.split('("[^"]*"|\+)',self.C.headers[self.col_sel].formula) if form_col.startswith(("c","C"))}:
                    try:
                        self.C.headers[idx].cols.discard(self.col_sel)
                    except:
                        pass
            else:
                old_formula_col_indexes = {int("".join(re.findall("([0-9]+)",form_col))) - 1 for form_col in re.findall("([cC][0-9]+)",self.C.headers[self.col_sel].formula)}
                for idx in old_formula_col_indexes:
                    try:
                        self.C.headers[idx].cols.discard(self.col_sel)
                    except:
                        pass
        self.C.headers[self.col_sel].formula_only_apply_if_cols = only_apply_result
        self.C.headers[self.col_sel].formula = formula
        self.repopulate()
        self.actions.add("formula")
        self.enable_cols_view_treeview()
        self.cols_view.select_row(str(self.col_sel))

    def cols_view_double_b1(self,event):
        region = self.cols_view.identify_region(event)
        if region in ("header", "top left"):
            return
        column = self.cols_view.identify_column(event, allow_end = False)
        header = self.cols_view.identify_row(event, allow_end = False)    
        if column is not None and header is not None:
            self.col_sel = int(header)
            if region == "table":
                self.cols_view.select_cell(row = header, column = column)
            elif region == "index":
                self.cols_view.select_row(header)
            typ = self.C.headers[self.col_sel].type_
            self.C.treecolsel = self.col_sel
            if column == 0:
                self.popup_rename_col()
            if typ in ("Text Detail","Numerical Detail","Date Detail"):
                if column == 1:
                    if typ == "Text Detail":
                        self.ss_rc_coltypes_menu.entryconfig("Text Detail",state="disabled")
                        self.ss_rc_coltypes_menu.entryconfig("Numerical Detail",state="normal")
                        self.ss_rc_coltypes_menu.entryconfig("Date Detail",state="normal")
                    elif typ == "Numerical Detail":
                        self.ss_rc_coltypes_menu.entryconfig("Text Detail",state="normal")
                        self.ss_rc_coltypes_menu.entryconfig("Numerical Detail",state="disabled")
                        self.ss_rc_coltypes_menu.entryconfig("Date Detail",state="normal")
                    elif typ == "Date Detail":
                        self.ss_rc_coltypes_menu.entryconfig("Text Detail",state="normal")
                        self.ss_rc_coltypes_menu.entryconfig("Numerical Detail",state="normal")
                        self.ss_rc_coltypes_menu.entryconfig("Date Detail",state="disabled")
                    self.ss_rc_coltypes_menu.tk_popup(event.x_root,event.y_root)
                if column == 2:
                    self.popup_edit_formula()
                elif column == 4:
                    self.edit_validation()
            if typ in ("ID", "Parent", "Text Detail", "Numerical Detail", "Date Detail"):
                if column == 3:
                    self.go_to_formatting_view()
    
    def cols_view_rc(self,event):
        region = self.cols_view.identify_region(event)
        if region in ("header", "top left"):
            return
        column = self.cols_view.identify_column(event, allow_end = False)
        header = self.cols_view.identify_row(event, allow_end = False)
        if header is not None:
            selectedcols = self.cols_view.get_selected_rows()
            if len(selectedcols) > 1 and header in selectedcols:
                self.ss_rc_popup_menu_cols_multiple.entryconfig("Delete columns",state="normal")
                if self.C.ic in selectedcols or self.C.pc in selectedcols:
                    self.ss_rc_popup_menu_cols_multiple.entryconfig("Delete columns",state="disabled")
                self.ss_rc_popup_menu_cols_multiple.tk_popup(event.x_root,event.y_root)
            else:
                if region == "table":
                    self.cols_view.select_row(row = header)
                elif region == "index":
                    self.cols_view.select_row(header)
                self.col_sel = int(header)
                self.C.treecolsel = self.col_sel
                if self.col_sel == self.C.ic or self.col_sel == self.C.pc:
                    self.ss_rc_popup_menu_cols_delete_menu.entryconfig("Delete Column",state="disabled")
                else:
                    self.ss_rc_popup_menu_cols_delete_menu.entryconfig("Delete Column",state="normal")
                typ = self.C.headers[self.col_sel].type_
                self.ss_rc_popup_menu_cols_edit_menu.entryconfig("Edit Formatting",state="normal")
                if typ not in ("Text Detail","Numerical Detail","Date Detail"):
                    self.ss_rc_popup_menu_cols.entryconfig("Change column type",state="disabled")
                    self.ss_rc_popup_menu_cols_edit_menu.entryconfig("Edit Formula",state="disabled")
                    self.ss_rc_popup_menu_cols_edit_menu.entryconfig("Edit Validation",state="disabled")
                else:
                    self.ss_rc_popup_menu_cols.entryconfig("Change column type",state="normal")
                    if not self.C.headers[self.col_sel].validation:
                        self.ss_rc_popup_menu_cols_edit_menu.entryconfig("Edit Formula",state="normal")
                    else:
                        self.ss_rc_popup_menu_cols_edit_menu.entryconfig("Edit Formula",state="disabled")
                    if not self.C.headers[self.col_sel].formula:
                        self.ss_rc_popup_menu_cols_edit_menu.entryconfig("Edit Validation",state="normal")
                    else:
                        self.ss_rc_popup_menu_cols_edit_menu.entryconfig("Edit Validation",state="disabled")
                    if typ == "Text Detail":
                        self.ss_rc_coltypes_menu.entryconfig("Text Detail",state="disabled")
                        self.ss_rc_coltypes_menu.entryconfig("Numerical Detail",state="normal")
                        self.ss_rc_coltypes_menu.entryconfig("Date Detail",state="normal")
                    elif typ == "Numerical Detail":
                        self.ss_rc_coltypes_menu.entryconfig("Text Detail",state="normal")
                        self.ss_rc_coltypes_menu.entryconfig("Numerical Detail",state="disabled")
                        self.ss_rc_coltypes_menu.entryconfig("Date Detail",state="normal")
                    elif typ == "Date Detail":
                        self.ss_rc_coltypes_menu.entryconfig("Text Detail",state="normal")
                        self.ss_rc_coltypes_menu.entryconfig("Numerical Detail",state="normal")
                        self.ss_rc_coltypes_menu.entryconfig("Date Detail",state="disabled")
                self.ss_rc_popup_menu_cols.tk_popup(event.x_root,event.y_root)
        elif header is None:
            self.col_sel = int(self.total)
            self.C.treecolsel = self.col_sel
            self.ss_rc_popup_menu_end.tk_popup(event.x_root,event.y_root)

    def formatting_view_double_b1(self,event):
        region = self.formatting_view.identify_region(event)
        if region == "table":
            column = self.formatting_view.identify_column(event, allow_end=False)
            condition = self.formatting_view.identify_row(event, allow_end=False)
            if column is not None and condition is not None:
                self.formatting_view.select_row(condition)
                self.cond_sel = int(condition)
                self.edit_condition()
    
    def formatting_view_rc(self,event):
        region = self.formatting_view.identify_region(event)
        if region == "table" or "index":
            column = self.formatting_view.identify_column(event, allow_end=False)
            condition = self.formatting_view.identify_row(event, allow_end=False)
            if column is not None and condition is not None:
                self.formatting_view.select_row(condition)
                self.cond_sel = int(condition)
                self.formatting_view_rc_menu.entryconfig("Edit condition",state="normal")
                self.formatting_view_rc_menu.entryconfig("Del condition",state="normal")
                if len(self.C.headers[self.col_sel].formatting) < 35:
                    self.formatting_view_rc_menu.entryconfig("Add condition",state="normal")
                else:
                    self.formatting_view_rc_menu.entryconfig("Add condition",state="disabled")
                self.formatting_view_rc_menu.tk_popup(event.x_root,event.y_root)
            elif condition is None:
                self.formatting_view.deselect("all")
                self.cond_sel = int(len(self.C.headers[self.col_sel].formatting))
                self.formatting_view_rc_menu.entryconfig("Edit condition",state="disabled")
                self.formatting_view_rc_menu.entryconfig("Del condition",state="disabled")
                self.formatting_view_rc_menu.tk_popup(event.x_root,event.y_root)
        
    def edit_condition(self,event=None):
        self.disable_formatting_view_treeview()
        header = self.C.headers[self.col_sel]
        if header.formatting:
            cond_tuple = header.formatting[self.cond_sel]
        else:
            cond_tuple = ("",self.displayed_colors[0])
        self.new_frame = edit_condition_frame(self.formatting_view_FRAME,
                                              condition=cond_tuple[0],
                                              colors=self.displayed_colors,
                                              color=self.internal_colors[cond_tuple[1]],
                                              coltype=header.type_,
                                              theme = self.C.C.theme)
        self.new_frame.grid(row=3,column=1,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if self.new_frame.result == False:
            self.enable_formatting_view_treeview()
            return
        condition = self.C.check_condition_validity(self.col_sel,self.new_frame.new_condition)
        if condition.startswith("Error:"):
            self.new_frame = error_frame(self.formatting_view_FRAME,"".join((" ",condition,"   See 'Help' under the 'File' menu for instructions on conditional formatting in Tree Surgeon   ")),
                                         theme = self.C.C.theme)
            self.new_frame.grid(row=3,column=1,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_formatting_view_treeview()
            return
        color = self.displayed_colors_dct[self.new_frame.color]
        self.C.headers[self.col_sel].formatting[self.cond_sel] = (condition, color)
        self.populate_formatting_view(col = self.col_sel)
        self.enable_formatting_view_treeview()
        self.formatting_view.select_row(f"{self.cond_sel}")

    def add_auto_conditions(self,num_or_date="num"):
        self.disable_formatting_view_treeview()
        header = self.C.headers[self.col_sel]
        if num_or_date == "num":
            self.new_frame = auto_add_condition_num_frame(self.formatting_view_FRAME,self.col_sel,self.C.sheet, theme = self.C.C.theme)
        else:
            self.new_frame = auto_add_condition_date_frame(self.formatting_view_FRAME,self.col_sel,self.C.sheet,self.C.DATE_FORM, theme = self.C.C.theme)
        self.new_frame.grid(row=3,column=1,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if self.new_frame.result == False:
            self.enable_formatting_view_treeview()
            return
        if num_or_date == "num":
            ac = {"0","1","2","3","4","5","6","7","8","9","-","."}
            min_v = "".join(c for c in self.new_frame.min_val if c in ac)
            max_v = "".join(c for c in self.new_frame.max_val if c in ac)
            if not min_v and not max_v:
                self.enable_formatting_view_treeview()
                return
            try:
                if not min_v:
                    min_v = 0
                else:
                    min_v = float(min_v)
                if not max_v:
                    max_v = 0
                else:
                    max_v = float(max_v)
            except:
                self.enable_formatting_view_treeview()
                return
            if min_v >= max_v:
                self.new_frame = error_frame(self.formatting_view_FRAME,
                                             "".join(("Error: Minimum value greater than or equal to maximum value"," - see 'Help' under the 'File' menu for instructions on conditional formatting in Tree Surgeon   ")),
                                             theme = self.C.C.theme)
                self.new_frame.grid(row=3,column=1,columnspan=2,sticky="nswe")
                self.bind("<Return>",self.new_frame.confirm)
                self.new_frame.wait_window()
                if self.window_destroyed:
                    return
                self.unbind("<Return>")
                self.enable_formatting_view_treeview()
                return
            self.C.headers[self.col_sel].formatting = []
            step = (max_v - min_v) / 20
            if header.type_ == "Numerical Detail":
                if self.new_frame.order == "ASCENDING":
                    v = float(min_v)
                    for i in range(1,21):
                        if not i % 20:
                            self.C.headers[self.col_sel].formatting.append(("".join((">= ",str(v)," and <= ",str(v + step))),self.scale_colors[i-1]))
                        else:
                            self.C.headers[self.col_sel].formatting.append(("".join((">= ",str(v)," and < ",str(v + step))),self.scale_colors[i-1]))
                            v += step
                elif self.new_frame.order == "DESCENDING":
                    v = float(max_v)
                    for i in range(1,21):
                        if not i % 20:
                            self.C.headers[self.col_sel].formatting.append(("".join(("<= ",str(v)," and >= ",str(v - step))),self.scale_colors[i-1]))
                        else:
                            self.C.headers[self.col_sel].formatting.append(("".join(("<= ",str(v)," and > ",str(v - step))),self.scale_colors[i-1]))
                            v -= step
            elif header.type_ == "Date Detail":
                if self.new_frame.order == "ASCENDING":
                    v = min_v
                    for i in range(1,21):
                        if not i % 20:
                            self.C.headers[self.col_sel].formatting.append(("".join((">= ",str(round(v))," and <= ",str(round(v + step)))),self.scale_colors[i-1]))
                        else:
                            self.C.headers[self.col_sel].formatting.append(("".join((">= ",str(round(v))," and < ",str(round(v + step)))),self.scale_colors[i-1]))
                            v += step
                elif self.new_frame.order == "DESCENDING":
                    v = max_v
                    for i in range(1,21):
                        if not i % 20:
                            self.C.headers[self.col_sel].formatting.append(("".join(("<= ",str(round(v))," and >= ",str(round(v - step)))),self.scale_colors[i-1]))
                        else:
                            self.C.headers[self.col_sel].formatting.append(("".join(("<= ",str(round(v))," and > ",str(round(v - step)))),self.scale_colors[i-1]))
                            v -= step
        elif num_or_date == "date":
            ac = {"0","1","2","3","4","5","6","7","8","9","/","-"}
            min_v = "".join(c for c in self.new_frame.min_val if c in ac).replace("-","/")
            max_v = "".join(c for c in self.new_frame.max_val if c in ac).replace("-","/")
            if not min_v and not max_v:
                self.enable_formatting_view_treeview()
                return
            DATE_FORM = self.C.convert_hyphen_to_slash_date_form(self.C.DATE_FORM)
            try:
                min_v = datetime.datetime.strptime(min_v, DATE_FORM)
                max_v = datetime.datetime.strptime(max_v, DATE_FORM)
            except:
                self.enable_formatting_view_treeview()
                return
            if min_v >= max_v:
                self.new_frame = error_frame(self.formatting_view_FRAME,
                                             "".join(("Error: Minimum value greater than or equal to maximum value"," - see 'Help' under the 'File' menu for instructions on conditional formatting in Tree Surgeon   ")),
                                             theme = self.C.C.theme)
                self.new_frame.grid(row=3,column=1,columnspan=2,sticky="nswe")
                self.bind("<Return>",self.new_frame.confirm)
                self.new_frame.wait_window()
                if self.window_destroyed:
                    return
                self.unbind("<Return>")
                self.enable_formatting_view_treeview()
                return
            self.C.headers[self.col_sel].formatting = []
            step = ((max_v - min_v).days) / 20
            step = datetime.timedelta(days=step)
            if self.new_frame.order == "ASCENDING":
                v = min_v # strptime
                for i in range(1,21):
                    s1 = datetime.datetime.strftime(v, DATE_FORM)
                    s2 = datetime.datetime.strftime(v + step, DATE_FORM)
                    if not i % 20:
                        self.C.headers[self.col_sel].formatting.append(("".join((">= ",s1," and <= ",s2)),self.scale_colors[i-1]))
                    else:
                        self.C.headers[self.col_sel].formatting.append(("".join((">= ",s1," and < ",s2)),self.scale_colors[i-1]))
                        v = v + step
            elif self.new_frame.order == "DESCENDING":
                v = max_v # strptime
                for i in range(1,21):
                    s1 = datetime.datetime.strftime(v, DATE_FORM)
                    s2 = datetime.datetime.strftime(v - step, DATE_FORM)
                    if not i % 20:
                        self.C.headers[self.col_sel].formatting.append(("".join(("<= ",s1," and >= ",s2)),self.scale_colors[i-1]))
                    else:
                        self.C.headers[self.col_sel].formatting.append(("".join(("<= ",s1," and > ",s2)),self.scale_colors[i-1]))
                        v = v - step
        self.populate_formatting_view(col=self.col_sel)
        self.enable_formatting_view_treeview()

    def add_condition(self,event=None):
        self.disable_formatting_view_treeview()
        header = self.C.headers[self.col_sel]
        cond_tuple = ("", self.displayed_colors[0])
        self.new_frame = edit_condition_frame(self.formatting_view_FRAME,
                                              condition=cond_tuple[0],
                                              colors=self.displayed_colors,
                                              color=cond_tuple[1],
                                              coltype=header.type_,
                                              confirm_text="Add condition",
                                              theme = self.C.C.theme)
        self.new_frame.grid(row=3,column=1,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if self.new_frame.result == False:
            self.enable_formatting_view_treeview()
            return
        condition = self.C.check_condition_validity(self.col_sel,self.new_frame.new_condition)
        if condition.startswith("Error:"):
            self.new_frame = error_frame(self.formatting_view_FRAME,"".join((" ",condition,"   See 'Help' under the 'File' menu for instructions on conditional formatting in Tree Surgeon   ")),
                                         theme = self.C.C.theme)
            self.new_frame.grid(row=3,column=1,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            self.enable_formatting_view_treeview()
            return
        color = self.displayed_colors_dct[self.new_frame.color]
        self.C.headers[self.col_sel].formatting.insert(self.cond_sel, (condition, color))
        self.populate_formatting_view(col = self.col_sel)
        self.enable_formatting_view_treeview()
        self.formatting_view.select_row(self.cond_sel)

    def del_condition(self,event=None):
        elements = self.formatting_view.get_selected_rows(get_cells_as_rows = True, return_tuple = True)
        if not elements:
            return
        self.C.headers[self.col_sel].formatting[elements[0]:elements[-1] + 1] = []
        self.cond_sel = None
        self.populate_formatting_view(col=self.col_sel)
        
    def disable_paste(self):
        self.ss_rc_popup_menu_cols.entryconfig("Paste columns",state="disabled")
        self.ss_rc_popup_menu_end.entryconfig("Paste columns",state="disabled")
 
    def popup_rename_col(self):
        self.disable_cols_view_treeview()
        currcolname = self.C.headers[self.col_sel].name
        if self.col_sel in self.C.hiers:
            self.new_frame = rename_column_frame(self,currcolname,"hierarchy", theme = self.C.C.theme)
        elif self.col_sel == self.C.ic:
            self.new_frame = rename_column_frame(self,currcolname,"ID", theme = self.C.C.theme)
        else:
            self.new_frame = rename_column_frame(self,currcolname,"detail", theme = self.C.C.theme)
        self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if not self.new_frame.result:
            self.enable_cols_view_treeview()
            return
        x = self.new_frame.result.lower()
        for i,h in enumerate(self.C.headers):
            if x == h.name.lower() and i != self.col_sel:
                self.new_frame = error_frame(self,"".join(("Name: ",h.name," already exists")), theme = self.C.C.theme)
                self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
                self.bind("<Return>",self.new_frame.confirm)
                self.new_frame.wait_window()
                if self.window_destroyed:
                    return
                self.unbind("<Return>")
                self.enable_cols_view_treeview()
                return
        self.C.rename_col(self.new_frame.result)
        self.actions.add("rename")
        self.repopulate()
        self.enable_cols_view_treeview()
        self.cols_view.select_row(str(self.col_sel))
        
    def popup_add_col(self):
        self.disable_cols_view_treeview()
        self.new_frame = add_detail_column_frame(self)
        self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.cols_view.bind("<1>",self.add_col_to_formula)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        self.cols_view.unbind("<1>")
        if not self.new_frame.result:
            self.enable_cols_view_treeview()
            return
        x = self.new_frame.result.lower()
        for h in self.C.headers:
            if x == h.name.lower():
                self.new_frame = error_frame(self,"".join(("Name: ",h.name," already exists")), theme = self.C.C.theme)
                self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
                self.bind("<Return>",self.new_frame.confirm)
                self.new_frame.wait_window()
                if self.window_destroyed:
                    return
                self.unbind("<Return>")
                self.enable_cols_view_treeview()
                return
        self.C.add_col(self.new_frame.result,self.new_frame.type_)
        self.disable_paste()
        self.actions.add("anyother")
        self.repopulate()
        self.enable_cols_view_treeview()
        self.cols_view.select_row(str(self.col_sel))
        
    def popup_add_hier_col(self):
        self.disable_cols_view_treeview()
        self.new_frame = add_hierarchy_frame(self)
        self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.new_frame.confirm)
        self.new_frame.wait_window()
        if self.window_destroyed:
            return
        self.unbind("<Return>")
        if not self.new_frame.result:
            self.enable_cols_view_treeview()
            return
        x = self.new_frame.result.lower()
        for h in self.C.headers:
            if x == h.name.lower():
                self.new_frame = error_frame(self,"".join(("Name: ",h.name," already exists")), theme = self.C.C.theme)
                self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
                self.bind("<Return>",self.new_frame.confirm)
                self.new_frame.wait_window()
                if self.window_destroyed:
                    return
                self.unbind("<Return>")
                self.enable_cols_view_treeview()
                return
        self.C.add_hier_col(self.new_frame.result)
        self.disable_paste()
        self.actions.add("anyother")
        self.repopulate()
        self.enable_cols_view_treeview()
        self.cols_view.select_row(str(self.col_sel))
        
    def popup_paste_col(self):
        names = tuple(f"{self.C.headers[h].name}" for h in self.C.cut_columns)
        self.C.snapshot_drag_cols_col_manager(self.C.cut_columns, self.C.treecolsel)
        self.disable_paste()
        self.actions.add("anyother")
        self.repopulate()
        for name in names:
            try:
                self.cols_view.add_row_selection(r = next(i for i, h in enumerate(self.C.headers) if h.name == name), redraw = False, run_binding_func = False)
            except:
                continue
        self.cols_view.refresh()
        
    def popup_cut_col(self):
        self.C.cut_cols(self.cols_view.get_selected_rows(return_tuple = True))
        self.ss_rc_popup_menu_cols.entryconfig("Paste columns",state="normal")
        self.ss_rc_popup_menu_end.entryconfig("Paste columns",state="normal")
        
    def popup_del_col(self,event=None):
        headers = self.cols_view.get_selected_rows(return_tuple = True)
        if headers:
            self.col_sel = headers[0]
            headers = set(headers)
        else:
            return
        if self.col_sel == self.C.ic or self.col_sel == self.C.pc or self.C.ic in headers or self.C.pc in headers or not headers:
            return
        self.disable_cols_view_treeview()
        if len(headers) > 1:
            self.new_frame = askconfirm_frame(self,"Delete columns: " + ", ".join([self.C.headers[idx].name for idx in headers]),
                                              bgcolor="yellow",
                                              fgcolor="black",
                                              confirm_text="Delete", theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            if self.new_frame.boolean == False:
                self.enable_cols_view_treeview()
                return
            for i, header in enumerate(headers):
                self.col_sel = int(header) - i
                self.C.treecolsel = self.col_sel
                self.C.del_col()
        else:
            self.new_frame = askconfirm_frame(self,"Delete column: "+self.C.headers[self.col_sel].name,
                                              bgcolor="yellow",
                                              fgcolor="black",
                                              confirm_text="Delete", theme = self.C.C.theme)
            self.new_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
            self.bind("<Return>",self.new_frame.confirm)
            self.new_frame.wait_window()
            if self.window_destroyed:
                return
            self.unbind("<Return>")
            if self.new_frame.boolean == False:
                self.enable_cols_view_treeview()
                return
            self.C.treecolsel = self.col_sel
            self.C.del_col()
        self.disable_paste()
        self.cols_view.deselect("all")
        self.actions.add("delcol")
        self.repopulate()
        self.col_sel = None
        self.enable_cols_view_treeview()


class view_id_popup(tk.Toplevel):
    def __init__(self,
                 C,
                 ids_row,
                 width=800,
                 height=800,
                 theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.title(f"{self.C.sheet[ids_row['rn']][self.C.ic]} - Click the X button or press escape to go back")
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)

        self.ids_rn = ids_row['rn']
        self.changes_made = 0

        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                  row_index_width = 150,
                                  row_index_align = "w",
                                  outline_thickness=0)
        self.sheetdisplay.headers(newheaders=["Column Value"])
        self.sheetdisplay.row_index(newindex = [f"{c}    {' ' * (len(str(len(self.C.headers))) - len(str(c)))}{hdr.name}" for c, hdr in enumerate(self.C.headers, 1)])
        self.redo_display()
        for c, hdr in enumerate(self.C.headers):
            if hdr.validation and hdr.validation not in (["only uk working days"],
                                                         {"only uk working days"},
                                                         ["only england working days"],
                                                         {"only england working days"},
                                                         ["only wales working days"],
                                                         {"only wales working days"},
                                                         ["only scotland working days"],
                                                         {"only scotland working days"},
                                                         ["only ni working days"],
                                                         {"only ni working days"}):
                self.sheetdisplay.create_dropdown(r = c,
                                                  c = 0,
                                                  values = hdr.validation,
                                                  set_value = self.C.sheet[self.ids_rn][c],
                                                  destroy_on_leave = False,
                                                  destroy_on_select = False,
                                                  set_cell_on_select = True,
                                                  redraw = False,
                                                  recreate_selection_boxes = False)
        self.sheetdisplay.set_width_of_index_to_text()
        self.sheetdisplay.set_xview(0.0)
        self.sheetdisplay.set_yview(0.0)
        self.enable_bindings()
        self.sheetdisplay.grid(row=1,column=0,sticky="nswe")
        self.status_bar = StatusBar(self,
                                    text = f"ID - {self.C.sheet[self.ids_rn][self.C.ic]} concise view",
                                    theme = theme)
        self.status_bar.grid(row=2,column=0,sticky="nswe")
        self.bind("<Escape>",self.cancel)
        center(self,width,height)
        self.deiconify()
        self.wait_window()

    def redo_display(self, event = None):
        self.sheetdisplay.data_reference(newdataref = [[v] for v in self.C.sheet[self.ids_rn]],
                                         reset_col_positions=False,
                                         reset_row_positions=False,
                                         redraw=False)
        self.sheetdisplay.dehighlight_cells(all_ = True)
        for tup1, tup2 in self.C.sheetdisplay.get_highlighted_cells().items():
            if tup1[0] == self.ids_rn:
                self.sheetdisplay.highlight_cells(row = tup1[1],
                                                  column = 0,
                                                  bg = tup2[0],
                                                  fg = tup2[1])
        self.sheetdisplay.set_all_dropdown_values_to_sheet()
        self.sheetdisplay.resize_dropdowns()
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.sheetdisplay.refresh()
        self.sheetdisplay.recreate_all_selection_boxes()

    def cut(self, event = None):
        pass

    def copy(self, event = None):
        currently_selected = self.sheetdisplay.get_currently_selected()
        if currently_selected:
            s = io.StringIO()
            writer = csv_module.writer(s, dialect = csv_module.excel_tab, lineterminator = "\n")
            if isinstance(currently_selected[0], int):
                boxes, maxrows = self.sheetdisplay.get_ctrl_x_c_boxes()
                row = []
                for rn in range(maxrows):
                    for r1, c1, r2, c2 in boxes:
                        if r2 - r1 < maxrows:
                            continue
                        data_ref_rn = r1 + rn
                        for c in range(c1, c2):
                            try:
                                row.append(self.C.sheet[self.ids_rn][data_ref_rn])
                            except:
                                row.append("")
                writer.writerow(row)
            elif currently_selected[0] == "column":
                boxes, maxrows = self.sheetdisplay.get_ctrl_x_c_boxes()
                row = []
                for rn in range(maxrows):
                    for r1, c1, r2, c2 in boxes:
                        if r2 - r1 < maxrows:
                            continue
                        data_ref_rn = r1 + rn
                        for c in range(c1, c2):
                            try:
                                row.append(self.C.headers[data_ref_rn].name)
                            except:
                                row.append("")
                writer.writerow(row)
                row = []
                for rn in range(maxrows):
                    for r1, c1, r2, c2 in boxes:
                        if r2 - r1 < maxrows:
                            continue
                        data_ref_rn = r1 + rn
                        for c in range(c1, c2):
                            try:
                                row.append(self.C.sheet[self.ids_rn][data_ref_rn])
                            except:
                                row.append("")
                writer.writerow(row)
            elif currently_selected[0] == "row":
                row = []
                boxes = self.sheetdisplay.get_ctrl_x_c_boxes()
                for r1, c1, r2, c2 in boxes:
                    for rn in range(r2 - r1):
                        data_ref_rn = r1 + rn
                        for c in range(c1, c2):
                            try:
                                row.append(self.C.headers[data_ref_rn].name)
                            except:
                                row.append("")
                writer.writerow(row)
                row = []
                for r1, c1, r2, c2 in boxes:
                    for rn in range(r2 - r1):
                        data_ref_rn = r1 + rn
                        for c in range(c1, c2):
                            try:
                                row.append(self.C.sheet[self.ids_rn][data_ref_rn])
                            except:
                                row.append("")
                writer.writerow(row)
            for r1, c1, r2, c2 in boxes:
                self.sheetdisplay.show_ctrl_outline(canvas = "table", start_cell = (c1, r1), end_cell = (c2, r2))
            self.clipboard_clear()
            s = s.getvalue().rstrip()
            self.clipboard_append(s)
            self.update()

    def paste(self, event = None):
        pass

    def undo(self, event = None):
        if not self.changes_made:
            return
        self.C.undo()
        self.redo_display()
        self.changes_made -= 1

    def delete(self, event = None):
        pass

    def ss_edit_cell(self, event = None):
        if not self.sheetdisplay.anything_selected():
            return
        r, c = self.sheetdisplay.get_currently_selected(True, True)
        if c is None or r is None:
            return
        if self.C.headers[r].type_ == "ID" or self.C.headers[r].formula or self.C.showing_all_hierarchies:
            return
        if self.C.headers[r].validation and self.C.headers[r].validation not in (["only uk working days"],
                                                                                 {"only uk working days"},
                                                                                 ["only england working days"],
                                                                                 {"only england working days"},
                                                                                 ["only wales working days"],
                                                                                 {"only wales working days"},
                                                                                 ["only scotland working days"],
                                                                                 {"only scotland working days"},
                                                                                 ["only ni working days"],
                                                                                 {"only ni working days"}):
            return
        self.unbind("<Escape>")
        st = set("qazwsxedcrfvtgbyhnujmikolpQAZWSXEDCRFVTGBYHNUJMIKOLP0987654321")
        ID = self.C.sheet[self.ids_rn][self.C.ic]
        ik = ID.lower()
        currentdetail = self.C.sheet[self.ids_rn][r]
        heading = self.C.headers[r].name
        if event.char in st:
            text = event.char
        elif event.keysym == "BackSpace":
            text = ""
        else:
            text = f"{currentdetail}"
        self.sheetdisplay.select_cell(row = r, column = c)
        self.sheetdisplay.see(row = r, column = c, keep_yscroll = False, keep_xscroll = False, bottom_right_corner = False, check_cell_visibility = True)
        self.sheetdisplay.RI.set_row_height(r, only_set_if_too_small = True)
        self.sheetdisplay.CH.set_col_width(0, only_set_if_too_small = True)
        self.sheetdisplay.refresh()
        self.sheetdisplay.create_text_editor(row = r, column = 0, text = text, set_data_ref_on_destroy = False)
        self.sheetdisplay.bind_text_editor_set(self.ss_edit_cell_destroy_entry, row = r, column = 0)

    def ss_edit_cell_destroy_entry(self, event = None):
        r = event[0]
        c = event[1]
        curr_ = self.sheetdisplay.get_currently_selected(True, True)
        newtext = self.sheetdisplay.get_text_editor_value(r = r,
                                                          c = c,
                                                          set_data_ref_on_destroy = False,
                                                          move_down = False if event[2] == "Escape" else c == curr_[1] and r == curr_[0],
                                                          redraw = False,
                                                          recreate = True)
        if event[2] == "Escape":
            self.bind("<Escape>", self.cancel)
            return
        y1 = int(self.ids_rn)
        x1 = int(r)
        sheetdisplay_col = int(x1)
        ID = self.C.sheet[self.ids_rn][self.C.ic]
        ik = ID.lower()
        currentdetail = self.C.sheet[self.ids_rn][r]
        heading = self.C.headers[r].name
        if newtext == currentdetail:
            self.bind("<Escape>",self.cancel)
            return
        if self.C.headers[x1].type_ == "ID":
            self.bind("<Escape>",self.cancel)
            return
        successful = False
        if self.C.headers[x1].type_ == "Parent":
            self.C.snapshot_paste_id()
            oldparent = f"{self.C.sheet[y1][x1]}"
            if self.C.cut_paste_edit_cell(self.C.sheet[y1][self.C.ic], oldparent, x1, newtext):
                successful = True
            if not successful:
                self.C.vs.pop()
                self.C.vp -= 1
                self.C.set_undo_label()
            else:
                self.C.changelog.append((self.C.get_datetime_changelog(),
                                       self.C.user_name,
                                       "Cut and paste ID + children" if self.C.nodes[ik].cn[x1] else "Cut and paste ID",
                                       self.C.sheet[y1][self.C.ic],
                                       f"Old parent: {oldparent if oldparent else 'n/a - Top ID'} old column #{x1 + 1} named: {self.C.headers[x1].name}",
                                       f"New parent: {newtext if newtext else 'n/a - Top ID'} new column #{x1 + 1} named: {self.C.headers[x1].name}"))
                self.C.prnt_tree()
                self.C.refresh_all_formulas_and_formatting(rows = [y1])
                self.C.sheetdisplay.refresh()
                try:
                    self.C.treedisplay.selection_set(self.sheet[y1][self.ic])
                    self.C.see_item(self.sheet[y1][self.ic])
                except:
                    pass
                self.C.disable_paste()
                self.C.C.status_bar.change_text(self.C.set_status_bar())
                self.redo_display()
                self.changes_made += 1
                self.bind("<Escape>",self.cancel)
                return
        if not successful and self.C.headers[x1].type_ not in ("Text Detail", "Numerical Detail", "Date Detail"):
            self.C.changelog.append((self.C.get_datetime_changelog(),
                                   self.C.user_name,
                                   f"Edit cell",
                                   f"ID: {ID} column #{x1 + 1} named: {self.C.headers[x1].name} with type: {self.C.headers[x1].type_}",
                                   f"{self.C.sheet[y1][x1]}",
                                   f"{newtext}"))
            self.C.snapshot_ctrl_x_v_del_key_id_par()
            self.C.sheet[y1][x1] = newtext
            self.C.nodes = {}
            self.C.disable_paste()
            self.C.clear_copied_details()
            self.C.auto_sort_nodes_bool.set(True)
            self.C.build_tree_start(add_warnings = False)
            self.C.fix_associate_sort_edit_cells()
            self.C.rns = {r[self.C.ic].lower(): i for i,r in enumerate(self.C.sheet)}
            self.C.sheetdisplay.deselect("all")
            self.C.sheetdisplay.data_reference(newdataref = self.C.sheet, reset_col_positions = False)
            self.C.sheetdisplay.display_subset_of_columns(indexes = self.C.displayed_columns, enable = not self.C.all_columns_displayed,
                                                        reset_col_positions = False, set_col_positions = False)
            self.C.reset_tagged_ids_dropdown()
            self.C.reset_tagged_ids_sheet()
            self.C.prnt_tree()
            self.C.refresh_all_formulas_and_formatting()
            self.C.sheetdisplay.refresh()
            self.C.C.status_bar.change_text(self.C.set_status_bar())
            self.redo_display()
            self.changes_made += 1
            self.bind("<Escape>",self.cancel)
        else:
            if not self.C.detail_is_valid_for_col(x1, newtext):
                self.bind("<Escape>",self.cancel)
                return
            if self.C.headers[x1].type_ == "Date Detail":
                newtext = self.C.convert_date(newtext, self.C.DATE_FORM)
            currentdetail = self.C.sheet[y1][x1]
            self.C.changelog.append((self.C.get_datetime_changelog(),
                                   self.C.user_name,
                                   "Edit cell",
                                   f"ID: {ID} column #{x1 + 1} named: {self.C.headers[x1].name} with type: {self.C.headers[x1].type_}",
                                   f"{self.C.sheet[y1][x1]}",
                                   f"{newtext}"))
            self.C.snapshot_ctrl_x_v_del_key()
            self.C.vs[-1]['cells'][(y1, x1)] = f"{self.C.sheet[y1][x1]}"
            self.C.sheet[y1][x1] = f"{newtext}"
            self.C.sheetdisplay.RI.set_row_height(y1)
            self.C.sheetdisplay.CH.set_col_width(sheetdisplay_col, only_set_if_too_small = True)
            self.C.refresh_all_formulas_and_formatting(rows = [y1])
            self.C.refresh_treedisplay_item(ID)
            self.C.disable_paste()
            self.C.sheetdisplay.refresh()
            self.C.C.status_bar.change_text(self.C.set_status_bar())
            self.redo_display()
            self.changes_made += 1
            self.bind("<Escape>",self.cancel)

    def cell_was_edited(self, event = None):
        currentdetail = self.C.sheet[self.ids_rn][event[0]]
        if currentdetail == self.sheetdisplay.get_cell_data(event[0], event[1]):
            return
        newtext = self.sheetdisplay.get_cell_data(event[0], event[1])
        ID = self.C.sheet[self.ids_rn][self.C.ic]
        ik = ID.lower()
        y1 = self.ids_rn
        x1 = event[0]
        sheetdisplay_col = x1
        self.C.changelog.append((self.C.get_datetime_changelog(),
                               self.C.user_name,
                               "Edit cell",
                               f"ID: {ID} column #{x1 + 1} named: {self.C.headers[x1].name} with type: {self.C.headers[x1].type_}",
                               f"{self.C.sheet[y1][x1]}",
                               f"{newtext}"))
        self.C.snapshot_ctrl_x_v_del_key()
        self.C.vs[-1]['cells'][(y1, x1)] = f"{self.C.sheet[y1][x1]}"
        self.C.sheet[y1][x1] = f"{newtext}"
        self.C.sheetdisplay.RI.set_row_height(y1)
        self.C.sheetdisplay.CH.set_col_width(sheetdisplay_col, only_set_if_too_small = True)
        self.C.refresh_all_formulas_and_formatting(rows = [y1])
        self.C.refresh_treedisplay_item(ID)
        self.C.disable_paste()
        self.C.sheetdisplay.refresh()
        self.C.C.status_bar.change_text(self.C.set_status_bar())
        self.redo_display()
        self.changes_made += 1

    def enable_bindings(self, event = None):
        self.sheetdisplay.basic_bindings(True)
        self.sheetdisplay.enable_bindings(("single",
                                           "copy",
                                           "drag_select",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                           "column_select",
                                           "row_select",
                                           "arrowkeys"))
        self.sheetdisplay.extra_bindings([("edit_cell", self.cell_was_edited)])
        #self.sheetdisplay.bind("<Control-x>", self.cut)
        #self.sheetdisplay.bind("<Control-X>", self.cut)
        self.sheetdisplay.bind("<Control-c>", self.copy)
        self.sheetdisplay.bind("<Control-C>", self.copy)
        #self.sheetdisplay.bind("<Control-v>", self.paste)
        #self.sheetdisplay.bind("<Control-V>", self.paste)
        self.sheetdisplay.bind("<Control-z>", self.undo)
        self.sheetdisplay.bind("<Control-Z>", self.undo)
        #self.sheetdisplay.bind("<Delete>", self.delete)
        for i in range(10):
            self.sheetdisplay.bind(f"{i}", self.ss_edit_cell)
        for c in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ":
            self.sheetdisplay.bind(f"<{c}>", self.ss_edit_cell)
        self.sheetdisplay.bind("<F2>", self.ss_edit_cell)
        self.sheetdisplay.bind("<Return>", self.ss_edit_cell)
        self.sheetdisplay.bind("<BackSpace>", self.ss_edit_cell)
        self.sheetdisplay.bind("<Double-Button-1>", self.ss_edit_cell)

    def disable_bindings(self, event = None):
        self.sheetdisplay.basic_bindings(False)
        self.sheetdisplay.disable_bindings(("disable_all", ))
        self.sheetdisplay.extra_bindings([("edit_cell", None)])
        #self.sheetdisplay.unbind("<Control-x>")
        #self.sheetdisplay.unbind("<Control-X>")
        self.sheetdisplay.unbind("<Control-c>")
        self.sheetdisplay.unbind("<Control-C>")
        #self.sheetdisplay.unbind("<Control-v>")
       # self.sheetdisplay.unbind("<Control-V>")
        self.sheetdisplay.unbind("<Control-z>")
        self.sheetdisplay.unbind("<Control-Z>")
        #self.sheetdisplay.unbind("<Delete>")
        for i in range(10):
            self.sheetdisplay.unbind(f"{i}")
        for c in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ":
            self.sheetdisplay.unbind(f"<{c}>")
        self.sheetdisplay.unbind("<F2>")
        self.sheetdisplay.unbind("<Return>")
        self.sheetdisplay.unbind("<BackSpace>")
        self.sheetdisplay.unbind("<Double-Button-1>")

    def USER_HAS_CLOSED_WINDOW(self, callback = None):
        self.USER_HAS_QUIT = True
        self.destroy()
        
    def cancel(self,event=None):
        self.USER_HAS_CLOSED_WINDOW()


class merge_sheets_popup(tk.Toplevel):
    def __init__(self, C, theme = "dark", add_rows = False):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Merge sheets - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.grid_rowconfigure(0, weight = 1)

        self.l_frame = frame(self, theme = theme)
        self.l_frame.grid(row = 0, column = 0, sticky = "nswe")
        self.r_frame = frame(self, theme = theme)
        self.r_frame.grid(row = 0, column = 1, sticky = "nswe")
        self.l_frame.grid_rowconfigure(3, weight = 1)
        self.r_frame.grid_rowconfigure(0, weight = 1)
        self.l_frame.grid_columnconfigure(0, weight = 1)
        self.r_frame.grid_columnconfigure(1, weight = 1)

        self.open_file_display = readonly_entry_with_scrollbar(self.l_frame,font=EF, theme = theme)
        self.open_file_display.grid(row=0,column=0,padx=2,pady=2,sticky="nswe")
        self.open_file_button = button(self.l_frame,
                                        text="⯇ Open file",
                                        style="EF.Std.TButton",
                                        command=self.open_file)
        self.open_file_button.grid(row=0,column=1,padx=2,pady=2,sticky="nswe")
        self.sheet_dropdown = ez_dropdown(self.l_frame,font=EF)
        self.sheet_dropdown.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
        self.sheet_dropdown.grid(row=1,column=0,padx=2,pady=2,sticky="nswe")
        self.select_sheet_button = button(self.l_frame,
                                          text="⯇ Load sheet",
                                          style="EF.Std.TButton",
                                          state="disabled",
                                          command=self.select_sheet)
        self.select_sheet_button.grid(row=1,column=1,padx=2,pady=2,sticky="nswe")

        self.selector = id_and_parent_column_selector(self.l_frame, theme = theme)
        self.selector.grid(row=2,column=0,rowspan=2,sticky="nswe")
        
        self.clipboard_button = button(self.l_frame,
                                          text=" Get data from clipboard ",
                                          style="EF.Std.TButton",
                                          state="normal",
                                          command=self.get_clipboard_data)
        self.clipboard_button.grid(row=2,column=1,padx=2,pady=(2,20),sticky="nswe")

        self.options_frame = frame(self.l_frame, theme = theme)
        self.options_frame.grid(row=3,column=1,sticky="nswe")
        
        self.add_new_ids_button = x_checkbutton(self.options_frame,
                                                text="Add any new IDs     ",
                                                style="x_button.Std.TButton",
                                                compound="right",
                                                checked=True)
        self.add_new_ids_button.grid(row=0,column=0,padx=10,pady=5,sticky="we")
        self.add_new_dcols_button = x_checkbutton(self.options_frame,
                                                  text="Add any new detail columns ",
                                                  style="x_button.Std.TButton",
                                                  compound="right")
        self.add_new_dcols_button.grid(row=1,column=0,padx=10,pady=5,sticky="we")
        self.add_new_pcols_button = x_checkbutton(self.options_frame,
                                                  text="Add any new parent columns ",
                                                  style="x_button.Std.TButton",
                                                  compound="right")
        self.add_new_pcols_button.grid(row=2,column=0,padx=10,pady=5,sticky="we")
        self.overwrite_details_button = x_checkbutton(self.options_frame,
                                                      text="Overwrite details for same IDs ",
                                                      style="x_button.Std.TButton",
                                                      compound="right")
        self.overwrite_details_button.grid(row=3,column=0,padx=10,pady=5,sticky="we")
        self.overwrite_parents_button = x_checkbutton(self.options_frame,
                                                      text="Overwrite parents for same IDs ",
                                                      style="x_button.Std.TButton",
                                                      compound="right")
        self.overwrite_parents_button.grid(row=4,column=0,padx=10,pady=5,sticky="we")
        
        self.button_frame = frame(self.l_frame, theme = theme)
        self.button_frame.grid(row=4,column=0,columnspan = 2, sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="b")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="b")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text="Confirm merge",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,35),pady=(20,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(35,20),pady=(20,20))
        self.status = StatusBar(self.l_frame,text="Open a file to import data", theme = theme)
        self.status.grid(row=5,column=0, columnspan = 2, sticky="ew")
        self.result = False
        self.add_new_ids = True
        self.add_new_dcols = False
        self.add_new_pcols = False
        self.overwrite_details = False
        self.overwrite_parents = False
        self.file_opened = ""
        self.sheet_opened = "n/a"
        self.row_len = 0
        self.ic = None
        self.pcols = []
        self.wb_ = None
        self.C.new_sheet = []
        self.rowsel = 0
        self.colsel = 0
        self.region = "header"

        self.showing_left = True
        self.toggle_left_button = button(self.r_frame,
                                        text="⯇",
                                        style="BF.Std.TButton",
                                        command=self.toggle_left_panel)
        self.toggle_left_button.grid(row=0,column=0, sticky = "ns")
        self.toggle_left_button.config(width = 10)
        
        self.sheetdisplay = Sheet(self.r_frame,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                  outline_thickness=0,
                                  row_drag_and_drop_perform = False,
                                  column_drag_and_drop_perform = False)
        self.sheetdisplay.enable_bindings(("enable_all"))
        self.sheetdisplay.extra_bindings([("row_index_drag_drop", self.drag_row),
                                          ("column_header_drag_drop", self.drag_col),
                                          ("rc_insert_column", self.reset_selectors),
                                          ("rc_insert_row", self.reset_selectors),
                                          ("rc_delete_column", self.reset_selectors),
                                          ("rc_delete_row", self.reset_selectors),
                                          ("ctrl_x", self.ctrl_x_in_sheet),
                                          ("delete_key", self.del_in_sheet),
                                          ("ctrl_v", self.ctrl_v_in_sheet),
                                          ("ctrl_z", self.ctrl_z_in_sheet),
                                          ("begin_edit_cell_use_keypress", self.begin_edit_cell),
                                          ("escape_edit_cell", self.escape_edit_cell),
                                          ("edit_cell", self.edit_cell_in_sheet)
                                          ])
        self.sheetdisplay.headers(newheaders = 0)
        self.C.new_sheet = [[h.name for h in self.C.headers]] + [list(repeat("", len(self.C.headers))) for r in range(2000)]
        self.C.new_sheet = self.sheetdisplay.set_sheet_data(self.C.new_sheet, verify = False)
        self.selector.set_columns([h for h in self.C.new_sheet[0]] if self.C.new_sheet else [])
        self.selector.detect_id_col()
        self.selector.detect_par_cols()
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.sheetdisplay.grid(row=0,column=1,sticky="nswe")
        if add_rows:
            self.toggle_left_panel()
            self.toggle_left_button.config(text = "⯈\nShow\nOptions\nand\nConfirm\n⯈")
        self.bind("<Escape>",self.cancel)
        center(self,1280,620)
        self.deiconify()
        self.wait_window()

    def begin_edit_cell(self, event = None):
        self.unbind("<Escape>")

    def escape_edit_cell(self, event = None):
        self.bind("<Escape>",self.cancel)

    def toggle_left_panel(self, event = None):
        if self.showing_left:
            self.grid_columnconfigure(0, weight = 0, uniform = "y")
            self.l_frame.grid_forget()
            self.showing_left = False
            self.toggle_left_button.config(text = "⯈")

        else:
            self.grid_columnconfigure(0, weight = 1, uniform = "x")
            self.l_frame.grid(row = 0, column = 0, sticky = "nswe")
            self.showing_left = True
            self.toggle_left_button.config(text = "⯇")

    def drag_col(self, selected_cols, c):
        c = int(c)
        colsiter = list(selected_cols)
        colsiter.sort()
        stins = colsiter[0]
        endins = colsiter[-1] + 1
        totalcols = len(colsiter)
        if stins > c:
            for rn in range(len(self.C.new_sheet)):
                self.C.new_sheet[rn] = (self.C.new_sheet[rn][:c] +
                                              self.C.new_sheet[rn][stins:stins + totalcols] +
                                              self.C.new_sheet[rn][c:stins] +
                                              self.C.new_sheet[rn][stins + totalcols:])
        else:
            for rn in range(len(self.C.new_sheet)):
                self.C.new_sheet[rn] = (self.C.new_sheet[rn][:stins] +
                                              self.C.new_sheet[rn][stins + totalcols:c + 1] +
                                              self.C.new_sheet[rn][stins:stins + totalcols] +
                                              self.C.new_sheet[rn][c + 1:])
        self.sheetdisplay.MT.data_ref = self.C.new_sheet
        self.selector.set_columns([h for h in self.C.new_sheet[0]])
        self.selector.detect_id_col()
        self.selector.detect_par_cols()

    def drag_row(self, selected_rows, r):
        r = int(r)
        rowsiter = list(selected_rows)
        rowsiter.sort()
        stins = rowsiter[0]
        endins = rowsiter[-1] + 1
        totalrows = len(rowsiter)
        if stins > r:
            self.C.new_sheet = (self.C.new_sheet[:r] +
                                      self.C.new_sheet[stins:stins + totalrows] +
                                      self.C.new_sheet[r:stins] +
                                      self.C.new_sheet[stins + totalrows:])
        else:
            self.C.new_sheet = (self.C.new_sheet[:stins] +
                                      self.C.new_sheet[stins + totalrows:r + 1] +
                                      self.C.new_sheet[stins:stins + totalrows] +
                                      self.C.new_sheet[r + 1:])
        self.sheetdisplay.MT.data_ref = self.C.new_sheet
        if endins == 0 or r == 0 or stins == 0:
            self.selector.set_columns([h for h in self.C.new_sheet[0]])
            self.selector.detect_id_col()
            self.selector.detect_par_cols()

    def del_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_x_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_v_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_z_in_sheet(self, event = None):
        self.reset_selectors()

    def edit_cell_in_sheet(self, event = None):
        self.bind("<Escape>",self.cancel)
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        if event[1] == idcol or event[1] in parcols or event[0] == 0:
            self.reset_selectors()

    def get_clipboard_data(self,event=None):
        self.start_work("Loading...")
        self.reset()
        try:
            temp_data = self.C.clipboard_get()
        except:
            self.stop_work("Error: Error getting clipboard data")
            return
        try:
            if temp_data.startswith("{") and temp_data.endswith("}"):
                self.C.new_sheet = self.C.json_to_sheet(json.loads(temp_data))
            else:
                delimiter_,quotechar_ = csv_delimiter_quotechar(temp_data)
                if delimiter_ is None:
                    self.stop_work("Error: Clipboard contained no appropriate data")
                    return
                for rn,r in enumerate(csv_module.reader(io.StringIO(temp_data),delimiter=delimiter_,quotechar=quotechar_,skipinitialspace=True)):
                    try:
                        self.C.new_sheet.append(r[:len(r) - next(i for i,c in enumerate(reversed(r)) if c)])
                    except:
                        continue
                    if not rn % 500:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status.change_text("Loading...  rows: " + str(rn))
        except:
            self.stop_work("Error: Error parsing clipboard data")
            return
        if not self.C.new_sheet:
            self.stop_work("Error: Clipboard contained no appropriate data")
            return
        rl = len(max(self.C.new_sheet, key = len))
        self.C.new_sheet[:] = [r + list(repeat("",rl - len(r))) for r in self.C.new_sheet]
        self.ic = None
        self.pcols = []
        for i,c in enumerate(self.C.new_sheet[0]):
            cell = c.lower()
            if cell == "id" or cell.startswith("id"):
                self.ic = i
            elif cell.startswith("parent"):
                self.pcols.append(i)
        if self.ic is None or not self.pcols:
            self.load_display([h for h in self.C.new_sheet[0]])
            self.stop_work("Select ID column and Parent columns")
        else:
            self.selector.set_columns([h for h in self.C.new_sheet[0]])
            self.selector.set_id_col(self.ic)
            self.selector.set_par_cols(self.pcols)
            self.stop_work("Ready to merge sheets")
        self.sheetdisplay.deselect("all")
        self.sheetdisplay.data_reference(newdataref=self.C.new_sheet,reset_col_positions=False,reset_row_positions=False,redraw=False)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.sheetdisplay.refresh()
        self.file_opened = "n/a - Data obtained from clipboard"
        self.sheet_opened = "n/a"

    def reset_selectors(self, event = None):
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        self.selector.set_columns([h for h in self.C.new_sheet[0]] if self.C.new_sheet else [])
        if idcol is not None and self.C.new_sheet:
            self.selector.set_id_col(idcol)
        if parcols and self.C.new_sheet:
            self.selector.set_par_cols(parcols)

    def return_wb_file(self,filepath):
        with open(filepath,"rb") as fh:
            in_mem = io.BytesIO(fh.read())
        return in_mem

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except:
            pass
        try:
            self.wb_ = None
        except:
            pass
            
    def USER_HAS_CLOSED_WINDOW(self,callback=None):
        self.C.new_sheet = []
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except:
            pass
        self.destroy()

    def open_file(self):
        self.start_work("Loading...   ")
        self.reset()
        filepath = filedialog.askopenfilename(parent=self,title="Select file")
        if not filepath:
            self.stop_work("Open a file to import data")
            return
        try:
            filepath = os.path.normpath(filepath)
        except:
            self.stop_work("Error: filepath invalid")
            return
        if not filepath.lower().endswith((".json",".xlsx",".xls",".xlsm",".csv",".tsv")):
            self.stop_work("Error: please select json/excel/csv   ")
            return
        check = os.path.isfile(filepath)
        if check == False:
            self.stop_work("Error: filepath invalid")
            return
        try:
            if filepath.lower().endswith((".csv",".tsv")):
                with open(filepath,"r") as fh:
                    temp_data = fh.read()
                delimiter_,quotechar_ = csv_delimiter_quotechar(temp_data)
                if delimiter_ is None:
                    self.stop_work("Error: File contained no appropriate data")
                    return
                for rn,r in enumerate(csv_module.reader(io.StringIO(temp_data),delimiter=delimiter_,quotechar=quotechar_,skipinitialspace=True)):
                    try:
                        self.C.new_sheet.append(r[:len(r) - next(i for i,c in enumerate(reversed(r)) if c)])
                    except:
                        continue
                    if not rn % 500:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status.change_text("Loading...  rows: " + str(rn))
                rl = len(max(self.C.new_sheet, key = len))
                self.C.new_sheet[:] = [r + list(repeat("",rl - len(r))) for r in self.C.new_sheet]
                self.load_display([h for h in self.C.new_sheet[0]])
                self.stop_work("Ready to merge sheets")
            elif filepath.lower().endswith(".json"):
                j = self.C.get_json_from_file(filepath)
                json_format = self.C.get_json_format(j)
                if not json_format:
                    self.C.new_sheet = []
                    self.stop_work("Error opening file, could not find data of correct format")
                    return
                self.C.new_sheet = self.C.json_to_sheet(j,format_=json_format[0],key=json_format[1],get_format=False,return_rowlen=False)
                if not self.C.new_sheet:
                    self.stop_work("Error: File contained no data")
                    self.select_sheet_button.config(state="disabled")
                    return
                rl = len(max(self.C.new_sheet, key = len))
                self.C.new_sheet[:] = [r + list(repeat("",rl - len(r))) for r in self.C.new_sheet]
                self.load_display([h for h in self.C.new_sheet[0]])
                self.stop_work("Ready to merge sheets")
            elif filepath.lower().endswith((".xlsx",".xls",".xlsm")):
                in_mem = self.return_wb_file(filepath)
                self.wb_ = load_workbook(in_mem,read_only=True,data_only=True)
                wbsheets = self.wb_.sheetnames
                if not wbsheets:
                    self.stop_work("Error: File/sheet contained no data")
                    return
                sheetnames = set(self.wb_.sheetnames)
                if "Treesurgeon Data" in sheetnames:
                    ws = self.wb_["Treesurgeon Data"]
                    ws.reset_dimensions()
                    try:
                        d = self.C.C.decompress_str_return_obj("".join("" if r[0].value is None else f"{r[0].value}" for r in islice(ws.rows, 1, None)),
                                                                               basetype = "32",
                                                                               dec = True)
                        self.C.new_sheet = [[h['name'] for h in d['headers']]] + d['records']
                        self.wb_.close()
                        self.select_sheet_button.config(state="disabled")
                        self.load_display([h for h in self.C.new_sheet[0]])
                        self.stop_work("Ready to merge sheets")
                    except:
                        self.C.new_sheet = []
                        self.wb_.close()
                        self.wb_ = load_workbook(in_mem,read_only=True,data_only=True)
                        self.stop_work("Error: Error opening program data")
                        self.sheet_dropdown['values'] = wbsheets
                        self.sheet_dropdown.set_my_value(wbsheets[0])
                        self.stop_work("Error: Error opening program data. Select a sheet to open")
                        self.select_sheet_button.config(state="normal")
                else:
                    self.sheet_dropdown['values'] = wbsheets
                    self.sheet_dropdown.set_my_value(wbsheets[0])
                    self.stop_work("Select a sheet to open")
                    self.select_sheet_button.config(state="normal")
        except Exception as error_msg:
            self.try_to_close_wb()
            self.C.new_sheet = []
            self.stop_work("Error: " + str(error_msg))
            return
        if not self.C.new_sheet and not filepath.lower().endswith((".xlsx",".xls",".xlsm")):
            self.C.new_sheet = []
            self.stop_work("Error: File/sheet contained no data")
            return
        self.sheetdisplay.data_reference(newdataref=self.C.new_sheet,reset_col_positions=True,reset_row_positions=True,redraw=False)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.open_file_display.set_my_value(filepath)
        self.file_opened = os.path.basename(self.open_file_display.get_my_value())

    def select_sheet(self):
        self.start_work("Loading...   ")
        self.sheet_opened = self.sheet_dropdown.get_my_value()
        ws = self.wb_[self.sheet_opened]
        ws.reset_dimensions()
        dapp = self.C.new_sheet.append
        for rn,r in enumerate(ws.rows):
            try:
                dapp(["" if x.value is None else f"{x.value}" for x in islice(r,0,len(r) - next(i for i,c in enumerate(reversed(r)) if c.value is not None))])
            except:
                continue
            if not rn % 500:
                self.update()
                if self.USER_HAS_QUIT:
                    return
                self.status.change_text("Loading...  rows: " + str(rn))
        self.try_to_close_wb()
        self.stop_work("Ready to merge sheets")
        if not self.C.new_sheet:
            self.stop_work("Error: File/sheet contained no data")
            self.select_sheet_button.config(state="disabled")
            return
        rl = len(max(self.C.new_sheet, key = len))
        self.C.new_sheet[:] = [r + list(repeat("",rl - len(r))) for r in self.C.new_sheet]
        self.select_sheet_button.config(state="disabled")
        self.load_display([h for h in self.C.new_sheet[0]])
        self.sheetdisplay.data_reference(newdataref=self.C.new_sheet,reset_col_positions=True,reset_row_positions=True,redraw=False)
        self.sheetdisplay.set_all_cell_sizes_to_text()

    def load_display(self,cols):
        self.selector.set_columns(cols)
        self.selector.detect_id_col()
        self.selector.detect_par_cols()

    def start_work(self,msg=""):
        self.status.change_text(msg)
        self.disable_widgets()

    def stop_work(self,msg=""):
        self.status.change_text(msg)
        self.enable_widgets()

    def enable_widgets(self):
        self.open_file_display.change_my_state("readonly")
        self.open_file_button.config(state="normal")
        self.sheet_dropdown.config(state="readonly")
        self.selector.enable_me()
        self.add_new_ids_button.config(state="normal")
        self.add_new_dcols_button.config(state="normal")
        self.add_new_pcols_button.config(state="normal")
        self.overwrite_details_button.config(state="normal")
        self.overwrite_parents_button.config(state="normal")
        self.confirm_button.config(state="normal")
        self.sheetdisplay.enable_bindings(("enable_all"))
        self.sheetdisplay.extra_bindings([("row_index_drag_drop", self.drag_row),
                                          ("column_header_drag_drop", self.drag_col),
                                          ("rc_insert_column", self.reset_selectors),
                                          ("rc_insert_row", self.reset_selectors),
                                          ("rc_delete_column", self.reset_selectors),
                                          ("rc_delete_row", self.reset_selectors),
                                          ("ctrl_x", self.ctrl_x_in_sheet),
                                          ("delete_key", self.del_in_sheet),
                                          ("ctrl_v", self.ctrl_v_in_sheet),
                                          ("ctrl_z", self.ctrl_z_in_sheet),
                                          ("edit_cell", self.edit_cell_in_sheet),
                                          ("begin_edit_cell_use_keypress", self.begin_edit_cell),
                                          ("escape_edit_cell", self.escape_edit_cell)
                                          ])

    def disable_widgets(self):
        self.open_file_display.change_my_state("disabled")
        self.open_file_button.config(state="disabled")
        self.sheet_dropdown.config(state="disabled")
        self.select_sheet_button.config(state="disabled")
        self.selector.disable_me()
        self.add_new_ids_button.config(state="disabled")
        self.add_new_dcols_button.config(state="disabled")
        self.add_new_pcols_button.config(state="disabled")
        self.overwrite_details_button.config(state="disabled")
        self.overwrite_parents_button.config(state="disabled")
        self.confirm_button.config(state="disabled")
        self.sheetdisplay.disable_bindings(("disable_all"))
        self.sheetdisplay.extra_bindings([("row_index_drag_drop", None),
                                          ("column_header_drag_drop", None),
                                          ("rc_insert_column", None),
                                          ("rc_insert_row", None),
                                          ("rc_delete_column", None),
                                          ("rc_delete_row", None),
                                          ("ctrl_x", None),
                                          ("delete_key", None),
                                          ("ctrl_v", None),
                                          ("ctrl_z", None),
                                          ("edit_cell", None),
                                          ("begin_edit_cell_use_keypress", None),
                                          ("escape_edit_cell", None)
                                          ])
        self.update()

    def reset(self):
        self.try_to_close_wb()
        self.row_len = 0
        self.ic = None
        self.pcols = []
        self.C.new_sheet = []
        self.open_file_display.set_my_value("")
        self.sheet_dropdown['values'] = []
        self.sheet_dropdown.set("")
        self.select_sheet_button.config(state="disabled")
        self.selector.clear_displays()

    def confirm(self,event=None):
        self.add_new_ids = self.add_new_ids_button.get_checked()
        self.add_new_dcols = self.add_new_dcols_button.get_checked()
        self.add_new_pcols = self.add_new_pcols_button.get_checked()
        self.overwrite_details = self.overwrite_details_button.get_checked()
        self.overwrite_parents = self.overwrite_parents_button.get_checked()
        self.ic = self.selector.get_id_col()
        self.pcols = self.selector.get_par_cols()
        if not self.C.new_sheet:
            self.status.change_text("Please open a file to load data")
            return
        self.row_len = len(max(self.C.new_sheet,key=len))
        if all(x == False for x in (self.add_new_ids,self.add_new_dcols,self.add_new_pcols,self.overwrite_details,self.overwrite_parents)):
            self.status.change_text("Please select at least one option")
            return
        if self.ic in set(self.pcols):
            self.status.change_text("ID column must be different to all parent columns")
            return
        if self.ic is None:
            self.status.change_text("Please select an ID column")
            return
        self.result = True
        self.destroy()

    def cancel(self,event=None):
        self.USER_HAS_CLOSED_WINDOW()


class get_clipboard_data_popup(tk.Toplevel):
    def __init__(self, C, cols, row_len, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Get data from clipboard - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        #self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(0,weight=1)
        #self.grid_rowconfigure(2,weight=1,uniform="y")
        self.selector = id_and_parent_column_selector(self)
        self.selector.grid(row=0,column=0,sticky="nsew")

        self.flattened_choices = flattened_base_ids_choices(self,command=self.flattened_mode_toggle, theme = theme)
        self.flattened_choices.change_theme(theme)
        self.flattened_choices.grid(row=1,column=0,padx=5,pady=5,sticky="nsew")
        self.flattened_selector = flattened_column_selector(self)
        self.flattened_selector.set_columns(cols)
        self.selector.change_theme(theme)
        self.flattened_selector.change_theme(theme)
        self.selector.set_columns(cols)
        self.selector.detect_id_col()
        self.selector.detect_par_cols()
        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                  column_drag_and_drop_perform = False,
                                  row_drag_and_drop_perform = False)
        self.sheetdisplay.enable_bindings(("enable_all"))
        self.sheetdisplay.extra_bindings([("row_index_drag_drop", self.drag_row),
                                          ("column_header_drag_drop", self.drag_col),
                                          ("rc_insert_column", self.reset_selectors),
                                          ("rc_insert_row", self.reset_selectors),
                                          ("rc_delete_column", self.reset_selectors),
                                          ("rc_delete_row", self.reset_selectors),
                                          ("ctrl_x", self.ctrl_x_in_sheet),
                                          ("delete_key", self.del_in_sheet),
                                          ("ctrl_v", self.ctrl_v_in_sheet),
                                          ("ctrl_z", self.ctrl_z_in_sheet),
                                          ("edit_cell", self.edit_cell_in_sheet),
                                          ("begin_edit_cell_use_keypress", self.begin_edit_cell),
                                          ("escape_edit_cell", self.escape_edit_cell),
                                          ("edit_cell", self.escape_edit_cell)
                                          ])
        self.sheetdisplay.grid(row=0,column=1,rowspan=4,sticky="nswe")
        self.sheetdisplay.data_reference(newdataref=self.C.new_sheet,redraw=False)
        self.sheetdisplay.headers(newheaders=0)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=2,column=0,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        
        self.confirm_button = button(self.button_frame,text="Overwrite existing data",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(20,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(20,20))
        self.status = StatusBar(self,text="Select ID and Parent columns", theme = theme)
        self.status.grid(row=3,column=0,sticky="we")
        self.result = False
        self.ic = None
        self.pcols = []
        self.bind("<Escape>",self.cancel)
        center(self,1280,620)
        self.selector.grid_forget()
        self.flattened_selector.grid(row=0,column=0,pady=(0,35),sticky="nsew")
        self.flattened_selector.grid_forget()
        self.selector.grid(row=0,column=0,sticky="nsew")
        self.deiconify()
        self.wait_window()

    def begin_edit_cell(self, event = None):
        self.unbind("<Escape>")

    def escape_edit_cell(self, event = None):
        self.bind("<Escape>",self.cancel)

    def reset_selectors(self, event = None):
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        ancparcols = self.flattened_selector.get_par_cols()
        self.selector.set_columns([h for h in self.C.new_sheet[0]] if self.C.new_sheet else [])
        self.flattened_selector.set_columns([h for h in self.C.new_sheet[0]] if self.C.new_sheet else [])
        if idcol is not None and self.C.new_sheet:
            self.selector.set_id_col(idcol)
        if parcols and self.C.new_sheet:
            self.selector.set_par_cols(parcols)
        if ancparcols and self.C.new_sheet:
            self.flattened_selector.set_par_cols(ancparcols)

    def drag_col(self, selected_cols, c):
        c = int(c)
        colsiter = list(selected_cols)
        colsiter.sort()
        stins = colsiter[0]
        endins = colsiter[-1] + 1
        totalcols = len(colsiter)
        if stins > c:
            for rn in range(len(self.C.new_sheet)):
                self.C.new_sheet[rn] = (self.C.new_sheet[rn][:c] +
                                              self.C.new_sheet[rn][stins:stins + totalcols] +
                                              self.C.new_sheet[rn][c:stins] +
                                              self.C.new_sheet[rn][stins + totalcols:])
        else:
            for rn in range(len(self.C.new_sheet)):
                self.C.new_sheet[rn] = (self.C.new_sheet[rn][:stins] +
                                              self.C.new_sheet[rn][stins + totalcols:c + 1] +
                                              self.C.new_sheet[rn][stins:stins + totalcols] +
                                              self.C.new_sheet[rn][c + 1:])
        self.sheetdisplay.MT.data_ref = self.C.new_sheet
        self.selector.set_columns([h for h in self.C.new_sheet[0]])
        self.flattened_selector.set_columns([h for h in self.C.new_sheet[0]])
        self.selector.detect_id_col()
        self.selector.detect_par_cols()

    def drag_row(self, selected_rows, r):
        r = int(r)
        rowsiter = list(selected_rows)
        rowsiter.sort()
        stins = rowsiter[0]
        endins = rowsiter[-1] + 1
        totalrows = len(rowsiter)
        if stins > r:
            self.C.new_sheet = (self.C.new_sheet[:r] +
                                      self.C.new_sheet[stins:stins + totalrows] +
                                      self.C.new_sheet[r:stins] +
                                      self.C.new_sheet[stins + totalrows:])
        else:
            self.C.new_sheet = (self.C.new_sheet[:stins] +
                                      self.C.new_sheet[stins + totalrows:r + 1] +
                                      self.C.new_sheet[stins:stins + totalrows] +
                                      self.C.new_sheet[r + 1:])
        self.sheetdisplay.MT.data_ref = self.C.new_sheet
        if endins == 0 or r == 0 or stins == 0:
            self.selector.set_columns([h for h in self.C.new_sheet[0]])
            self.flattened_selector.set_columns([h for h in self.C.new_sheet[0]])
            self.selector.detect_id_col()
            self.selector.detect_par_cols()

    def del_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_x_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_v_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_z_in_sheet(self, event = None):
        self.reset_selectors()

    def edit_cell_in_sheet(self, event = None):
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        ancparcols = self.flattened_selector.get_par_cols()
        if event[1] == idcol or event[1] in parcols or event[1] in ancparcols or event[0] == 0:
            self.reset_selectors()

    def flattened_mode_toggle(self):
        x = self.flattened_choices.get_choices()[0]
        if x:
            self.selector.grid_forget()
            self.flattened_selector.grid(row=0,column=0,pady=(0,35),sticky="nsew")
        else:
            self.flattened_selector.grid_forget()
            self.selector.grid(row=0,column=0,sticky="nsew")

    def confirm(self,event=None):
        self.ic = self.selector.get_id_col()
        self.pcols = self.selector.get_par_cols()
        self.flattened_pcols = self.flattened_selector.get_par_cols()
        self.flattened = self.flattened_choices.get_choices()
        self.C.new_sheet = self.sheetdisplay.get_sheet_data()
        if self.flattened[0]:
            if not self.flattened_pcols:
                self.status.change_text("Please select hierarchy columns")
                return
        else:
            if self.ic in set(self.pcols):
                self.status.change_text("ID column must be different to all parent columns")
                return
            if self.ic is None:
                self.status.change_text("Please select an ID column")
                return
            if not self.pcols:
                self.status.change_text("Please select parent columns")
                return
        self.result = True
        self.destroy()

    def cancel(self,event=None):
        self.destroy()


class sheet_column_display_chooser_popup(tk.Toplevel):
    def __init__(self, C, headers, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Show/Hide columns - Click on columns to toggle")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(0,
                                  weight = 1)
        self.grid_rowconfigure(0,
                               weight = 1)
        
        self.headers =  headers
        self.chosen_indexes = []
        self.result = False
        self.toggled = None

        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                  outline_thickness = 0)
        self.sheetdisplay.headers(newheaders = ["Column Name"])
        self.sheetdisplay.data_reference(newdataref = [[hdr.name] for hdr in self.C.headers],
                                         reset_col_positions = False,
                                         reset_row_positions = False,
                                         redraw = False)
        self.sheetdisplay.set_all_cell_sizes_to_text()

        disp = set(self.C.displayed_columns)
        for c in range(len(self.C.headers)):
            if c in disp:
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#8cba66",
                                                  fg = "black")
            else:
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#fc8c55",
                                                  fg = "black")
        self.sheetdisplay.grid(row = 0,
                               column = 0,
                               sticky = "nswe")

        self.button_f1 = frame(self, theme = theme)
        self.button_f1.grid_columnconfigure(0,
                                            weight = 1,
                                            uniform = "x")
        self.button_f1.grid_columnconfigure(1,
                                            weight = 1,
                                            uniform = "x")
        self.button_f1.grid_rowconfigure(0,
                                         weight = 1)
        self.button_f1.grid(row = 1,
                            column = 0,
                            sticky = "nswe")
        self.add_all_button = button(self.button_f1,
                                     text = "Show all",
                                     style = "EF.Std.TButton",
                                     command = self.add_all)
        self.add_all_button.grid(row = 0,
                                 column = 0,
                                 sticky = "nswe",
                                 padx = (20, 10),
                                 pady = (5, 10))
        self.remove_all_button = button(self.button_f1,
                                         text = "Hide all",
                                         style = "EF.Std.TButton",
                                         command = self.remove_all)
        self.remove_all_button.grid(row = 0,
                                     column = 1,
                                     sticky = "nswe",
                                     padx = (10, 20),
                                     pady = (5, 10))

        self.button_f2 = frame(self, theme = theme)
        self.button_f2.grid_columnconfigure(0,
                                            weight = 1,
                                            uniform = "x")
        self.button_f2.grid_columnconfigure(1,
                                            weight = 1,
                                            uniform = "x")
        self.button_f2.grid_rowconfigure(0,
                                         weight = 1)
        self.button_f2.grid(row = 2,
                            column = 0,
                            sticky = "nswe")
        self.confirm_button = button(self.button_f2,
                                     text = "Confirm selections",
                                     style = "EF.Std.TButton",
                                     command = self.confirm)
        self.confirm_button.grid(row = 0,
                                 column = 0,
                                 sticky = "nswe",
                                 padx = (20, 10),
                                 pady = (20, 20))
        self.cancel_button = button(self.button_f2,
                                     text = "Cancel",
                                     style = "EF.Std.TButton",
                                     command = self.cancel)
        self.cancel_button.grid(row = 0,
                                 column = 1,
                                 sticky = "nswe",
                                 padx = (10, 20),
                                 pady = (20, 20))

        self.info = StatusBar(self,
                              text = "** Hidden columns will be excluded when copying to clipboard **",
                              theme = self.C.C.theme)
        self.info.grid(row = 3,
                       column = 0,
                       sticky = "nswe")

        self.sheetdisplay.bind("<Button-1>", self.b1)
        self.sheetdisplay.bind("<B1-Motion>", self.b1_motion)
        self.bind("<Escape>",self.cancel)
        center(self, 500, 700)
        self.deiconify()
        self.wait_window()

    def b1(self, event = None):
        c = self.sheetdisplay.identify_row(event, allow_end = False)
        if c is not None:
            self.toggled = c
            highs = self.sheetdisplay.get_highlighted_cells()
            if highs[(c, 0)][0] == "#8cba66":
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#fc8c55",
                                                  fg = "black")
            else:
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#8cba66",
                                                  fg = "black")
        self.sheetdisplay.refresh()

    def b1_motion(self, event = None):
        c = self.sheetdisplay.identify_row(event, allow_end = False)
        if c is not None and self.toggled != c:
            self.toggled = c
            highs = self.sheetdisplay.get_highlighted_cells()
            if highs[(c, 0)][0] == "#8cba66":
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#fc8c55",
                                                  fg = "black")
            else:
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#8cba66",
                                                  fg = "black")
        self.sheetdisplay.refresh()

    def add_all(self,event=None):
        self.sheetdisplay.dehighlight_cells("all")
        for c in range(len(self.C.headers)):
            self.sheetdisplay.highlight_cells(row = c,
                                              column = 0,
                                                bg = "#8cba66",
                                                fg = "black")
        self.sheetdisplay.refresh()
            
    def remove_all(self,event=None):
        self.sheetdisplay.dehighlight_cells("all")
        for c in range(len(self.C.headers)):
            self.sheetdisplay.highlight_cells(row = c,
                                              column = 0,
                                                bg = "#fc8c55",
                                                fg = "black")
        self.sheetdisplay.refresh()

    def confirm(self,event=None):
        sels = [k[0] for k, v in self.sheetdisplay.get_highlighted_cells().items() if v[0] == "#8cba66"]
        if not sels or len(sels) == len(self.headers):
            self.C.all_columns_displayed = True
        else:
            self.C.all_columns_displayed = False
            self.chosen_indexes = sorted(sels)
        self.result = True
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class askconfirm_three(tk.Toplevel):
    def __init__(self,
                 C,
                 action,
                 text1 = "Confirm 1",
                 text2 = "Confirm 2",
                 text3 = "Cancel",
                 bgcolor = "green",
                 fgcolor = "white",
                 theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format = "gif", data = top_left_icon))
        self.title("Confirm Action - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.action_label = label(self, text = "Action:", font = EF)
        self.action_label.config(background=bgcolor,foreground=fgcolor)
        self.action_label.grid(row=0,column=0,sticky="nswe",pady=(20,5),padx=20)
        self.action_display = display_text(parent = self, text = action)
        self.action_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.action_display.config(height = 75)
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(10,20))
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(2,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.button1 = button(self.button_frame,text=text1,style="EF.Std.TButton",command=self.button1)
        self.button1.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(20,10))
        self.button2 = button(self.button_frame,text=text2,style="EF.Std.TButton",command=self.button2)
        self.button2.grid(row=0,column=1,sticky="nswe",padx=(10,10),pady=(20,10))
        self.button3 = button(self.button_frame,text=text3,style="EF.Std.TButton",command=self.cancel)
        self.button3.grid(row=0,column=2,sticky="nswe",padx=(10,20),pady=(20,10))
        self.choice = None
        self.bind("<Escape>",self.cancel)
        self.action_display.place_cursor()
        center(self,600,200)
        self.deiconify()
        self.wait_window()
    def button1(self,event=None):
        self.choice = 1
        self.destroy()
    def button2(self,event=None):
        self.choice = 2
        self.destroy()
    def cancel(self,event=None):
        self.destroy()


class askconfirm(tk.Toplevel):
    def __init__(self,C,action,confirm_text="Confirm",cancel_text="Cancel",bgcolor="green",fgcolor="white", theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format = "gif", data = top_left_icon))
        self.title("Confirm Action - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.action_label = label(self, text = "Action:", font = EF)
        self.action_label.config(background=bgcolor,foreground=fgcolor)
        self.action_label.grid(row=0,column=0,sticky="nswe",pady=(20,5),padx=20)
        self.action_display = display_text(parent = self, text = action, theme = theme)
        self.action_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.action_display.config(height = 75)
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(10,20))
        self.button_frame.grid_columnconfigure(0, weight = True, uniform = "x")
        self.button_frame.grid_columnconfigure(1, weight = True, uniform = "x")
        self.button_frame.grid_rowconfigure(0, weight = True)
        self.confirm_button = button(self.button_frame,text=confirm_text,style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row = 0, column = 0, sticky = "nswe", padx = (0, 20))
        self.cancel_button = button(self.button_frame, text=cancel_text,style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe")
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.boolean = False
        self.action_display.place_cursor()
        center(self, 530, 155)
        self.deiconify()
        self.wait_window()
    def confirm(self,event=None):
        self.boolean = True
        self.destroy()
    def cancel(self,event=None):
        self.destroy()


class license_key_entry_popup(tk.Toplevel):
    def __init__(self, C, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False, False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title(f"Enter license key - Contact {contact_email} for more information")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(2,weight=1)

        self.label2 = label(self, text = "License key:", font = EF, theme = theme)
        self.label2.config(anchor = "e")
        self.label2.grid(row=1,column=0,sticky="nswe",pady=(0,17),padx=(20,10))
        
        self.display2 = entry_with_scrollbar(self, theme = theme)
        self.display2.set_my_value("")
        self.display2.grid(row=1,column=1,sticky="nswe",pady=(18,20),padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 2, column = 0, columnspan = 2, sticky = "nswe")
        self.confirm_button = button(self.bf,text="Confirm",style="EF.Std.TButton",
                                     command = self.confirm)
        self.confirm_button.grid(row=0, column = 0, sticky = "nswe",
                                 padx = 60,
                                 pady = 10)
        self.cancel_button = button(self.bf, text = "Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe",
                                padx = 60,
                                pady = 10)
        self.status_bar = StatusBar(self, text = "Enter your 48 character license key, case insensitive", theme = theme)
        self.status_bar.grid(row=3,column=0,columnspan=2,sticky="nswe")
        self.bind("<Return>",self.confirm)
        self.display2.bind("<Return>", self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.display2.place_cursor()
        center(self,670,150)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        cset = set("ABCDEFGHIJKLMNPQRSTUVWXYZ0123456789")
        self.license_key = "".join(c for c in "".join(self.display2.get_my_value().upper().split("-")) if c in cset)
        valid = self.C.license_key_valid(self.license_key)
        if valid:
            self.result = True
            self.destroy()
        elif valid is None:
            self.status_bar.change_text(f"License key invalid, contact {contact_email} for help   ")
        elif not valid:
            self.status_bar.change_text(f"License key expired, contact {contact_email} for help   ")
        if not valid:
            self.C.configsettings["License"] = ""
            self.C.setup_config()
        
    def cancel(self,event=None):
        self.destroy()
        

class save_new_version_presave_popup(tk.Toplevel):
    def __init__(self, C, file_location, theme = "dark"):
        tk.Toplevel.__init__(self, C, width = "1", height = "1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm", "iconphoto", self._w,tk.PhotoImage(format = "gif", data = top_left_icon))
        self.title("Save new version - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.file_loc_label = label(self, text = "Your new version\nwill be saved here:", font = EF, theme = theme)
        self.file_loc_label.grid(row=0,column=0,sticky="nswe",padx=(20,10))
        self.file_loc_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.file_loc_display.set_my_value(file_location)
        self.file_loc_display.grid(row=0,column=1,sticky="nswe",pady=(20,20),padx=(0,20))
        self.choose_loc_button = button(self,text="Choose where to save a new version",style="EF.Std.TButton",command=self.choose_loc)
        self.choose_loc_button.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(0,10))
        self.confirm_button = button(self,text="Confirm save a new version here",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=2,column=0,columnspan=2,sticky="nswe",padx=20,pady=(5,20))
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.file_loc_display.place_cursor()
        center(self,550,170)
        self.deiconify()
        self.wait_window()
        
    def choose_loc(self,event=None):
        folder = os.path.normpath(filedialog.askdirectory(parent=self,title="Select a folder to save new version in"))
        if folder == ".":
            return
        self.file_loc_display.set_my_value(folder)
        
    def confirm(self,event=None):
        self.result = os.path.normpath(self.file_loc_display.get_my_value())
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class save_new_version_postsave_popup(tk.Toplevel):
    def __init__(self, C, file_location, filename, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1",background = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Success! New version saved - Click the Okay/X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.file_loc_label = label(self,text="Your new version\nwas saved here:",font=EF, theme = theme)
        self.file_loc_label.grid(row=0,column=0,sticky="nswe",padx=(20,10))
        self.file_loc_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.file_loc_display.set_my_value(file_location)
        self.file_loc_display.grid(row=0,column=1,sticky="nswe",pady=(20,20),padx=(0,20))
        self.file_name_label = label(self,text="This is the\n filename:",font=EF, theme = theme)
        self.file_name_label.grid(row=1,column=0,sticky="nswe",padx=(20,10))
        self.file_name_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.file_name_display.set_my_value(filename)
        self.file_name_display.grid(row=1,column=1,sticky="nswe",pady=(0,20),padx=(0,20))
        self.confirm_button = button(self,text="Okay",style="EF.Std.TButton",command=self.cancel)
        self.confirm_button.grid(row=2,column=0,columnspan=2,sticky="nswe",padx=20,pady=(0,20))
        self.bind("<Return>",self.cancel)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.file_name_display.place_cursor()
        center(self, 550, 185)
        self.deiconify()
        self.wait_window()
        
    def cancel(self,event=None):
        self.destroy()


class save_new_version_error_popup(tk.Toplevel):
    def __init__(self, C, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1",background= theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Error locating folder - Click the Okay/X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(0,weight=1)
        self.info_label = label(self,text="There was an error locating the saving folder. \n - Would you like to choose where to save a new version?",font=EF, theme = theme)
        self.info_label.grid(row=0,column=0,sticky="nswe",padx=20,pady=20)
        self.confirm_button = button(self,text="Choose where to save a new version",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(0,20))
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        center(self, 550, 130)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = True
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class sort_sheet_popup(tk.Toplevel):
    def __init__(self,C,headers,theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1",background=theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Sort sheet - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_columnconfigure(0,weight=1)
        self.sort_decision = {'type': None,
                              'col': None,
                              'order': None}
        self.sort_by_col_button = button(self,style="EF.Std.TButton",
                                         text="Sort by column",
                                         command=self.sort_by_col)
        self.sort_icon = tk.PhotoImage(format="gif",data=sort_icon)
        self.sort_by_col_button.config(image=self.sort_icon,compound="left")
        self.sort_by_col_button.grid(row=0,column=1, sticky="nswe",pady=(15,5),padx=70)
        self.order_label = label(self,text="Order:",font=EF, theme = theme)
        self.order_label.grid(row=1,column=0,sticky="nswe",pady=5,padx=(50,10))
        self.order_dropdown = ez_dropdown(self,EF)
        self.order_dropdown['values'] = ["ASCENDING","DESCENDING"]
        self.order_dropdown.set_my_value("ASCENDING")
        self.order_dropdown.grid(row=1,column=1,sticky="nswe",pady=5,padx=(0,70))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",pady=5,padx=(50,10))
        self.col_dropdown = ez_dropdown(self,EF)
        self.col_dropdown['values'] = headers
        self.col_dropdown.set_my_value(headers[0])
        self.col_dropdown.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,70))
        self.divider = frame(self)
        self.divider.config(bg = theme_fg(theme))
        self.divider.config(height = 5)
        self.divider.grid(row = 3, column = 0, columnspan = 2, padx = 20, pady = (10, 15), sticky = "ew")
        self.sort_by_tree_button = button(self,style="EF.Std.TButton",
                                          text="Sort by tree walk",
                                          command=self.sort_by_tree)
        self.tree_icon = tk.PhotoImage(format="gif",data=tree_icon)
        self.sort_by_tree_button.config(image=self.tree_icon,compound="left")
        self.sort_by_tree_button.grid(row=4,column=1, sticky="nswe",pady=(20,20),padx=70)
        self.bind("<Escape>",self.go_back)
        self.order_dropdown.bind("<<ComboboxSelected>>",lambda event: self.focus_set())
        self.col_dropdown.bind("<<ComboboxSelected>>",lambda event: self.focus_set())
        center(self, 550, 215)
        self.deiconify()
        self.wait_window()
        
    def sort_by_col(self,event=None):
        self.sort_decision = {'type': "by column",
                              'col': self.col_dropdown.get_my_value(),
                              'order': self.order_dropdown.get_my_value()}
        self.destroy()
        
    def sort_by_tree(self,event=None):
        self.sort_decision['type'] = "by tree"
        self.destroy()
        
    def go_back(self,event=None):
        self.destroy()


class edit_detail_date_popup(tk.Toplevel):
    def __init__(self,C,ID,column,current_detail,DATE_FORM,validation_values=[],set_value=None, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1")
        self.withdraw()
        self.resizable(False, False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Change date detail - Click the X button or press escape to go back")
        self.C = C
        self.config(bg = theme_bg(theme))
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_label = label(self,text="ID:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",padx=20)
        self.col_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe")
        
        if validation_values:
            self.validation_dropdown = ez_dropdown(self,font=EF)
            self.validation_dropdown['values'] = validation_values
            if set_value is not None:
                self.validation_dropdown.set_my_value(set_value)
            else:
                self.validation_dropdown.set_my_value(validation_values[0])
            self.validation_dropdown.grid(row=3,column=0,columnspan=2,sticky="nswe",padx=20,pady=10)
            self.validation_dropdown.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
            width_ = 600
            height_ = 225
            self.bind("<Return>",self.confirm_validation)
        else:
            self.entries_frame = frame(self, theme = theme)
            self.entries_frame.grid_columnconfigure(3,weight=1)
            self.entries_frame.grid(row=3,column=0,columnspan=2,sticky="nswe",pady=10)
            if DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
                self.date_label = label(self.entries_frame,text="Set date DD/MM/YYYY:",font=EF, theme = theme)
            elif DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
                self.date_label = label(self.entries_frame,text="Set date YYYY/MM/DD:",font=EF, theme = theme)
            elif DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
                self.date_label = label(self.entries_frame,text="Set date MM/DD/YYYY:",font=EF, theme = theme)
            self.date_label.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=10)
            self.date_entry_widget = date_entry(self.entries_frame,DATE_FORM, theme = theme)
            self.date_entry_widget.grid(row=0,column=1,sticky="nswe",padx=(0,30),pady=10)
            self.numerical_label = label(self.entries_frame,text="OR set Number:",font=EF, theme = theme)
            self.numerical_label.grid(row=0,column=2,sticky="nswe",padx=(0,10),pady=10)
            self.numerical_entry_widget = numerical_entry_with_scrollbar(self.entries_frame, theme = theme)
            self.numerical_entry_widget.grid(row=0,column=3,sticky="nswe",padx=(0,20),pady=15)
            if "/" in current_detail or "-" in current_detail:
                self.date_entry_widget.set_my_value(current_detail)
            else:
                self.numerical_entry_widget.set_my_value(current_detail)
            self.numerical_entry_widget.my_entry.bind("<Return>",self.confirm_normal)
            self.date_entry_widget.entry_1.bind("<Return>",self.confirm_normal)
            self.date_entry_widget.entry_2.bind("<Return>",self.confirm_normal)
            self.date_entry_widget.entry_3.bind("<Return>",self.confirm_normal)
            self.date_entry_widget.place_cursor()
            width_ = 850
            height_ = 280

        self.confirm_button = button(self.bf,text="Save",style="EF.Std.TButton",
                                     command=self.confirm_validation if validation_values else self.confirm_normal)
        self.confirm_button.grid(row=0, column = 0, sticky = "nswe",
                                 padx = 70 if validation_values else 100,
                                 pady = 20)
        self.cancel_button = button(self.bf, text = "Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe",
                                padx = 70 if validation_values else 100,
                                pady = 20)
        
        self.result = False
        center(self,width_,height_)
        self.deiconify()
        self.bind("<Escape>",self.cancel)
        self.wait_window()
        
    def confirm_normal(self,event=None):
        self.result = True
        x1 = self.date_entry_widget.get_my_value()
        x2 = self.numerical_entry_widget.get_my_value()
        if not all(c in ("/", "-") for c in x1):
            self.saved_string = x1
        elif x2:
            self.saved_string = x2
        else:
            self.saved_string = ""
        self.destroy()
        
    def confirm_validation(self,event=None):
        self.result = True
        self.saved_string = self.validation_dropdown.get_my_value()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class edit_detail_numerical_popup(tk.Toplevel):
    def __init__(self,C,ID,column,current_detail,validation_values=[],set_value=None, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Change numerical detail - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_label = label(self,text="ID:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",padx=20)
        self.col_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe")

        if validation_values:
            self.validation_dropdown = ez_dropdown(self,font=EF)
            self.validation_dropdown['values'] = validation_values
            if set_value is not None:
                self.validation_dropdown.set_my_value(set_value)
            else:
                self.validation_dropdown.set_my_value(validation_values[0])
            self.validation_dropdown.grid(row=3,column=0,columnspan=2,sticky="nswe",padx=20,pady=10)
            self.validation_dropdown.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
            width_ = 600
            height_ = 225
            self.bind("<Return>",self.confirm_validation)
        else:
            self.entry_widget = numerical_entry_with_scrollbar(self, theme = theme)
            self.entry_widget.set_my_value(current_detail)
            self.entry_widget.grid(row=3,column=0,columnspan=2,sticky="nswe",padx=20,pady=10)
            self.entry_widget.my_entry.bind("<Return>",self.confirm_normal)
            width_ = 600
            height_ = 240
            self.entry_widget.place_cursor()
            
        self.confirm_button = button(self.bf,text="Save",style="EF.Std.TButton",
                                     command=self.confirm_validation if validation_values else self.confirm_normal)
        self.confirm_button.grid(row=0, column = 0, sticky = "nswe", padx = 70, pady = 20)
        self.cancel_button = button(self.bf, text = "Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = 70, pady = 20)
        
        self.result = False
        center(self,width_,height_)
        self.deiconify()
        self.bind("<Escape>",self.cancel)
        self.wait_window()
        
    def confirm_normal(self,event=None):
        self.result = True
        self.saved_string = self.entry_widget.get_my_value()
        self.destroy()
        
    def confirm_validation(self,event=None):
        self.result = True
        self.saved_string = self.validation_dropdown.get_my_value()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()
        

class edit_detail_text_popup(tk.Toplevel):
    def __init__(self,C,ID,column,current_detail,validation_values=[],set_value=None, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1")
        self.withdraw()
        self.resizable(False, False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Edit cell - Click the X button or press escape to go back")
        self.C = C
        self.config(bg = theme_bg(theme))
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_label = label(self,text="ID:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",padx=20)
        self.col_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe")
        
        if validation_values:
            self.validation_dropdown = ez_dropdown(self,font=EF)
            self.validation_dropdown['values'] = validation_values
            if set_value is not None:
                self.validation_dropdown.set_my_value(set_value)
            else:
                self.validation_dropdown.set_my_value(validation_values[0])
            self.validation_dropdown.grid(row=3,column=0,columnspan=2,sticky="nswe",padx=20,pady=10)
            self.validation_dropdown.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
            width_ = 620
            height_ = 225
            self.confirm_button = button(self.bf,text="Save",style="EF.Std.TButton",command=self.confirm_validation)
            self.bind("<Return>",self.confirm_validation)
        else:
            self.grid_rowconfigure(3,weight=1)
            self.text_widget = wrapped_text_with_find_and_yscroll(self,current_detail,"normal",15, theme = theme)
            self.text_widget.grid(row=3,column=0,sticky="nswe",columnspan=2)
            self.text_widget.place_cursor()
            width_ = 800
            height_ = 595
            self.confirm_button = button(self.bf,text="Save",style="EF.Std.TButton",command=self.confirm_normal)
        self.confirm_button.grid(row=0,column=0,sticky="nswe", padx = 75, pady = 20)
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = 75, pady = 20)
        center(self,width_,height_)
        self.result = False
        self.deiconify()
        self.bind("<Escape>",self.cancel)
        self.wait_window()
        
    def confirm_normal(self,event=None):
        self.result = True
        self.saved_string = self.text_widget.get_my_value().rstrip()
        self.destroy()
        
    def confirm_validation(self,event=None):
        self.result = True
        self.saved_string = self.validation_dropdown.get_my_value()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class view_column_text_popup(tk.Toplevel):
    def __init__(self,C,ID,column,text, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1")
        self.withdraw()
        self.resizable(False, False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("View text - Click the X button or press escape to go back")
        self.C = C
        self.config(bg = theme_bg(theme))
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_label = label(self,text="ID:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",padx=20)
        self.col_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,20))
        self.text_widget = wrapped_text_with_find_and_yscroll(self,text,"disabled",15, theme = theme)
        self.text_widget.grid(row=3,column=0,sticky="nswe",columnspan=2)
        self.cancel_button = button(self, text = "Close", style = "EF.Std.TButton", command = self.cancel)
        self.cancel_button.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe", padx = 220, pady = (25, 20))
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.text_widget.place_cursor()
        center(self,850,545)
        self.deiconify()
        self.wait_window()
        
    def cancel(self,event=None):
        self.destroy()


class add_top_id_popup(tk.Toplevel):
    def __init__(self, C, ss_selection, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Add top ID - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_name_label = label(self,text="ID name:",font=EF, theme = theme)
        self.id_name_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_name_display = entry_with_scrollbar(self, theme = theme)
        self.id_name_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        if self.C.tv_label_col != self.C.ic:
            self.id_tv_label = label(self,text="ID Treeview Label:",font=EF, theme = theme)
            self.id_tv_label.grid(row=1,column=0,sticky="nswe",padx=20)
            self.id_tv_display = entry_with_scrollbar(self, theme = theme)
            self.id_tv_display.grid(row=1,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.enter_ss_sel_button = button(self,text="Enter current sheet selection",style="EF.Std.TButton",command=self.enter_ss_sel)
        self.enter_ss_sel_button.grid(row=2,column=1,sticky="nswe",padx= (0, 20), pady=(10,5))
        if not ss_selection:
            self.enter_ss_sel_button.config(state="disabled")
        else:
            self.ss_sel = ss_selection
        
        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 3, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Add",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (75, 50), pady = (30, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 75), pady = (30, 20))
        
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.id_name_display.place_cursor()
        center(self,600,186)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = "".join(self.id_name_display.get_my_value().strip().split())
        try:
            self.id_label = self.id_tv_display.get_my_value().strip().split("\n")[0]
        except:
            pass
        self.destroy()
        
    def enter_ss_sel(self,event=None):
        self.id_name_display.set_my_value(self.ss_sel)
        if self.C.tv_label_col != self.C.ic:
            detail = self.C.sheet[self.C.rns[self.ss_sel.lower()]][self.C.tv_label_col]
            ni = detail.find("\n")
            if ni == -1:
                self.id_tv_display.set_my_value(detail)
            else:
                self.id_tv_display.set_my_value(detail[:ni])
                
    def cancel(self,event=None):
        self.destroy()


class add_child_or_sibling_id_popup(tk.Toplevel):
    def __init__(self, C, chld_or_sib, desired_parent, ss_selection, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", background = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        if chld_or_sib == "sibling":
            self.title("Add sibling ID - Click the X button or press escape to go back")
        elif chld_or_sib == "child":
            self.title("Add child ID - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.parent_label = label(self,text="Desired\nparent:",font=EF, theme = theme)
        self.parent_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.parent_display = readonly_entry_with_scrollbar(self, theme = theme)
        if desired_parent:
            self.parent_display.set_my_value(desired_parent)
        else:
            self.parent_display.set_my_value("** No parent - Top ID **")
        self.parent_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.id_name_label = label(self,text="ID name:",font=EF, theme = theme)
        self.id_name_label.grid(row=1,column=0,sticky="nswe",padx=20)
        self.id_name_display = entry_with_scrollbar(self, theme = theme)
        self.id_name_display.grid(row=1,column=1,sticky="nswe",pady=(5,10),padx=(0,20))
        if self.C.tv_label_col != self.C.ic:
            self.id_tv_label = label(self,text="ID Treeview Label:",font=EF, theme = theme)
            self.id_tv_label.grid(row=2,column=0,sticky="nswe",padx=20)
            self.id_tv_display = entry_with_scrollbar(self, theme = theme)
            self.id_tv_display.grid(row=2,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.enter_ss_sel_button = button(self,text="Enter current sheet selection",style="EF.Std.TButton",command=self.enter_ss_sel)
        self.enter_ss_sel_button.grid(row=3,column=1,sticky="nswe", padx= (0, 20), pady=(10,5))
        if not ss_selection:
            self.enter_ss_sel_button.config(state="disabled")
        else:
            self.ss_sel = ss_selection
        
        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Add",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (75, 50), pady = (30, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 75), pady = (30, 20))
        
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.id_name_display.place_cursor()
        center(self,600,237)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = "".join(self.id_name_display.get_my_value().strip().split())
        try:
            self.id_label = self.id_tv_display.get_my_value().strip().split("\n")[0]
        except:
            pass
        self.destroy()
        
    def enter_ss_sel(self,event=None):
        self.id_name_display.set_my_value(self.ss_sel)
        if self.C.tv_label_col != self.C.ic:
            detail = self.C.sheet[self.C.rns[self.ss_sel.lower()]][self.C.tv_label_col]
            ni = detail.find("\n")
            if ni == -1:
                self.id_tv_display.set_my_value(detail)
            else:
                self.id_tv_display.set_my_value(detail[:ni])
                
    def cancel(self,event=None):
        self.destroy()


class rename_id_popup(tk.Toplevel):
    def __init__(self, C, ID, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Rename ID - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(2,weight=1)
        self.id_label = label(self,text="Current ID\nname:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.new_name_label = label(self,text="New ID\nname:",font=EF, theme = theme)
        self.new_name_label.grid(row=1,column=0,sticky="nswe",padx=20)
        self.new_name_display = entry_with_scrollbar(self, theme = theme)
        self.new_name_display.grid(row=1,column=1,sticky="nswe",pady=5,padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 2, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Rename",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (75, 50), pady = (20, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 75), pady = (20, 20))
        
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.new_name_display.place_cursor()
        center(self,600,185)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = "".join(self.new_name_display.get_my_value().strip().split())
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class enter_sheet_name_popup(tk.Toplevel):
    def __init__(self, C, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Enter desired sheet name - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.sheet_name_label = label(self,text="New sheet\nname:",font=EF, theme = theme)
        self.sheet_name_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.sheet_entry = entry_with_scrollbar(self, theme = theme)
        self.sheet_entry.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        
        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 1, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Confirm",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (75, 50), pady = (20, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 75), pady = (20, 20))
        
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.sheet_entry.place_cursor()
        center(self,600,137)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = self.sheet_entry.get_my_value()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class error(tk.Toplevel):
    def __init__(self, C, msg, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("*** Error *** - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.errorlabel = label(self,text="Error\nmessage:",font=EF, theme = theme)
        self.errorlabel.config(background="red",foreground="white")
        self.errorlabel.grid(row=0,column=0,sticky="nswe",pady=(20,5),padx=20)
        self.error_display = display_text(parent = self, text = msg, theme = theme)
        self.error_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.error_display.config(height = 75)
        self.confirm_button = button(self,text="Okay",style="EF.Std.TButton",command=self.cancel)
        self.confirm_button.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(10,20))
        self.bind("<Return>",self.cancel)
        self.bind("<Escape>",self.cancel)
        center(self,600,180)
        self.deiconify()
        self.wait_window()
        
    def cancel(self,event=None):
        self.destroy()


class treeview_id_finder(tk.Toplevel):
    def __init__(self, C, hiers, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("ID is in multiple hierarchies, choose which hierarchy to go to - Click the X button or press escape to go back")
        self.C = C
        self.GO = False
        self.selected = 0
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1,weight=1)
        self.dd_1 = ez_dropdown(self,EF)
        self.dd_1['values'] = hiers
        self.dd_1.set_my_value(hiers[0])
        self.dd_1.grid(row=0,column=0,sticky="nswe",columnspan=2,padx=20,pady=(20,5))
        
        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 1, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Go",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (90, 50), pady = (20, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 90), pady = (20, 20))
        
        self.bind("<Escape>",self.cancel)
        center(self,700,120)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.selected = self.dd_1.displayed.get()
        self.GO = True
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class ss_settings_chooser(tk.Toplevel):
    def __init__(self, C, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Choose sheet settings - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.changed = False
        self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1,weight=1)
        self.settings_frame = frame(self, theme = theme)
        self.settings_frame.grid(row=0,column=0,sticky="nswe",columnspan=2,padx=20,pady=(20,5))
        
        self.dd_2_label = label(self.settings_frame,"Main table align: ",BF, theme = theme)
        self.dd_2_label.grid(row=0,column=0,sticky="nswe")
        self.dd_2 = ez_dropdown(self.settings_frame,EF)
        self.dd_2['values'] = ("Left","Center")
        x = self.C.sheetdisplay.align()
        if x == "w":
            self.dd_2.set_my_value("Left")
        elif x == "center":
            self.dd_2.set_my_value("Center")
        self.dd_2.grid(row=0,column=1,sticky="nswe")
        
        self.dd_3_label = label(self.settings_frame,"Row index align: ",BF, theme = theme)
        self.dd_3_label.grid(row=1,column=0,sticky="nswe")
        self.dd_3 = ez_dropdown(self.settings_frame,EF)
        self.dd_3['values'] = ("Left","Center")
        x = self.C.sheetdisplay.row_index_align()
        if x == "w":
            self.dd_3.set_my_value("Left")
        elif x == "center":
            self.dd_3.set_my_value("Center")
        self.dd_3.grid(row=1,column=1,sticky="nswe")
        
        self.dd_4_label = label(self.settings_frame,"Headers align: ",BF, theme = theme)
        self.dd_4_label.grid(row=2,column=0,sticky="nswe")
        self.dd_4 = ez_dropdown(self.settings_frame,EF)
        self.dd_4['values'] = ("Left","Center")
        x = self.C.sheetdisplay.header_align()
        if x == "w":
            self.dd_4.set_my_value("Left")
        elif x == "center":
            self.dd_4.set_my_value("Center")
        self.dd_4.grid(row=2,column=1,sticky="nswe")

        self.dd_5_label = label(self.settings_frame,"Theme: ",BF, theme = theme)
        self.dd_5_label.grid(row=3,column=0,sticky="nswe")
        self.dd_5 = ez_dropdown(self.settings_frame,EF)
        self.dd_5['values'] = ("Light","Dark")
        if self.C.sheetdisplay.MT.table_background == "white":
            self.dd_5.set_my_value("Light")
        else:
            self.dd_5.set_my_value("Dark")
        self.dd_5.grid(row=3,column=1,sticky="nswe")
        
        self.confirm_button = button(self,text="Confirm",
                                     style="EF.Std.TButton",
                                     command=self.confirm)
        self.confirm_button.grid(row=1,column=0,sticky="nswe",padx=20,pady=(15,20))
        self.cancel_button = button(self,text="Cancel",
                                    style="EF.Std.TButton",
                                    command=self.cancel)
        self.cancel_button.grid(row=1,column=1,sticky="nswe",padx=20,pady=(15,20))
        self.bind("<Escape>",self.cancel)
        center(self,500,185)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        x = self.dd_2.displayed.get()
        if x == "Left":
            self.C.sheetdisplay.align("w",redraw=False)
        elif x == "Center":
            self.C.sheetdisplay.align("center",redraw=False)

        x = self.dd_3.displayed.get()
        if x == "Left":
            self.C.sheetdisplay.row_index_align("w",redraw=False)
        elif x == "Center":
            self.C.sheetdisplay.row_index_align("center",redraw=False)

        x = self.dd_4.displayed.get()
        if x == "Left":
            self.C.sheetdisplay.header_align("w",redraw=False)
        elif x == "Center":
            self.C.sheetdisplay.header_align("center",redraw=False)
        self.C.change_theme(self.dd_5.displayed.get().lower())
        self.changed = True
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class textpopup(tk.Toplevel):
    def __init__(self,C,text,width_=700,height_=650, theme = "dark", use_entry_bg = False):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.theme = theme
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.word = ""
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.find_frame = frame(self, theme = theme)
        self.find_frame.grid(row=0,column=0,columnspan=2,sticky="nswe")
        self.find_icon = tk.PhotoImage(format="gif",data=find_icon)
        self.search_button = button(self.find_frame,
                                    text=" Find:",
                                    command=self.find)
        self.search_button.config(image=self.find_icon,compound="left")
        self.search_button.pack(side="left",fill="x")
        self.find_window = normal_entry(self.find_frame,font=BF, theme = theme)
        self.find_window.bind("<Return>",self.find)
        self.find_window.pack(side="left",fill="x",expand=True)
        self.find_reset_button = button(self.find_frame,text="X",command=self.find_reset)
        self.find_reset_button.pack(side="left",fill="x")
        self.find_results_label = label(self.find_frame,"0/0",BF, theme = theme)
        self.find_results_label.pack(side="left",fill="x")
        self.find_up_button = button(self.find_frame,text="▲",command=self.find_up)
        self.find_up_button.pack(side="left",fill="x")
        self.find_down_button = button(self.find_frame,text="▼",command=self.find_down)
        self.find_down_button.pack(side="left",fill="x")
        self.textbox = working_text(self,
                                    wrap="none",
                                    theme = theme,
                                    use_entry_bg = use_entry_bg,
                                    override_bg = theme_treeview_bg(theme) if theme != "dark" else None)
        self.yscrollb = scrollbar(self,
                                  self.textbox.yview,
                                  "vertical",
                                  self.textbox)
        self.xscrollb = scrollbar(self,
                                  self.textbox.xview,
                                  "horizontal",
                                  self.textbox)
        self.textbox.delete(1.0,"end")
        self.textbox.insert(1.0,text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=1,column=0,sticky="nswe")
        self.yscrollb.grid(row=1,column=1,sticky="nswe")
        self.xscrollb.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.buttonframe = frame(self, theme = theme)
        self.buttonframe.grid(row=3,column=0,columnspan=2,sticky="nswe")
        self.save_text_button = button(self.buttonframe,text="Save as",
                                       style="EF.Std.TButton",
                                       command=lambda: self.save_text(text))
        self.save_text_button.pack(side="left",fill="x",padx=(20,40),pady=20)
        self.save_text_button.config(width=24)
        self.cancel_button = button(self.buttonframe,text="Close popup",
                                    style="EF.Std.TButton",
                                    command=self.cancel)
        self.cancel_button.pack(side="right",fill="x",padx=(40,20),pady=20)
        self.cancel_button.config(width=24)
        self.bind("<Escape>",self.cancel)
        center(self,width_,height_)
        self.deiconify()
        self.wait_window()
        
    def save_text(self,text):
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save text on popup window",
                                               filetypes=[('Text File','.txt'),('CSV File','.csv')],
                                               defaultextension=".txt",
                                               confirmoverwrite=True)
        if not newfile:
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv",".txt")):
            errorpopup = error(self,"Can only save .csv/.txt files", theme = self.theme)
            self.grab_set()
            return
        try:
            with open(newfile,"w") as fh:
                for line in text:
                    fh.write(line)
        except:
            errorpopup = error(self,"Error saving file", theme = self.theme)
            self.grab_set()
            return
        
    def find(self,event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        self.addon = "+" + str(len(self.word)) + "c"
        start = "1.0"
        while start:
            start = self.textbox.search(self.word,index=start,nocase=1,stopindex="end")
            if start:
                end = start + self.addon
                self.find_results.append(start)
                self.textbox.tag_add("i",start,end)
                start = end
        if self.find_results:
            self.textbox.tag_config("i",background="Yellow")
            self.find_results_label.config(text="1/"+str(len(self.find_results)))
            self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
            self.textbox.tag_config("c",background="Orange")
            self.textbox.see(self.find_results[self.results_number])
            
    def find_up(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_down(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_reset(self,newfind=False):
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        if newfind == False:
            self.find_window.delete(0,"end")
        for tag in self.textbox.tag_names():
            self.textbox.tag_delete(tag)
        self.find_results_label.config(text="0/0")
        
    def cancel(self,event=None):
        self.destroy()


class terms_popup(tk.Toplevel):
    def __init__(self, C, text, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Tree Surgeon© - EULA - ")
        self.resizable(False,False)
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.has_agreed = False
        self.grid_columnconfigure(0,weight=1,uniform="y")
        self.grid_columnconfigure(1,weight=1,uniform="y")
        self.grid_rowconfigure(0,weight=1)
        self.textbox = working_text(self,
                                    wrap="word",
                                    theme = theme,
                                    use_entry_bg = False,
                                    override_bg = theme_treeview_bg(theme) if theme != "dark" else None)
        self.textbox.config(height=18)
        self.yscrollb = scrollbar(self,
                                  self.textbox.yview,
                                  "vertical",
                                  self.textbox)
        self.textbox.delete(1.0,"end")
        self.textbox.insert(1.0,text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=0,column=0,columnspan=2,padx=(20,0),pady=20,sticky="nswe")
        self.yscrollb.grid(row=0,column=2,padx=(0,20),pady=20,sticky="nswe")
        self.agree_button = button(self,text="Agree",
                                   underline = 0,
                                   style="EF.Std.TButton",
                                   command=self.agree)
        self.agree_button.grid(row=1,column=0,sticky="nswe",padx=(20,40),pady=20)
        self.disagree_button = button(self,text="Disagree",
                                      underline = 0,
                                      style="EF.Std.TButton",
                                      command=self.disagree)
        self.disagree_button.grid(row=1,column=1,sticky="nswe",padx=(40,20),pady=20)
        self.bind("<Escape>", self.disagree)
        self.bind("<A>", self.agree)
        self.bind("<a>", self.agree)
        self.bind("<D>", self.disagree)
        self.bind("<d>", self.disagree)
        center(self,700,650)
        self.deiconify()
        self.wait_window()
        
    def agree(self, event = None):
        self.has_agreed = True
        self.destroy()
        
    def disagree(self, event = None):
        self.destroy()


class helppopup(tk.Toplevel):
    def __init__(self, C, text, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1",bg = theme_bg(theme))
        self.withdraw()
        self.title(" Tree Surgeon© - Help")
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.word = ""
        self.findpos_start = "1.0"
        self.findpos_end = "1.0"
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.find_frame = frame(self, theme = theme)
        self.find_frame.grid(row=0,column=0,columnspan=3,sticky="nswe")
        self.find_icon = tk.PhotoImage(format="gif",data=find_icon)
        self.search_button = button(self.find_frame,
                                    text=" Find:",
                                    command=self.find)
        self.search_button.config(image=self.find_icon,compound="left")
        self.search_button.pack(side="left",fill="x")
        self.find_window = normal_entry(self.find_frame,font=BF, theme = theme)
        self.find_window.bind("<Return>",self.find)
        self.find_window.pack(side="left",fill="x",expand=True)
        self.find_reset_button = button(self.find_frame,text="X",command=self.find_reset)
        self.find_reset_button.pack(side="left",fill="x")
        self.find_results_label = label(self.find_frame,"0/0",BF, theme = theme)
        self.find_results_label.pack(side="left",fill="x")
        self.find_up_button = button(self.find_frame,text="▲",command=self.find_up)
        self.find_up_button.pack(side="left",fill="x")
        self.find_down_button = button(self.find_frame,text="▼",command=self.find_down)
        self.find_down_button.pack(side="left",fill="x")
        self.buttonframe = frame(self, theme = theme)
        self.buttonframe.grid(row=1,column=0,rowspan=2,sticky="nswe")
        
        self.basics = button(self.buttonframe,text="Program\nBasics",
                             style="EF.Std.TButton",
                             command=lambda:self.scrollto("programbasics"))
        self.basics.pack(side="top",padx=2,pady=2,fill="x")
        
        self.tutorials = button(self.buttonframe,text="Tips &\nTutorials",
                             style="EF.Std.TButton",
                             command=lambda:self.scrollto("tutorials"))
        self.tutorials.pack(side="top",padx=2,pady=2,fill="x")
        
        self.xlsx_files = button(self.buttonframe,text=".XLSX\nFiles",
                                 style="EF.Std.TButton",
                                 command=lambda:self.scrollto("xlsxfiles"))
        self.xlsx_files.pack(side="top",padx=2,pady=2,fill="x")
        
        self.treecomparehelp = button(self.buttonframe,text="Tree\nCompare",
                                      style="EF.Std.TButton",
                                      command=lambda:self.scrollto("treecompare"))
        self.treecomparehelp.pack(side="top",padx=2,pady=2,fill="x")
        
        self.menubar = button(self.buttonframe,text="Treeview\nMenu",
                              style="EF.Std.TButton",
                              command=lambda:self.scrollto("treeviewmenu"))
        self.menubar.pack(side="top",padx=2,pady=2,fill="x")
        
        self.managecolumns = button(self.buttonframe,text="Manage\nColumns",
                                    style="EF.Std.TButton",
                                    command=lambda:self.scrollto("managecolumns"))
        self.managecolumns.pack(side="top",padx=2,pady=2,fill="x")
        
        self.buttons = button(self.buttonframe,text="Treeview\nButtons",
                              style="EF.Std.TButton",
                              command=lambda:self.scrollto("treeviewbuttons"))
        self.buttons.pack(side="top",padx=2,pady=2,fill="x")
        
        self.functions = button(self.buttonframe,text="Treeview\nFunctions",
                                style="EF.Std.TButton",
                                command=lambda:self.scrollto("treeviewfunctions"))
        self.functions.pack(side="top",padx=2,pady=2,fill="x")
        
        self.tsrgnfiles = button(self.buttonframe,text=".JSON FILES",
                                style="EF.Std.TButton",
                                command=lambda:self.scrollto("tsrgnfiles"))
        self.tsrgnfiles.pack(side="top",padx=2,pady=2,fill="x")
        
        self.api = button(self.buttonframe,text="Using the\nAPI",
                          style="EF.Std.TButton",
                          command=lambda:self.scrollto("api"))
        self.api.pack(side="top",padx=2,pady=2,fill="x")
        
        self.textbox = working_text(self,
                                    font = ("Calibri",12),
                                    wrap="word",
                                    theme = theme,
                                    use_entry_bg = False,
                                    override_bg = theme_treeview_bg(theme) if theme != "dark" else None)
        self.yscrollb = scrollbar(self,
                                  self.textbox.yview,
                                  "vertical",
                                  self.textbox)
        self.textbox.delete(1.0,"end")
        self.textbox.insert(1.0,text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=1,column=1,sticky="nswe")
        self.yscrollb.grid(row=1,column=2,sticky="nswe")
        self.textbox.focus_set()
        self.bind("<Escape>",self.cancel)
        center(self,975,650)
        self.deiconify()
        self.wait_window()
        
    def scrollto(self,option):
        if option == "programbasics":
            self.textbox.see(self.textbox.search("    PROGRAM BASICS    ","1.0").split(".")[0] + ".0")
        elif option == "xlsxfiles":
            self.textbox.see(self.textbox.search("    XLSX FILES    ","1.0").split(".")[0] + ".0")
        elif option == "tutorials":
            self.textbox.see(self.textbox.search("    HELPFUL TIPS AND TUTORIALS    ","1.0").split(".")[0] + ".0")
        elif option == "treecompare":
            self.textbox.see(self.textbox.search("    TREE COMPARE    ","1.0").split(".")[0] + ".0")
        elif option == "treeviewmenu":
            self.textbox.see(self.textbox.search("    TREEVIEW MENUBAR    ","1.0").split(".")[0] + ".0")
        elif option == "managecolumns":
            self.textbox.see(self.textbox.search("    MANAGE COLUMNS    ","1.0").split(".")[0] + ".0")
        elif option == "treeviewbuttons":
            self.textbox.see(self.textbox.search("    TREEVIEW BUTTONS    ","1.0").split(".")[0] + ".0")
        elif option == "treeviewfunctions":
            self.textbox.see(self.textbox.search("    TREEVIEW FUNCTIONS    ","1.0").split(".")[0] + ".0")
        elif option == "tsrgnfiles":
            self.textbox.see(self.textbox.search("    JSON FILES    ","1.0").split(".")[0] + ".0")
        elif option == "api":
            self.textbox.see(self.textbox.search("    USING THE API    ","1.0").split(".")[0] + ".0")
            
    def find(self,event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        self.addon = "+" + str(len(self.word)) + "c"
        start = "1.0"
        while start:
            start = self.textbox.search(self.word,index=start,nocase=1,stopindex="end")
            if start:
                end = start + self.addon
                self.find_results.append(start)
                self.textbox.tag_add("i",start,end)
                start = end
        if self.find_results:
            self.textbox.tag_config("i",background="Yellow")
            self.find_results_label.config(text="1/"+str(len(self.find_results)))
            self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
            self.textbox.tag_config("c",background="Orange")
            self.textbox.see(self.find_results[self.results_number])
            
    def find_up(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_down(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_reset(self,newfind=False):
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        if newfind == False:
            self.find_window.delete(0,"end")
        for tag in self.textbox.tag_names():
            self.textbox.tag_delete(tag)
        self.find_results_label.config(text="0/0")
        
    def cancel(self,event=None):
        self.destroy()



