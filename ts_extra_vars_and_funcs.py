#Copyright © 2020 R. A. Gardner

import tkinter as tk
from tkinter import ttk, filedialog
import csv as csv_module
from collections import defaultdict, deque, Counter
import os
import datetime
import re
from itertools import islice, repeat, chain, cycle
import json
import pickle
import zlib
import lzma
from base64 import b32encode as b32e, b32decode as b32d
import csv as csv_module
from openpyxl import Workbook, load_workbook
import io
import datetime
from tksheet import Sheet
from math import floor
from ast import literal_eval
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from platform import system as get_os
from fastnumbers import isint, isintlike, isfloat, isreal

software_version_number = "2.64251"
software_version_full = "version: " + software_version_number
app_title = " Tree Surgeon©"
contact_email = "ragardner@protonmail.com"
website1 = "github.com/ragardner"
website2 = "ragardner.github.io/Tree-Surgeon"

def theme_button_fg(theme = "dark blue"):
    return "#2b2b2b"

def theme_bg(theme = "dark blue"):
    if theme == "light blue":
        return "#f0f0f0"
    elif theme == "light green":
        return "#f3f2f1"
    elif theme == "dark blue":
        return "#2f2f2f"
    elif theme == "dark green":
        return "#2f2f2f"

def theme_fg(theme = "dark blue"):
    if theme == "light blue":
        return "#2b2b2b"
    elif theme == "light green":
        return "#2b2b2b"
    elif theme == "dark blue":
        return "white"
    elif theme == "dark green":
        return "white"

def theme_entry_bg(theme = "dark blue"):
    if theme == "light blue":
        return "white"
    elif theme == "light green":
        return "white"
    elif theme == "dark blue":
        return "#3d3d3d"
    elif theme == "dark green":
        return "#3d3d3d"

def theme_entry_fg(theme = "dark blue"):
    if theme == "light blue":
        return "black"
    elif theme == "light green":
        return "#2b2b2b"
    elif theme == "dark blue":
        return "#f0f0f0"
    elif theme == "dark green":
        return "#f0f0f0"

def theme_entry_dbg(theme = "dark blue"):
    if theme == "light blue":
        return "#f0f0f0"
    elif theme == "light green":
        return "#edebe9"
    elif theme == "dark blue":
        return "#2f2f2f"
    elif theme == "dark green":
        return "#2f2f2f"

def theme_entry_dfg(theme = "dark blue"):
    if theme == "light blue":
        return "gray55"
    elif theme == "light green":
        return "#373737"
    elif theme == "dark blue":
        return "#9fa6ad"
    elif theme == "dark green":
        return "#9fa6ad"

def theme_entry_cursor(theme = "dark blue"):
    if theme == "light blue":
        return "black"
    elif theme == "light green":
        return "#2b2b2b"
    elif theme == "dark blue":
        return "#f0f0f0"
    elif theme == "dark green":
        return "#f0f0f0"

def theme_red_bg(theme = "dark blue"):
    if theme == "light blue":
        return "#ffeef0"
    elif theme == "light green":
        return "#ffeef0"
    elif theme == "dark blue":
        return "#6e3d3d"
    elif theme == "dark green":
        return "#6e3d3d"

def theme_red_fg(theme = "dark blue"):
    if theme == "light blue":
        return "black"
    elif theme == "light green":
        return "#373737"
    elif theme == "dark blue":
        return "white"
    elif theme == "dark green":
        return "white"

def theme_green_bg(theme = "dark blue"):
    if theme == "light blue":
        return "#e6ffed"
    elif theme == "light green":
        return "#e6ffed"
    elif theme == "dark blue":
        return "#3d6e3d"
    elif theme == "dark green":
        return "#3d6e3d"

def theme_green_fg(theme = "dark blue"):
    if theme == "light blue":
        return "black"
    elif theme == "light green":
        return "#373737"
    elif theme == "dark blue":
        return "white"
    elif theme == "dark green":
        return "white"

def theme_status_fg(theme = "dark blue"):
    if theme == "light blue":
        return "#191919"
    elif theme == "light green":
        return "#191919"
    elif theme == "dark blue":
        return "#f48224"
    elif theme == "dark green":
        return "white"
    
def theme_treeview_selected_bg(theme = "dark blue"):
    if theme == "light blue":
        return "#1a73e8"
    elif theme == "light green":
        return "#217346"
    elif theme == "dark blue":
        return "#1a73e8"
    elif theme == "dark green":
        return "#36897f"

def theme_treeview_selected_fg(theme = "dark blue"):
    if theme == "light blue":
        return "white"
    elif theme == "light green":
        return "white"
    elif theme == "dark blue":
        return "white"
    elif theme == "dark green":
        return "white"

def theme_treeview_fg(theme = "dark blue"):
    if theme == "dark blue":
        return "#ecf0f2"
    elif theme == "light green":
        return "#232323"
    elif theme == "light blue":
        return "#232323"
    elif theme == "dark green":
        return "#ecf0f2"

def theme_treeview_bg(theme = "dark blue"):
    if theme == "light blue":
        return "#fafafa"
    elif theme == "light green":
        return "#fafafa"
    elif theme == "dark blue":
        return "#3d3d3d"
    elif theme == "dark green":
        return "#3d3d3d"

def theme_treeview_heading_bg(theme = "dark blue"):
    if theme == "dark blue":
        return "#2d2d2d"
    elif theme == "light green":
        return "#f3f2f1"
    elif theme == "light blue":
        return "#f8f9fa"
    elif theme == "dark green":
        return "#2d2d2d"

def theme_treeview_heading_fg(theme = "dark blue"):
    if theme == "dark blue":
        return "#c2c9cf"
    elif theme == "light green":
        return "#373737"
    elif theme == "light blue":
        return "#1e3442"
    elif theme == "dark green":
        return "#c2c9cf"

def theme_treeview_relief(theme = "dark blue"):
    if theme == "dark blue":
        return "sunken"
    elif theme == "light green":
        return "raised"
    elif theme == "light blue":
        return "raised"
    elif theme == "dark green":
        return "sunken"

def fixed_map(option, style):
    return [elm for elm in style.map('Treeview', query_opt=option) if elm[:2] != ('!disabled', '!selected')]

def csv_delimiter_quotechar(data):
    d = Counter(m.group() for m in re.finditer(r"""\t|,|\t'|'\t|\t"|"\t|,'|',|,"|",""", data))
    if not d['\t'] and not d[',']:
        if isinstance(data, str):
            return '\t', '"'
        else:
            return None, None
    if d['\t'] >= d[',']:
        delimiter_ = "\t"
    elif d['\t'] < d[',']:
        delimiter_ = ","
    if d['\t"'] + d[',"'] + d['"\t'] + d['",'] >= d["\t'"] + d[",'"] + d["'\t"] + d["',"]:
        quotechar_ = '"'
    elif d['\t"'] + d[',"'] + d['"\t'] + d['",'] < d["\t'"] + d[",'"] + d["'\t"] + d["',"]:
        quotechar_ = "'"
    return delimiter_, quotechar_

colors = ('white', 'gray93', 'LightSkyBlue1',
          'light grey', 'antique white',
          'papaya whip', 'bisque', 'peach puff',
          'navajo white', 'NavajoWhite2',
          'wheat1', 'wheat2',
          'khaki', 'pale goldenrod',
          'gold', 'LightGoldenrod1', 'LightGoldenrod2',
          'goldenrod1', 'goldenrod2', 
          'LightYellow2', 'LightYellow3',
          'RosyBrown1', 'burlywood1',
          'LemonChiffon2', 'LemonChiffon3',
          'cornsilk2', 'ivory2',
          'sky blue', 'light sky blue', 
          'light blue', 'LightBlue1', 'powder blue', 'CadetBlue1',
          'pale turquoise', 'medium turquoise', 'turquoise',
          'medium aquamarine', 'aquamarine2', 
          'medium sea green', 'dark sea green', 'DarkSeaGreen2',
          'DarkOliveGreen2',
          'salmon1', 'coral1', 'pink', 'pink1', 'orchid1', 'plum1', 
          'LavenderBlush2', 'MistyRose2',
          'thistle', 'thistle1', 'thistle2', 'thistle3')

menu_kwargs = {'font': ("Calibri", 11, "normal"),
               'background': "#f2f2f2",
               'foreground': "gray2",
               'activebackground': "#91c9f7",
               'activeforeground': "black"}

openpyxl_thin_border = Border(left=Side(style='thin'),
                              right=Side(style='thin'), 
                              top=Side(style='thin'), 
                              bottom=Side(style='thin'))

def case_insensitive_replace(find_, repl, text):
    return re.sub("(?i)" + re.escape(find_), lambda m: repl, text)

def openpyxl_get_fill(color):
    try:
        return PatternFill(start_color=color,
                           end_color=color,
                           fill_type='solid')
    except:
        return PatternFill(start_color="white",
                           end_color="white",
                           fill_type='solid')
orange_fill = openpyxl_get_fill('FFA51E')
slate_fill  = openpyxl_get_fill('E1E1E1')
tan_fill = openpyxl_get_fill('EDEBE1')
green_add_fill = openpyxl_get_fill('E6FFED')
red_remove_fill = openpyxl_get_fill('FFEEF0')
openpyxl_left_align = Alignment(horizontal='left')
openpyxl_center_align = Alignment(horizontal='center')

def xl_column_string(n):
    s = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        s = chr(65 + remainder) + s
    return s

TF = ("Calibri",15,"bold")
BF = ("Calibri",11)
STSF = ("Calibri",11)
EF = ("Calibri",12)
EFB = ("Calibri",13,"bold")
ERR_ASK_FNT = ("Calibri",12,"bold")

def center(toplevel, desired_width = None, desired_height = None, move_left = False):
    toplevel.update_idletasks()
    w = toplevel.winfo_screenwidth() - 20
    h = toplevel.winfo_screenheight() - 100
    if desired_width is not None:
        if desired_width >= w:
            desired_width = w - 20
    if desired_height is not None:
        if desired_height >= h:
            desired_height = h - 20
    if desired_width and desired_height:
        size = (desired_width,desired_height)
    else:
        size = tuple(int(Q) for Q in toplevel.geometry().split("+")[0].split("x"))
    x = (w/2 - size[0]/2) if not move_left else (w/2 - size[0]/2) - 400
    y = h/2 - size[1]/2
    toplevel.geometry("%dx%d+%d+%d" % (size + (x, y)))

def get_platform_rc_binding():
    x = str(get_os())
    if x == "Darwin":
        return "<2>"
    else:
        return "<3>"

def xl_x(d, k):
    return bytes(b ^ k for b, k in zip(d, cycle(k)))

safe_builtins = set()
class RestrictedUnpickler(pickle.Unpickler):
    def find_class(self, module, name):
        if module == "builtins" and name in safe_builtins:
            return getattr(builtins, name)
        raise pickle.UnpicklingError(f"global {module}.{name} is forbidden")

def restricted_loads(s):
    return RestrictedUnpickler(io.BytesIO(s)).load()

class UK_Workday_Checker:
    def __init__(self):
        self.d = pickle.loads(lzma.decompress(b32d(("7U3XUWC2AAAAJZWWWRDAEABBAELAAAAAOQX6LI7AOCRQSY25ABAACDTXNZB4EZIGZMPAPWSBZ3VGKS2T7JYJ5RBUPQG2NI4TC77Q"
"VK57MP3M3B2NPGKSTYM6JREGXH637T253J7UXD2WGDLYM6T5663DJ42N324XLZGPMKZRHUHYA3EUFFQ64RGPVBSPRQNFKLNSHEV2"
"GBYIAGANUN675DBJDJ2HOBS2Q4CQSZKMF7WJNPIKQV7KVXBBYIVQTUOJXGWWO6ZGA5RZLREAR7O5YCMTY4NEHE3NCRA3HV6PHXNH"
"OCJZFJTGSLNRCBFFWR27HKSDESI5SL5EQASBI4JMGSJDCEOIUOXW4OGVIUJGR5J32WP4RPEB63FSO42RK6Y5CDVKZ7LGYRYGUTJZ"
"NWZWJRBQHCD3HIN2LK3E6FO3QSESIFAO6KGTHKXLL2FCHGRFMQDXAPLRDCMI53BUKCWLTGLL2RCPUBLFJKZBCCDYETWGW5O565Q3"
"ZWNSZTM267IUNECO3I7XRVPEF5QNMQMB27GFDGJAQOGBYVMR2Y32CFJYM4J2PBSVKZTYG34FLLL7BWOSJYKIGKAVYQFPVMMSVCEN"
"OIPEV2L4FTWLBSE4NX7Y5P3JW4VAXQGMR5TOR2BFPRPSOJIK7AXUJ2UJ24UH4GDJ5N7KZA2S7DGM7XK743OSRWCJSJD4TXS52NL3"
"7NQQGCJLFQTBJIAUKAZYDNEGGJ4TUBRYOY3YZOK3NHICBJOLULLW2TVL6G45IM3W42ERPCVWTWNYVGZL4KQYON627NZKEGY3HZYK"
"ZZNPGX22J2XPWXZII7HLSSWD3YQFJ7INBU4YXPU6YN3GK7DTJXMDVXC274VKPIPQ2S27XAKIEDZJIMVHHQTCS7TA723HENLWUCVP"
"DAWLNHZM5ZEXXGLJX3UYNBZBL6RQDGHI5S2LQB2PANCOP7EJZCDKM43MS56TCYPFM7HDMMCJMKOET3MJDNDD6S47L36UC4WJP6GY"
"KM2QVE6NRX5AX54NI2HTYXINMSYFMJC26SNX545ZAIA2ONHPX2NOQUWEE5HOULYQLKCMCZU2I2CD7QPWSXPCZ3OIKVRNVP4XZRN5"
"TVSTIZFFCPCUPOVZLF5PBRL2D5EQCMWVWDVQOA4BVCXSXJDXZOKWSZYHMCVKP5BZVRZ2XSNKGNG4C3Y6WSIZAEWWX7XMHK26QIWA"
"O5OEPAVSGPU5YMQCMB4PUKFDPOCLZBR5EPL6F5E3ARYFHS7BTG6ZDDB4HVVCQ6IVZMIMUWGTSXYMKPPQ6SWZ2FW5G6NC7VMU563P"
"TU2UHFFPZYF7MZ6DYXC6QTO35LXAL7XBXNNZGJ7WHRWGUI43QEE4GVNSLCOC5KZXMBMYJTJRR6RM4ZP7FJGOGRC7MPIGS65BHBPQ"
"5URJQOQ6XXDW3XM2HUWNS6KATZLIBDQDZEKR4C7HDYMWHOCTQTMZN52U3XUXXKKCIERZYKPN4MD4DXIGJ6N6CDAKIM37RYCNBCCM"
"XL57WGL2SG2ALJNJ2GRYFWQGOSSSAUY4HY7ABXJ3ZKZXNJBII4S6FV4J7IIGXUVJ4US3Q7SZM6U6WWLV5OHOMXQJRSZAHUKKRDO5"
"AZBLF6JOBVSNZVKIHSGCLQWR6ZIR5A3EZQ5HDDXTK2CBGJXDFTUAPVNOR4DSJBN2RUDPX2NDC5HLA5CV3FDFR55QMYXYW53ZMYGD"
"KRF3E25XPVPO3QEPJIWR3QX6QPJVXVKP73F53LD7KUOR52G7OBKTBEUTE4AQKD2E4LSYKEWH2MRIIPVRDLUHACXZJZ57RTLDV7GL"
"AU7ABK6W252NOY4XIP4ZS3HMM3E2AUYOL35ND2LEZSEJ433LLA3EMVKQMA3UAZVR3SVPPDIL3OTAKRP5BW3RC3GLHHTMKYBOX25J"
"2VLQRZEHSWDAN2ZHL3ISEIONFN5RBAL3NS2QPVZMQEBJXADCRL5ESQ2WPZGV6K5OVN4JNYZLLF743LMC2JON42664AS6ETUZS5EH"
"KRT3ZNA5N6LNMTIP6WONRUXJXHMFD3WQBVER3BKBTNBMEWKCWH3NVPJAXDUC3STKKKCK5VNUSQ4LWP3L7CWIB6WGL3ITNWV6AIP4"
"A4QJEKFTCM2K337WLFY77V3L3WN5GC4QWCYZZGFLFLXJCPIPJD4GQ5WT2UB3AWT436X5VSYVUZAYZQ3EBM3MZC72U3Q2LIYHCF6G"
"IQ2D5S45NHNGHVCHOFPQKLG3XO7KYFADKCS5OD55OFYGMEWX7VL2JYRYKH7HBOVVH5RTJ33T67LZI5YAOCY5TAUS34MOG36C4TTE"
"FVCB55HXZWOSI75RIOPIFMBR4QYW673WVVX7YM5FK73O5CK4FF3UYQWAOPD2FM7DPDSR7KATNLQTDXWYROHA6ZXUT6GKHSPCZD2E"
"DNSUJAENPLZNJNUAWAPA6R5IJXQKJXHQFQN6I7XFUSDPSHWTN4J6DDDWX5EXFPLN5X3HPWANVBL7TW2G7HZSVAIDSF64CV4F2W25"
"PHEIPQDTSESMUOJBDZ6XQRCBCXLQ527ZP5OGGAZXUUMZ2P4X72A5EX5BQACO4FDZKU6QERL56ZUI5OD4A4HFKHJS4YGIZEKQEIVF"
"LUC5S2GYJP3BLNWJ2EFQ4L4FYCY6GP2V76KP36374I3ODVAO5XNLSVSLHFVALPODGNZVEX4JGLONVBIYSXAAKUSZFQHO27VL7MYC"
"57CVVHZK77BKJMA6H3BT3QQHN3K3V3QTZ46MHP3SWBG4FJB3D5IS7744HHXYR5PI4EU7IICBDG3VEZX3KCHR5HYE7UDN73OTQDDU"
"NCO2J25HOZNEH2PNFESR7BMCB57HSYXRRCQZT3UJEZ4TH2CH3EHIEZUDXXYURE7G3W2URMZBQNQJRCOZSNR27U7ED3IV55DQ3CGM"
"AJ32OLLK3QFAAAC4WZ5KD62YGA5SJ6ILC35N72JXACWOYZBC4RTK7LNAAPFT5OGHEJEIG2AVIPILLRHZNY4AL46JQ3BJRXKZCJSM"
"JCUHOU7X2IWYYWBTECQESDJ7POYRRWAFACVQX2C3KAXN2I7KATLTORK4L44ZAMCYH3ZPC5E6KUDPB3FUMMSJYS3HNKEXACFETYN4"
"HKA73WGMAR32PM7YABDYYOK34MPBQYB6ZB7EAR6EUFC5XWAPFRI7MWDLMS6FSXFWIQKYEYDL6LX3FJAMFPXLLLW5524O57RRI6JG"
"EANBSWMQZMCPPXHAHIIKJP7USRERYDRUBJEF6YGZHFXIQRSJKQYNWGZYBAXJZJVYAT2QT22TR72G5ULPN5S5XDEDNBQLZOT7EG22"
"A7QW3KWXIHSFEVKDN5NT6QKZVPXLKVTEHN7O7YPZRL46NNYMGSCWRAWFZYLLUQGOBXJ7V7COBYOIYZUM2UK43EUJVEXKO6CZRPWU"
"NP5LCTATHAS2FJ6S4BQCRRRZYCQ6AEPVBBE2GKMLEXKBBHQQ7BD5LDUIJRVW2IAG64MV23BN7XF3XLG4C5IV5IK3KMY65QXBKWB2"
"LJWM5VNKDMAWVDBKWKZUKTVZ6DLOT2SU2S7JIOFS6BWR3ZK6TJIZ7UV2XYKJINN7XDONBPJ5LJ3OZTBKPQYO6NZPGITEQJM5OPUC"
"37WSRFELJOZDIVUC6PDHLRYR6WHA3KTAMFUCTCLCNAWLV7U6FPEXSRE75YAB2QZLQX475LVKML4R2HXW5AY5TMYUKL3YNNYEE2GU"
"K2U5W4DNY4XFXPHXNHSK22YRN6K756K4INPQ7F6TTAEVPZS5OQVA4N4HEMCCKFM5QMOZXCMXIIPCRJ5HH4ATQDKOQCIXCNQSRBLJ"
"CVBEMXXOIEYUR3JLQOOU4KMH65BLZEHDJBYGCI2XI624UYVIUER5S5QMV65ODG5MM7Q3ZBKYFGARHYY7ZKY6G55Z4KMFUAAAAC7R"
"XP7OF5EKLSYAAH7RFJHBAEAHPPUDJ2Y4IZ73AIAAAAAAARMVU===").encode())))
        
    def workday(self, date_str, region = "uk"):
        try:
            if isinstance(date_str, str):
                date_obj = datetime.date(int(date_str[:4]), int(date_str[5:7]), int(date_str[8:10]))
            else:
                date_obj = date_str
            if date_obj.weekday() in (5, 6):
                return False
            if region == "uk":
                v = self.d["uk"]
            elif region == "england":
                v = self.d["england"]
            elif region == "scotland":
                v = self.d["scotland"]
            elif region == "ni":
                v = self.d["ni"]
            elif region == "wales":
                v = self.d["wales"]
            if date_obj in v:
                return False
            return True
        except:
            return False


top_left_icon = (
"""R0lGODlhZABkAHAAACH5BAEAAPwALAAAAABkAGQAhwAAAAAAMwAAZgAAmQAAzAAA/wArAAArMwArZgArmQArzAAr/wBVAABVMwBVZgBVmQBVzABV/wCAAACAMwCAZgC
AmQCAzACA/wCqAACqMwCqZgCqmQCqzACq/wDVAADVMwDVZgDVmQDVzADV/wD/AAD/MwD/ZgD/mQD/zAD//zMAADMAMzMAZjMAmTMAzDMA/zMrADMrMzMrZjMrmTMrzD
Mr/zNVADNVMzNVZjNVmTNVzDNV/zOAADOAMzOAZjOAmTOAzDOA/zOqADOqMzOqZjOqmTOqzDOq/zPVADPVMzPVZjPVmTPVzDPV/zP/ADP/MzP/ZjP/mTP/zDP//2YAA
GYAM2YAZmYAmWYAzGYA/2YrAGYrM2YrZmYrmWYrzGYr/2ZVAGZVM2ZVZmZVmWZVzGZV/2aAAGaAM2aAZmaAmWaAzGaA/2aqAGaqM2aqZmaqmWaqzGaq/2bVAGbVM2bV
ZmbVmWbVzGbV/2b/AGb/M2b/Zmb/mWb/zGb//5kAAJkAM5kAZpkAmZkAzJkA/5krAJkrM5krZpkrmZkrzJkr/5lVAJlVM5lVZplVmZlVzJlV/5mAAJmAM5mAZpmAmZm
AzJmA/5mqAJmqM5mqZpmqmZmqzJmq/5nVAJnVM5nVZpnVmZnVzJnV/5n/AJn/M5n/Zpn/mZn/zJn//8wAAMwAM8wAZswAmcwAzMwA/8wrAMwrM8wrZswrmcwrzMwr/8
xVAMxVM8xVZsxVmcxVzMxV/8yAAMyAM8yAZsyAmcyAzMyA/8yqAMyqM8yqZsyqmcyqzMyq/8zVAMzVM8zVZszVmczVzMzV/8z/AMz/M8z/Zsz/mcz/zMz///8AAP8AM
/8AZv8Amf8AzP8A//8rAP8rM/8rZv8rmf8rzP8r//9VAP9VM/9VZv9Vmf9VzP9V//+AAP+AM/+AZv+Amf+AzP+A//+qAP+qM/+qZv+qmf+qzP+q///VAP/VM//VZv/V
mf/VzP/V////AP//M///Zv//mf//zP///wAAAAAAAAAAAAAAAAj/APcJHEiwoMGDCBMOVHbjRjKFECNKnEixosBomGzYuMHjzA2NFkOKHBkxWaYb+2yEOcPj4w0wDR2
SnEkzYqZkygZ+DJPMzUYwPYDa6HjGBsGHNZPOzLjxDJqPG9Gg6WGjKNWiLsNk2jfMAJp90JSKnYgmDJqNNqjyCNPSjEq0LlvKtaHsKhoYY/Mm9Jg2LUwzH632xYq2KO
GibiQx0AeWYM6cepMqC3O17w3AgTe2NFy1cEuhVZVptJFJ2TAGGsNEnqnMzQ2PgDfGdqlyrVyOH6nmDgyVdkuVPRqecWMA8mqJymDYgDn0ckO0uuHC/OscbecbuoWGa
e62o0ZMxyn2/2DLcbvu33Blc+xxO/3zn1SvclTJG+Vx4wlFWxXaGy5g70G9VJ11tGH3FlXA8SDUGWfsowwa+M10RiaZhPHaVgrpU55LMHEYV4GaObfdcy7FVhlfBr7G
V1VJubGddSAVxJhAJ6VlW38bBQjYbOkRqNFLAf4G4opYlUYTDG3gtpZuAz11gwHgDeMTf5alpxuIuaUHA21YCfljYSFqRCGGIRWlHg9v2YCGmVVdiOKON7wYQ2ppuYV
lenNqFl2BlwXoV5CjlekRGpgIaRVWnKU1YWfnbXTAjxu69yVUIQooqaRmBuiRRSe9pOZQ5KX5Gm6e/hbUqb0ppyRtksIQQwPbEf+206SSGqjWRxYxKBdltYGIoFW4Qa
eelcxthGUDo/VH6Z3uwWQmRWeBuWeYhs3HFnlC/VedgKzasOV1xvKG5keP2jDnh+kV1RJSEzF0qmbk9WabbGl5uNGI/FFp7r5xXYtmbC2NeMO3UAUIlJmsWoTJWZyBa
OFn2V1KW3ydSVrZhps5d6NG5wJVoG5ukSkRw5lIN3Ff8bbka4G2mdfQjQHzBipUHpm6EWrB/Yimn69VhMkZL1JcIFD5DqssXNMCey+axmLWHJiEnQvXtRqtOWEmYUGk
jBg1Oycr0hr1gGrYPZYHLqt3bmvlcgHeSZUbDlonskGiaVYxiPRSBgaBaKP/Feq0adf6I3nolV1xiDBlqFmHxn459bheHt2v5D7ibflaaEmNZXmbfZXfDUkyCheJqMW
5ZOUWN26dYDiy/DCvUH1LcNkwMQ0RTqtS+a3Uvzbeul+SA0llziVy1CVbcz5arq+Ej3oDeAdNxiC2F/+oamxff5TnwL7LFh9z4zke+Eaqeospo39p5FT0T7dlmcowGO
DSu9mr7B+6LodIGIJcdu+Sr2EiFWagN5CSvcxSpwKDYMgnLLS9xDZdMhOvWEIb6tXqM777jayIhqXPwI0gJWuD6FBUwYEhqzZUy0x8LASXzRgOXUOqyp5M1CZ90awz6
yMINCRhOhX6iGIwIJjQ/7DCowoKrnhvsZSsggUiiMFFiBsxiEYgZsNaMSBLsQKVl/gnF8I0bk8nex+OCqMx2higaZ0hoECMdZvA+cpj6XIgYDrkJYmNrirT4dOk4JTB
j2DiBvjJxM/YlkQ7mo1X0BEK5t6DHcq96H9lawsYJ2W/ILHFWT0wyFnWA5qhVal5zCkYcO5mtJ/M5znMmV1m+uY/9VVqMGFQ4z58gC1IPumOGnsYn2pHKuvAbjtX/Jb
/OmhLyRHukVRSzUCScZ7X1Es5yisfEtMWmzsSqHyaS5ei9MWWF96LfFcMUyYFogwGORNNYZCf/GDQgPiByXJk7J7snvQRdsKgdO6RmRsNx/+A7alqVzpRk508woAtqR
MGBxiYklAkLOD1yAaP4l4/vaVQ5cgPdTsh5ispOtF9kUeNokEDehoy0XVetIF1JNGXJlpPZMUgnBdtp0IZmKfyjVQ6cLnonNgJLzUVJBOw6UsM1GnC+M1Jb1ARmCtn1
RCYnHGnA9tdP8MJzZvF4AZTVakpraOcqj6nS840jiR+xDMbHHRgBihXCj/CuHqhJzobmVMDyEU+hbJ0S1XdnuEgOlNKUUVvuiGgGxjWQBNC1FwGQGiYbkZR5Rw1fPPT
iEW1hxpVqUqdMj3jAQjmWNSc657ncuaLGoqfaBUqjxtRZz8TKz+XTDRPyzvjj/jCU7n/OtZc7bRBbnNrz3Ze8YzrNBdqTipKejLgkh+J0GvOwp4NGVW4Mw0SXQdWULo
qJ3MMPGxBhSu7dsrUnp/laUW5K7W6upUuBckIQz/C2jmtE1lbWuQTtectl070AIkNprFm99LP+jZPDVCBUb1KvnLN6aWpUdBH5rYPN8AGrQAYqj2flLwgfstPVNGpZK
OqAgP0F68NAIANSkq+xD6xujDoMF9FvNm0ctesMalXyAzyYB4MFaLtZS0MAGBhC2MniA0BgGwbsFMLJzbF8QviZpN8Yw/XV1UHEHD8bqBiDycWqvt1pW7GSRBzqkfCL
d4sfm8QgyCqYLM2OHOJh1rmASM0/8lKTmxDPBxEA6wgsSpgQIcPwGMf/698v+Wub8wjyzXyB00S1nFiW9JjH+9Ox1LeQp3LrIJEo2XJcDZqfoMYAxVjZap4RcvsDjUM
hDSlTVVJK0KftCXtGUALFhawSwxgZSNbuM3xo/NzeNzmMuNX020OroBanVpim01FCskEhPQTJ1rrkczmgoGkzwjZK8PgztKO9Ztd/KMjG5nNiYV1klUN1/+1JTdOiVF
EWtOcF7GJo9yDgRYiatY33zqI4k5yi1NE5zrnOttv3rf3RBerX0FDGRFSyKxkGFfh7mtLkg4iZVz16jq/OuJlVvUZqxNEjHv4BvlOMbK4BxcGfecT5P+8S6FLMhxBup
a+PFVOxKFpbYp3VdKVTmyUW0u+HkeV0xG9rULR2YNyGoRdI9mkU+sLbxiEwcz1hnWHOZ3vWstPxC8BuJm3rtsfBbqHNkD6WJi4ORto4emu0vqjYiDujMdvs/XUAtvxL
eUUu8RVZLyVSlajj9EgbDZ7a4jWOX1bV+X711GVe49xnU2/Ma0o2/lgZPSxjGRQ5qM2cEMmyidpxd862m2X85HFrQUqI3lOPslSD1DUg0is/gYzCo9A+NKggWjCAGJ4
yRbMDnCsWjgMUn8q1U3P6cGUs5D6+ETgBSl7ilBIEmjXQpplW2bP41XeMQADklUAAK2UjExTQcv/QLLWfIVlfwU2kLQkMlFk4G995mY2ADEUcgZMkL/8I8kErM8OAzE
oI7GSBgaV1mMzBwBigH/4FwaS9nQC9gnD0HG8J23Sd20phoD4NwwDA3xUFg3KIGBPx3hSBwAJZ4Gr4X5otwX7QA/yhn2lV3UkiH+SIGlboH1OlxxaB2udBgCS8IIJqH
1zh2Sw9nTWpwIjyIN6wQgWpnhgwHYDI3fSJ2AGYIT4h28gh28dl4SdpgJSWH4PeIW7N3cDs3sCBgBbiH+dlm2kt4Kxdn9leBz6IGDmonhltgWmR4ZtWH6bpwURt38r2
GHzd4eyRwxAhoNoZ3oqsHKAOBYPOHfi1nlB/6SDiRgeeZh9lAhrhshgkSgWgggDSyhtvKd4AOA5magXXTiD2NdoMDAJoxgZyfF0W9CCS4iDkLiKerGCrph2dIeJtEgT
KgB8T+d+OBg/B7iLY/GB2od2VihgxKiJZ3d2M+iIbaaFy5gUySBv6ed0qLhj05gUFeJ0QaiHFEhporiNIwF8W3CCVthjdkiOIxGEYBCAX3iKKRYD7CgS0OeEcXiF9wY
AuliPCQF8S/iOsLaHuCZgKOiPE6F/GSh36ectimd9IoiQErEI5yhv7veDISdglSaRWgMG7keHKyh3GHeKPMaRCsEI/CeE2Chp0YZksTaMJlkQrYiDsVhm2td2hDj3iK
UWkwQRhALGh1LXY1KmkQbAff1Yj5KACZOACesnCZOwfk+ZCU4pCVOJBpOABlSZBpIAk5EREAA7"""
)

sort_icon = (
"""R0lGODlhEwATAHAAACH5BAEAAPwALAAAAAATABMAhwAAAAAAMwAAZgAAmQAAzAAA/wArAAArMwArZgArmQArzAAr/wBVAABVMwBVZgBVmQBVzABV/wCAAACAMwCAZgCAmQCAzACA/wCqAACqMwCqZgCqmQCqzACq/wDVAADVM
wDVZgDVmQDVzADV/wD/AAD/MwD/ZgD/mQD/zAD//zMAADMAMzMAZjMAmTMAzDMA/zMrADMrMzMrZjMrmTMrzDMr/zNVADNVMzNVZjNVmTNVzDNV/zOAADOAMzOAZjOAmTOAzDOA/zOqADOqMzOqZjOqmTOqzDOq/zPVADPVMz
PVZjPVmTPVzDPV/zP/ADP/MzP/ZjP/mTP/zDP//2YAAGYAM2YAZmYAmWYAzGYA/2YrAGYrM2YrZmYrmWYrzGYr/2ZVAGZVM2ZVZmZVmWZVzGZV/2aAAGaAM2aAZmaAmWaAzGaA/2aqAGaqM2aqZmaqmWaqzGaq/2bVAGbVM2b
VZmbVmWbVzGbV/2b/AGb/M2b/Zmb/mWb/zGb//5kAAJkAM5kAZpkAmZkAzJkA/5krAJkrM5krZpkrmZkrzJkr/5lVAJlVM5lVZplVmZlVzJlV/5mAAJmAM5mAZpmAmZmAzJmA/5mqAJmqM5mqZpmqmZmqzJmq/5nVAJnVM5nV
ZpnVmZnVzJnV/5n/AJn/M5n/Zpn/mZn/zJn//8wAAMwAM8wAZswAmcwAzMwA/8wrAMwrM8wrZswrmcwrzMwr/8xVAMxVM8xVZsxVmcxVzMxV/8yAAMyAM8yAZsyAmcyAzMyA/8yqAMyqM8yqZsyqmcyqzMyq/8zVAMzVM8zVZ
szVmczVzMzV/8z/AMz/M8z/Zsz/mcz/zMz///8AAP8AM/8AZv8Amf8AzP8A//8rAP8rM/8rZv8rmf8rzP8r//9VAP9VM/9VZv9Vmf9VzP9V//+AAP+AM/+AZv+Amf+AzP+A//+qAP+qM/+qZv+qmf+qzP+q///VAP/VM//VZv
/Vmf/VzP/V////AP//M///Zv//mf//zP///wAAAAAAAAAAAAAAAAhTAPcJHEiwoEGBahIaVHMQoUKCCRkejAgxosSCFB1avDgw40aLGB92FBmS476MC0meVDnSJMqSMCeqfFnR5cyPD2mu3NhS5kudGmNOTGmyodF9AQEAOw=="""
)

tree_icon = (
"""R0lGODlhEwATAHAAACH5BAEAAPwALAAAAAATABMAhwAAAAAAMwAAZgAAmQAAzAAA/wArAAArMwArZgArmQArzAAr/wBVAABVMwBVZgBVmQBVzABV/wCAAACAMwCAZgCAmQCAzACA/wCqAACqMwCqZgCqmQCq
zACq/wDVAADVMwDVZgDVmQDVzADV/wD/AAD/MwD/ZgD/mQD/zAD//zMAADMAMzMAZjMAmTMAzDMA/zMrADMrMzMrZjMrmTMrzDMr/zNVADNVMzNVZjNVmTNVzDNV/zOAADOAMzOAZjOAmTOAzDOA/zOqADOq
MzOqZjOqmTOqzDOq/zPVADPVMzPVZjPVmTPVzDPV/zP/ADP/MzP/ZjP/mTP/zDP//2YAAGYAM2YAZmYAmWYAzGYA/2YrAGYrM2YrZmYrmWYrzGYr/2ZVAGZVM2ZVZmZVmWZVzGZV/2aAAGaAM2aAZmaAmWaA
zGaA/2aqAGaqM2aqZmaqmWaqzGaq/2bVAGbVM2bVZmbVmWbVzGbV/2b/AGb/M2b/Zmb/mWb/zGb//5kAAJkAM5kAZpkAmZkAzJkA/5krAJkrM5krZpkrmZkrzJkr/5lVAJlVM5lVZplVmZlVzJlV/5mAAJmA
M5mAZpmAmZmAzJmA/5mqAJmqM5mqZpmqmZmqzJmq/5nVAJnVM5nVZpnVmZnVzJnV/5n/AJn/M5n/Zpn/mZn/zJn//8wAAMwAM8wAZswAmcwAzMwA/8wrAMwrM8wrZswrmcwrzMwr/8xVAMxVM8xVZsxVmcxV
zMxV/8yAAMyAM8yAZsyAmcyAzMyA/8yqAMyqM8yqZsyqmcyqzMyq/8zVAMzVM8zVZszVmczVzMzV/8z/AMz/M8z/Zsz/mcz/zMz///8AAP8AM/8AZv8Amf8AzP8A//8rAP8rM/8rZv8rmf8rzP8r//9VAP9V
M/9VZv9Vmf9VzP9V//+AAP+AM/+AZv+Amf+AzP+A//+qAP+qM/+qZv+qmf+qzP+q///VAP/VM//VZv/Vmf/VzP/V////AP//M///Zv//mf//zP///wAAAAAAAAAAAAAAAAiHAPcJHEiwoEGDQzQMOXgwAy+H
QyQtZLgv4RAiD3kpVHhQIa8PRBxmIDKEV8iCJUc+JKlEZEkNAzdazAhRQxKFE/fxWplxp8OQOwtCFJnRYwaDQHmlVMpLIpGDSYBqoHmS4dGlT3NS3Bcp6NaCTXd9BVtL0liCkniJPStwUVO2OnlFssUL7teAADs="""
)

find_icon = (
"""R0lGODlhEwATAHAAACH5BAEAAPwALAAAAAATABMAhwAAAAAAMwAAZgAAmQAAzAAA/wArAAArMwArZgArmQArzAAr/wBVAABVMwBVZgBVmQBVzABV/wCAAACAMwCAZgCAmQCAzACA/wCqAACqMwCqZgCqmQCqzACq/wDVAADVMw
DVZgDVmQDVzADV/wD/AAD/MwD/ZgD/mQD/zAD//zMAADMAMzMAZjMAmTMAzDMA/zMrADMrMzMrZjMrmTMrzDMr/zNVADNVMzNVZjNVmTNVzDNV/zOAADOAMzOAZjOAmTOAzDOA/zOqADOqMzOqZjOqmTOqzDOq/zPVADPVMzPV
ZjPVmTPVzDPV/zP/ADP/MzP/ZjP/mTP/zDP//2YAAGYAM2YAZmYAmWYAzGYA/2YrAGYrM2YrZmYrmWYrzGYr/2ZVAGZVM2ZVZmZVmWZVzGZV/2aAAGaAM2aAZmaAmWaAzGaA/2aqAGaqM2aqZmaqmWaqzGaq/2bVAGbVM2bVZm
bVmWbVzGbV/2b/AGb/M2b/Zmb/mWb/zGb//5kAAJkAM5kAZpkAmZkAzJkA/5krAJkrM5krZpkrmZkrzJkr/5lVAJlVM5lVZplVmZlVzJlV/5mAAJmAM5mAZpmAmZmAzJmA/5mqAJmqM5mqZpmqmZmqzJmq/5nVAJnVM5nVZpnV
mZnVzJnV/5n/AJn/M5n/Zpn/mZn/zJn//8wAAMwAM8wAZswAmcwAzMwA/8wrAMwrM8wrZswrmcwrzMwr/8xVAMxVM8xVZsxVmcxVzMxV/8yAAMyAM8yAZsyAmcyAzMyA/8yqAMyqM8yqZsyqmcyqzMyq/8zVAMzVM8zVZszVmc
zVzMzV/8z/AMz/M8z/Zsz/mcz/zMz///8AAP8AM/8AZv8Amf8AzP8A//8rAP8rM/8rZv8rmf8rzP8r//9VAP9VM/9VZv9Vmf9VzP9V//+AAP+AM/+AZv+Amf+AzP+A//+qAP+qM/+qZv+qmf+qzP+q///VAP/VM//VZv/Vmf/V
zP/V////AP//M///Zv//mf//zP///wAAAAAAAAAAAAAAAAhaAPcJHEiwoMGDAtMoTIOwoEJlEJUpbJgwYkSGDdNYVLYPIsaDGjd2/GgwpMWRGUV6zGjyIsWFFiciXLRo4UKKNBdRdEhzJ8GbPgXSJLkzTc+g+4AilYm0KcGAADs="""
)

unchecked_icon = (
"""R0lGODlhEwATAHAAACH5BAEAAPwALAAAAAATABMAhwAAAAAAMwAAZgAAmQAAzAAA/wArAAArMwArZgArmQArzAAr/wBVAABVMwBVZgBVmQBVzABV/wCAAACAMwCA
ZgCAmQCAzACA/wCqAACqMwCqZgCqmQCqzACq/wDVAADVMwDVZgDVmQDVzADV/wD/AAD/MwD/ZgD/mQD/zAD//zMAADMAMzMAZjMAmTMAzDMA/zMrADMrMzMrZjMr
mTMrzDMr/zNVADNVMzNVZjNVmTNVzDNV/zOAADOAMzOAZjOAmTOAzDOA/zOqADOqMzOqZjOqmTOqzDOq/zPVADPVMzPVZjPVmTPVzDPV/zP/ADP/MzP/ZjP/mTP/
zDP//2YAAGYAM2YAZmYAmWYAzGYA/2YrAGYrM2YrZmYrmWYrzGYr/2ZVAGZVM2ZVZmZVmWZVzGZV/2aAAGaAM2aAZmaAmWaAzGaA/2aqAGaqM2aqZmaqmWaqzGaq
/2bVAGbVM2bVZmbVmWbVzGbV/2b/AGb/M2b/Zmb/mWb/zGb//5kAAJkAM5kAZpkAmZkAzJkA/5krAJkrM5krZpkrmZkrzJkr/5lVAJlVM5lVZplVmZlVzJlV/5mA
AJmAM5mAZpmAmZmAzJmA/5mqAJmqM5mqZpmqmZmqzJmq/5nVAJnVM5nVZpnVmZnVzJnV/5n/AJn/M5n/Zpn/mZn/zJn//8wAAMwAM8wAZswAmcwAzMwA/8wrAMwr
M8wrZswrmcwrzMwr/8xVAMxVM8xVZsxVmcxVzMxV/8yAAMyAM8yAZsyAmcyAzMyA/8yqAMyqM8yqZsyqmcyqzMyq/8zVAMzVM8zVZszVmczVzMzV/8z/AMz/M8z/
Zsz/mcz/zMz///8AAP8AM/8AZv8Amf8AzP8A//8rAP8rM/8rZv8rmf8rzP8r//9VAP9VM/9VZv9Vmf9VzP9V//+AAP+AM/+AZv+Amf+AzP+A//+qAP+qM/+qZv+q
mf+qzP+q///VAP/VM//VZv/Vmf/VzP/V////AP//M///Zv//mf//zP///wAAAAAAAAAAAAAAAAg/AGMIHEiwYMF9CBMqXIhQIMOHCh1ChChxIsOKFiPGyHhxI0eN
H0GGbOhxJMaPJzmmzLjSYsuJLymWDGmwZs2AADs="""
)

checked_icon = (
"""R0lGODlhEwATAHAAACH5BAEAAPwALAAAAAATABMAhwAAAAAAMwAAZgAAmQAAzAAA/wArAAArMwArZgArmQArzAAr/wBVAABVMwBVZgBVmQBVzABV/wCAAACAMwCAZgCAmQCA
zACA/wCqAACqMwCqZgCqmQCqzACq/wDVAADVMwDVZgDVmQDVzADV/wD/AAD/MwD/ZgD/mQD/zAD//zMAADMAMzMAZjMAmTMAzDMA/zMrADMrMzMrZjMrmTMrzDMr/zNVADNV
MzNVZjNVmTNVzDNV/zOAADOAMzOAZjOAmTOAzDOA/zOqADOqMzOqZjOqmTOqzDOq/zPVADPVMzPVZjPVmTPVzDPV/zP/ADP/MzP/ZjP/mTP/zDP//2YAAGYAM2YAZmYAmWYA
zGYA/2YrAGYrM2YrZmYrmWYrzGYr/2ZVAGZVM2ZVZmZVmWZVzGZV/2aAAGaAM2aAZmaAmWaAzGaA/2aqAGaqM2aqZmaqmWaqzGaq/2bVAGbVM2bVZmbVmWbVzGbV/2b/AGb/
M2b/Zmb/mWb/zGb//5kAAJkAM5kAZpkAmZkAzJkA/5krAJkrM5krZpkrmZkrzJkr/5lVAJlVM5lVZplVmZlVzJlV/5mAAJmAM5mAZpmAmZmAzJmA/5mqAJmqM5mqZpmqmZmq
zJmq/5nVAJnVM5nVZpnVmZnVzJnV/5n/AJn/M5n/Zpn/mZn/zJn//8wAAMwAM8wAZswAmcwAzMwA/8wrAMwrM8wrZswrmcwrzMwr/8xVAMxVM8xVZsxVmcxVzMxV/8yAAMyA
M8yAZsyAmcyAzMyA/8yqAMyqM8yqZsyqmcyqzMyq/8zVAMzVM8zVZszVmczVzMzV/8z/AMz/M8z/Zsz/mcz/zMz///8AAP8AM/8AZv8Amf8AzP8A//8rAP8rM/8rZv8rmf8r
zP8r//9VAP9VM/9VZv9Vmf9VzP9V//+AAP+AM/+AZv+Amf+AzP+A//+qAP+qM/+qZv+qmf+qzP+q///VAP/VM//VZv/Vmf/VzP/V////AP//M///Zv//mf//zP///wAAAAAA
AAAAAAAAAAhVAGMIHEiwYMF9CBMqXIhQ4D6HDBU6nAgx4sCHMRpWlAix4kWOGTWCTPhR5MiSJkFuTDmS4UqMKFnCnMmxJc2UMUt2fEmTYkiLFDFGrMlzqEaDSAsGBAA7"""
)
contact_info = f" Copyright © 2020 R. A. Gardner.\n {contact_email}\n {website1}\n {website2}"

about_basic = (
"""This program is for management of hierarchy based master data which is stored in table
format. It was written in Python and utilizes the following libraries:
 - fastnumbers
 - lxml
 - openpyxl
 - cython
 - pyinstaller"""
)

help_progbasics = (
"""
       _________________    PROGRAM BASICS    _________________

This program is for management of hierarchy based master data which is stored in table
format. Supported file formats are:
 - .xlsx, .xls, .xlsm
 - .json Javascript object notation where the full table is under the key 'records'
 - .csv/.tsv (comma or tab delimited)

Any sheets opened with Tree Surgeon must contain a single header row at the top of the
sheet.

Additional settings and data such as the changelog, formatting, formulas and column types
can be saved with the formats .xlsx and .json.

Any changes made when using this program will not be saved unless you choose to do so. This
can be done by going to the main menubar and choosing a save option under the 'File' menu
while in the Treeview.

Sheets must have an ID column and atleast one parent column, it does not matter in which
order. e.g.

    ID     Parent    Detail
    ID1    Par1      Detail 1
    ID2    Par2      Detail 2
    
Sheet can have multiple parent columns (hierarchies) and multiple detail columns but must
have only one ID column. In the ID column each ID must be unique else they will be renamed
with '_DUPLICATE_'.

Sheets can have an unlimited number of parent columns (hierarchies) and an unlimited
number of detail columns.

The columns can be in any order and multiple columns of the same type can be separated by
other types of columns.

If the headers are not unique they will be renamed with a duplicate number. Any missing
headers will have names created for them.

Header names are not case sensitive.

There is no limit to the number of characters allowed for headers, details or ID names. Any
ID and header names with spaces, newlines or carriage returns will have them removed.
Details are exempt from this rule.

Any mistakes in the sheet such as infinite loops of children, IDs appearing in a parent
column but not in the ID column and duplications will be corrected upon creating the tree.

The corrections will not be made to the original sheet unless you choose to save the sheet.
Such corrections will appear as warnings when you first view the treeview window.

When first viewing the treeview window there are two separate panels, the one on the left
is the tree panel, where you can expand and collapse IDs and their children and manage the
hierarchies. The panel on the right is the sheet, there are some functions available on
this panel as well.

If an ID has no parents or children in any hierarchy it will be placed in the first
hierarchy (in order of the columns). The same will happen if you Undo a change made to
anything but details.

The code used for displaying sheets can be found here: github.com/ragardner/tksheet


       _________________    HELPFUL TIPS AND TUTORIALS    _________________

- Setting only uk, england, wales, scotland, northern ireland working days in date detail
  validation

In column manager you can right click on a column to change its type or create a new date
detail column. Please note that changing a columns type might delete some invalidated cell
contents. After getting a date detail column you can then edit its validation by right
clicking on the column or double left clicking on its validation cell. To set up working
day date validation you then type in one of the following options (ni stands for northern
ireland):
        - only uk working days
        - only england working days
        - only scotland working days
        - only wales working days
        - only ni working days
        
This should then prevent any non working dates for the chosen region being entered into
that column, this includes dates which fall on Saturdays or Sundays.

- Date column formulas and conditional formatting

When entering formulas and conditional formatting in Date Detail columns, use forward
slash dates e.g. DD/MM/YYYY. This is because hyphens will be interpreted as subtractions.
If you want to enter a specific date, for current date use the letters: cd

- Difference between treeview panel Shift + Click and Ctrl + Click

In the treeview panel when selecting multiple IDs there is a difference in the behavior
of Shift and Ctrl click. Shift + Click will select everything between the highest /
lowest selected ID, including all children. Whereas Ctrl + Click will only select the
IDs you click on.

- Moving IDs between hierarchies

To move an ID to another hierarchy or add an ID to another hierarchy you can right
click on the ID in the treeview panel and go to Cut or Copy and then either Cut ID or
Copy ID. Then using the dropdown box at the top of the treeview panel, labeled "Hierarchy"
select the hierarchy you would like to move / add the ID to. Then go to the position
or ID where you would like to place the Cut / Copied ID and right click and select a
paste option.

To move multiple IDs in one go you can use Shift + Left Click or Ctrl + Left Click to
select multiple IDs then use Ctrl + X (Cut) or Ctrl + C (Copy) or Right Click on one of
the selected IDs.

- Copying only a specific hierarchy or specific columns to the clipboard

To copy a subset of columns to the clipboard go to Options -> Show/Hide columns and
then select the hierarchies / columns to copy.

Then go to Edit -> Copy to clipboard and select an option. This will copy the all rows
with only your chosen columns.

To only copy a specific subset of IDs use Shift + Left Click or Ctrl + Left Click to
multi-select and Right Click on a selected ID and go to Copy -> Clipboard IDs row or
Clipboard IDs + childrens rows.

Alternatively you can use Export -> Export flattened sheet and Right Click on the
columns in the resulting popup window and select Delete to remove unwanted columns;
use the hierarchy selector on the left of the popup to change the output to a different
hierarchy.

- Deleting IDs

When using Delete on an ID in the sheet panel or Delete in all hierarchies in the
treeview panel it will Delete an ID completely; across all hierarchies.

When using any other delete option it will only delete an ID in the currently selected
hierarchy. However, if that ID is the last appearance of the ID across all hierarchies then
it will completely delete it, just like with Delete in all hierarchies.

- Deleting a column

To delete a column, not just its cells, go to Format -> Manage columns and right click
on the column you wish to delete and select Delete -> Delete column. Note you cannot
delete a hierarchy if it is the only hierarchy in the sheet and you cannot a delete a
hierarchy if you are currently viewing it.

- Adding a text formula to a column

To add a formula to a text column, go to Format -> Manage columns, right click on the
column you would like to have set by a formula and select Edit -> Edit Formula. As an
example, a formula to combine the values of two columns with a space in between and an
additional value would be:

c1+" "+c2+" my additional text"

c1 and c2 denote the column numbers, or positions, c1 being the first column.

- Adding multiple new rows

To add multiple new rows you can use Import -> Merge sheets and then either opening a file
or using the clipboard or just using the sheet display in the popup to paste / insert new
rows. Right clicking in the header or index will result in a popup box where you can insert
a new row or column. You can use Ctrl + V to paste data in, as long as it's in the form of
tab delimited text.

- Getting all information on an ID

An easy way to get an IDs complete information within the sheet, including parents and
children across all hierarchies and all details is to select an ID in the treeview or
sheet panel and then go to View -> IDs details."""
)

help_xlsx_files = (
"""
       _________________    XLSX FILES    _________________

The default save format for Tree Surgeon is .xlsx.

When saving .xlsx files you can also save Tree Surgeon data to keep your changelog, row
heights, column widths, formulas, formatting, validation, treeview ID order and more.

When loading a file saved with program data the sheet and changelog in the program data,
not the visible sheet, will take precedent. This means any edits in the viewable sheet
will not be loaded into Tree Surgeon.

To disable saving with program data go to Options -> XLSX save options -> Save xlsx with
program data.

If choosing to save program data any sheet named "Treesurgeon Data" will be overwritten
when saving a workbook.

You can also save a viewable changelog sheet. Sheets with "Changelog" in their name will be
overwritten when this option is chosen.

When saving .xlsx files you can also save the flattened format of the currently viewed
hierarchy any sheets with "flattened" in their name will be overwritten. If viewing all
hierarchies when saving then the first hierarchy will be saved.

When comparing or merging if the workbook contains program data then it will take
precedent, else a sheet will need to be selected to load data."""
)

help_treecompare = (
"""
       _________________    TREE COMPARE    _________________

Accessible from the "File" menu, this window allows comparison of trees and sheets. You can
open files using the "Open file" button which will open a file dialog. Once open the file
name or path will display next to this button and if the file is an excel file and was
opened from the file dialog then you will have to select a sheet from the drop down box
next to "Load sheet". Once you have done that you can select your sheets ID column and
parent column numbers and do the same with the 2nd panel on the right. After you are happy
with your selections click the "Create Report" button to compare. A report will be
generated and you have the option to save it as a .xlsx file which will have 2 sheets.

You can mix different file types when comparing."""
)

help_treemenu = (
"""
       _________________    TREEVIEW MENUBAR    _________________

Under the File Menu there are the following options:
- Open: Opens an excel or .csv file.
- Compare sheets: This option takes you to a tree comparison window. For more information
  click the help section "Tree Compare".
- Create new: Create a new sheet from scratch.
- Save: Options are Save (Ctrl + S), Save as, Save as with username-date- time (saves with
  the users login name, current date and time added to end of the filename) and Save new
  version (adds one to any detected file of the same name found in the chosen folder).
- Quit: Quits the program.

Under the Edit Menu there are the following options:
- Find and replace(Ctrl + F / Ctrl + H), pressing Ctrl + G when this window has focus with
  the find in sheet panel showing will use the windows search results, not the search
  results in the main windows dropdown boxes.
- Find next (Ctrl + G) goes to the next appropriate search result.
- Undo (Ctrl + Z): you get 50 undos. Although the changelog can be saved with any filetype
  other than .csv the changes cannot be undone across saves.
- Copy sheet to clipboard copies the displayed sheet to your computers clipboard to be
  pasted as a string, copying this way is the ONLY way to get just the displayed columns
  in the sheet display, saving will always save all columns, regardless of which are being
  displayed. To choose to display certain columns go to Options -> Choose sheet columns.
  Copy as json will follow the json format you have selected under Options.
- Clear copied/cut clears the copied/cut ID
- Clear panel selections deselects both the treeview and sheet selections.
- Clear all tagged IDs clears all tagged IDs and the associated drop-down  box. This is not
  an Undo-able action.

Please note that when you undo a change not related to details such as copying or deleting
an ID any IDs without parents and children in any hierarchy will be placed into the FIRST
hierarchy.

Under the Format Menu there are the following options:
- Manage columns is the only way to add, delete, cut and paste columns. You can add new
  detail or hierarchy columns, delete detail or hierarchy columns and change their order.
  You can do all this by right clicking in the popup, to change the column order quickly
  you can drag and drop column headers. When using drag and drop you can use your
  mousewheel to scroll down if the desired drop location is further down.
- Sort sheet gives you two options for sorting the sheet:
   - Sort by tree: This button sorts the sheet in the order that the IDs occur in the tree.
   - Sort by column: Using this button and the two drop-down boxes to its right you can
     sort the sheet using a basic natrual sorting order, numbers taking priority.
- Autosort treeview IDs re-sorts the order of all treeview IDs and children, this is on by
  default but if you disable it the order of IDs will not longer be automatically
  alphanumerically maintained. You can manually set your own treeview IDs order by uncheck-
  ing this option and then pressing right click and holding over the desired ID and then
  dragging it to the desired location.
- Show treeview levels makes the treeview put a number on the left hand side of every ID
  in the treeview which represents the depth of the ID.
- Date format switches the date format for the program, it will try to change formats for
  formulas, conditional formatting and details for all date columns.

Under the View Menu there are 9 options and 4 panel configurations:
- View build warnings shows all warnings and issues that occurred and were fixed during
  first construction of the tree.
- View changelog shows an enumerated view of all changes made to the sheet, it is bound
  to Ctrl + L.
- Panel 1 IDs details shows the treedisplays currently selected IDs full information.
- Panel 2 IDs details shows the sheets currently selected IDs full information.
- Expand all opens all IDs in the tree panel so that all children are visible. It is bound
  to the E key.
- Collapse all closes all IDs in the tree panel so that only the top IDs are visible. It is
  bound to the Q key.
- Save position saves the current scroll position in the tree panel.
- Go to saved scrolls to the previously saved tree panel scroll position.
- View all hierarchies is bound to the 0 key.
- Set all column widths changes the size of the columns in the tree and sheet panels to be
  wide enough to show the whole of the widest cell.

Under the Import Menu there are the following options:
- Import changes allows an exported/saved Tree Surgeon changelog to be imported and the
  individual changes are then attempted on the currently open sheet. Supported changes are:
  
  Edit cell
  Edit cell |
  Move columns
  Add new hierarchy column
  Add new detail column
  Delete hierarchy column
  Delete detail column
  Column rename
  Edit formula
  Edit validation
  Change detail column type
  Date format change
  Cut and paste ID
  Cut and paste ID |
  Cut and paste ID + children
  Cut and paste ID + children |
  Cut and paste children
  Copy and paste ID
  Copy and paste ID |
  Copy and paste ID + children
  Copy and paste ID + children |
  Add ID
  Rename ID
  Delete ID
  Delete ID |
  Delete ID, orphan children
  Delete ID + all children
  Delete ID from all hierarchies
  Delete ID from all hierarchies |
  Delete ID from all hierarchies, orphan children
  Sort sheet

  You can also recycle the imported changes, importing them again into another file.
  There are certain things that may stop a change from being imported, for example if
  the change was made to a column with a different name or number than the column in the
  open sheet or if an IDs parent is different. Unfortunately at this time it does not tell
  you why a change has not been imported successfully, this may be improved in a future
  version.

- Get sheet from clipboard and overwrite allows you to get copied data from your devices
  clipboard and overwrite all current data. This action can be undone.
- Merge sheets allows you to merge one sheet with another, you have options to overwrite
  details, parents, add new ids etc. You also can simply add multiple additional rows
  by pasting into the sheet on the right hand side of the pop-up.

Under the Export Menu there are the following options:
- Export changes gives a view of the changelog and allows saving/exporting of changes.
- Export flattened sheet allows you to add all IDs flattened levels for any hierarchies
  to a sheet and then gives options for saving as .xlsx or .csv or copying to clipboard.

Under the Options Menu there are the following options:
- Show/Hide columns allows you to select which columns to display in the sheet and also
  which columns will be copied to the clipboard when using clipboarding functions.
- XLSX save options gives options for what data is saved within .xlsx files.
- JSON output format gives four choices for the format of ALL json output, saving and
  copying to clipboard. They are explained in more detail under the JSON Help section.
- Mirror selections toggles auto selecting of a row in the sheet when you select it in the
  treeview. It is bound to the M key.
- Choose sheet columns allows you to choose which columns to show in the sheet panel, you
  can put them in any order, to change their order once you have add selections drag and
  drop a column header to it's desired location.
- Choose sheet settings allows changing of sheet text alignment.
- Display options gives four choices for viewing the treeview/sheet.
  

       _________________    MANAGE COLUMNS    _________________

In the column manager popup you can set column types, add formulas to columns, add
conditional formatting to columns and set validation for columns. Settings can be changed
by either double clicking on a column under the setting you want to change or right
clicking on a column.

Column types:

A detail column can have one of three different types, Text, Numerical and Date. Text
details can be any text, Numerical details can be any number and Date details can be either
a date one of three formats (YYYY/MM/DD, DD/MM/YYYY, MM/DD/YYYY) or a whole number
(integer).

Changing a column type will result in any details, formulas, formatting or validation being
evaluated and potentially deleted if they do not meet the column types requirements.

Formulas:

Formulas are only allowed in Detail columns. You can set columns automatically based on the
values of another column. Columns in formulas are represented by a 'c' character followed
by the column number e.g. the formula for c4 could be c2+c3.

The limitations change depending on what column type you are adding a formula to.

For Text Detail columns you can add together columns and text. Any user entered text must
be surrounded in double quotes in the formula e.g. c1+" my text here"

For Numerical Detail columns the following characters are allowed:
c    Column e.g. c5
0-9 Any number
+   Plus
-   Minus
/   Divide
*   Multiply
**  Exponent
//  Integer division
%   Modulus
^   Binary XOR
()  Brackets

A simple numerical detail formula would be (c2/c3)*100

For Date Detail columns +, -, c, 0-9, cd (meaning current date) and dates are allowed. A
simple date formula for days remaining (if the end date is in column 2) c2-cd. You can
also use whole numbers in your date formulas, e.g. cd+100

When entering specific dates use forward slash dates e.g. DD/MM/YYYY. This is because
hyphens are interpreted as subtraction.

Conditional Formatting:

You can add conditional formatting to columns, meaning when certain conditions are met the
cells in that column will be filled with a chosen color. You can set a maximum of 36
conditions.

For Text Detail columns conditions are limited to text matching, e.g. if the cell contains
exactly the user input then fill cell. Text conditions are not case sensitive.

For Numerical Detail columns the following characters are allowed:
c   Column e.g. c5
0-9 Any number
.   Decimal place
-   Negative number
>   Greater than
<   Less than
==  Equal to
>=  Greater than or equal to
<=  Less than or equal to
and Used to add extra condition e.g. > 5 and < 10
or  Used to add extra condition e.g. == 5 or == 6

For Date Detail columns the following characters are allowed:
c   Column e.g. c5
cd  Current date
0-9 Any number
.   Decimal place
-   Negative number
>   Greater than
<   Less than
==  Equal to
>=  Greater than or equal to
<=  Less than or equal to
and Used to add extra condition e.g. > 5 and < 10
or  Used to add extra condition e.g. == 5 or == 6

Some examples of conditional formatting for a date detail column:
> 20/06/2019
== 100

Conditions must have spaces in between statements, unlike formulas.

"""
)

help_treebuts = (
"""
       _________________    TREEVIEW BUTTONS    _________________

In the tree panel:

Find: Clicking the find button will attempt to find either an ID or detail (depending on
which you has selected in the drop-down box on the right of "Find") with letters, numbers
etc. you have typed into the entry box on the right of the drop-down box. The drop-down
box below the Find button will display any results found within the CURRENTLY viewed
hierarchy. If you are currently viewing all hierarchies then Find will search all
hierarchies. All finds are NOT case sensitive, including "exact match".

Hierarchy: This is the drop-down box where you can select which hierarchy to view. The
hierarchies are bound to the number keys, 1 being the first in the drop-down box, 2 being
the 2nd and so on up to 9. 0 is bound to View all hierarchies, as is H. To switch beyond
9 you'll have to use the drop-down box.

Detail: This is the drop-down box where you can select which column to display next to the
IDs in the tree panel.

In the sheet panel:

Auto-select ID: The same as mirror selections under the view menu.

Tag/Untag ID (Ctrl + T): Allows you to tag IDs, tagged IDs show up in the dropdown box
next to the button and persist through saving as .json and .xlsx.

Column Manager: Allows quick access to the column manager panel, same as Manage columns
under the Format menu.

Find: Works the same way that the Find button for the Tree panel works except it searches
the sheet instead."""
)

help_functions = (
"""
       _________________    TREEVIEW FUNCTIONS    _________________

By right clicking on an ID in the tree panel you can select various functions. The main
functions are Cut, Copy and delete.

Right clicking is disabled when viewing all hierarchies, as are a few of the other
functions.

To cut or copy an ID between different hierarchies simply right click on the ID and select
whichever option you want then switch hierarchy and right click in empty space or on the
ID you want to paste the cut/ copied ID to as a sibling or child. If you want to paste an
ID as an ID without a parent right click on a top ID and choose paste as sibling.

You can also cut all of an IDs children, including grandchildren and so on, and paste them
under where you right click.

Using shift click you can select multiple up or down of an existing selection. Using
control click you can make multiple selections.

There is a difference in the behavior of Shift and Ctrl click. Shift + Click will select
everything between the highest / lowest selected ID, including all children. Whereas Ctrl
+ Click will only select the IDs you click on.

When using the Control X, C and V keys to cut/copy and paste they will work on the selected
ID, not on the position where the mouse is hovering, unless pasting over empty space using
Control V. Cutting and copying using this method will only perform on IDs that are on the
same level as the top most (index-wise) ID, after pressing Control X or C it will deselect
any selections that were not cut or copied. Pressing the Delete key on multiple selections
will work the same way, except performing a Delete immediately. The delete key uses the
typical Delete ID function, not deleting its children.

In the tree panel there are 5 delete ID options. Delete ID only removes the ID from the
hierarchy you're currently viewing IF the ID occurs in another hierarchy, if it does not
then it totally removes the ID. Del all of ID totally removes the ID. Del ID+children is
the same as Delete ID but for every child and child of that child and so on recursively
under the selected ID.

You can quickly edit a detail by double clicking on the detail/cell you want to edit. To
delete a detail press Confirm when editing a detail with the cell empty.

Double clicking will also allow you to fully view the columns full text.

Pasting a detail or details will work between both panels. You can drag and drop rows in
the sheet panel to change their order.

When using drag and drop you can use your mousewheel to scroll down, just move your mouse
a bit after scrolling to cause the selection to move.

"""
)

help_tsrgnfiles = (
"""
       _________________    JSON FILES    _________________

There are four json formats which Tree Surgeon can utilise, with each one the entire sheet
is kept under the key "records". However the program will also look for the keys: sheet,
data and table. The first format, also the first option under "Options -> JSON output
format" is displayed as an example below:


A dictionary of key (column header) and value (list of column cells)

{"records":
    {
        "ID":
                    [
                     "ID_1",
                     "ID_2"
                    ],
        "DETAIL_1":
                    [
                     "",
                     ""
                    ],
        "PARENT_1":
                    [
                     "ID_1s_Parent",
                     "ID_2s_Parent"
                    ]
    }
}


The second json format option example is displayed below:

A list of dictionaries (rows) where inside each dictionary the key is the header and the
value is the cell

{
 "records": [
        {
         "ID":       "ID_1",
         "DETAIL_1": "",
         "PARENT_1": "ID1s_Parent"
         },
        {
         "ID":       "ID1s_Parent",
         "DETAIL_1": "",
         "PARENT_1": ""
         }
            ]
}


The third json format option is displayed below:

A list of lists (rows) where each row simply contains values that are the cells

{
 "records":
    [
        [
         "ID",
         "DETAIL_1",
         "PARENT_1"
        ],
        [
         "ID_1",
         "",
         "ID_1s_parent"
        ]
    ]
}


The fourth json format option is displayed below:

A tab delimited csv stored as a string under the key 'records', this format is really non-
typical so only use it if you really need to.

{
 "records":
    "ID\\tDetail-1\\tParent-1\\nID_1\\t\\tID_1s_Parent"
}


Program data is only included if Save is used as opposed to Copy to clipboard
it is json in the following format:


{"version": "2.00",
 "records": <full sheet including headers stored here>,
 "changelog": [],
 "Treesurgeon_Data": "base32string"}

"""
)

help_using_api = (
"""
       _________________    USING THE API    _________________

Tree Surgeon can be run as an API run without triggering a user interface to get different
outputs and file conversions.

Currently the only service available is to flatten a hierarchy and the only output available
is a tab delimited csv file.

The input file must be either .xlsx, .xls, .xlsm, .csv or .tsv.

ALSO! Please note that if any of the parameters include spaces then they may need to be
surrounded by double quotes e.g. "my xlsx sheet name" depending on how you choose to start
the API.

Tree Surgeon must be run with the following paramters, ignore the <> symbols:

Required parameters:
 1# Type of service - <flatten> is the only option currently
 2# Input filepath - <the full filepath including name of the file to open>
 3# Name of sheet (must put something even if opening a csv) - <name of xlsx sheet to open>
 4# Output filepath - <the full filepath including name of the file to create>
 5# Output csv delimiter - <tab> or <comma>
 6# Overwrite files of the same name - <True or False>
 7# The column position of the ID column e.g. 1 is the first column <integer>
 8# The column position of the hierarchy to work on - <integer>
 9# The positions of all the hierarchy columns seperated by commas without spaces e.g. <4,5>

Optional parameters:
10# The column position of the detail column to split if flattening, can be <integer or False>
    (defaults to the first detail column available if any are, False if not)
11# Reverse order of flattened hierarchy columns - <True or False>
12# Only base IDs as rows - <True or False> default is True
13# Rename ID column - <True or False> default is True
14# Justify flattened columns - <True or False> default is True
15# Remove detail columns and other hierarchy columns - <True or False> default is False
16# Add an index column in position 1 - <True or False> default is False

Examples of a full run command would be:
"Tree Surgeon.exe" flatten DEMO_FILE.xlsx sheet1 api_test_csv.csv tab True 1 7 7,8 6 True True True True True False

or...

python "Tree Surgeon.pyw" flatten DEMO_FILE.xlsx sheet1 api_test_csv.csv tab True 1 7 7,8 6 True True True True True False


"""
)

