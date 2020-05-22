#Copyright © 2020 R. A. Gardner

from ts_extra_vars_and_funcs import *
from ts_toplevels import *
from ts_widgets import *
from ts_classes_d import *

import tkinter as tk
from tkinter import ttk, filedialog
from tksheet import Sheet
import os
from sys import argv
from platform import system as get_os
import datetime
import io
from ast import literal_eval
import re
from operator import itemgetter
from itertools import chain, islice, repeat, cycle
from collections import defaultdict, deque, Counter
from math import floor
import json
import pickle
import zlib
import lzma
from base64 import urlsafe_b64encode as b64e, urlsafe_b64decode as b64d
from base64 import b32encode as b32e, b32decode as b32d
import csv as csv_module
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import builtins
from fastnumbers import isint, isintlike, isfloat, isreal


class treecompare(tk.Frame):
    def __init__(self,parent,C):
        tk.Frame.__init__(self,parent)
        self.C = C
        self.heads1 = []
        self.heads2 = []
        self.data1 = []
        self.data2 = []
        self.sheet1 = []
        self.sheet2 = []
        self.nodes1 = {}
        self.nodes2 = {}
        self.rns1 = {}
        self.rns2 = {}
        self.report = []
        self.ic1 = 0
        self.ic2 = 0
        self.parent_cols1 = []
        self.parent_cols2 = []
        self.row_len1 = 0
        self.row_len2 = 0
        self.shkeys = []

        self.l_frame = frame(self)
        self.l_frame.config(highlightthickness=1,highlightbackground="black")
        self.l_frame_btns = frame(self.l_frame)
        
        self.r_frame = frame(self)
        self.r_frame.config(highlightthickness=1,highlightbackground="black")
        self.r_frame_btns = frame(self.r_frame)

        self.l_frame.grid(row=0,column=0,sticky="nswe")
        self.r_frame.grid(row=0,column=1,sticky="nswe")
        self.grid_rowconfigure(0,weight=1)
        self.grid_columnconfigure(0,weight=1,uniform="x")
        self.grid_columnconfigure(1,weight=1,uniform="x")

        self.l_frame_btns.pack(side="top",fill="both")
        self.r_frame_btns.pack(side="top",fill="both")

        self.ss_filename1 = readonly_entry(self.l_frame_btns,font=EF, theme = self.C.theme)
        self.ss_filename1.grid(row=1,column=0,padx=2,pady=2,sticky="nswe")

        self.open_ss1 = button(self.l_frame_btns,text="⯇ Open file",
                               style="EF.Std.TButton",
                               command=self.open_ss1)
        self.open_ss1.grid(row=1,column=1,padx=20,pady=2,sticky="nsw")
        self.open_ss1.config(width = 30)

        self.sheet_dropdown_displayed1 = tk.StringVar(self.l_frame_btns)
        self.sheet_dropdown1 = ttk.Combobox(self.l_frame_btns,textvariable=self.sheet_dropdown_displayed1,
                                                font=EF,state="readonly",background="white")
        self.sheet_dropdown1.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
        self.sheet_dropdown1.grid(row=2,padx=2,pady=2,column=0,sticky="nswe")

        self.load_sheet1 = button(self.l_frame_btns,text="⯇ Load sheet",
                                  style="EF.Std.TButton",
                                  command=self.load_sheet1)
        self.load_sheet1.config(state="disabled")
        self.load_sheet1.config(width = 30)
        self.load_sheet1_STATE = "disabled"
        self.load_sheet1.grid(row=2,column=1,padx=20,pady=2,sticky="nsw")

        self.selector_1 = id_and_parent_column_selector(self.l_frame_btns, expand = True)
        self.selector_1.config(width = 400, height = 330)
        self.selector_1.grid(row = 3, column = 0, sticky = "nswe")
        
        self.run_compare_button = button(self.l_frame_btns,
                                          text="Create Report",
                                          style="TF.Std.TButton",
                                          command = self.run_comparison)
        self.run_compare_button.config(width = 40)
        self.run_compare_button.grid(row=3,column=1,padx=10,pady=20, sticky = "e")
        
        self.sheetdisplay1 = Sheet(self.l_frame,
                                   theme = self.C.theme,
                                   header_font = ("Calibri", 13, "normal"),
                                   outline_thickness=1,
                                   outline_color="black")
        self.sheetdisplay1.enable_bindings(("single",
                                            "drag_select",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                            "column_select",
                                           "row_select",
                                           "arrowkeys"))
        self.sheetdisplay1.pack(side="top",fill="both",expand=True)

        #__________________ R FRAME ______________________________________________________________

        self.ss_filename2 = readonly_entry(self.r_frame_btns,font=EF, theme = self.C.theme)
        self.ss_filename2.grid(row=1,column=0,padx=2,pady=2,sticky="nswe")

        self.open_ss2 = button(self.r_frame_btns,text="⯇ Open file",
                               style="EF.Std.TButton",
                               command=self.open_ss2)
        self.open_ss2.config(width = 30)
        self.open_ss2.grid(row=1,column=1,padx=20,pady=2,sticky="nsw")

        self.sheet_dropdown_displayed2 = tk.StringVar(self.r_frame_btns)
        self.sheet_dropdown2 = ttk.Combobox(self.r_frame_btns,textvariable=self.sheet_dropdown_displayed2,
                                                font=EF,state="readonly",background="white")
        self.sheet_dropdown2.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
        self.sheet_dropdown2.grid(row=2,column=0,padx=2,pady=2,sticky="nswe")

        self.load_sheet2 = button(self.r_frame_btns,text="⯇ Load sheet",
                                  style="EF.Std.TButton",
                                  command=self.load_sheet2)
        self.load_sheet2.config(state="disabled")
        self.load_sheet2.config(width = 30)
        self.load_sheet2_STATE = "disabled"
        self.load_sheet2.grid(row=2,column=1,padx=20,pady=2,sticky="nsw")

        self.selector_2 = id_and_parent_column_selector(self.r_frame_btns, expand = True)
        self.selector_2.config(width = 400, height = 330)
        self.selector_2.grid(row=3,column=0,sticky="nswe")
        
        self.sheetdisplay2 = Sheet(self.r_frame,
                                   theme = self.C.theme,
                                   header_font = ("Calibri", 13, "normal"),
                                   outline_thickness=1,
                                   outline_color="black")
        self.sheetdisplay2.enable_bindings(("single",
                                            "drag_select",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                            "column_select",
                                           "row_select",
                                           "arrowkeys"))
        self.sheetdisplay2.pack(side="top",fill="both",expand=True)

    def enable_widgets(self):
        self.C.menubar_state("normal",start=True)
        self.ss_filename1.config(state="readonly")
        self.open_ss1.config(state="normal")
        self.sheet_dropdown1.config(state="readonly")
        self.sheet_dropdown1.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
        self.load_sheet1.config(state=self.load_sheet1_STATE)
        self.selector_1.enable_me()
        self.sheetdisplay1.enable_bindings(("single",
                                            "drag_select",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                            "column_select",
                                           "row_select",
                                           "arrowkeys"))
        self.sheetdisplay1.basic_bindings(True)
        
        self.ss_filename2.config(state="readonly")
        self.open_ss2.config(state="normal")
        self.sheet_dropdown2.config(state="readonly")
        self.sheet_dropdown2.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
        self.load_sheet2.config(state=self.load_sheet2_STATE)
        self.selector_2.enable_me()
        self.sheetdisplay2.enable_bindings(("single",
                                            "drag_select",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                            "column_select",
                                           "row_select",
                                           "arrowkeys"))
        self.run_compare_button.config(state="normal")
        self.sheetdisplay2.basic_bindings(True)

    def disable_widgets(self):
        self.C.menubar_state("disabled")
        self.ss_filename1.config(state="disabled")
        self.open_ss1.config(state="disabled")
        self.sheet_dropdown1.config(state="disabled")
        self.sheet_dropdown1.unbind("<<ComboboxSelected>>")
        self.load_sheet1_STATE = str(self.load_sheet1['state'])
        self.load_sheet1.config(state="disabled")
        self.selector_1.disable_me()
        self.sheetdisplay1.disable_bindings(("single",
                                            "drag_select",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                            "column_select",
                                           "row_select",
                                           "arrowkeys"))
        self.sheetdisplay1.basic_bindings(False)

        self.ss_filename2.config(state="disabled")
        self.open_ss2.config(state="disabled")
        self.sheet_dropdown2.config(state="disabled")
        self.sheet_dropdown2.unbind("<<ComboboxSelected>>")
        self.load_sheet2_STATE = str(self.load_sheet2['state'])
        self.load_sheet2.config(state="disabled")
        self.selector_2.disable_me()
        self.sheetdisplay2.disable_bindings(("single",
                                            "drag_select",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                            "column_select",
                                           "row_select",
                                           "arrowkeys"))
        self.run_compare_button.config(state="disabled")
        self.sheetdisplay2.basic_bindings(False)

    def start_work(self,msg=""):
        self.C.status_bar.change_text(msg)
        self.disable_widgets()

    def stop_work(self,msg=""):
        self.C.status_bar.change_text(msg)
        self.enable_widgets()

    def populate(self):
        self.C.change_app_title(title="Comparing sheets")
        self.sheetdisplay1.change_theme(self.C.theme)
        self.sheetdisplay2.change_theme(self.C.theme)
        self.C.show_frame("treecompare")

    def de_populate(self):
        self.reset_vars()

    def reset_vars(self):
        try:
            self.C.wb.close()
        except:
            pass
        self.reset_1(staying_on_compare=False)
        self.reset_2(staying_on_compare=False)

    def reset_1(self,staying_on_compare=True):
        try:
            self.C.wb.close()
        except:
            pass
        self.C.wb = None
        self.heads1 = []
        self.data1 = []
        self.sheet1 = []
        self.sheet2 = []
        self.nodes1 = {}
        self.nodes2 = {}
        self.rns1 = {}
        self.rns2 = {}
        self.report = []
        self.ic1 = 0
        self.parent_cols1 = []
        self.row_len1 = 0
        self.ss_filename1.set_my_value("")
        self.sheet_dropdown1['values'] = []
        self.sheet_dropdown_displayed1.set("")
        self.load_sheet1.config(state="disabled")
        self.selector_1.clear_displays()
        self.sheetdisplay1.dehighlight_cells(all_=True,redraw=False)
        self.sheetdisplay1.dehighlight_cells(canvas="row_index",all_=True,redraw=False)
        self.sheetdisplay1.dehighlight_cells(canvas="header",all_=True,redraw=False)
        self.sheetdisplay2.dehighlight_cells(all_=True,redraw=False)
        self.sheetdisplay2.dehighlight_cells(canvas="row_index",all_=True,redraw=False)
        self.sheetdisplay2.dehighlight_cells(canvas="header",all_=True,redraw=False)
        self.sheetdisplay1.data_reference(newdataref=[],redraw=True)

    def reset_2(self,staying_on_compare=True):
        try:
            self.C.wb.close()
        except:
            pass
        self.C.wb = None
        self.heads2 = []
        self.data2 = []
        self.sheet1 = []
        self.sheet2 = []
        self.nodes1 = {}
        self.nodes2 = {}
        self.rns1 = {}
        self.rns2 = {}
        self.report = []
        self.ic2 = 0
        self.parent_cols2 = []
        self.row_len2 = 0
        self.ss_filename2.set_my_value("")
        self.sheet_dropdown2['values'] = []
        self.sheet_dropdown_displayed2.set("")
        self.load_sheet2.config(state="disabled")
        self.selector_2.clear_displays()
        self.sheetdisplay1.dehighlight_cells(all_=True,redraw=False)
        self.sheetdisplay1.dehighlight_cells(canvas="row_index",all_=True,redraw=False)
        self.sheetdisplay1.dehighlight_cells(canvas="header",all_=True,redraw=False)
        self.sheetdisplay2.dehighlight_cells(all_=True,redraw=False)
        self.sheetdisplay2.dehighlight_cells(canvas="row_index",all_=True,redraw=False)
        self.sheetdisplay2.dehighlight_cells(canvas="header",all_=True,redraw=False)
        self.sheetdisplay2.data_reference(newdataref=[],redraw=True)
        
    def open_ss1(self):
        if self.data1:
            confirm = askconfirm(self,"Note: Opening resets",
                                 theme = self.C.theme)
            if confirm.boolean == False:
                return
        self.start_work("Loading...   ")
        self.reset_1(False)
        filepath = filedialog.askopenfilename(parent=self.C,title="Select file")
        if not filepath:
            self.stop_work("Program ready")
            return
        try:
            filepath = os.path.normpath(filepath)
        except:
            errorpopup = error(self,"Filepath invalid   ", theme = self.C.theme)
            self.stop_work("Program ready")
            return
        if not filepath.lower().endswith((".json",".xlsx",".xls",".xlsm",".csv",".tsv")):
            errorpopup = error(self,"Please select excel/csv/json   ", theme = self.C.theme)
            self.stop_work("Program ready")
            return
        check = os.path.isfile(filepath)
        if check == False:
            errorpopup = error(self,"Filepath invalid   ", theme = self.C.theme)
            self.stop_work("Program ready")
            return
        try:
            if filepath.lower().endswith((".csv",".tsv")):
                with open(filepath,"r") as fh:
                    temp_data = fh.read()
                delimiter_,quotechar_ = csv_delimiter_quotechar(temp_data)
                if delimiter_ is None:
                    errorpopup = error(self,"No appropriate data in file   ", theme = self.C.theme)
                    self.stop_work("Program ready")
                    return
                for rn,r in enumerate(csv_module.reader(io.StringIO(temp_data),delimiter=delimiter_,quotechar=quotechar_,skipinitialspace=True)):
                    try:
                        self.data1.append(r[:len(r) - next(i for i,c in enumerate(reversed(r)) if c)])
                    except:
                        pass
                    if not rn % 500:
                        self.C.update()
                        if self.C.USER_HAS_QUIT:
                            return
                        self.C.status_bar.change_text(f"Loading...  rows: {rn}")
                if not self.data1:
                    errorpopup = error(self,"File contains no data   ", theme = self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.load_display1()
                self.stop_work("Program ready")
            elif filepath.lower().endswith(".json"):
                j = self.C.treeframe.get_json_from_file(filepath)
                json_format = self.C.treeframe.get_json_format(j)
                if not json_format:
                    raise Exception
                self.data1 = self.C.treeframe.json_to_sheet(j,format_=json_format[0],key=json_format[1],get_format=False)
                if not self.data1:
                    errorpopup = error(self,"File contains no data   ", theme = self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.load_display1()
                self.stop_work("Program ready")
            else:
                in_mem = self.C.return_wb_file(filepath)
                self.C.wb = load_workbook(in_mem,read_only=True,data_only=True)
                if len(self.C.wb.sheetnames) < 1:
                    errorpopup = error(self,"File contains no data   ", theme = self.C.theme)
                    self.stop_work("Program ready")
                    return
                sheetnames = set(self.C.wb.sheetnames)
                if "Treesurgeon Data" in sheetnames:
                    ws = self.C.wb["Treesurgeon Data"]
                    ws.reset_dimensions()
                    try:
                        d = self.C.decompress_str_return_obj("".join(["" if r[0].value is None else f"{r[0].value}" for r in islice(ws.rows, 1, None)]),
                                                                       basetype = "32",
                                                                       dec = True)
                        self.data1 = [[h['name'] for h in d['headers']]] + d['records']
                        self.C.wb.close()
                        self.load_display1()
                        self.stop_work("Program ready")
                    except:
                        self.data1 = []
                        self.C.wb.close()
                        self.C.wb = load_workbook(in_mem,read_only=True,data_only=True)
                        errorpopup = error(self,"Error opening program data, select a sheet   ", theme = self.C.theme)
                        self.sheet_dropdown1['values'] = self.C.wb.sheetnames
                        self.sheet_dropdown_displayed1.set(self.C.wb.sheetnames[0])
                        self.stop_work("Program ready")
                        self.open_ss2.config(state="disabled")
                        self.load_sheet1.config(state="normal")
                else:
                    self.sheet_dropdown1['values'] = self.C.wb.sheetnames
                    self.sheet_dropdown_displayed1.set(self.C.wb.sheetnames[0])
                    self.stop_work("Program ready")
                    self.open_ss2.config(state="disabled")
                    self.load_sheet1.config(state="normal")
            self.ss_filename1.set_my_value(filepath)
        except Exception as error_msg:
            errorpopup = error(self,"Error: " + str(error_msg), theme = self.C.theme)
            self.stop_work("Program ready")

    def load_sheet1(self):
        self.start_work("Loading...   ")
        ws = self.C.wb[self.sheet_dropdown_displayed1.get()]
        ws.reset_dimensions()
        for rn,r in enumerate(ws.rows):
            try:
                self.data1.append(["" if x.value is None else f"{x.value}" for x in islice(r,0,len(r) - next(i for i,c in enumerate(reversed(r)) if c.value is not None))])
            except:
                pass
            if not rn % 500:
                self.C.update()
                if self.C.USER_HAS_QUIT:
                    return
                self.C.status_bar.change_text(f"Loading...  rows: {rn}")
        self.C.wb.close()
        self.stop_work("Program ready")
        if not self.data1:
            errorpopup = error(self,"Sheet contains no data   ", theme = self.C.theme)
            self.load_sheet1.config(state="disabled")
            return
        self.load_sheet1.config(state="disabled")
        self.load_display1()

    def load_display1(self):
        self.row_len1 = len(max(self.data1,key=len))
        self.sheetdisplay1.data_reference(newdataref=self.data1,redraw=False)
        self.sheetdisplay1.set_all_cell_sizes_to_text()
        self.selector_1.set_columns([h for h in self.data1[0]])
        self.selector_1.detect_id_col()
        self.selector_1.detect_par_cols()
        
    def open_ss2(self):
        if self.data2:
            confirm = askconfirm(self,"Note: Opening resets",
                                 theme = self.C.theme)
            if confirm.boolean == False:
                return
        self.start_work("Loading...   ")
        self.reset_2(False)
        filepath = filedialog.askopenfilename(parent=self.C,title="Select file")
        if not filepath:
            self.stop_work("Program ready")
            return
        try:
            filepath = os.path.normpath(filepath)
        except:
            errorpopup = error(self,"Filepath invalid   ", theme = self.C.theme)
            self.stop_work("Program ready")
            return
        if not filepath.lower().endswith((".json",".xlsx",".xls",".xlsm",".csv",".tsv")):
            errorpopup = error(self,"Please select excel/csv/json   ", theme = self.C.theme)
            self.stop_work("Program ready")
            return
        check = os.path.isfile(filepath)
        if check == False:
            errorpopup = error(self,"Filepath invalid   ", theme = self.C.theme)
            self.stop_work("Program ready")
            return
        try:
            if filepath.lower().endswith((".csv",".tsv")):
                with open(filepath,"r") as fh:
                    temp_data = fh.read()
                delimiter_,quotechar_ = csv_delimiter_quotechar(temp_data)
                if delimiter_ is None:
                    errorpopup = error(self,"No appropriate data in file   ", theme = self.C.theme)
                    self.stop_work("Program ready")
                    return
                for rn,r in enumerate(csv_module.reader(io.StringIO(temp_data),delimiter=delimiter_,quotechar=quotechar_,skipinitialspace=True)):
                    try:
                        self.data2.append(r[:len(r) - next(i for i,c in enumerate(reversed(r)) if c)])
                    except:
                        pass
                    if not rn % 500:
                        self.C.update()
                        if self.C.USER_HAS_QUIT:
                            return
                        self.C.status_bar.change_text(f"Loading...  rows: {rn}")
                if not self.data2:
                    errorpopup = error(self,"File contains no data   ", theme = self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.load_display2()
                self.stop_work("Program ready")
            elif filepath.lower().endswith(".json"):
                j = self.C.treeframe.get_json_from_file(filepath)
                json_format = self.C.treeframe.get_json_format(j)
                if not json_format:
                    raise Exception
                self.data2 = self.C.treeframe.json_to_sheet(j,format_=json_format[0],key=json_format[1],get_format=False)
                if not self.data2:
                    errorpopup = error(self,"File contains no data   ", theme = self.C.theme)
                    self.stop_work("Program ready")
                    return
                self.load_display2()
                self.stop_work("Program ready")
            else:
                in_mem = self.C.return_wb_file(filepath)
                try:
                    self.C.wb = load_workbook(in_mem,read_only=True,data_only=True)
                except:
                    errorpopup = error(self,"Error opening file   ", theme = self.C.theme)
                    self.stop_work("Program ready")
                    return
                if len(self.C.wb.sheetnames) < 1:
                    errorpopup = error(self,"File contains no data   ", theme = self.C.theme)
                    self.stop_work("Program ready")
                    return
                sheetnames = set(self.C.wb.sheetnames)
                if "Treesurgeon Data" in sheetnames:
                    ws = self.C.wb["Treesurgeon Data"]
                    ws.reset_dimensions()
                    try:
                        d = self.C.decompress_str_return_obj("".join(["" if r[0].value is None else f"{r[0].value}" for r in islice(ws.rows, 1, None)]),
                                                                       basetype = "32",
                                                                       dec = True)
                        self.data2 = [[h['name'] for h in d['headers']]] + d['records']
                        self.C.wb.close()
                        self.load_display2()
                        self.stop_work("Program ready")
                    except:
                        self.data2 = []
                        self.C.wb.close()
                        self.C.wb = load_workbook(in_mem,read_only=True,data_only=True)
                        errorpopup = error(self,"Error opening program data, select a sheet   ", theme = self.C.theme)
                        self.sheet_dropdown2['values'] = self.C.wb.sheetnames
                        self.sheet_dropdown_displayed2.set(self.C.wb.sheetnames[0])
                        self.stop_work("Program ready")
                        self.open_ss1.config(state="disabled")
                        self.load_sheet2.config(state="normal")
                else:
                    self.sheet_dropdown2['values'] = self.C.wb.sheetnames
                    self.sheet_dropdown_displayed2.set(self.C.wb.sheetnames[0])
                    self.stop_work("Program ready")
                    self.open_ss1.config(state="disabled")
                    self.load_sheet2.config(state="normal")
            self.ss_filename2.set_my_value(filepath)
        except Exception as error_msg:
            errorpopup = error(self,"Error: " + str(error_msg), theme = self.C.theme)
            self.stop_work("Program ready")

    def load_sheet2(self):
        self.start_work("Loading...   ")
        ws = self.C.wb[self.sheet_dropdown_displayed2.get()]
        ws.reset_dimensions()
        for rn,r in enumerate(ws.rows):
            try:
                self.data2.append(["" if x.value is None else f"{x.value}" for x in islice(r,0,len(r) - next(i for i,c in enumerate(reversed(r)) if c.value is not None))])
            except:
                pass
            if not rn % 500:
                self.C.update()
                if self.C.USER_HAS_QUIT:
                    return
                self.C.status_bar.change_text(f"Loading...  rows: {rn}")
        self.C.wb.close()
        self.stop_work("Program ready")
        if not self.data2:
            errorpopup = error(self,"Sheet contains no data   ", theme = self.C.theme)
            self.load_sheet2.config(state="disabled")
            return
        self.load_sheet2.config(state="disabled")
        self.load_display2()

    def load_display2(self):
        self.row_len2 = len(max(self.data2,key=len))
        self.sheetdisplay2.data_reference(newdataref=self.data2,redraw=False)
        self.sheetdisplay2.set_all_cell_sizes_to_text()
        self.selector_2.set_columns([h for h in self.data2[0]])
        self.selector_2.detect_id_col()
        self.selector_2.detect_par_cols()

    def heads_comparison(self, heads, datavar, addition):
        if datavar == 1:
            row_len = self.row_len1
        elif datavar == 2:
            row_len = self.row_len2
        if len(heads) < row_len:
            heads += list(repeat("",row_len - len(heads)))
        tally_of_heads = defaultdict(lambda:-1)
        for coln in range(len(heads)):
            cell = heads[coln]
            if not cell:
                cell = f"_MISSING_{coln + 1}"
                addition.append((f" - Missing header in column #{coln + 1}", ))
            else:
                if " " in cell:
                    addition.append((f" - Spaces in header column #{coln + 1}", ))
                if "\n" in cell:
                    addition.append((f" - Newlines in header column #{coln + 1}", ))
                if "\r" in cell:
                    addition.append((f" - Carriage returns in header column #{coln + 1}", ))
            cell = "".join(cell.strip().split())
            hk = cell.lower()
            tally_of_heads[hk] += 1
            if tally_of_heads[hk] > 0:
                orig = cell
                x = 1
                while hk in tally_of_heads:
                    cell = f"{orig}_DUPLICATED_{x}"
                    hk = cell.lower()
                    x += 1
                tally_of_heads[hk] += 1
                addition.append((f" - Duplicate header in column #{coln + 1}", ))
            heads[coln] = cell
        return heads, addition

    def run_comparison(self):
        self.ic1 = self.selector_1.get_id_col()
        self.parent_cols1 = list(self.selector_1.get_par_cols())
        self.ic2 = self.selector_2.get_id_col()
        self.parent_cols2 = list(self.selector_2.get_par_cols())
        if self.ic1 is None or self.ic2 is None:
            return
        if not self.parent_cols1 or not self.parent_cols2:
            return
        if self.ic1 in self.parent_cols1 or self.ic2 in self.parent_cols2:
            return
        self.start_work("Creating comparison...   ")
        sheetname_1 = os.path.basename(self.ss_filename1.get())
        sheetname_2 = os.path.basename(self.ss_filename2.get())
        if sheetname_1 == sheetname_2:
            sheetname_1 = "Sheet 1 - Left Panel"
            sheetname_2 = "Sheet 2 - Right Panel"
        self.sheet1 = []
        self.sheet2 = []
        self.nodes1 = {}
        self.nodes2 = {}
        self.rns1 = {}
        self.rns2 = {}
        self.report = {"ids": [],
                       "info": []}
        dt = datetime.datetime.today()
        self.report['info'].append((f"Report Created: {dt.strftime('%A %d %B %Y %H-%M-%S')}", ))
        self.heads1, addition = self.heads_comparison(self.data1[0].copy(), 1, self.report['info'])
        treebuilder = TreeBuilder()
        self.sheet1, self.nodes1, addition, self.rns1 = treebuilder.build(self.data1,
                                                                          self.sheet1,
                                                                          self.row_len1,
                                                                          self.ic1,
                                                                          self.parent_cols1,
                                                                          self.nodes1,
                                                                          [],
                                                                          self.rns1,
                                                                          add_warnings = True,
                                                                          skip_1st = True,
                                                                          compare = True,
                                                                          fix_associate = True)
        if addition:
            self.report['info'].append((f"WARNINGS - {sheetname_1} - ", ))
            self.report['info'].extend(addition)
            self.report['info'].append(("", ))
        self.heads2, addition = self.heads_comparison(self.data2[0].copy(), 2, self.report['info'])
        self.sheet2, self.nodes2, addition, self.rns2 = treebuilder.build(self.data2,
                                                                          self.sheet2,
                                                                          self.row_len2,
                                                                          self.ic2,
                                                                          self.parent_cols2,
                                                                          self.nodes2,
                                                                          [],
                                                                          self.rns2,
                                                                          add_warnings = True,
                                                                          skip_1st = True,
                                                                          compare = True,
                                                                          fix_associate = True)
        if addition:
            self.report['info'].append((f"WARNINGS - {sheetname_2} - ", ))
            self.report['info'].extend(addition)
            self.report['info'].append(("", ))
        qhst1 = set(self.parent_cols1)
        qhst2 = set(self.parent_cols2)
        pcold = defaultdict(list)
        for i,h in enumerate(self.heads1):
            if i in qhst1:
                pcold[h].append(i)
        for i,h in enumerate(self.heads2):
            if i in qhst2:
                pcold[h].append(i)
        detcold = defaultdict(list)
        qhst1.add(self.ic1)
        qhst2.add(self.ic2)
        for i,h in enumerate(self.heads1):
            if i not in qhst1:
                detcold[h].append(i)
        for i,h in enumerate(self.heads2):
            if i not in qhst2:
                detcold[h].append(i)
        matching_hrs_names = [k for k,v in pcold.items() if len(v) > 1]
        matching_hrs_names.sort(key=self.srtkey)
        matching_details_names = [k for k,v in detcold.items() if len(v) > 1]
        matching_details_names.sort(key=self.srtkey)
        self.report['info'].append(("", ))
        self.report['info'].append(("GENERAL INFORMATION", ))
        if self.row_len1 > self.row_len2:
            self.report['info'].append((f" - {sheetname_1} has {self.row_len1 - self.row_len2} more columns than {sheetname_2}", ))
            self.report['info'].append((f"       {sheetname_1} total columns: {self.row_len1}", ))
            self.report['info'].append((f"       {sheetname_2} total columns: {self.row_len2}", ))
        elif self.row_len2 > self.row_len1:
            self.report['info'].append((f" - {sheetname_2} has {self.row_len2 - self.row_len1} more columns than {sheetname_1}", ))
            self.report['info'].append((f"       {sheetname_1} total columns: {self.row_len1}", ))
            self.report['info'].append((f"       {sheetname_2} total columns: {self.row_len2}", ))
        else:
            self.report['info'].append((f" - Sheets have the same number of columns ({self.row_len1})", ))
        if len(self.nodes1) > len(self.nodes2):
            self.report['info'].append((f" - {sheetname_1} has {len(self.nodes1) - len(self.nodes2)} more IDs than {sheetname_2}", ))
            self.report['info'].append((f"       {sheetname_1} total IDs: {len(self.nodes1)}", ))
            self.report['info'].append((f"       {sheetname_2} total IDs: {len(self.nodes2)}", ))
        elif len(self.nodes2) > len(self.nodes1):
            self.report['info'].append((f" - {sheetname_2} has {len(self.nodes2) - len(self.nodes1)} more IDs than {sheetname_1}", ))
            self.report['info'].append((f"       {sheetname_1} total IDs: {len(self.nodes1)}", ))
            self.report['info'].append((f"       {sheetname_2} total IDs: {len(self.nodes2)}", ))
        else:
            self.report['info'].append((f" - Sheets have the same number of IDs ({len(self.sheet1)})", ))
        self.report['info'].append(("", ))
        self.report['info'].append(("HEADERS", ))
        if self.ic1 == self.ic2 and self.heads1[self.ic1] == self.heads2[self.ic2]:
            self.report['info'].append((f" - Sheets have the same ID column names and indexes", ))
        else:
            self.report['info'].append((f" - {sheetname_1} has ID column: {self.ic1 + 1} - {self.heads1[self.ic1]}", ))
            self.report['info'].append((f" - {sheetname_2} has ID column: {self.ic2 + 1} - {self.heads2[self.ic2]}", ))
        if self.parent_cols1 == self.parent_cols2 and [self.heads1[pcol_1] for pcol_1 in self.parent_cols1] == [self.heads2[pcol_2] for pcol_2 in self.parent_cols2]:
            self.report['info'].append((f" - Sheets have the same Parent column names and indexes", ))
        else:
            self.report['info'].append((f" - {sheetname_1} has parent columns: ", ))
            for pcol in self.parent_cols1:
                self.report['info'].append((f"       Column: {pcol+1} - {self.heads1[pcol]}", ))
            self.report['info'].append((f" - {sheetname_2} has parent columns: ", ))
            for pcol in self.parent_cols2:
                self.report['info'].append((f"       Column: {pcol+1} - {self.heads2[pcol]}", ))
        if len(matching_details_names) > 0:
            hdset1 = {h for i,h in enumerate(self.heads1) if i not in qhst1}
            hdset2 = {h for i,h in enumerate(self.heads2) if i not in qhst2}
            if all(detcold[n][0] == detcold[n][1] for n in matching_details_names):
                self.report['info'].append((" - Sheets have the same detail column names and indexes", ))
            else:
                self.report['info'].append((" - Sheets have following matching detail column names:", ))
                for n in matching_details_names:
                    self.report['info'].append((f"   - {n}", ))
                    self.report['info'].append((f"       Column {sheetname_1}: {detcold[n][0]}", ))
                    self.report['info'].append((f"       Column {sheetname_2}: {detcold[n][1]}", ))
            if any(h not in hdset2 for h in hdset1):
                self.report['info'].append((f" - {sheetname_1} has following detail columns that {sheetname_2} doesn't:", ))
                for h in hdset1:
                    if h not in hdset2:
                        self.report['info'].append((f"   - {h}", ))
            if any(h not in hdset1 for h in hdset2):
                self.report['info'].append((f" - {sheetname_2} has following detail columns that {sheetname_1} doesn't:", ))
                for h in hdset2:
                    if h not in hdset1:
                        self.report['info'].append((f"   - {h}", ))
        else:
            self.report['info'].append((" - Sheets have no matching detail column names", ))
        shared_ids = False
        if any(node in self.nodes2 for node in self.nodes1):
            shared_ids = True
        if not shared_ids:
            if any(node in self.nodes1 for node in self.nodes2):
                shared_ids = True
        if shared_ids:
            if matching_hrs_names:
                if self.row_len1 >= self.row_len2:
                    for rn,row in enumerate(self.sheet2):
                        ID = row[self.ic2]
                        ik = ID.lower()
                        if ik in self.nodes1:
                            rnstr = f"{rn + 2}"
                            for nx in matching_hrs_names:
                                h1 = pcold[nx][0]
                                h2 = pcold[nx][1]
                                if self.nodes1[ik].ps[h1]:
                                    p1 = self.nodes1[ik].ps[h1].k
                                else:
                                    p1 = self.nodes1[ik].ps[h1]
                                if self.nodes2[ik].ps[h2]:
                                    p2 = self.nodes2[ik].ps[h2].k
                                else:
                                    p2 = self.nodes2[ik].ps[h2]
                                if p1 != p2 and p1 is None:
                                    if p2 == "":
                                        self.report['ids'].append((f"{ID}",
                                                                   f"Present in hierarchy: {nx} in {sheetname_2} and not {sheetname_1}",
                                                                   f"Not present",
                                                                   f"Appears as top ID"))
                                    elif p2:
                                        self.report['ids'].append((f"{ID}",
                                                                   f"Present in hierarchy: {nx} in {sheetname_2} and not {sheetname_1}",
                                                                   f"Not present",
                                                                   f"{self.nodes2[ik].ps[h2].name}"))
                                elif p1 != p2 and p2 is None:
                                    if p1 == "":
                                        self.report['ids'].append((f"{ID}",
                                                                   f"Present in hierarchy: {nx} in {sheetname_1} and not {sheetname_2}",
                                                                   f"Appears as top ID",
                                                                   f"Not present"))
                                    elif p1:
                                        self.report['ids'].append((f"{ID}",
                                                                   f"Present in hierarchy: {nx} in {sheetname_1} and not {sheetname_2}",
                                                                   f"{self.nodes1[ik].ps[h1].name}",
                                                                   f"Not present"))
                                elif p1 != p2 and p1 == "":
                                    self.report['ids'].append((f"{ID}",
                                                               f"Parents in hierarchy: {nx}",
                                                               f"Appears as top ID",
                                                               f"{self.nodes2[ik].ps[h2].name}"))
                                elif p1 != p2 and p2 == "":
                                    self.report['ids'].append((f"{ID}",
                                                               f"Parents in hierarchy: {nx}",
                                                               f"{self.nodes1[ik].ps[h1].name}",
                                                               f"Appears as top ID"))
                                elif p1 != p2:
                                    self.report['ids'].append((f"{ID}",
                                                               f"Parents in hierarchy: {nx}",
                                                               f"{self.nodes1[ik].ps[h1].name}",
                                                               f"{self.nodes2[ik].ps[h2].name}"))
                            for nx in matching_details_names:
                                c1 = self.sheet1[self.rns1[ik]][detcold[nx][0]]
                                c2 = row[detcold[nx][1]]
                                if c1.lower() != c2.lower():
                                    self.report['ids'].append((f"{ID}",
                                                               f"Details in column: {nx}",
                                                               f"{c1}",
                                                               f"{c2}"))
                elif self.row_len1 < self.row_len2:
                    for rn,row in enumerate(self.sheet1):
                        ID = row[self.ic1]
                        ik = ID.lower()
                        if ik in self.nodes2:
                            rnstr = f"{rn + 2}"
                            for nx in matching_hrs_names:
                                h1 = pcold[nx][0]
                                h2 = pcold[nx][1]
                                if self.nodes1[ik].ps[h1]:
                                    p1 = self.nodes1[ik].ps[h1].k
                                else:
                                    p1 = self.nodes1[ik].ps[h1]
                                if self.nodes2[ik].ps[h2]:
                                    p2 = self.nodes2[ik].ps[h2].k
                                else:
                                    p2 = self.nodes2[ik].ps[h2]
                                if p1 != p2 and p1 is None:
                                    if p2 == "":
                                        self.report['ids'].append((f"{ID}",
                                                                   f"Present in hierarchy: {nx} in {sheetname_2} and not {sheetname_1}",
                                                                   f"Not present",
                                                                   f"Appears as top ID"))
                                    elif p2:
                                        self.report['ids'].append((f"{ID}",
                                                                   f"Present in hierarchy: {nx} in {sheetname_2} and not {sheetname_1}",
                                                                   f"Not present",
                                                                   f"{self.nodes2[ik].ps[h2].name}"))
                                elif p1 != p2 and p2 is None:
                                    if p1 == "":
                                        self.report['ids'].append((f"{ID}",
                                                                   f"Present in hierarchy: {nx} in {sheetname_1} and not {sheetname_2}",
                                                                   f"Appears as top ID",
                                                                   f"Not present"))
                                    elif p1:
                                        self.report['ids'].append((f"{ID}",
                                                                   f"Present in hierarchy: {nx} in {sheetname_1} and not {sheetname_2}",
                                                                   f"{self.nodes1[ik].ps[h1].name}",
                                                                   f"Not present"))
                                elif p1 != p2 and p1 == "":
                                    self.report['ids'].append((f"{ID}",
                                                               f"Parents in hierarchy: {nx}",
                                                               f"Appears as top ID",
                                                               f"{self.nodes2[ik].ps[h2].name}"))
                                elif p1 != p2 and p2 == "":
                                    self.report['ids'].append((f"{ID}",
                                                               f"Parents in hierarchy: {nx}",
                                                               f"{self.nodes1[ik].ps[h1].name}",
                                                               f"Appears as top ID"))
                                elif p1 != p2:
                                    self.report['ids'].append((f"{ID}",
                                                               f"Parents in hierarchy: {nx}",
                                                               f"{self.nodes1[ik].ps[h1].name}",
                                                               f"{self.nodes2[ik].ps[h2].name}"))
                            for nx in matching_details_names:
                                c1 = row[detcold[nx][0]]
                                c2 = self.sheet2[self.rns2[ik]][detcold[nx][1]]
                                if c1.lower() != c2.lower():
                                    self.report['ids'].append((f"{ID}",
                                                               f"Details in column: {nx}",
                                                               f"{c1}",
                                                               f"{c2}"))
            elif not matching_hrs_names:
                if self.row_len1 >= self.row_len2:
                    for rn,row in enumerate(self.sheet2):
                        ID = row[self.ic2]
                        ik = ID.lower()
                        if ik in self.nodes1:
                            rnstr = f"{rn + 2}"
                            for nx in matching_details_names:
                                c1 = self.sheet1[self.rns1[ik]][detcold[nx][0]]
                                c2 = row[detcold[nx][1]]
                                if c1.lower() != c2.lower():
                                    self.report['ids'].append((f"{ID}",
                                                               f"Details in column: {nx}",
                                                               f"{c1}",
                                                               f"{c2}"))
                elif self.row_len1 < self.row_len2:
                    for rn,row in enumerate(self.sheet1):
                        ID = row[self.ic1]
                        ik = ID.lower()
                        if ik in self.nodes2:
                            rnstr = f"{rn + 2}"
                            for nx in matching_details_names:
                                c1 = row[detcold[nx][0]]
                                c2 = self.sheet2[self.rns2[ik]][detcold[nx][1]]
                                if c1.lower() != c2.lower():
                                    self.report['ids'].append((f"{ID}",
                                                               f"Details in column: {nx}",
                                                               f"{c1}",
                                                               f"{c2}"))
            missids1 = False
            missids2 = False
            if any(ik not in self.nodes2 for ik in self.nodes1):
                missids1 = True
            if any(ik not in self.nodes1 for ik in self.nodes2):
                missids2 = True
            if missids1 or missids2:
                self.report['info'].append(("", ))
                self.report['info'].append(("DELETED OR ADDED IDS", ))
            if missids1:
                self.report['info'].append((f" - {sheetname_1} has the following IDs that {sheetname_2} doesn't:", ))
                self.report['info'].extend([(f"   - {self.nodes1[ik].name}", ) for ik in self.nodes1 if ik not in self.nodes2])
            if missids2:
                self.report['info'].append((f" - {sheetname_2} has the following IDs that {sheetname_1} doesn't:", ))
                self.report['info'].extend([(f"   - {self.nodes2[ik].name}", ) for ik in self.nodes2 if ik not in self.nodes1])
        self.stop_work("Program ready")
        self.sheetname_1 = sheetname_1
        self.sheetname_2 = sheetname_2
        report_window = compare_report_popup(self, theme = self.C.theme)

    def srtkey(self,e):
        return [int(c) if c.isdigit() else c.lower() for c in re.split("([0-9]+)",e)]
