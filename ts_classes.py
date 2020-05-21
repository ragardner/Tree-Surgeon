#Copyright © 2020 R. A. Gardner

from ts_extra_vars_and_funcs import *
import tkinter as tk
from tkinter import ttk, filedialog
import csv as csv_module
from collections import defaultdict, deque, Counter
import os
import datetime
import re
from itertools import islice, repeat, chain, cycle
import json
import pickle
import zlib
import lzma
from base64 import b32encode as b32e, b32decode as b32d
import csv as csv_module
from openpyxl import Workbook, load_workbook
import io
import datetime
from tksheet import Sheet
from math import floor
from ast import literal_eval
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from platform import system as get_os
from fastnumbers import isint, isintlike, isfloat, isreal


class id_and_parent_column_selector(tk.Frame):
    def __init__(self,
                 parent,
                 headers=[[]],
                 show_disp_1=True,
                 show_disp_2=True,
                 theme = "light",
                 expand = False):
        tk.Frame.__init__(self,parent,background=theme_bg(theme))
        self.grid_propagate(False)
        self.grid_rowconfigure(1,weight=1)
        if show_disp_1:
            self.grid_columnconfigure(0,weight=1,uniform="x")
        if show_disp_2:
            self.grid_columnconfigure(1,weight=1,uniform="x")
        self.C = parent
        self.headers = headers
        self.id_col = None
        self.par_cols = set()
        self.id_col_display = readonly_entry_with_scrollbar(self, font = EFB, theme = theme)
        self.id_col_display.set_my_value("   ID column:   ")
        if show_disp_1:
            self.id_col_display.grid(row=0,column=0,sticky="nswe")
        self.id_col_selection = Sheet(self,
                                      height = 280 if not expand else None,
                                      width = 250 if not expand else None,
                                      theme = theme,
                                      show_selected_cells_border = False,
                                              align="center",
                                              header_align="center",
                                              row_index_align="center",
                                            selected_cells_background="#0078d7",
                                      selected_cells_foreground = "white",
                                      header_select_foreground="white",
                                         row_index_select_foreground="white",
                                              header_select_background="#0078d7",
                                              row_index_select_background="#0078d7",
                                              header_font = ("Calibri", 13, "normal"),
                                              column_width=170,
                                              row_index_width=60,
                                              headers=["SELECT ID"])
        self.id_col_selection.data_reference(newdataref=self.headers)
        self.id_col_selection.enable_bindings(("single",
                                                "column_width_resize",
                                                "double_click_column_resize"))
        self.id_col_selection.extra_bindings([("cell_select",self.id_col_selection_B1),
                                              ("deselect",self.id_col_selection_B1)])
        if show_disp_1:
            self.id_col_selection.grid(row=1,column=0,sticky="nswe")
        self.par_col_selection = Sheet(self,
                                       height = 280 if not expand else None,
                                       width = 250 if not expand else None,
                                        theme = theme,
                                              align="center",
                                       show_selected_cells_border = False,
                                              header_align="center",
                                              row_index_align="center",
                                              selected_cells_background="#8cba66",
                                       selected_cells_foreground = "white",
                                       header_select_foreground="white",
                                         row_index_select_foreground="white",
                                              header_select_background="#8cba66",
                                              row_index_select_background="#8cba66",
                                              header_font = ("Calibri", 13, "normal"),
                                              column_width=170,
                                              row_index_width=60,
                                              headers=["SELECT PARENTS"])
        self.par_col_selection.data_reference(newdataref=self.headers)
        self.par_col_selection.extra_bindings([("cell_select",self.par_col_selection_B1),
                                               ("deselect",self.par_col_deselection_B1)])
        self.par_col_selection.enable_bindings(("toggle",
                                                "column_width_resize",
                                                "double_click_column_resize"))
        if show_disp_2:
            self.par_col_selection.grid(row=1,column=1,sticky="nswe")
        self.par_col_display = readonly_entry_with_scrollbar(self, font = EFB, theme = theme)
        self.par_col_display.set_my_value("   Parent columns:   ")
        if show_disp_2:
            self.par_col_display.grid(row=0,column=1,sticky="nswe")
        self.detect_id_col_button = button(self,text="Detect ID column",style="EFB.Std.TButton",
                                           command=self.detect_id_col)
        if show_disp_1:
            self.detect_id_col_button.grid(row=2,column=0,padx=2,pady=2,sticky="ns")
        self.detect_par_cols_button = button(self,text="Detect parent columns",style="EFB.Std.TButton",
                                             command=self.detect_par_cols)
        if show_disp_2:
            self.detect_par_cols_button.grid(row=2,column=1,padx=2,pady=2,sticky="ns")

    def reset_size(self,width=500,height=350):
        self.config(width=width,height=height)
        self.update_idletasks()
        self.par_col_selection.refresh()
        self.id_col_selection.refresh()

    def set_columns(self,columns):
        self.clear_displays()
        self.set_par_cols([])
        self.headers = [[h] for h in columns]
        self.id_col_selection.data_reference(newdataref=self.headers,redraw=True)
        self.par_col_selection.data_reference(newdataref=self.headers,redraw=True)

    def disable_me(self):
        self.id_col_selection.basic_bindings(enable=False)
        self.par_col_selection.basic_bindings(enable=False)
        self.detect_id_col_button.config(state="disabled")
        self.detect_par_cols_button.config(state="disabled")

    def enable_me(self):
        self.id_col_selection.basic_bindings(enable=True)
        self.par_col_selection.basic_bindings(enable=True)
        self.detect_id_col_button.config(state="normal")
        self.detect_par_cols_button.config(state="normal")

    def detect_id_col(self):
        for i,e in enumerate(self.headers):
            if not e:
                continue
            x = e[0].lower().strip()
            if x == "id" or x.startswith("id"):
                self.set_id_col(i)
                break

    def detect_par_cols(self):
        parent_cols = []
        for i,e in enumerate(self.headers):
            if not e:
                continue
            x = e[0].lower().strip()
            if x.startswith("parent"):
                parent_cols.append(i)
        if parent_cols:
            self.set_par_cols(parent_cols)

    def id_col_selection_B1(self,event=None):
        if event is not None:
            self.id_col = tuple(tup[0] for tup in self.id_col_selection.get_selected_cells())
            if self.id_col:
                self.id_col = self.id_col[0]
                self.id_col_display.set_my_value(f"   ID column:   {self.id_col + 1}")
            else:
                self.id_col = None
                self.id_col_display.set_my_value("   ID column:   ")

    def par_col_selection_B1(self,event=None):
        if event is not None:
            self.par_cols = set(tup[0] for tup in self.par_col_selection.get_selected_cells())
            self.par_col_display.set_my_value("   Parent columns:   " + ", ".join([str(n) for n in sorted(p + 1 for p in self.par_cols)]))

    def par_col_deselection_B1(self,event=None):
        if event is not None:
            self.par_cols = set(tup[0] for tup in self.par_col_selection.get_selected_cells())
            self.par_col_display.set_my_value("   Parent columns:   " + ", ".join([str(n) for n in sorted(p + 1 for p in self.par_cols)]))

    def clear_displays(self):
        self.headers = [[]]
        self.id_col = None
        self.id_col_selection.deselect("all")
        self.id_col_display.set_my_value("   ID column:   ")
        self.par_cols = set()
        self.par_col_selection.deselect("all")
        self.par_col_display.set_my_value("   Parent columns:   ")
        self.id_col_selection.data_reference(newdataref=[[]],redraw=True)
        self.par_col_selection.data_reference(newdataref=[[]],redraw=True)

    def set_id_col(self,col):
        self.id_col = col
        self.id_col_selection.deselect("all")
        self.id_col_selection.refresh()
        self.id_col_selection.select_cell(row=col,column=0,redraw=True)
        self.id_col_selection.see(row=col,column=0)
        self.id_col_display.set_my_value("   ID column:   " + str(col + 1))

    def set_par_cols(self, cols):
        self.par_col_selection.deselect("all")
        self.par_col_selection.refresh()
        if cols:
            self.par_cols = set(cols)
            for r in cols:
                self.par_col_selection.toggle_select_cell(r, 0, redraw = False)
            self.par_col_selection.see(row = cols[0], column = 0)
            self.par_col_display.set_my_value("   Parent columns:   " + ", ".join([f"{n}" for n in sorted(p + 1 for p in self.par_cols)]))

    def get_id_col(self):
        return self.id_col

    def get_par_cols(self):
        return tuple(sorted(self.par_cols))

    def change_theme(self, theme = "dark"):
        self.id_col_selection.change_theme(theme)
        self.par_col_selection.change_theme(theme)
        self.id_col_selection.set_options(selected_cells_background="#0078d7",
                                            selected_cells_foreground = "white",
                                            header_select_foreground="white",
                                            row_index_select_foreground="white",
                                            header_select_background="#0078d7",
                                            row_index_select_background="#0078d7")
        self.par_col_selection.set_options(selected_cells_background="#8cba66",
                                            selected_cells_foreground = "white",
                                            header_select_foreground="white",
                                            row_index_select_foreground="white",
                                            header_select_background="#8cba66",
                                            row_index_select_background="#8cba66")
        self.config(background=theme_bg(theme))
        self.id_col_display.my_entry.config(background = theme_entry_bg(theme),
                                  foreground = theme_entry_fg(theme),
                                  disabledbackground = theme_entry_dbg(theme),
                                  disabledforeground = theme_entry_dfg(theme),
                                  insertbackground = theme_entry_cursor(theme),
                                  readonlybackground = theme_entry_dbg(theme)
                                  )
        self.par_col_display.my_entry.config(background = theme_entry_bg(theme),
                                  foreground = theme_entry_fg(theme),
                                  disabledbackground = theme_entry_dbg(theme),
                                  disabledforeground = theme_entry_dfg(theme),
                                  insertbackground = theme_entry_cursor(theme),
                                  readonlybackground = theme_entry_dbg(theme)
                                  )

class flattened_base_ids_choices(tk.Frame):
    def __init__(self,
                 parent,
                 command,
                 theme = "dark"):
        tk.Frame.__init__(self,parent,background = theme_bg(theme))
        self.C = parent
        self.extra_func = command
        self.only_base_ids_button = x_checkbutton(self,
                                                  text="Sheet is flattened with base IDs   ",
                                                  style="x_button.Std.TButton",
                                                  compound="right",
                                                  command=self.flattened_mode_toggle,
                                                  checked=False)
        self.only_base_ids_button.grid(row=0,column=0,sticky="nswe",padx=(0,5),pady=(0,5))
        self.order_dropdown = ez_dropdown(self,font=EF)
        self.order_dropdown.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
        self.order_dropdown['values'] = ["Order: Top → Base","Order: Base → Top"]
        self.order_dropdown.set_my_value("Order: Top → Base")
        self.order_dropdown.grid(row=0,column=1,sticky="nswe")
        self.flattened_mode_toggle(func=False)

    def disable_me(self):
        self.only_base_ids_buttons.config(state="disabled")
        self.order_dropdown.config(state="disabled")

    def enable_me(self):
        self.only_base_ids_buttons.config(state="normal")
        self.flattened_mode_toggle(func=False)

    def flattened_mode_toggle(self,event=None,func=True):
        if self.only_base_ids_button.get_checked():
            self.order_dropdown.config(state="readonly")
        else:
            self.order_dropdown.config(state="disabled")
        if func:
            self.extra_func()

    def set_checked(self,s=True):
        self.only_base_ids_button.set_checked(s)

    def get_choices(self):
        return self.only_base_ids_button.get_checked(), self.order_dropdown.get_my_value(), True

    def change_theme(self, theme = "dark"):
        self.config(bg = theme_bg(theme))


class flattened_column_selector(tk.Frame):
    def __init__(self,
                 parent,
                 headers=[[]],
                 theme = "dark"):
        tk.Frame.__init__(self, parent, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.grid_rowconfigure(1,weight=1)
        self.grid_columnconfigure(0,weight=1)
        self.C = parent
        self.headers = headers
        self.par_cols = set()
        self.par_col_display = readonly_entry_with_scrollbar(self,font=EFB, theme = theme)
        self.par_col_display.set_my_value("   Hierarchy columns:   ")
        self.par_col_display.grid(row=0,column=0,sticky="nswe")
        self.par_col_selection = Sheet(self,
                                       width = 500,
                                       height = 300,
                                       theme = theme,
                                              align="center",
                                       show_selected_cells_border = False,
                                              header_align="center",
                                              row_index_align="center",
                                              selected_cells_background="#ffa51e",
                                              header_select_background="#ffa51e",
                                              row_index_select_background="#ffa51e",
                                             header_select_foreground="black",
                                             row_index_select_foreground="black",
                                             selected_cells_foreground="black",
                                              header_font = ("Calibri", 13, "normal"),
                                              column_width=350,
                                              row_index_width=60,
                                              headers=["SELECT ALL HIERARCHY COLUMNS"])
        self.par_col_selection.data_reference(newdataref=self.headers)
        self.par_col_selection.extra_bindings([("cell_select",self.par_col_selection_B1),
                                               ("deselect",self.par_col_deselection_B1)])
        self.par_col_selection.enable_bindings(("toggle",
                                                "column_width_resize",
                                                "double_click_column_resize"))
        self.par_col_selection.grid(row=1,column=0,sticky="nswe")
        
    def set_columns(self,columns):
        self.clear_displays()
        self.set_par_cols([])
        self.headers = [[h] for h in columns]
        self.par_col_selection.data_reference(newdataref=self.headers,redraw=True)

    def disable_me(self):
        self.par_col_selection.basic_bindings(enable=False)
        self.par_col_selection.extra_bindings([("cell_select",None),
                                               ("deselect",None)])

    def enable_me(self):
        self.par_col_selection.basic_bindings(enable=True)
        self.par_col_selection.extra_bindings([("cell_select",self.par_col_selection_B1),
                                               ("deselect",self.par_col_deselection_B1)])

    def detect_par_cols(self):
        parent_cols = []
        for i,e in enumerate(self.headers):
            if not e:
                continue
            x = e[0].lower().strip()
            if x.startswith("parent"):
                parent_cols.append(i)
        if parent_cols:
            self.set_par_cols(parent_cols)

    def par_col_selection_B1(self,event=None):
        if event is not None:
            self.par_cols = set(tup[0] for tup in self.par_col_selection.get_selected_cells())
            self.par_col_display.set_my_value("   Parent columns:   " + ", ".join([str(n) for n in sorted(p + 1 for p in self.par_cols)]))

    def par_col_deselection_B1(self,event=None):
        if event is not None:
            self.par_cols = set(tup[0] for tup in self.par_col_selection.get_selected_cells())
            self.par_col_display.set_my_value("   Parent columns:   " + ", ".join([str(n) for n in sorted(p + 1 for p in self.par_cols)]))
            
    def clear_displays(self):
        self.headers = [[]]
        self.par_col_selection.data_reference(newdataref=[[]],redraw=True)
        self.par_cols = set()
        self.par_col_selection.deselect("all")
        self.par_col_display.set_my_value("   Hierarchy columns:   ")

    def set_par_cols(self,cols):
        self.par_col_selection.deselect("all")
        self.par_col_selection.refresh()
        if cols:
            self.par_cols = set(cols)
            for r in cols:
                self.par_col_selection.toggle_select_cell(r, 0, redraw = False)
            self.par_col_selection.see(row=cols[0],column=0)
            self.par_col_display.set_my_value("   Hierarchy columns:   " + ",".join([str(n) for n in sorted(p + 1 for p in self.par_cols)]))
        self.par_col_selection.refresh()

    def get_par_cols(self):
        return tuple(sorted(self.par_cols))

    def change_theme(self, theme = "dark"):
        self.config(bg = theme_bg(theme))
        self.par_col_selection.change_theme(theme)
        self.par_col_selection.set_options(selected_cells_background="#ffa51e",
                                              header_select_background="#ffa51e",
                                              row_index_select_background="#ffa51e",
                                             header_select_foreground="black",
                                             row_index_select_foreground="black",
                                             selected_cells_foreground="black")
        self.par_col_display.my_entry.config(background = theme_entry_bg(theme),
                                  foreground = theme_entry_fg(theme),
                                  disabledbackground = theme_entry_dbg(theme),
                                  disabledforeground = theme_entry_dfg(theme),
                                  insertbackground = theme_entry_cursor(theme),
                                  readonlybackground = theme_entry_dbg(theme)
                                  )


class single_column_selector(tk.Frame):
    def __init__(self,
                 parent,
                 headers=[[]],
                 width=250,
                 height=350,
                 theme = "dark"):
        tk.Frame.__init__(self,parent,width=width,height=height, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.grid_rowconfigure(1,weight=1)
        self.grid_columnconfigure(0,weight=1)
        self.C = parent
        self.headers = headers
        self.col = None
        self.col_display = readonly_entry_with_scrollbar(self, font = EFB, theme = theme)
        self.col_display.set_my_value("  Column:   ")
        self.col_display.grid(row=0,column=0,sticky="nswe")
        self.col_selection = Sheet(self,
                                   theme = theme,
                                   show_selected_cells_border = False,
                                         align="center",
                                          header_align="center",
                                          row_index_align="center",
                                          selected_cells_background="#ffa51e",
                                          header_select_background="#ffa51e",
                                          row_index_select_background="#ffa51e",
                                         header_select_foreground="black",
                                         row_index_select_foreground="black",
                                         selected_cells_foreground="black",
                                          header_font = ("Calibri", 13, "normal"),
                                          column_width=180,
                                          row_index_width=50,
                                          headers=["SELECT A COLUMN"])
        self.col_selection.data_reference(newdataref=self.headers)
        self.col_selection.extra_bindings([("cell_select",self.col_selection_B1),
                                           ("deselect", self.col_deselect)])
        self.col_selection.enable_bindings(("single",
                                            "column_width_resize",
                                            "double_click_column_resize"))
        self.col_selection.grid(row=1,column=0,sticky="nswe")
        
    def set_columns(self,columns):
        self.clear_displays()
        self.headers = [[h] for h in columns]
        self.col_selection.data_reference(newdataref=self.headers,redraw=True)

    def disable_me(self):
        self.col_selection.basic_bindings(enable=False)

    def enable_me(self):
        self.col_selection.basic_bindings(enable=True)

    def col_selection_B1(self,event=None):
        if event is not None:
            self.col = event[1]
            self.col_display.set_my_value("   Column:   " + str(event[1] + 1))

    def col_deselect(self, event = None):
        if event is not None:
            self.col = None
            self.col_display.set_my_value("  Column:   ")

    def par_col_deselection_B1(self,event=None):
        if event is not None:
            self.par_cols = set(tup[0] for tup in self.par_col_selection.get_selected_cells())
            self.par_col_display.set_my_value("   Parent columns:   " + ", ".join([str(n) for n in sorted(p + 1 for p in self.par_cols)]))

    def clear_displays(self):
        self.headers = [[]]
        self.col_selection.data_reference(newdataref=[[]],redraw=True)
        self.col = 0
        self.col_selection.deselect("all")
        self.col_display.set_my_value("   Hierarchy columns:   ")

    def set_col(self,col = None):
        if col is not None:
            self.col = int(col)
            self.col_selection.deselect("all")
            self.col_selection.select_cell(col, 0, redraw = False)
            self.col_selection.see(row=col,column=0)
            self.col_selection.refresh()
            self.col_display.set_my_value("   Column:   " + str(col + 1))

    def get_col(self):
        return int(self.col)


class export_flattened_popup(tk.Toplevel):
    def __init__(self,C,
                 width=1280,
                 height=800,
                 theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.title("Export flattened sheet - Click the X button or press escape to go back")
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.wb_ = None

        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(0,weight=1)

        self.selector = single_column_selector(self, theme = theme)
        self.selector.enable_me()
        
        self.selector.set_columns([self.C.headers[h].name for h in self.C.hiers])
        if self.C.pc == -1:
            self.selector.set_col(0)
        else:
            self.selector.set_col(self.C.hiers.index(self.C.pc))
        self.selector.grid(row=0,column=0,sticky="nwe",pady=(10,20),padx=10)

        self.rename_id_col_button = x_checkbutton(self,
                                              text="Rename ID column  ",
                                              style="x_button.Std.TButton",
                                              compound="right",
                                              checked=self.C.xlsx_flattened_rename_id_col.get())
        self.rename_id_col_button.grid(row=1,column=0,sticky="new",pady=(10,5),padx=10)

        self.add_index_button = x_checkbutton(self,
                                              text="Add index column  ",
                                              style="x_button.Std.TButton",
                                              compound="right",
                                              checked=self.C.xlsx_flattened_add_index.get())
        self.add_index_button.grid(row=2,column=0,sticky="new",pady=(10,5),padx=10)

        self.remove_details_button = x_checkbutton(self,
                                                      text="Remove other columns  ",
                                                      style="x_button.Std.TButton",
                                                      compound="right",
                                                      checked=self.C.xlsx_flattened_details.get())
        self.remove_details_button.grid(row=3,column=0,sticky="new",pady=(10,5),padx=10)

        self.split_1st_det_col_button = x_checkbutton(self,
                                                      text="Split treeview detail column  ",
                                                      style="x_button.Std.TButton",
                                                      compound="right",
                                                      checked=self.C.xlsx_flattened_split_1st_det_col.get())
        self.split_1st_det_col_button.grid(row=4,column=0,sticky="new",pady=(10,5),padx=10)

        self.only_base_ids_button = x_checkbutton(self,
                                                  text="Base IDs mode  ",
                                                  style="x_button.Std.TButton",
                                                  compound="right",
                                                  checked=self.C.xlsx_flattened_base_ids.get())
        self.only_base_ids_button.grid(row=5,column=0,sticky="new",pady=(10,5),padx=10)

        self.justify_rows_button = x_checkbutton(self,
                                                     text="Justify rows  ",
                                                     style="x_button.Std.TButton",
                                                     compound="right",
                                                     checked=self.C.xlsx_flattened_justify.get())
        self.justify_rows_button.grid(row=6,column=0,sticky="new",pady=5,padx=10)

        self.order_button = x_checkbutton(self,
                                            text="Reverse Order  ",
                                            style="x_button.Std.TButton",
                                            compound="right",
                                           checked=self.C.xlsx_flattened_reverse_order.get())
        self.order_button.grid(row=7,column=0,sticky="new",pady=5,padx=10)

        self.build_button = button(self,text="  Flatten sheet  ",
                                       style="EF.Std.TButton",
                                       command=self.build_flattened)
        self.build_button.grid(row=8,column=0,pady=10,padx=10,sticky="nsew")
        
        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                  outline_thickness = 0)
        self.sheetdisplay.enable_bindings("enable_all")
        self.sheetdisplay.extra_bindings("begin_edit_cell_use_keypress", self.begin_edit)
        self.sheetdisplay.extra_bindings("escape_edit_cell", self.escape_edit)
        self.sheetdisplay.extra_bindings("edit_cell", self.escape_edit)
        self.sheetdisplay.headers(newheaders = 0)
        self.sheetdisplay.grid(row=0,column=1,rowspan=7,sticky="nswe")

        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(2,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(3,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.button_frame.grid(row=8,column=1,sticky="nswe")
        self.save_button = button(self.button_frame,text="Save as",
                                       style="EF.Std.TButton",
                                       command=self.save_as)
        self.save_button.grid(row=0,column=0,padx=5,pady=20,sticky="nswe")
        self.clipboard_json_button = button(self.button_frame,text="Clipboard as json",
                                               style="EF.Std.TButton",
                                               command=self.clipboard_json)
        self.clipboard_json_button.grid(row=0,column=1,padx=5,pady=20,sticky="nswe")
        self.clipboard_indent_button = button(self.button_frame,text="Clipboard (indent separated)",
                                               style="EF.Std.TButton",
                                               command=self.clipboard_indent)
        self.clipboard_indent_button.grid(row=0,column=2,padx=5,pady=20,sticky="nswe")
        self.clipboard_comma_button = button(self.button_frame,text="Clipboard (comma separated)",
                                               style="EF.Std.TButton",
                                               command=self.clipboard_comma)
        self.clipboard_comma_button.grid(row=0,column=3,padx=5,pady=20,sticky="nswe")
        self.status_bar = StatusBar(self, text = "Use the parent column selector to change hierarchy output", theme = theme)
        self.status_bar.grid(row=9,column=0,columnspan=2,sticky="nswe")
        
        self.bind("<Escape>",self.cancel)
        self.build_flattened()
        
        center(self,width,height)
        self.deiconify()
        self.wait_window()

    def escape_edit(self, event = None):
        self.bind("<Escape>",self.cancel)
        
    def begin_edit(self, event = None):
        self.unbind("<Escape>")

    def start_work(self,msg=""):
        self.status_bar.change_text(msg)
        self.disable_widgets()

    def stop_work(self,msg=""):
        self.status_bar.change_text(msg)
        self.enable_widgets()

    def enable_widgets(self):
        self.sheetdisplay.enable_bindings("enable_all")
        self.sheetdisplay.extra_bindings("begin_edit_cell_use_keypress", self.begin_edit)
        self.sheetdisplay.extra_bindings("escape_edit_cell", self.escape_edit)
        self.sheetdisplay.extra_bindings("edit_cell", self.escape_edit)
        self.sheetdisplay.basic_bindings(True)
        self.save_button.config(state="normal")
        self.clipboard_indent_button.config(state="normal")
        self.clipboard_json_button.config(state="normal")
        self.clipboard_comma_button.config(state="normal")
        self.build_button.config(state="normal")
        self.selector.enable_me()

    def disable_widgets(self):
        self.build_button.config(state="disabled")
        self.sheetdisplay.disable_bindings("disable_all")
        self.sheetdisplay.extra_bindings("begin_edit_cell_use_keypress", None)
        self.sheetdisplay.extra_bindings("escape_edit_cell", None)
        self.sheetdisplay.extra_bindings("edit_cell", None)
        self.sheetdisplay.basic_bindings(False)
        self.save_button.config(state="disabled")
        self.clipboard_json_button.config(state="disabled")
        self.clipboard_indent_button.config(state="disabled")
        self.clipboard_comma_button.config(state="disabled")
        self.selector.disable_me()
        self.update()

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except:
            pass
        try:
            self.wb_ = None
        except:
            pass

    def USER_HAS_CLOSED_WINDOW(self,callback=None):
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except:
            pass
        self.destroy()

    def clipboard_json(self):
        self.start_work("Copying to clipboard...")
        self.C.C.clipboard_clear()
        self.C.C.clipboard_append(json.dumps(self.C.dump_full_sheet_to_json(self.sheetdisplay.get_sheet_data()[0],self.sheetdisplay.get_sheet_data()[1:],include_headers=True)))
        self.C.C.update()
        self.stop_work("Sheet successfully copied to clipboard as json!")

    def clipboard_indent(self):
        self.start_work("Copying to clipboard...")
        s = io.StringIO()
        writer = csv_module.writer(s,dialect=csv_module.excel_tab,lineterminator="\n")
        for row in self.sheetdisplay.get_sheet_data():
            writer.writerow(row)
        s = s.getvalue().rstrip()
        self.C.C.clipboard_clear()
        self.C.C.clipboard_append(s)
        self.C.C.update()
        self.stop_work("Sheet successfully copied to clipboard (indent separated)!")

    def clipboard_comma(self):
        self.start_work("Copying to clipboard...")
        s = io.StringIO()
        writer = csv_module.writer(s,dialect=csv_module.excel,lineterminator="\n")
        for row in self.sheetdisplay.get_sheet_data():
            writer.writerow(row)
        s = s.getvalue().rstrip()
        self.C.C.clipboard_clear()
        self.C.C.clipboard_append(s)
        self.C.C.update()
        self.stop_work("Sheet successfully copied to clipboard (comma separated)!")

    def build_flattened(self):
        self.start_work("Flattening sheet...")
        self.sheetdisplay.deselect("all")
        self.sheetdisplay.set_sheet_data(data = self.C.build_flattened(self.C.sheet,
                                                                     self.sheetdisplay.get_sheet_data(),
                                                                     [f"{hdr.name}" for hdr in self.C.headers],
                                                                     int(self.C.ic),
                                                                     int(self.C.hiers[self.selector.get_col()]),
                                                                     list(self.C.hiers),
                                                                     self.justify_rows_button.get_checked(),
                                                                     self.order_button.get_checked(),
                                                                     self.only_base_ids_button.get_checked(),
                                                                     self.remove_details_button.get_checked(),
                                                                     self.split_1st_det_col_button.get_checked(),
                                                                     self.add_index_button.get_checked(),
                                                                     self.rename_id_col_button.get_checked()),
                                         verify = False)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.stop_work("Sheet successfully flattened!")
        
    def save_as(self):
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save flattened sheet as",
                                               filetypes=[('Excel file','.xlsx'),('JSON File','.json'),('CSV File','.csv')],
                                               defaultextension=".xlsx",
                                               confirmoverwrite=True)
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv",".xlsx",".json")):
            self.grab_set()
            self.stop_work("Can only save .json/.csv/.xlsx file types")
            return
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook()
                ws = self.wb_.active
                for rn,row in enumerate(self.sheetdisplay.get_sheet_data()):
                    ws.append(row)
                    if not rn % 50:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status_bar.change_text("".join(("Saving...  rows: ",str(rn))))
                ws.freeze_panes = "A2"
                if self.sheetdisplay.get_sheet_data():
                    for i in range(1, len(self.sheetdisplay.get_sheet_data()[0]) + 1):
                        ws.cell(row = 1, column = i).fill = orange_fill
                        ws.cell(row = 1, column = i).border = openpyxl_thin_border
                        ws.column_dimensions[xl_column_string(i)].width = 25
                self.wb_.save(newfile)
                self.try_to_close_wb()
            elif newfile.lower().endswith(".json"):
                with open(newfile,"w",newline="") as fh:
                    fh.write(json.dumps(self.C.dump_full_sheet_to_json(self.sheetdisplay.get_sheet_data()[0],self.sheetdisplay.get_sheet_data()[1:],include_headers=True)))             
            elif newfile.lower().endswith(".csv"):
                with open(newfile,"w",newline="") as fh:
                    writer = csv_module.writer(fh,dialect=csv_module.excel_tab,lineterminator="\n")
                    for rn,row in enumerate(self.sheetdisplay.get_sheet_data()):
                        writer.writerow(row)
                        if not rn % 50:
                            self.update()
                            if self.USER_HAS_QUIT:
                                return
                            self.status_bar.change_text("".join(("Saving...  rows: ",str(rn))))
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work("Error saving file: " + str(error_msg))
            return
        self.stop_work("Success! Flattened sheet saved")
        
    def cancel(self,event=None):
        self.USER_HAS_CLOSED_WINDOW()


class post_import_changes_popup(tk.Toplevel):
    def __init__(self, C, changes, successful, width = 1200, height = 800, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.title("Successful / Unsuccessful Changes - Click the X button or press escape to go back")
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.total_changes = "Total changes: " + str(len(changes))
        
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(0,weight=1)
        
        self.changes = changes
        self.successful = successful
        
        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                    outline_thickness=0,
                                    row_index_width=245)
        self.sheetdisplay.enable_bindings(("single",
                                           "copy",
                                           "drag_select",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.sheetdisplay.headers(newheaders=["Date","User","Type","ID/Name/Number","Old Value","New Value"])
        self.sheetdisplay.row_index(0)
        self.sheetdisplay.data_reference(newdataref=self.changes,reset_col_positions=False,reset_row_positions=False,redraw=False)
        self.sheetdisplay.display_subset_of_columns(indexes=[1,2,3,4,5],enable=True,reset_col_positions=False)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        for i, b in enumerate(reversed(self.successful)):
            if b:
                self.sheetdisplay.highlight_cells(row = i, canvas = "row_index", bg = theme_green_bg(theme), fg = theme_green_fg(theme))
                for c in range(6):
                    self.sheetdisplay.highlight_cells(row = i, column = c, bg = theme_green_bg(theme), fg = theme_green_fg(theme))
            else:
                self.sheetdisplay.highlight_cells(row = i, canvas = "row_index", bg = theme_red_bg(theme), fg = theme_red_fg(theme))
                for c in range(6):
                    self.sheetdisplay.highlight_cells(row = i, column = c, bg = theme_red_bg(theme), fg = theme_red_fg(theme))
        self.sheetdisplay.grid(row=0,column=0,sticky="nswe")
        self.status_bar = StatusBar(self, text = self.total_changes, theme = theme)
        self.status_bar.grid(row=1,column=0,sticky="nswe")
        self.bind("<Escape>",self.cancel)
        center(self,width,height)
        self.deiconify()
        self.wait_window()

    def cancel(self,event=None):
        self.destroy()


class changelog_popup(tk.Toplevel):
    def __init__(self,C,width=999,height=800, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.title("Changelog - Click the X button or press escape to go back")
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.find_results = []
        self.results_number = 0
        self.wb_ = None
        self.total_changes = "Total changes: " + str(len(self.C.changelog)) + " | "
        
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        
        self.find_frame = frame(self, theme = theme)
        self.find_frame.grid(row=0,column=0,columnspan=2,sticky="nswe")
        self.find_icon = tk.PhotoImage(format="gif",data=find_icon)
        self.search_button = button(self.find_frame,
                                    text=" Find:",
                                    command=self.find)
        self.search_button.config(image=self.find_icon,compound="left")
        self.search_button.pack(side="left",fill="x")
        self.find_window = normal_entry(self.find_frame,font=BF, theme = theme)
        self.find_window.bind("<Return>",self.find)
        self.find_window.pack(side="left",fill="x",expand=True)
        self.find_reset_button = button(self.find_frame,text="X",command=self.find_reset)
        self.find_reset_button.pack(side="left",fill="x")
        self.find_results_label = label(self.find_frame,"0/0",BF, theme = theme)
        self.find_results_label.pack(side="left",fill="x")
        self.find_up_button = button(self.find_frame,text="▲",command=self.find_up)
        self.find_up_button.pack(side="left",fill="x")
        self.find_down_button = button(self.find_frame,text="▼",command=self.find_down)
        self.find_down_button.pack(side="left",fill="x")
        self.changelog = self.C.changelog[::-1]
        
        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  row_index_align = "w",
                                  header_font = ("Calibri", 13, "normal"),
                                       outline_thickness=0)
        self.sheetdisplay.enable_bindings(("single",
                                           "copy",
                                           "drag_select",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.sheetdisplay.headers(newheaders=["Date","User","Type","ID/Name/Number","Old Value","New Value"])
        self.sheetdisplay.row_index(0)
        self.sheetdisplay.data_reference(newdataref=self.changelog,reset_col_positions=False,reset_row_positions=True,redraw=False)
        self.sheetdisplay.display_subset_of_columns(indexes=[1,2,3,4,5],enable=True,reset_col_positions=False)
        self.red_bg = theme_red_bg(theme)
        self.green_bg = theme_green_bg(theme)
        self.red_fg = theme_red_fg(theme)
        self.green_fg = theme_green_fg(theme)
        self.sheetdisplay.highlight_cells(column = 4,
                                          canvas = "header",
                                          bg = self.red_bg,
                                          fg = self.red_fg)
        self.sheetdisplay.highlight_cells(column = 5,
                                          canvas = "header",
                                          bg = self.green_bg,
                                          fg = self.green_fg)
        for r in range(len(self.changelog)):
            self.sheetdisplay.highlight_cells(row = r,
                                              column = 4,
                                              bg = self.red_bg,
                                              fg = self.red_fg)
            self.sheetdisplay.highlight_cells(row = r,
                                              column = 5,
                                              bg = self.green_bg,
                                              fg = self.green_fg)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.sheetdisplay.set_width_of_index_to_text()
        self.sheetdisplay.grid(row=1,column=0,sticky="nswe")
        self.status_bar = StatusBar(self, text = self.total_changes, theme = theme)
        self.status_bar.grid(row=2,column=0,sticky="nswe")
        
        self.buttonframe = frame(self, theme = theme)
        self.buttonframe.grid_columnconfigure(0, weight = 1, uniform = "b")
        self.buttonframe.grid_columnconfigure(1, weight = 1, uniform = "b")
        self.buttonframe.grid_columnconfigure(2, weight = 1, uniform = "b")
        
        self.buttonframe.grid(row=3,column=0,sticky="nswe")
        self.save_text_button = button(self.buttonframe,text="Export all",
                                       style="EF.Std.TButton",
                                       command=self.save_as)
        self.save_text_button.grid(row = 0, column = 0, padx = 25, pady = 20, sticky = "nswe")

        self.export_selected_button = button(self.buttonframe,text="Export selected as",
                                               style="EF.Std.TButton",
                                               command=self.save_selected_as)
        self.export_selected_button.grid(row = 0, column = 1, padx = 25, pady = 20, sticky = "nswe")
        
        self.prune_button = button(self.buttonframe,text="Prune from selected",
                                    style="EF.Std.TButton",
                                    command=self.prune)
        self.prune_button.grid(row = 0, column = 2, padx = 25, pady = 20, sticky = "nswe")
        
        self.bind("<Escape>",self.cancel)
        center(self,width,height)
        self.deiconify()
        self.wait_window()

    def prune(self, event = None):
        selectedrows = self.sheetdisplay.get_selected_rows(get_cells_as_rows = True, return_tuple = True)
        if not selectedrows:
            return
        num = len(selectedrows)
        self.start_work(f"Pruning {num} changes...")
        up_to = len(self.C.changelog) - min(selectedrows) - 1
        if self.C.changelog[up_to][2].endswith(("|", "| ")):
            for i, entry in enumerate(islice(self.C.changelog, up_to, None), up_to):
                if not entry[2].endswith(("|", "| ")):
                    up_to = i
                    break
        self.C.snapshot_prune_changelog(up_to)
        self.C.changelog[:up_to + 1] = []
        self.changelog = self.C.changelog[::-1]
        self.sheetdisplay.headers(newheaders=["Date","User","Type","ID/Name/Number","Old Value","New Value"])
        self.sheetdisplay.row_index(newindex=0)
        self.sheetdisplay.data_reference(newdataref=self.changelog,reset_col_positions=False,reset_row_positions=True,redraw=False)
        self.sheetdisplay.display_subset_of_columns(indexes=[1,2,3,4,5],enable=True,reset_col_positions=False)
        self.sheetdisplay.dehighlight_cells(all_ = True)
        for r in range(len(self.changelog)):
            self.sheetdisplay.highlight_cells(row = r, column = 4, bg = self.red_bg, fg = self.red_fg)
            self.sheetdisplay.highlight_cells(row = r, column = 5, bg = self.green_bg, fg = self.green_fg)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.total_changes = "Total changes: " + str(len(self.C.changelog)) + " | "
        self.status_bar.config(text = self.total_changes)
        self.C.C.status_bar.change_text(self.C.set_status_bar())
        self.sheetdisplay.refresh()
        self.stop_work("Success! Changelog pruned")

    def start_work(self,msg=""):
        self.status_bar.change_text(self.total_changes + msg)
        self.disable_widgets()

    def stop_work(self,msg=""):
        self.status_bar.change_text(self.total_changes + msg)
        self.enable_widgets()

    def enable_widgets(self):
        self.sheetdisplay.enable_bindings(("single",
                                           "copy",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.find_window.bind("<Return>",self.find)
        self.find_reset_button.config(state="normal")
        self.find_up_button.config(state="normal")
        self.find_down_button.config(state="normal")
        self.save_text_button.config(state="normal")

    def disable_widgets(self):
        self.sheetdisplay.disable_bindings(("single",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                           "row_width_resize",
                                           "row_select",
                                           "arrowkeys"))
        self.find_window.unbind("<Return>")
        self.find_reset_button.config(state="disabled")
        self.find_up_button.config(state="disabled")
        self.find_down_button.config(state="disabled")
        self.save_text_button.config(state="disabled")
        self.update()

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except:
            pass
        try:
            self.wb_ = None
        except:
            pass

    def USER_HAS_CLOSED_WINDOW(self,callback=None):
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except:
            pass
        self.destroy()       
        
    def save_as(self):
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save changes as",
                                               filetypes=[('CSV File','.csv'),('Excel file','.xlsx'),('JSON File','.json')],
                                               defaultextension=".csv",
                                               confirmoverwrite=True)
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv",".xlsx",".json")):
            self.grab_set()
            self.stop_work("Can only save .csv/.xlsx/.json file types")
            return
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook()
                ws = self.wb_.active
                ws.append(["Date","User","Type","ID/Name/Number","Old Value","New Value"])
                for rn,row in enumerate(self.changelog):
                    ws.append(row)
                    if not rn % 20:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status_bar.change_text("".join((self.total_changes,"Saving...  changes: ",str(rn))))
                ws.freeze_panes = "A2"
                for i in range(1, 7):
                    if i < 5:
                        ws.cell(row = 1, column = i).fill = orange_fill
                    elif i == 5:
                        ws.cell(row = 1, column = i).fill = red_remove_fill
                    else:
                        ws.cell(row = 1, column = i).fill = green_add_fill
                    ws.cell(row = 1, column = i).border = openpyxl_thin_border
                ws.column_dimensions["A"].width = 37
                ws.column_dimensions["B"].width = 20
                ws.column_dimensions["C"].width = 37
                ws.column_dimensions["D"].width = 52
                ws.column_dimensions["E"].width = 60
                ws.column_dimensions["F"].width = 60
                self.wb_.save(newfile)
                self.try_to_close_wb()
            elif newfile.lower().endswith(".csv"):
                with open(newfile,"w",newline="") as fh:
                    writer = csv_module.writer(fh,dialect=csv_module.excel_tab,lineterminator="\n")
                    writer.writerow(["Date","User","Type","ID/Name/Number","Old Value","New Value"])
                    for rn,row in enumerate(self.changelog):
                        writer.writerow(row)
                        if not rn % 20:
                            self.update()
                            if self.USER_HAS_QUIT:
                                return
                            self.status_bar.change_text("".join((self.total_changes,"Saving...  changes: ",str(rn))))
            elif newfile.lower().endswith(".json"):
                with open(newfile,"w",newline="") as fh:
                    fh.write(json.dumps(self.C.dump_full_sheet_to_json(["Date","User","Type","ID/Name/Number","Old Value","New Value"],self.changelog,include_headers=True)))
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work("Error saving file: " + str(error_msg))
            return
        self.stop_work("Success! Changelog saved")

    def save_selected_as(self):
        selectedrows = self.sheetdisplay.get_selected_rows(get_cells_as_rows = True, return_tuple = True)
        if not selectedrows:
            return
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save selected changes as",
                                               filetypes=[('CSV File','.csv'),('Excel file','.xlsx'),('JSON File','.json')],
                                               defaultextension=".csv",
                                               confirmoverwrite=True)
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv",".xlsx",".json")):
            self.grab_set()
            self.stop_work("Can only save .csv/.xlsx/.json file types")
            return
        from_row = min(selectedrows)
        to_row   = max(selectedrows) + 1
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook()
                ws = self.wb_.active
                ws.append(["Date","User","Type","ID/Name/Number","Old Value","New Value"])
                for rn,row in enumerate(islice(self.changelog, from_row, to_row)):
                    ws.append(row)
                    if not rn % 20:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status_bar.change_text("".join((self.total_changes,"Saving...  changes: ",str(rn))))
                ws.freeze_panes = "A2"
                for i in range(1, 7):
                    if i < 5:
                        ws.cell(row = 1, column = i).fill = orange_fill
                    elif i == 5:
                        ws.cell(row = 1, column = i).fill = red_remove_fill
                    else:
                        ws.cell(row = 1, column = i).fill = green_add_fill
                    ws.cell(row = 1, column = i).border = openpyxl_thin_border
                ws.column_dimensions["A"].width = 37
                ws.column_dimensions["B"].width = 20
                ws.column_dimensions["C"].width = 37
                ws.column_dimensions["D"].width = 52
                ws.column_dimensions["E"].width = 60
                ws.column_dimensions["F"].width = 60
                self.wb_.save(newfile)
                self.try_to_close_wb()
            elif newfile.lower().endswith(".csv"):
                with open(newfile,"w",newline="") as fh:
                    writer = csv_module.writer(fh,dialect=csv_module.excel_tab,lineterminator="\n")
                    writer.writerow(["Date","User","Type","ID/Name/Number","Old Value","New Value"])
                    for rn,row in enumerate(islice(self.changelog, from_row, to_row)):
                        writer.writerow(row)
                        if not rn % 20:
                            self.update()
                            if self.USER_HAS_QUIT:
                                return
                            self.status_bar.change_text("".join((self.total_changes,"Saving...  changes: ",str(rn))))
            elif newfile.lower().endswith(".json"):
                with open(newfile,"w",newline="") as fh:
                    fh.write(json.dumps(self.C.dump_full_sheet_to_json(["Date","User","Type","ID/Name/Number","Old Value","New Value"],self.changelog[from_row:to_row],include_headers=True)))
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work("Error saving file: " + str(error_msg))
            return
        self.stop_work("Success! Changelog saved")

    def find(self,event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        x = self.word.lower()
        for rn,row in enumerate(self.changelog):
            for colno,cell in enumerate(row):
                if x in cell.lower():
                    if colno == 0:
                        self.find_results.append((rn,6))
                        break
                    else:
                        self.find_results.append((rn,colno))
        if self.find_results:
            for rn,colno in islice(self.find_results,1,len(self.find_results)):
                if colno == 6:
                    for i in range(1,6):
                        self.sheetdisplay.highlight_cells(row = rn, column = i, bg = "yellow",fg="black")
                else:
                    self.sheetdisplay.highlight_cells(row=rn,column=colno,bg="yellow",fg="black")
            if self.find_results[self.results_number][1] == 6:
                for i in range(1,6):
                    self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="orange",fg="black")
            else:
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange",fg="black")
            self.find_results_label.config(text="1/"+str(len(self.find_results)))
            self.sheetdisplay.see(row=self.find_results[0][0],column=0,keep_xscroll=True)
        self.sheetdisplay.refresh()
            
    def find_up(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        if self.find_results[self.results_number][1] == 6:
            for i in range(1,6):
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="yellow",fg="black")
        else:
            self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="yellow",fg="black")
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        if self.find_results[self.results_number][1] == 6:
            for i in range(1,6):
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="orange",fg="black")
        else:
            self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange",fg="black")
        self.sheetdisplay.see(row=self.find_results[self.results_number][0],column=0,keep_xscroll=True)
        self.sheetdisplay.refresh()
        
    def find_down(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        if self.find_results[self.results_number][1] == 6:
            for i in range(1,6):
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="yellow",fg="black")
        else:
            self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="yellow",fg="black")
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        if self.find_results[self.results_number][1] == 6:
            for i in range(1,6):
                self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=i,bg="orange",fg="black")
        else:
            self.sheetdisplay.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange",fg="black")
        self.sheetdisplay.see(row=self.find_results[self.results_number][0],column=0,keep_xscroll=True)
        self.sheetdisplay.refresh()
        
    def find_reset(self,newfind=False):
        self.find_results = []
        self.results_number = 0
        self.sheetdisplay.dehighlight_cells(all_=True,redraw=False)
        if newfind == False:
            self.find_window.delete(0,"end")
        self.find_results_label.config(text="0/0")
        for r in range(len(self.changelog)):
            self.sheetdisplay.highlight_cells(row = r,
                                              column = 4,
                                              bg = self.red_bg,
                                              fg = self.red_fg)
            self.sheetdisplay.highlight_cells(row = r,
                                              column = 5,
                                              bg = self.green_bg,
                                              fg = self.green_fg)
        self.sheetdisplay.refresh()
        
    def cancel(self,event=None):
        self.USER_HAS_CLOSED_WINDOW()


class compare_report_popup(tk.Toplevel):
    def __init__(self,
                 C,
                 width=1200,
                 height=800,
                 theme = "dark"):
        tk.Toplevel.__init__(self,
                             C,
                             width="1",
                             height="1",
                             bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.title("Comparison Report - Click the X button or press escape to go back")
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.find_results = []
        self.results_number = 0
        self.wb_ = None
        report = self.C.report
        self.sheet1name = self.C.sheetname_1
        self.sheet2name = self.C.sheetname_2

        self.open_tab = 1
        
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)

        self.notebook = ttk.Notebook(self)
        self.notebook.grid(row=1,column=0,sticky="nswe")
        
        self.f1 = frame(self, theme = theme)
        self.f1.grid_columnconfigure(0,weight=1)
        self.f1.grid_rowconfigure(1,weight=1)
        self.notebook.add(self.f1, text = "Matching IDs Differences")

        self.f2 = frame(self, theme = theme)
        self.f2.grid_columnconfigure(0,weight=1)
        self.f2.grid_rowconfigure(1,weight=1)
        self.notebook.add(self.f2, text = "Other Differences")
        self.notebook.select(self.f1)
        self.notebook.enable_traversal()
        self.notebook.bind("<<NotebookTabChanged>>", self.tab_change)
        
        self.find_frame = frame(self, theme = theme)
        self.find_frame.grid(row=0,column=0,columnspan=2,sticky="nswe")
        self.find_icon = tk.PhotoImage(format="gif",data=find_icon)
        self.search_button = button(self.find_frame,
                                    text=" Find:",
                                    command=self.find)
        self.search_button.config(image=self.find_icon,compound="left")
        self.search_button.pack(side="left",fill="x")
        self.find_window = normal_entry(self.find_frame,font=BF, theme = theme)
        self.find_window.bind("<Return>",self.find)
        self.find_window.pack(side="left",fill="x",expand=True)
        self.find_reset_button = button(self.find_frame,text="X",command=self.find_reset)
        self.find_reset_button.pack(side="left",fill="x")
        self.find_results_label = label(self.find_frame,"0/0",BF, theme = theme)
        self.find_results_label.pack(side="left",fill="x")
        self.find_up_button = button(self.find_frame,text="▲",command=self.find_up)
        self.find_up_button.pack(side="left",fill="x")
        self.find_down_button = button(self.find_frame,text="▼",command=self.find_down)
        self.find_down_button.pack(side="left",fill="x")
        
        self.sheetdisplay1 = Sheet(self.f1,
                                   theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                       outline_thickness=0)
        self.sheetdisplay1.enable_bindings(("single",
                                            "copy",
                                           "drag_select",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.sheetdisplay1.headers(newheaders=["ID","Difference",self.sheet1name,self.sheet2name])
        self.sheetdisplay1.data_reference(newdataref = report['ids'],
                                         reset_col_positions = False,
                                         reset_row_positions = False,
                                         redraw = False)
            
        self.sheetdisplay1.set_all_cell_sizes_to_text()
        self.sheetdisplay1.grid(row=1,column=0,sticky="nswe")

        self.sheetdisplay2 = Sheet(self.f2,
                                   theme = theme,
                                   outline_thickness=0)
        self.sheetdisplay2.enable_bindings(("single",
                                            "copy",
                                           "drag_select",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                           "row_width_resize",
                                           "row_select",
                                           "arrowkeys"))
        self.sheetdisplay2.data_reference(newdataref = report['info'],
                                          reset_col_positions = False,
                                          reset_row_positions = False,
                                          redraw = False)

        self.sheetdisplay2.set_all_cell_sizes_to_text()
        self.sheetdisplay2.grid(row=1,column=0,sticky="nswe")
        
        self.buttonframe = frame(self, theme = theme)
        self.buttonframe.grid(row=3,column=0,sticky="nswe")
        self.cancel_button = button(self.buttonframe,text="Done",
                                       style="EF.Std.TButton",
                                       command=self.cancel)
        self.cancel_button.pack(side = "right", padx = (20,100), pady = 20)
        self.save_text_button = button(self.buttonframe,text="Save Report",
                                       style="EF.Std.TButton",
                                       command=self.save_report)
        self.save_text_button.pack(side = "right", padx = (50, 30), pady = 20)
        

        self.bind("<Escape>",self.cancel)
        center(self,width,height)
        self.deiconify()
        self.wait_window()

    def tab_change(self, event = None):
        self.find_reset(True)
        self.open_tab = self.notebook.index(self.notebook.select()) + 1

    def start_work(self,msg=""):
        self.disable_widgets()

    def stop_work(self,msg=""):
        self.enable_widgets()

    def enable_widgets(self):
        self.sheetdisplay1.enable_bindings(("single",
                                           "copy",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.sheetdisplay2.enable_bindings(("single",
                                           "copy",
                                          "column_width_resize",
                                          "double_click_column_resize",
                                          "row_height_resize",
                                          "double_click_row_resize",
                                          "row_width_resize",
                                          "row_select",
                                          "arrowkeys"))
        self.find_window.bind("<Return>",self.find)
        self.find_reset_button.config(state="normal")
        self.find_up_button.config(state="normal")
        self.find_down_button.config(state="normal")
        self.save_text_button.config(state="normal")

    def disable_widgets(self):
        self.sheetdisplay1.disable_bindings(("single",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                           "row_width_resize",
                                           "row_select",
                                           "arrowkeys"))
        self.sheetdisplay2.disable_bindings(("single",
                                            "copy",
                                           "column_width_resize",
                                           "double_click_column_resize",
                                           "row_height_resize",
                                           "double_click_row_resize",
                                           "row_width_resize",
                                           "row_select",
                                           "arrowkeys"))
        self.find_window.unbind("<Return>")
        self.find_reset_button.config(state="disabled")
        self.find_up_button.config(state="disabled")
        self.find_down_button.config(state="disabled")
        self.save_text_button.config(state="disabled")
        self.update()

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except:
            pass
        try:
            self.wb_ = None
        except:
            pass

    def USER_HAS_CLOSED_WINDOW(self,callback=None):
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except:
            pass
        self.destroy()       
        
    def save_report(self):
        self.start_work("Opened save dialog")
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save as",
                                               filetypes=[('Excel file','.xlsx')],
                                               defaultextension=".xlsx",
                                               confirmoverwrite=True)
        if not newfile:
            self.stop_work()
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith(".xlsx"):
            self.grab_set()
            self.stop_work("Can only save .xlsx file type")
            return
        try:
            if newfile.lower().endswith(".xlsx"):
                self.wb_ = Workbook()
                ws = self.wb_.active
                ws.title = "Matching IDs Differences"
                ws.append(["ID", "Difference", self.sheet1name, self.sheet2name])
                for rn,row in enumerate(self.sheetdisplay1.get_sheet_data()):
                    ws.append(row)
                ws.freeze_panes = "A2"
                for i in range(1, 6):
                    ws.cell(row = 1, column = i).fill = orange_fill
                    ws.cell(row = 1, column = i).border = openpyxl_thin_border

                ws = self.wb_.create_sheet(title = "Other Differences")
                for rn,row in enumerate(self.sheetdisplay2.get_sheet_data()):
                    ws.append(row)
                
                self.wb_.save(newfile)
                self.try_to_close_wb()
        except Exception as error_msg:
            self.try_to_close_wb()
            self.grab_set()
            self.stop_work("Error saving file: " + str(error_msg))
            return
        self.stop_work("Success! Report saved")

    def find(self,event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        x = self.word.lower()
        if self.open_tab == 1:
            target_sheet = self.sheetdisplay1
            find_res = self.find_results
            res_num = self.results_number
        else:
            target_sheet = self.sheetdisplay2
        for rn, row in enumerate(target_sheet.get_sheet_data()):
            for colno, cell in enumerate(row):
                if x in cell.lower():
                    self.find_results.append((rn, colno))
        if self.find_results:
            for rn,colno in islice(self.find_results,1,len(self.find_results)):
                target_sheet.highlight_cells(row=rn,column=colno,bg="yellow")
            target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange")
            self.find_results_label.config(text="1/"+str(len(self.find_results)))
            target_sheet.see(row=self.find_results[0][0],column=0,keep_xscroll=True)
        target_sheet.refresh()
            
    def find_up(self,event=None):
        if self.open_tab == 1:
            target_sheet = self.sheetdisplay1
        else:
            target_sheet = self.sheetdisplay2
        if not self.find_results or len(self.find_results) == 1:
            return
        target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="yellow")
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange")
        target_sheet.see(row=self.find_results[self.results_number][0],column=0,keep_xscroll=True)
        target_sheet.refresh()
        
    def find_down(self,event=None):
        if self.open_tab == 1:
            target_sheet = self.sheetdisplay1
        else:
            target_sheet = self.sheetdisplay2
        if not self.find_results or len(self.find_results) == 1:
            return
        target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="yellow")
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        target_sheet.highlight_cells(row=self.find_results[self.results_number][0],column=self.find_results[self.results_number][1],bg="orange")
        target_sheet.see(row=self.find_results[self.results_number][0],column=0,keep_xscroll=True)
        target_sheet.refresh()
        
    def find_reset(self,newfind=False):
        try:
            self.find_results = []
            self.results_number = 0
            self.sheetdisplay1.dehighlight_cells(all_=True,redraw=True)
            self.sheetdisplay2.dehighlight_cells(all_=True,redraw=True)
            if newfind == False:
                self.find_window.delete(0,"end")
            self.find_results_label.config(text="0/0")
        except:
            pass
        
    def cancel(self,event=None):
        self.USER_HAS_CLOSED_WINDOW()


class x_checkbutton(ttk.Button):
    def __init__(self,parent,text="",style="Std.TButton",command=None,state="normal",checked=False,compound="right"):
        button.__init__(self,
                        parent,
                        text=text,
                        style=style,
                        command=command,
                        state=state)
        self.image_compound = compound
        self.on_icon = tk.PhotoImage(format="gif",data=checked_icon)
        self.off_icon = tk.PhotoImage(format="gif",data=unchecked_icon)
        self.checked = checked
        if checked:
            self.config(image=self.on_icon,compound=compound)
        else:
            self.config(image=self.off_icon,compound=compound)
        self.bind("<1>",self.B1)
    def set_checked(self,state="toggle"):
        if state == "toggle":
            self.checked = not self.checked
            if self.checked:
                self.config(image=self.on_icon,compound=self.image_compound)
            else:
                self.config(image=self.off_icon,compound=self.image_compound)
        elif state:
            self.checked = True
            self.config(image=self.on_icon,compound=self.image_compound)
        elif not state:
            self.checked = False
            self.config(image=self.off_icon,compound=self.image_compound)
    def get_checked(self):
        return bool(self.checked)
    def B1(self,event):
        x = str(self['state'])
        if "normal" in x:
            
            self.checked = not self.checked
            if self.checked:
                self.config(image=self.on_icon,compound=self.image_compound)
            else:
                self.config(image=self.off_icon,compound=self.image_compound)
        self.update_idletasks()
    def change_text(self,text):
        self.config(text=text)
        self.update_idletasks()


class merge_sheets_popup(tk.Toplevel):
    def __init__(self, C, theme = "dark", add_rows = False):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Merge sheets - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW",self.USER_HAS_CLOSED_WINDOW)
        self.USER_HAS_QUIT = False
        self.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.grid_rowconfigure(0, weight = 1)

        self.l_frame = frame(self, theme = theme)
        self.l_frame.grid(row = 0, column = 0, sticky = "nswe")
        self.r_frame = frame(self, theme = theme)
        self.r_frame.grid(row = 0, column = 1, sticky = "nswe")
        self.l_frame.grid_rowconfigure(3, weight = 1)
        self.r_frame.grid_rowconfigure(0, weight = 1)
        self.l_frame.grid_columnconfigure(0, weight = 1)
        self.r_frame.grid_columnconfigure(1, weight = 1)

        self.open_file_display = readonly_entry_with_scrollbar(self.l_frame,font=EF, theme = theme)
        self.open_file_display.grid(row=0,column=0,padx=2,pady=2,sticky="nswe")
        self.open_file_button = button(self.l_frame,
                                        text="⯇ Open file",
                                        style="EF.Std.TButton",
                                        command=self.open_file)
        self.open_file_button.grid(row=0,column=1,padx=2,pady=2,sticky="nswe")
        self.sheet_dropdown = ez_dropdown(self.l_frame,font=EF)
        self.sheet_dropdown.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
        self.sheet_dropdown.grid(row=1,column=0,padx=2,pady=2,sticky="nswe")
        self.select_sheet_button = button(self.l_frame,
                                          text="⯇ Load sheet",
                                          style="EF.Std.TButton",
                                          state="disabled",
                                          command=self.select_sheet)
        self.select_sheet_button.grid(row=1,column=1,padx=2,pady=2,sticky="nswe")

        self.selector = id_and_parent_column_selector(self.l_frame, theme = theme)
        self.selector.grid(row=2,column=0,rowspan=2,sticky="nswe")
        
        self.clipboard_button = button(self.l_frame,
                                          text=" Get data from clipboard ",
                                          style="EF.Std.TButton",
                                          state="normal",
                                          command=self.get_clipboard_data)
        self.clipboard_button.grid(row=2,column=1,padx=2,pady=(2,20),sticky="nswe")

        self.options_frame = frame(self.l_frame, theme = theme)
        self.options_frame.grid(row=3,column=1,sticky="nswe")
        
        self.add_new_ids_button = x_checkbutton(self.options_frame,
                                                text="Add any new IDs     ",
                                                style="x_button.Std.TButton",
                                                compound="right",
                                                checked=True)
        self.add_new_ids_button.grid(row=0,column=0,padx=10,pady=5,sticky="we")
        self.add_new_dcols_button = x_checkbutton(self.options_frame,
                                                  text="Add any new detail columns ",
                                                  style="x_button.Std.TButton",
                                                  compound="right")
        self.add_new_dcols_button.grid(row=1,column=0,padx=10,pady=5,sticky="we")
        self.add_new_pcols_button = x_checkbutton(self.options_frame,
                                                  text="Add any new parent columns ",
                                                  style="x_button.Std.TButton",
                                                  compound="right")
        self.add_new_pcols_button.grid(row=2,column=0,padx=10,pady=5,sticky="we")
        self.overwrite_details_button = x_checkbutton(self.options_frame,
                                                      text="Overwrite details for same IDs ",
                                                      style="x_button.Std.TButton",
                                                      compound="right")
        self.overwrite_details_button.grid(row=3,column=0,padx=10,pady=5,sticky="we")
        self.overwrite_parents_button = x_checkbutton(self.options_frame,
                                                      text="Overwrite parents for same IDs ",
                                                      style="x_button.Std.TButton",
                                                      compound="right")
        self.overwrite_parents_button.grid(row=4,column=0,padx=10,pady=5,sticky="we")
        
        self.button_frame = frame(self.l_frame, theme = theme)
        self.button_frame.grid(row=4,column=0,columnspan = 2, sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="b")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="b")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text="Confirm merge",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,35),pady=(20,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(35,20),pady=(20,20))
        self.status = StatusBar(self.l_frame,text="Open a file to import data", theme = theme)
        self.status.grid(row=5,column=0, columnspan = 2, sticky="ew")
        self.result = False
        self.add_new_ids = True
        self.add_new_dcols = False
        self.add_new_pcols = False
        self.overwrite_details = False
        self.overwrite_parents = False
        self.file_opened = ""
        self.sheet_opened = "n/a"
        self.row_len = 0
        self.ic = None
        self.pcols = []
        self.wb_ = None
        self.C.new_sheet = []
        self.rowsel = 0
        self.colsel = 0
        self.region = "header"

        self.showing_left = True
        self.toggle_left_button = button(self.r_frame,
                                        text="⯇",
                                        style="BF.Std.TButton",
                                        command=self.toggle_left_panel)
        self.toggle_left_button.grid(row=0,column=0, sticky = "ns")
        self.toggle_left_button.config(width = 10)
        
        self.sheetdisplay = Sheet(self.r_frame,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                  outline_thickness=0,
                                  row_drag_and_drop_perform = False,
                                  column_drag_and_drop_perform = False)
        self.sheetdisplay.enable_bindings(("enable_all"))
        self.sheetdisplay.extra_bindings([("row_index_drag_drop", self.drag_row),
                                          ("column_header_drag_drop", self.drag_col),
                                          ("rc_insert_column", self.reset_selectors),
                                          ("rc_insert_row", self.reset_selectors),
                                          ("rc_delete_column", self.reset_selectors),
                                          ("rc_delete_row", self.reset_selectors),
                                          ("ctrl_x", self.ctrl_x_in_sheet),
                                          ("delete_key", self.del_in_sheet),
                                          ("ctrl_v", self.ctrl_v_in_sheet),
                                          ("ctrl_z", self.ctrl_z_in_sheet),
                                          ("begin_edit_cell_use_keypress", self.begin_edit_cell),
                                          ("escape_edit_cell", self.escape_edit_cell),
                                          ("edit_cell", self.edit_cell_in_sheet)
                                          ])
        self.sheetdisplay.headers(newheaders = 0)
        self.C.new_sheet = [[h.name for h in self.C.headers]] + [list(repeat("", len(self.C.headers))) for r in range(2000)]
        self.C.new_sheet = self.sheetdisplay.set_sheet_data(self.C.new_sheet, verify = False)
        self.selector.set_columns([h for h in self.C.new_sheet[0]] if self.C.new_sheet else [])
        self.selector.detect_id_col()
        self.selector.detect_par_cols()
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.sheetdisplay.grid(row=0,column=1,sticky="nswe")
        if add_rows:
            self.toggle_left_panel()
            self.toggle_left_button.config(text = "⯈\nShow\nOptions\nand\nConfirm\n⯈")
        self.bind("<Escape>",self.cancel)
        center(self,1280,620)
        self.deiconify()
        self.wait_window()

    def begin_edit_cell(self, event = None):
        self.unbind("<Escape>")

    def escape_edit_cell(self, event = None):
        self.bind("<Escape>",self.cancel)

    def toggle_left_panel(self, event = None):
        if self.showing_left:
            self.grid_columnconfigure(0, weight = 0, uniform = "y")
            self.l_frame.grid_forget()
            self.showing_left = False
            self.toggle_left_button.config(text = "⯈")

        else:
            self.grid_columnconfigure(0, weight = 1, uniform = "x")
            self.l_frame.grid(row = 0, column = 0, sticky = "nswe")
            self.showing_left = True
            self.toggle_left_button.config(text = "⯇")

    def drag_col(self, selected_cols, c):
        c = int(c)
        colsiter = list(selected_cols)
        colsiter.sort()
        stins = colsiter[0]
        endins = colsiter[-1] + 1
        totalcols = len(colsiter)
        if stins > c:
            for rn in range(len(self.C.new_sheet)):
                self.C.new_sheet[rn] = (self.C.new_sheet[rn][:c] +
                                              self.C.new_sheet[rn][stins:stins + totalcols] +
                                              self.C.new_sheet[rn][c:stins] +
                                              self.C.new_sheet[rn][stins + totalcols:])
        else:
            for rn in range(len(self.C.new_sheet)):
                self.C.new_sheet[rn] = (self.C.new_sheet[rn][:stins] +
                                              self.C.new_sheet[rn][stins + totalcols:c + 1] +
                                              self.C.new_sheet[rn][stins:stins + totalcols] +
                                              self.C.new_sheet[rn][c + 1:])
        self.sheetdisplay.MT.data_ref = self.C.new_sheet
        self.selector.set_columns([h for h in self.C.new_sheet[0]])
        self.selector.detect_id_col()
        self.selector.detect_par_cols()

    def drag_row(self, selected_rows, r):
        r = int(r)
        rowsiter = list(selected_rows)
        rowsiter.sort()
        stins = rowsiter[0]
        endins = rowsiter[-1] + 1
        totalrows = len(rowsiter)
        if stins > r:
            self.C.new_sheet = (self.C.new_sheet[:r] +
                                      self.C.new_sheet[stins:stins + totalrows] +
                                      self.C.new_sheet[r:stins] +
                                      self.C.new_sheet[stins + totalrows:])
        else:
            self.C.new_sheet = (self.C.new_sheet[:stins] +
                                      self.C.new_sheet[stins + totalrows:r + 1] +
                                      self.C.new_sheet[stins:stins + totalrows] +
                                      self.C.new_sheet[r + 1:])
        self.sheetdisplay.MT.data_ref = self.C.new_sheet
        if endins == 0 or r == 0 or stins == 0:
            self.selector.set_columns([h for h in self.C.new_sheet[0]])
            self.selector.detect_id_col()
            self.selector.detect_par_cols()

    def del_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_x_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_v_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_z_in_sheet(self, event = None):
        self.reset_selectors()

    def edit_cell_in_sheet(self, event = None):
        self.bind("<Escape>",self.cancel)
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        if event[1] == idcol or event[1] in parcols or event[0] == 0:
            self.reset_selectors()

    def get_clipboard_data(self,event=None):
        self.start_work("Loading...")
        self.reset()
        try:
            temp_data = self.C.clipboard_get()
        except:
            self.stop_work("Error: Error getting clipboard data")
            return
        try:
            if temp_data.startswith("{") and temp_data.endswith("}"):
                self.C.new_sheet = self.C.json_to_sheet(json.loads(temp_data))
            else:
                delimiter_,quotechar_ = self.csv_delimiter_quotechar(temp_data)
                if delimiter_ is None:
                    self.stop_work("Error: Clipboard contained no appropriate data")
                    return
                for rn,r in enumerate(csv_module.reader(io.StringIO(temp_data),delimiter=delimiter_,quotechar=quotechar_,skipinitialspace=True)):
                    try:
                        self.C.new_sheet.append(r[:len(r) - next(i for i,c in enumerate(reversed(r)) if c)])
                    except:
                        continue
                    if not rn % 500:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status.change_text("Loading...  rows: " + str(rn))
        except:
            self.stop_work("Error: Error parsing clipboard data")
            return
        if not self.C.new_sheet:
            self.stop_work("Error: Clipboard contained no appropriate data")
            return
        rl = len(max(self.C.new_sheet, key = len))
        self.C.new_sheet[:] = [r + list(repeat("",rl - len(r))) for r in self.C.new_sheet]
        self.ic = None
        self.pcols = []
        for i,c in enumerate(self.C.new_sheet[0]):
            cell = c.lower()
            if cell == "id" or cell.startswith("id"):
                self.ic = i
            elif cell.startswith("parent"):
                self.pcols.append(i)
        if self.ic is None or not self.pcols:
            self.load_display([h for h in self.C.new_sheet[0]])
            self.stop_work("Select ID column and Parent columns")
        else:
            self.selector.set_columns([h for h in self.C.new_sheet[0]])
            self.selector.set_id_col(self.ic)
            self.selector.set_par_cols(self.pcols)
            self.stop_work("Ready to merge sheets")
        self.sheetdisplay.deselect("all")
        self.sheetdisplay.data_reference(newdataref=self.C.new_sheet,reset_col_positions=False,reset_row_positions=False,redraw=False)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.sheetdisplay.refresh()
        self.file_opened = "n/a - Data obtained from clipboard"
        self.sheet_opened = "n/a"

    def reset_selectors(self, event = None):
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        self.selector.set_columns([h for h in self.C.new_sheet[0]] if self.C.new_sheet else [])
        if idcol is not None and self.C.new_sheet:
            self.selector.set_id_col(idcol)
        if parcols and self.C.new_sheet:
            self.selector.set_par_cols(parcols)

    def csv_delimiter_quotechar(self,data):
        d = Counter(m.group() for m in re.finditer(r"""\t|,|\t'|'\t|\t"|"\t|,'|',|,"|",""",data))
        if not d['\t'] and not d[',']:
            return None,None
        if d['\t'] >= d[',']:
            delimiter_ = "\t"
        elif d['\t'] < d[',']:
            delimiter_ = ","
        if d['\t"'] + d[',"'] + d['"\t'] + d['",'] >= d["\t'"] + d[",'"] + d["'\t"] + d["',"]:
            quotechar_ = '"'
        elif d['\t"'] + d[',"'] + d['"\t'] + d['",'] < d["\t'"] + d[",'"] + d["'\t"] + d["',"]:
            quotechar_ = "'"
        return delimiter_,quotechar_

    def return_wb_file(self,filepath):
        with open(filepath,"rb") as fh:
            in_mem = io.BytesIO(fh.read())
        return in_mem

    def try_to_close_wb(self):
        try:
            self.wb_.close()
        except:
            pass
        try:
            self.wb_ = None
        except:
            pass
            
    def USER_HAS_CLOSED_WINDOW(self,callback=None):
        self.C.new_sheet = []
        self.USER_HAS_QUIT = True
        try:
            self.try_to_close_wb()
        except:
            pass
        self.destroy()

    def open_file(self):
        self.start_work("Loading...   ")
        self.reset()
        filepath = filedialog.askopenfilename(parent=self,title="Select file")
        if not filepath:
            self.stop_work("Open a file to import data")
            return
        try:
            filepath = os.path.normpath(filepath)
        except:
            self.stop_work("Error: filepath invalid")
            return
        if not filepath.lower().endswith((".json",".xlsx",".xls",".xlsm",".csv",".tsv")):
            self.stop_work("Error: please select json/excel/csv   ")
            return
        check = os.path.isfile(filepath)
        if check == False:
            self.stop_work("Error: filepath invalid")
            return
        try:
            if filepath.lower().endswith((".csv",".tsv")):
                with open(filepath,"r") as fh:
                    temp_data = fh.read()
                delimiter_,quotechar_ = self.csv_delimiter_quotechar(temp_data)
                if delimiter_ is None:
                    self.stop_work("Error: File contained no appropriate data")
                    return
                for rn,r in enumerate(csv_module.reader(io.StringIO(temp_data),delimiter=delimiter_,quotechar=quotechar_,skipinitialspace=True)):
                    try:
                        self.C.new_sheet.append(r[:len(r) - next(i for i,c in enumerate(reversed(r)) if c)])
                    except:
                        continue
                    if not rn % 500:
                        self.update()
                        if self.USER_HAS_QUIT:
                            return
                        self.status.change_text("Loading...  rows: " + str(rn))
                rl = len(max(self.C.new_sheet, key = len))
                self.C.new_sheet[:] = [r + list(repeat("",rl - len(r))) for r in self.C.new_sheet]
                self.load_display([h for h in self.C.new_sheet[0]])
                self.stop_work("Ready to merge sheets")
            elif filepath.lower().endswith(".json"):
                j = self.C.get_json_from_file(filepath)
                json_format = self.C.get_json_format(j)
                if not json_format:
                    self.C.new_sheet = []
                    self.stop_work("Error opening file, could not find data of correct format")
                    return
                self.C.new_sheet = self.C.json_to_sheet(j,format_=json_format[0],key=json_format[1],get_format=False,return_rowlen=False)
                if not self.C.new_sheet:
                    self.stop_work("Error: File contained no data")
                    self.select_sheet_button.config(state="disabled")
                    return
                rl = len(max(self.C.new_sheet, key = len))
                self.C.new_sheet[:] = [r + list(repeat("",rl - len(r))) for r in self.C.new_sheet]
                self.load_display([h for h in self.C.new_sheet[0]])
                self.stop_work("Ready to merge sheets")
            elif filepath.lower().endswith((".xlsx",".xls",".xlsm")):
                in_mem = self.return_wb_file(filepath)
                self.wb_ = load_workbook(in_mem,read_only=True,data_only=True)
                wbsheets = self.wb_.sheetnames
                if not wbsheets:
                    self.stop_work("Error: File/sheet contained no data")
                    return
                sheetnames = set(self.wb_.sheetnames)
                if "Treesurgeon Data" in sheetnames:
                    ws = self.wb_["Treesurgeon Data"]
                    ws.reset_dimensions()
                    try:
                        d = self.C.C.decompress_str_return_obj("".join(["" if r[0].value is None else f"{r[0].value}" for r in islice(ws.rows, 1, None)]),
                                                                               basetype = "32",
                                                                               dec = True)
                        self.C.new_sheet = [[h['name'] for h in d['headers']]] + d['records']
                        self.wb_.close()
                        self.select_sheet_button.config(state="disabled")
                        self.load_display([h for h in self.C.new_sheet[0]])
                        self.stop_work("Ready to merge sheets")
                    except:
                        self.C.new_sheet = []
                        self.wb_.close()
                        self.wb_ = load_workbook(in_mem,read_only=True,data_only=True)
                        self.stop_work("Error: Error opening program data")
                        self.sheet_dropdown['values'] = wbsheets
                        self.sheet_dropdown.set_my_value(wbsheets[0])
                        self.stop_work("Error: Error opening program data. Select a sheet to open")
                        self.select_sheet_button.config(state="normal")
                else:
                    self.sheet_dropdown['values'] = wbsheets
                    self.sheet_dropdown.set_my_value(wbsheets[0])
                    self.stop_work("Select a sheet to open")
                    self.select_sheet_button.config(state="normal")
        except Exception as error_msg:
            self.try_to_close_wb()
            self.C.new_sheet = []
            self.stop_work("Error: " + str(error_msg))
            return
        if not self.C.new_sheet and not filepath.lower().endswith((".xlsx",".xls",".xlsm")):
            self.C.new_sheet = []
            self.stop_work("Error: File/sheet contained no data")
            return
        self.sheetdisplay.data_reference(newdataref=self.C.new_sheet,reset_col_positions=True,reset_row_positions=True,redraw=False)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        self.open_file_display.set_my_value(filepath)
        self.file_opened = os.path.basename(self.open_file_display.get_my_value())

    def select_sheet(self):
        self.start_work("Loading...   ")
        self.sheet_opened = self.sheet_dropdown.get_my_value()
        ws = self.wb_[self.sheet_opened]
        ws.reset_dimensions()
        dapp = self.C.new_sheet.append
        for rn,r in enumerate(ws.rows):
            try:
                dapp(["" if x.value is None else f"{x.value}" for x in islice(r,0,len(r) - next(i for i,c in enumerate(reversed(r)) if c.value is not None))])
            except:
                continue
            if not rn % 500:
                self.update()
                if self.USER_HAS_QUIT:
                    return
                self.status.change_text("Loading...  rows: " + str(rn))
        self.try_to_close_wb()
        self.stop_work("Ready to merge sheets")
        if not self.C.new_sheet:
            self.stop_work("Error: File/sheet contained no data")
            self.select_sheet_button.config(state="disabled")
            return
        rl = len(max(self.C.new_sheet, key = len))
        self.C.new_sheet[:] = [r + list(repeat("",rl - len(r))) for r in self.C.new_sheet]
        self.select_sheet_button.config(state="disabled")
        self.load_display([h for h in self.C.new_sheet[0]])
        self.sheetdisplay.data_reference(newdataref=self.C.new_sheet,reset_col_positions=True,reset_row_positions=True,redraw=False)
        self.sheetdisplay.set_all_cell_sizes_to_text()

    def load_display(self,cols):
        self.selector.set_columns(cols)
        self.selector.detect_id_col()
        self.selector.detect_par_cols()

    def start_work(self,msg=""):
        self.status.change_text(msg)
        self.disable_widgets()

    def stop_work(self,msg=""):
        self.status.change_text(msg)
        self.enable_widgets()

    def enable_widgets(self):
        self.open_file_display.change_my_state("readonly")
        self.open_file_button.config(state="normal")
        self.sheet_dropdown.config(state="readonly")
        self.selector.enable_me()
        self.add_new_ids_button.config(state="normal")
        self.add_new_dcols_button.config(state="normal")
        self.add_new_pcols_button.config(state="normal")
        self.overwrite_details_button.config(state="normal")
        self.overwrite_parents_button.config(state="normal")
        self.confirm_button.config(state="normal")
        self.sheetdisplay.enable_bindings(("enable_all"))
        self.sheetdisplay.extra_bindings([("row_index_drag_drop", self.drag_row),
                                          ("column_header_drag_drop", self.drag_col),
                                          ("rc_insert_column", self.reset_selectors),
                                          ("rc_insert_row", self.reset_selectors),
                                          ("rc_delete_column", self.reset_selectors),
                                          ("rc_delete_row", self.reset_selectors),
                                          ("ctrl_x", self.ctrl_x_in_sheet),
                                          ("delete_key", self.del_in_sheet),
                                          ("ctrl_v", self.ctrl_v_in_sheet),
                                          ("ctrl_z", self.ctrl_z_in_sheet),
                                          ("edit_cell", self.edit_cell_in_sheet),
                                          ("begin_edit_cell_use_keypress", self.begin_edit_cell),
                                          ("escape_edit_cell", self.escape_edit_cell)
                                          ])

    def disable_widgets(self):
        self.open_file_display.change_my_state("disabled")
        self.open_file_button.config(state="disabled")
        self.sheet_dropdown.config(state="disabled")
        self.select_sheet_button.config(state="disabled")
        self.selector.disable_me()
        self.add_new_ids_button.config(state="disabled")
        self.add_new_dcols_button.config(state="disabled")
        self.add_new_pcols_button.config(state="disabled")
        self.overwrite_details_button.config(state="disabled")
        self.overwrite_parents_button.config(state="disabled")
        self.confirm_button.config(state="disabled")
        self.sheetdisplay.disable_bindings(("disable_all"))
        self.sheetdisplay.extra_bindings([("row_index_drag_drop", None),
                                          ("column_header_drag_drop", None),
                                          ("rc_insert_column", None),
                                          ("rc_insert_row", None),
                                          ("rc_delete_column", None),
                                          ("rc_delete_row", None),
                                          ("ctrl_x", None),
                                          ("delete_key", None),
                                          ("ctrl_v", None),
                                          ("ctrl_z", None),
                                          ("edit_cell", None),
                                          ("begin_edit_cell_use_keypress", None),
                                          ("escape_edit_cell", None)
                                          ])
        self.update()

    def reset(self):
        self.try_to_close_wb()
        self.row_len = 0
        self.ic = None
        self.pcols = []
        self.C.new_sheet = []
        self.open_file_display.set_my_value("")
        self.sheet_dropdown['values'] = []
        self.sheet_dropdown.set("")
        self.select_sheet_button.config(state="disabled")
        self.selector.clear_displays()

    def confirm(self,event=None):
        self.add_new_ids = self.add_new_ids_button.get_checked()
        self.add_new_dcols = self.add_new_dcols_button.get_checked()
        self.add_new_pcols = self.add_new_pcols_button.get_checked()
        self.overwrite_details = self.overwrite_details_button.get_checked()
        self.overwrite_parents = self.overwrite_parents_button.get_checked()
        self.ic = self.selector.get_id_col()
        self.pcols = self.selector.get_par_cols()
        if not self.C.new_sheet:
            self.status.change_text("Please open a file to load data")
            return
        self.row_len = len(max(self.C.new_sheet,key=len))
        if all(x == False for x in (self.add_new_ids,self.add_new_dcols,self.add_new_pcols,self.overwrite_details,self.overwrite_parents)):
            self.status.change_text("Please select at least one option")
            return
        if self.ic in set(self.pcols):
            self.status.change_text("ID column must be different to all parent columns")
            return
        if self.ic is None:
            self.status.change_text("Please select an ID column")
            return
        self.result = True
        self.destroy()

    def cancel(self,event=None):
        self.USER_HAS_CLOSED_WINDOW()


class get_clipboard_data_popup(tk.Toplevel):
    def __init__(self, C, cols, row_len, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Get data from clipboard - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        #self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(0,weight=1)
        #self.grid_rowconfigure(2,weight=1,uniform="y")
        self.selector = id_and_parent_column_selector(self)
        self.selector.grid(row=0,column=0,sticky="nsew")

        self.flattened_choices = flattened_base_ids_choices(self,command=self.flattened_mode_toggle, theme = theme)
        self.flattened_choices.change_theme(theme)
        self.flattened_choices.grid(row=1,column=0,padx=5,pady=5,sticky="nsew")
        self.flattened_selector = flattened_column_selector(self)
        self.flattened_selector.set_columns(cols)
        self.selector.change_theme(theme)
        self.flattened_selector.change_theme(theme)
        self.selector.set_columns(cols)
        self.selector.detect_id_col()
        self.selector.detect_par_cols()
        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                  column_drag_and_drop_perform = False,
                                  row_drag_and_drop_perform = False)
        self.sheetdisplay.enable_bindings(("enable_all"))
        self.sheetdisplay.extra_bindings([("row_index_drag_drop", self.drag_row),
                                          ("column_header_drag_drop", self.drag_col),
                                          ("rc_insert_column", self.reset_selectors),
                                          ("rc_insert_row", self.reset_selectors),
                                          ("rc_delete_column", self.reset_selectors),
                                          ("rc_delete_row", self.reset_selectors),
                                          ("ctrl_x", self.ctrl_x_in_sheet),
                                          ("delete_key", self.del_in_sheet),
                                          ("ctrl_v", self.ctrl_v_in_sheet),
                                          ("ctrl_z", self.ctrl_z_in_sheet),
                                          ("edit_cell", self.edit_cell_in_sheet),
                                          ("begin_edit_cell_use_keypress", self.begin_edit_cell),
                                          ("escape_edit_cell", self.escape_edit_cell),
                                          ("edit_cell", self.escape_edit_cell)
                                          ])
        self.sheetdisplay.grid(row=0,column=1,rowspan=4,sticky="nswe")
        self.sheetdisplay.data_reference(newdataref=self.C.new_sheet,redraw=False)
        self.sheetdisplay.headers(newheaders=0)
        self.sheetdisplay.set_all_cell_sizes_to_text()
        
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=2,column=0,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        
        self.confirm_button = button(self.button_frame,text="Overwrite existing data",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(20,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(20,20))
        self.status = StatusBar(self,text="Select ID and Parent columns", theme = theme)
        self.status.grid(row=3,column=0,sticky="we")
        self.result = False
        self.ic = None
        self.pcols = []
        self.bind("<Escape>",self.cancel)
        center(self,1280,620)
        self.selector.grid_forget()
        self.flattened_selector.grid(row=0,column=0,pady=(0,35),sticky="nsew")
        self.flattened_selector.grid_forget()
        self.selector.grid(row=0,column=0,sticky="nsew")
        self.deiconify()
        self.wait_window()

    def begin_edit_cell(self, event = None):
        self.unbind("<Escape>")

    def escape_edit_cell(self, event = None):
        self.bind("<Escape>",self.cancel)

    def reset_selectors(self, event = None):
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        ancparcols = self.flattened_selector.get_par_cols()
        self.selector.set_columns([h for h in self.C.new_sheet[0]] if self.C.new_sheet else [])
        self.flattened_selector.set_columns([h for h in self.C.new_sheet[0]] if self.C.new_sheet else [])
        if idcol is not None and self.C.new_sheet:
            self.selector.set_id_col(idcol)
        if parcols and self.C.new_sheet:
            self.selector.set_par_cols(parcols)
        if ancparcols and self.C.new_sheet:
            self.flattened_selector.set_par_cols(ancparcols)

    def drag_col(self, selected_cols, c):
        c = int(c)
        colsiter = list(selected_cols)
        colsiter.sort()
        stins = colsiter[0]
        endins = colsiter[-1] + 1
        totalcols = len(colsiter)
        if stins > c:
            for rn in range(len(self.C.new_sheet)):
                self.C.new_sheet[rn] = (self.C.new_sheet[rn][:c] +
                                              self.C.new_sheet[rn][stins:stins + totalcols] +
                                              self.C.new_sheet[rn][c:stins] +
                                              self.C.new_sheet[rn][stins + totalcols:])
        else:
            for rn in range(len(self.C.new_sheet)):
                self.C.new_sheet[rn] = (self.C.new_sheet[rn][:stins] +
                                              self.C.new_sheet[rn][stins + totalcols:c + 1] +
                                              self.C.new_sheet[rn][stins:stins + totalcols] +
                                              self.C.new_sheet[rn][c + 1:])
        self.sheetdisplay.MT.data_ref = self.C.new_sheet
        self.selector.set_columns([h for h in self.C.new_sheet[0]])
        self.flattened_selector.set_columns([h for h in self.C.new_sheet[0]])
        self.selector.detect_id_col()
        self.selector.detect_par_cols()

    def drag_row(self, selected_rows, r):
        r = int(r)
        rowsiter = list(selected_rows)
        rowsiter.sort()
        stins = rowsiter[0]
        endins = rowsiter[-1] + 1
        totalrows = len(rowsiter)
        if stins > r:
            self.C.new_sheet = (self.C.new_sheet[:r] +
                                      self.C.new_sheet[stins:stins + totalrows] +
                                      self.C.new_sheet[r:stins] +
                                      self.C.new_sheet[stins + totalrows:])
        else:
            self.C.new_sheet = (self.C.new_sheet[:stins] +
                                      self.C.new_sheet[stins + totalrows:r + 1] +
                                      self.C.new_sheet[stins:stins + totalrows] +
                                      self.C.new_sheet[r + 1:])
        self.sheetdisplay.MT.data_ref = self.C.new_sheet
        if endins == 0 or r == 0 or stins == 0:
            self.selector.set_columns([h for h in self.C.new_sheet[0]])
            self.flattened_selector.set_columns([h for h in self.C.new_sheet[0]])
            self.selector.detect_id_col()
            self.selector.detect_par_cols()

    def del_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_x_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_v_in_sheet(self, event = None):
        self.reset_selectors()

    def ctrl_z_in_sheet(self, event = None):
        self.reset_selectors()

    def edit_cell_in_sheet(self, event = None):
        idcol = self.selector.get_id_col()
        parcols = self.selector.get_par_cols()
        ancparcols = self.flattened_selector.get_par_cols()
        if event[1] == idcol or event[1] in parcols or event[1] in ancparcols or event[0] == 0:
            self.reset_selectors()

    def flattened_mode_toggle(self):
        x = self.flattened_choices.get_choices()[0]
        if x:
            self.selector.grid_forget()
            self.flattened_selector.grid(row=0,column=0,pady=(0,35),sticky="nsew")
        else:
            self.flattened_selector.grid_forget()
            self.selector.grid(row=0,column=0,sticky="nsew")

    def confirm(self,event=None):
        self.ic = self.selector.get_id_col()
        self.pcols = self.selector.get_par_cols()
        self.flattened_pcols = self.flattened_selector.get_par_cols()
        self.flattened = self.flattened_choices.get_choices()
        self.C.new_sheet = self.sheetdisplay.get_sheet_data()
        if self.flattened[0]:
            if not self.flattened_pcols:
                self.status.change_text("Please select hierarchy columns")
                return
        else:
            if self.ic in set(self.pcols):
                self.status.change_text("ID column must be different to all parent columns")
                return
            if self.ic is None:
                self.status.change_text("Please select an ID column")
                return
            if not self.pcols:
                self.status.change_text("Please select parent columns")
                return
        self.result = True
        self.destroy()

    def cancel(self,event=None):
        self.destroy()


class sheet_column_display_chooser_popup(tk.Toplevel):
    def __init__(self, C, headers, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Show/Hide columns - Click on columns to toggle")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(0,
                                  weight = 1)
        self.grid_rowconfigure(0,
                               weight = 1)
        
        self.headers =  headers
        self.chosen_indexes = []
        self.result = False
        self.toggled = None

        self.sheetdisplay = Sheet(self,
                                  theme = theme,
                                  header_font = ("Calibri", 13, "normal"),
                                  outline_thickness = 0)
        self.sheetdisplay.headers(newheaders = ["Column Name"])
        self.sheetdisplay.data_reference(newdataref = [[hdr.name] for hdr in self.C.headers],
                                         reset_col_positions = False,
                                         reset_row_positions = False,
                                         redraw = False)
        self.sheetdisplay.set_all_cell_sizes_to_text()

        disp = set(self.C.displayed_columns)
        for c in range(len(self.C.headers)):
            if c in disp:
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#8cba66",
                                                  fg = "black")
            else:
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#fc8c55",
                                                  fg = "black")
        self.sheetdisplay.grid(row = 0,
                               column = 0,
                               sticky = "nswe")

        self.button_f1 = frame(self, theme = theme)
        self.button_f1.grid_columnconfigure(0,
                                            weight = 1,
                                            uniform = "x")
        self.button_f1.grid_columnconfigure(1,
                                            weight = 1,
                                            uniform = "x")
        self.button_f1.grid_rowconfigure(0,
                                         weight = 1)
        self.button_f1.grid(row = 1,
                            column = 0,
                            sticky = "nswe")
        self.add_all_button = button(self.button_f1,
                                     text = "Show all",
                                     style = "EF.Std.TButton",
                                     command = self.add_all)
        self.add_all_button.grid(row = 0,
                                 column = 0,
                                 sticky = "nswe",
                                 padx = (20, 10),
                                 pady = (5, 10))
        self.remove_all_button = button(self.button_f1,
                                         text = "Hide all",
                                         style = "EF.Std.TButton",
                                         command = self.remove_all)
        self.remove_all_button.grid(row = 0,
                                     column = 1,
                                     sticky = "nswe",
                                     padx = (10, 20),
                                     pady = (5, 10))

        self.button_f2 = frame(self, theme = theme)
        self.button_f2.grid_columnconfigure(0,
                                            weight = 1,
                                            uniform = "x")
        self.button_f2.grid_columnconfigure(1,
                                            weight = 1,
                                            uniform = "x")
        self.button_f2.grid_rowconfigure(0,
                                         weight = 1)
        self.button_f2.grid(row = 2,
                            column = 0,
                            sticky = "nswe")
        self.confirm_button = button(self.button_f2,
                                     text = "Confirm selections",
                                     style = "EF.Std.TButton",
                                     command = self.confirm)
        self.confirm_button.grid(row = 0,
                                 column = 0,
                                 sticky = "nswe",
                                 padx = (20, 10),
                                 pady = (20, 20))
        self.cancel_button = button(self.button_f2,
                                     text = "Cancel",
                                     style = "EF.Std.TButton",
                                     command = self.cancel)
        self.cancel_button.grid(row = 0,
                                 column = 1,
                                 sticky = "nswe",
                                 padx = (10, 20),
                                 pady = (20, 20))

        self.info = StatusBar(self,
                              text = "** Hidden columns will be excluded when copying to clipboard **",
                              theme = self.C.C.theme)
        self.info.grid(row = 3,
                       column = 0,
                       sticky = "nswe")

        self.sheetdisplay.bind("<Button-1>", self.b1)
        self.sheetdisplay.bind("<B1-Motion>", self.b1_motion)
        self.bind("<Escape>",self.cancel)
        center(self, 500, 700)
        self.deiconify()
        self.wait_window()

    def b1(self, event = None):
        c = self.sheetdisplay.identify_row(event, allow_end = False)
        if c is not None:
            self.toggled = c
            highs = self.sheetdisplay.get_highlighted_cells()
            if highs[(c, 0)][0] == "#8cba66":
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#fc8c55",
                                                  fg = "black")
            else:
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#8cba66",
                                                  fg = "black")
        self.sheetdisplay.refresh()

    def b1_motion(self, event = None):
        c = self.sheetdisplay.identify_row(event, allow_end = False)
        if c is not None and self.toggled != c:
            self.toggled = c
            highs = self.sheetdisplay.get_highlighted_cells()
            if highs[(c, 0)][0] == "#8cba66":
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#fc8c55",
                                                  fg = "black")
            else:
                self.sheetdisplay.highlight_cells(row = c,
                                                  column = 0,
                                                  bg = "#8cba66",
                                                  fg = "black")
        self.sheetdisplay.refresh()

    def add_all(self,event=None):
        self.sheetdisplay.dehighlight_cells("all")
        for c in range(len(self.C.headers)):
            self.sheetdisplay.highlight_cells(row = c,
                                              column = 0,
                                                bg = "#8cba66",
                                                fg = "black")
        self.sheetdisplay.refresh()
            
    def remove_all(self,event=None):
        self.sheetdisplay.dehighlight_cells("all")
        for c in range(len(self.C.headers)):
            self.sheetdisplay.highlight_cells(row = c,
                                              column = 0,
                                                bg = "#fc8c55",
                                                fg = "black")
        self.sheetdisplay.refresh()

    def confirm(self,event=None):
        sels = [k[0] for k, v in self.sheetdisplay.get_highlighted_cells().items() if v[0] == "#8cba66"]
        if not sels or len(sels) == len(self.headers):
            self.C.all_columns_displayed = True
        else:
            self.C.all_columns_displayed = False
            self.chosen_indexes = sorted(sels)
        self.result = True
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class auto_add_condition_num_frame(tk.Frame):
    def __init__(self,
                 parent,
                 col_sel,
                 sheet,
                 theme = "dark"):
        tk.Frame.__init__(self, parent, height = 200, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.C = parent
        self.col_sel = col_sel
        self.sheet_ref = sheet
        self.grid_columnconfigure(1,weight=1)
        self.grid_columnconfigure(3,weight=1)
        self.grid_rowconfigure(3,weight=1)
        self.min_label = label(self,text="Min:",font=EFB, theme = theme)
        self.min_label.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(5,0))
        self.min_entry = numerical_entry_with_scrollbar(self, theme = theme)
        self.min_entry.grid(row=0,column=1,sticky="nswe",padx=(20,0),pady=(20,0))
        self.max_label = label(self,text="Max:",font=EFB, theme = theme)
        self.max_label.grid(row=0,column=2,sticky="nswe",padx=(10,10),pady=(5,0))
        self.max_entry = numerical_entry_with_scrollbar(self, theme = theme)
        self.max_entry.grid(row=0,column=3,sticky="nswe",padx=(10,20),pady=(20,0))
        self.get_col_min = button(self,text="Get column minimum",style="EF.Std.TButton",command=self.get_col_min_val)
        self.get_col_min.grid(row=1,column=1,sticky="nswe",padx=(20,0))
        self.get_col_max = button(self,text="Get column maximum",style="EF.Std.TButton",command=self.get_col_max_val)
        self.get_col_max.grid(row=1,column=3,sticky="nswe",padx=(10,20))
        self.asc_desc_dropdown = ez_dropdown(self,font=EF)
        self.asc_desc_dropdown['values'] = ("ASCENDING","DESCENDING")
        self.asc_desc_dropdown.set_my_value("ASCENDING")
        self.asc_desc_dropdown.grid(row=2,column=1,sticky="nswe",padx=(20,0))
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=3,column=0,columnspan=4,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text="Save conditions",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(20,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(20,20))
        self.result = False
        self.min_val = ""
        self.max_val = ""
        self.order = ""
        self.min_entry.place_cursor()

    def get_col_min_val(self,event=None):
        c = self.col_sel
        try:
            self.min_entry.set_my_value(str(min(float(r[c]) for r in self.sheet_ref if isreal(r[c]))))
        except:
            pass

    def get_col_max_val(self,event=None):
        c = self.col_sel
        try:
            self.max_entry.set_my_value(str(max(float(r[c]) for r in self.sheet_ref if isreal(r[c]))))
        except:
            pass

    def confirm(self,event=None):
        self.result = True
        self.min_val = self.min_entry.get_my_value()
        self.max_val = self.max_entry.get_my_value()
        self.order = self.asc_desc_dropdown.get_my_value()
        self.destroy()

    def cancel(self,event=None):
        self.destroy()


class auto_add_condition_date_frame(tk.Frame):
    def __init__(self,
                 parent,
                 col_sel,
                 sheet,
                 DATE_FORM,
                 theme = "dark"):
        tk.Frame.__init__(self, parent, height = 225, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.C = parent
        self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(2,weight=1)
        self.grid_rowconfigure(3,weight=1)
        self.col_sel = col_sel
        self.sheet_ref = sheet
        self.DATE_FORM = DATE_FORM
        if DATE_FORM == "%d/%m/%Y":
            label_form = "DD/MM/YYYY"
        elif DATE_FORM == "%m/%d/%Y":
            label_form = "MM/DD/YYYY"
        elif DATE_FORM == "%Y/%m/%d":
            label_form = "YYYY/MM/DD"
        else:
            self.DATE_FORM = "DD/MM/YYYY"
            label_form = "DD/MM/YYYY"
        self.min_label = label(self,text="Min  " + label_form,font=EFB, theme = theme)
        self.min_label.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(5,0))
        self.min_entry = date_entry(self,DATE_FORM, theme = theme)
        self.min_entry.grid(row=0,column=1,sticky="nswe",padx=(20,0),pady=(20,0))
        self.max_label = label(self,text="Max  " + label_form,font=EFB, theme = theme)
        self.max_label.grid(row=0,column=2,sticky="nswe",padx=(10,10),pady=(5,0))
        self.max_entry = date_entry(self,DATE_FORM, theme = theme)
        self.max_entry.grid(row=0,column=3,sticky="nswe",padx=(10,20),pady=(20,0))
        self.get_col_min = button(self,text="Get column minimum",style="EF.Std.TButton",command=self.get_col_min_val)
        self.get_col_min.grid(row=1,column=1,sticky="nswe",padx=(20,0))
        self.get_col_max = button(self,text="Get column maximum",style="EF.Std.TButton",command=self.get_col_max_val)
        self.get_col_max.grid(row=1,column=3,sticky="nswe",padx=(10,20))
        self.asc_desc_dropdown = ez_dropdown(self,font=EF)
        self.asc_desc_dropdown['values'] = ("ASCENDING","DESCENDING")
        self.asc_desc_dropdown.set_my_value("ASCENDING")
        self.asc_desc_dropdown.grid(row=2,column=1,sticky="nswe",padx=(20,0))
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=3,column=0,columnspan=4,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text="Save conditions",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(20,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(20,20))
        self.result = False
        self.min_val = ""
        self.max_val = ""
        self.order = ""
        self.min_entry.place_cursor()

    def detect_date_form(self,date):
        forms = []
        for form in ("%d/%m/%Y","%m/%d/%Y","%Y/%m/%d"):
            try:
                datetime.datetime.strptime(date,form).date()
                forms.append(form)
            except:
                pass
        if len(forms) == 1:
            return forms[0]
        return False

    def get_col_min_val(self,event=None):
        c = self.col_sel
        try:
            self.min_entry.set_my_value(datetime.datetime.strftime(min(datetime.datetime.strptime(r[c],self.DATE_FORM)
                                                                       for r in self.sheet_ref if self.detect_date_form(r[c]) == self.DATE_FORM),self.DATE_FORM))                                                   
        except:
            pass

    def get_col_max_val(self,event=None):
        c = self.col_sel
        try:
            self.max_entry.set_my_value(datetime.datetime.strftime(max(datetime.datetime.strptime(r[c],self.DATE_FORM)
                                                                       for r in self.sheet_ref if self.detect_date_form(r[c]) == self.DATE_FORM),self.DATE_FORM))
        except:
            pass

    def confirm(self,event=None):
        self.result = True
        self.min_val = self.min_entry.get_my_value()
        self.max_val = self.max_entry.get_my_value()
        self.order = self.asc_desc_dropdown.get_my_value()
        self.destroy()

    def cancel(self,event=None):
        self.destroy()
        

class edit_condition_frame(tk.Frame):
    def __init__(self,
                 parent,
                 condition,
                 colors,
                 color=None,
                 coltype="Text Detail",
                 confirm_text="Save condition",
                 theme = "dark"):
        tk.Frame.__init__(self,parent, height = 160, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.C = parent
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        if coltype in ("ID", "Parent", "Text Detail"):
            self.if_cell_label = label(self,text="If cell text is exactly:",font=EFB, theme = theme)
        else:
            self.if_cell_label = label(self,text="If cell value is:",font=EFB, theme = theme)
        self.if_cell_label.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(20,40))
        self.condition_display = condition_entry_with_scrollbar(self,coltype=coltype, theme = theme)
        self.condition_display.set_my_value(condition)
        self.condition_display.grid(row=0,column=1,sticky="nswe",pady=(20,20),padx=(0,0))
        self.color_dropdown = ez_dropdown(self,EF)
        self.color_dropdown['values'] = colors
        if color is None:
            self.color_dropdown.set_my_value(colors[0])
        else:
            self.color_dropdown.set_my_value(color)
        self.color_dropdown.grid(row=0,column=2,sticky="nswe",pady=(20,20),padx=(0,20))
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=1,column=0,columnspan=3,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text=confirm_text,style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(0,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(0,20))
        self.result = False
        self.new_condition = ""
        self.color_dropdown.bind("<<ComboboxSelected>>",self.disable_cancel)
        self.condition_display.place_cursor()
    def disable_cancel(self,event=None):
        try:
            self.cancel_button.config(state="disabled")
            self.after(300,self.enable_cancel)
        except:
            pass
    def enable_cancel(self,event=None):
        self.cancel_button.config(state="normal")
    def confirm(self,event=None):
        self.result = True
        self.new_condition = self.condition_display.get_my_value()
        self.color = self.color_dropdown.get_my_value()
        self.destroy()
    def cancel(self,event=None):
        self.destroy()


class edit_formula_frame(tk.Frame):
    def __init__(self,parent,colname,formula,type_,formula_apply_only,theme = "dark"):
        tk.Frame.__init__(self,parent,height=215, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.C = parent
        self.type_ = type_
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(2,weight=1)
        self.col_name_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_name_display.set_my_value(colname)
        self.col_name_display.grid(row=0,column=0,columnspan=4,sticky="nswe",pady=(20,20),padx=(20,20))
        self.equals_label = label(self,text=" = ",font=EFB, theme = theme)
        self.equals_label.grid(row=1,column=0,sticky="nswe",padx=(20,10),pady=(0,20))
        self.formula_display = formula_entry_with_scrollbar(self, theme = theme)
        self.formula_display.set_my_value(formula)
        self.formula_display.grid(row=1,column=1,sticky="nswe",pady=(0,20),padx=(0,5))
        self.formula_only_apply = label(self,text="Only apply formula\nif no cells are blank",font=EFB, theme = theme)
        self.formula_only_apply.grid(row=1,column=2,sticky="nswe",padx=(0,5),pady=(0,20))
        self.formula_only_apply = ez_dropdown(self,EF)
        self.formula_only_apply['values'] = ["True","False"]
        self.formula_only_apply.set_my_value(f"{formula_apply_only}")
        self.formula_only_apply.grid(row=1,column=3,sticky="nswe",pady=(0,20),padx=(0,20))
        self.formula_only_apply.bind("<<ComboboxSelected>>",lambda event: self.formula_display.focus_set())
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=2,column=0,columnspan=4,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text="Save formula",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(0,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(0,20))
        self.result = False
        self.new_formula = ""
        self.formula_display.place_cursor()
        
    def confirm(self, event = None):
        self.result = True
        self.formula_only_apply_result = self.formula_only_apply.get_my_value()
        self.new_formula = self.formula_display.get_my_value()
        self.destroy()

    def cancel(self, event = None):
        self.destroy()


class edit_validation_frame(tk.Frame):
    def __init__(self,parent,coltype,colname,validation, theme = "dark"):
        tk.Frame.__init__(self, parent, height = 210)
        self.grid_propagate(False)
        self.C = parent
        self.config(bg = theme_bg(theme))
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(2,weight=1)
        self.col_name_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_name_display.set_my_value(colname)
        self.col_name_display.grid(row=0,column=0,columnspan=2,sticky="nswe",pady=(20,20),padx=(20,20))
        self.valid_label = label(self,text=" Valid details:",font=EFB, theme = theme)
        self.valid_label.grid(row=1,column=0,sticky="nswe",padx=(20,10),pady=(0,40))
        if coltype == "Text Detail":
            self.validation_display = entry_with_scrollbar(self, theme = theme)
        else:
            self.validation_display = validation_entry_with_scrollbar(self, coltype, theme = theme)
        if validation:
            self.validation_display.set_my_value(",".join(validation))
            self.validation_display.place_cursor()
        elif not validation:
            try:
                self.validation_display.disable_checking()
            except:
                pass
            x = (
"""Type in values/text separated by commas ',' which will appear in a dropdown box when a user attempts to edit a cell in this column. To set an non-value
either put a comma at the start or a double comma wherever you want a non-value to show in the dropdown box.
For Numerical Detail columns you can set a range of values by using '_' like so: 'start_stop_step' e.g. '5_0_-1' which would create 5,4,3,2,1,0
For Date Detail columns and to limit user entry to only UK working day dates; enter: only uk working days"""
            )
            self.validation_display.set_my_value(" ".join(x.split("\n")))
        self.validation_display.grid(row=1,column=1,sticky="nswe",pady=(0,20),padx=(0,20))
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text="Save validation",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(0,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(0,20))
        self.result = False
        self.new_validation = ""
        if not validation:
            self.validation_display.my_entry.bind("<FocusIn>",self.show_remove_help)
            
    def show_remove_help(self,event=None):
        self.validation_display.set_my_value("")
        self.validation_display.my_entry.unbind("<FocusIn>")
        try:
            self.validation_display.enable_checking()
        except:
            pass
        
    def confirm(self,event=None):
        self.result = True
        self.new_validation = self.validation_display.get_my_value()
        if self.new_validation.lower() in ("only uk working days", "only england working days", "only wales working days", "only scotland working days", "only ni working days"):
            self.new_validation = self.new_validation.lower()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class validation_entry_with_scrollbar(tk.Frame):
    def __init__(self,parent,coltype,theme = "dark"):
        tk.Frame.__init__(self,parent)
        self.config(bg = theme_bg(theme))
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.my_entry = validation_normal_entry(self, font = EF, coltype = coltype, theme = theme)
        self.my_entry.grid(row=0,column=0,sticky="nswe")
        self.my_scrollbar = scrollbar(self,self.my_entry.xview,
                                      "horizontal",
                                      self.my_entry)
        self.my_scrollbar.grid(row=1,column=0,sticky="ew")
        
    def change_my_state(self,state,event=None):
        self.my_entry.config(state=state)
        
    def place_cursor(self,event=None):
        self.my_entry.focus_set()
        
    def get_my_value(self,event=None):
        return self.my_entry.get()
    
    def set_my_value(self,val,event=None):
        self.my_entry.set_my_value(val)
        
    def disable_checking(self):
        self.my_entry.disable_checking()
        
    def enable_checking(self):
        self.my_entry.enable_checking()


class validation_normal_entry(tk.Entry):
    def __init__(self, parent, font, coltype, width_ = None, theme = "dark"):
        tk.Entry.__init__(self, parent, font = font,
                          background = theme_entry_bg(theme),
                          foreground = theme_entry_fg(theme),
                          disabledbackground = theme_entry_dbg(theme),
                          disabledforeground = theme_entry_dfg(theme),
                          insertbackground = theme_entry_cursor(theme),
                          readonlybackground = theme_entry_dbg(theme))
        if width_:
            self.config(width=width_)
        if coltype == "Numerical Detail":
            self.allowed_chars = {"0","1","2","3","4","5","6","7","8","9",",","-","_","."}
        if coltype == "Date Detail":
            self.allowed_chars = {"0","1","2","3","4","5","6","7","8","9",",","/","-",
                                  "o","O","n","N","l","L","y","Y","u","U","k","K","c","C","t","T","w","W","e","E",
                                  "w","W","o","O","r","R","k","K","i","I","n","N",
                                  "g","G","d","D","a","A","y","Y","s","S"," "}
        self.sv = tk.StringVar()
        self.config(textvariable=self.sv)
        self.sv.trace("w", lambda name, index, mode, sv=self.sv: self.validate_(self.sv))
        self.rc_popup_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Select all",accelerator="Ctrl+A",command=self.select_all,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Cut",accelerator="Ctrl+X",command=self.cut,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Copy",accelerator="Ctrl+C",command=self.copy,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Paste",accelerator="Ctrl+V",command=self.paste,**menu_kwargs)
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def validate_(self,sv):
        self.sv.set("".join([c.lower() for c in self.sv.get() if c in self.allowed_chars]))
        
    def disable_checking(self):
        self.config(textvariable="")
        
    def enable_checking(self):
        self.config(textvariable=self.sv)
        self.sv.set("".join([c.lower() for c in self.sv.get() if c in self.allowed_chars]))
        
    def rc(self,event):
        self.focus_set()
        self.rc_popup_menu.tk_popup(event.x_root,event.y_root)
        
    def select_all(self,event=None):
        self.event_generate("<Control-a>")
        return "break"
    
    def cut(self,event=None):
        self.event_generate("<Control-x>")
        return "break"
    
    def copy(self,event=None):
        self.event_generate("<Control-c>")
        return "break"
    
    def paste(self,event=None):
        self.event_generate("<Control-v>")
        return "break"
    
    def set_my_value(self,newvalue):
        self.delete(0,"end")
        self.insert(0,str(newvalue))
        
    def enable_me(self):
        self.config(state="normal")
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def disable_me(self):
        self.config(state="disabled")
        self.unbind("<1>")
        self.unbind(get_platform_rc_binding())


class condition_entry_with_scrollbar(tk.Frame):
    def __init__(self,parent,coltype="Text Detail", theme = "dark"):
        tk.Frame.__init__(self,parent)
        self.config(bg = theme_bg(theme))
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.my_entry = condition_normal_entry(self,font=EF,coltype=coltype, theme = theme)
        self.my_entry.grid(row=0,column=0,sticky="nswe")
        self.my_scrollbar = scrollbar(self,self.my_entry.xview,
                                      "horizontal",
                                      self.my_entry)
        self.my_scrollbar.grid(row=1,column=0,sticky="ew")
        
    def change_my_state(self,state,event=None):
        self.my_entry.config(state=state)
        
    def place_cursor(self,event=None):
        self.my_entry.focus_set()
        
    def get_my_value(self,event=None):
        return self.my_entry.get()
    
    def set_my_value(self,val,event=None):
        self.my_entry.set_my_value(val)


class condition_normal_entry(tk.Entry):
    def __init__(self, parent, font, coltype = "Text Detail", width_ = None, theme = "dark"):
        tk.Entry.__init__(self, parent, font = font,
                          background = theme_entry_bg(theme),
                          foreground = theme_entry_fg(theme),
                          disabledbackground = theme_entry_dbg(theme),
                          disabledforeground = theme_entry_dfg(theme),
                          insertbackground = theme_entry_cursor(theme),
                          readonlybackground = theme_entry_dbg(theme))
        if width_:
            self.config(width=width_)
        self.coltype = coltype
        if self.coltype not in ("ID", "Parent", "Text Detail"):
            self.validate_text = True
        else:
            self.validate_text = False
        if coltype != "Date Detail":
            self.allowed_chars = {"a","n","d","o","r","A","N","D","O","R","0","1","2","3","4","5","6","7","8","9","!",">","<","="," ","-",".","C","c"}
        else:
            self.allowed_chars = {"a","n","d","o","r","A","N","D","O","R","0","1","2","3","4","5","6","7","8","9","!",">","<","="," ","/","C","c","-"}
        self.sv = tk.StringVar()
        self.config(textvariable=self.sv)
        self.sv.trace("w", lambda name, index, mode, sv=self.sv: self.validate_(self.sv))
        self.rc_popup_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Select all",accelerator="Ctrl+A",command=self.select_all,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Cut",accelerator="Ctrl+X",command=self.cut,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Copy",accelerator="Ctrl+C",command=self.copy,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Paste",accelerator="Ctrl+V",command=self.paste,**menu_kwargs)
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        self.set_my_value(" ")
        
    def validate_(self,sv):
        if self.validate_text:
            self.sv.set("".join([c.lower() for c in self.sv.get().replace("  "," ") if c in self.allowed_chars]))
            
    def rc(self,event):
        self.focus_set()
        self.rc_popup_menu.tk_popup(event.x_root,event.y_root)
        
    def select_all(self,event=None):
        self.event_generate("<Control-a>")
        return "break"
    
    def cut(self,event=None):
        self.event_generate("<Control-x>")
        return "break"
    
    def copy(self,event=None):
        self.event_generate("<Control-c>")
        return "break"
    
    def paste(self,event=None):
        self.event_generate("<Control-v>")
        return "break"
    
    def set_my_value(self,newvalue):
        self.delete(0,"end")
        self.insert(0,str(newvalue))
        
    def enable_me(self):
        self.config(state="normal")
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def disable_me(self):
        self.config(state="disabled")
        self.unbind("<1>")
        self.unbind(get_platform_rc_binding())


class formula_entry_with_scrollbar(tk.Frame):
    def __init__(self, parent, theme = "dark"):
        tk.Frame.__init__(self, parent, bg = theme_bg(theme))
        self.C = parent
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.my_entry = formula_normal_entry(self, font = EF, theme = theme)
        self.my_entry.grid(row=0,column=0,sticky="nswe")
        self.my_scrollbar = scrollbar(self,self.my_entry.xview,
                                      "horizontal",
                                      self.my_entry)
        self.my_scrollbar.grid(row=1,column=0,sticky="ew")
        
    def change_my_state(self,state,event=None):
        self.my_entry.config(state=state)
        
    def place_cursor(self,event=None):
        self.my_entry.focus_set()
        
    def get_my_value(self,event=None):
        return self.my_entry.get()
    
    def set_my_value(self,val,event=None):
        self.my_entry.set_my_value(val)


class formula_normal_entry(tk.Entry):
    def __init__(self, parent, font, width_ = None, theme = "dark"):
        tk.Entry.__init__(self, parent, font = font,
                          background = theme_entry_bg(theme),
                          foreground = theme_entry_fg(theme),
                          disabledbackground = theme_entry_dbg(theme),
                          disabledforeground = theme_entry_dfg(theme),
                          insertbackground = theme_entry_cursor(theme),
                          readonlybackground = theme_entry_dbg(theme))
        if width_:
            self.config(width = width_)
        self.C = parent
        self.allowed_chars = {"d","D","c","C","0","1","2","3","4","5","6","7","8","9","+","(",")","*","-","/","%",".","^"}
        self.sv = tk.StringVar()
        self.config(textvariable=self.sv)
        try:
            if self.C.C.type_ != "Text Detail":
                self.sv.trace("w", lambda name, index, mode, sv=self.sv: self.validate_(self.sv))
        except:
            self.sv.trace("w", lambda name, index, mode, sv=self.sv: self.validate_(self.sv))
        self.rc_popup_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Select all",accelerator="Ctrl+A",command=self.select_all,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Cut",accelerator="Ctrl+X",command=self.cut,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Copy",accelerator="Ctrl+C",command=self.copy,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Paste",accelerator="Ctrl+V",command=self.paste,**menu_kwargs)
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def validate_(self,sv):
        self.sv.set("".join([c.lower() for c in self.sv.get() if c in self.allowed_chars]))
        
    def rc(self,event):
        self.focus_set()
        self.rc_popup_menu.tk_popup(event.x_root,event.y_root)
        
    def select_all(self,event=None):
        self.event_generate("<Control-a>")
        return "break"
    
    def cut(self,event=None):
        self.event_generate("<Control-x>")
        return "break"
    
    def copy(self,event=None):
        self.event_generate("<Control-c>")
        return "break"
    
    def paste(self,event=None):
        self.event_generate("<Control-v>")
        return "break"
    
    def set_my_value(self,newvalue):
        self.delete(0,"end")
        self.insert(0,str(newvalue))
        
    def enable_me(self):
        self.config(state="normal")
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def disable_me(self):
        self.config(state="disabled")
        self.unbind("<1>")
        self.unbind(get_platform_rc_binding())


class askconfirm_frame(tk.Frame):
    def __init__(self,parent,action,confirm_text="Confirm",cancel_text="Cancel",bgcolor="green",fgcolor="white", theme = "dark"):
        tk.Frame.__init__(self,parent,height=150, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.C = parent
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.action_label = label(self,text="Action:",font=EF, theme = theme)
        self.action_label.config(background=bgcolor,foreground=fgcolor)
        self.action_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.action_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.action_display.my_entry.config(font=ERR_ASK_FNT)
        self.action_display.set_my_value(action)
        self.action_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=1,column=0,columnspan=2,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text=confirm_text,style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=20)
        self.cancel_button = button(self.button_frame,text=cancel_text,style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=20)
        self.boolean = False
        self.action_display.place_cursor()
    def confirm(self,event=None):
        self.boolean = True
        self.destroy()
    def cancel(self,event=None):
        self.destroy()


class askconfirm_three(tk.Toplevel):
    def __init__(self,
                 C,
                 action,
                 text1 = "Confirm 1",
                 text2 = "Confirm 2",
                 text3 = "Cancel",
                 bgcolor = "green",
                 fgcolor = "white",
                 theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format = "gif", data = top_left_icon))
        self.title("Confirm Action - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.action_label = label(self, text = "Action:", font = EF)
        self.action_label.config(background=bgcolor,foreground=fgcolor)
        self.action_label.grid(row=0,column=0,sticky="nswe",pady=(20,5),padx=20)
        self.action_display = display_text(parent = self, text = action)
        self.action_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.action_display.config(height = 75)
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(10,20))
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(2,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.button1 = button(self.button_frame,text=text1,style="EF.Std.TButton",command=self.button1)
        self.button1.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(20,10))
        self.button2 = button(self.button_frame,text=text2,style="EF.Std.TButton",command=self.button2)
        self.button2.grid(row=0,column=1,sticky="nswe",padx=(10,10),pady=(20,10))
        self.button3 = button(self.button_frame,text=text3,style="EF.Std.TButton",command=self.cancel)
        self.button3.grid(row=0,column=2,sticky="nswe",padx=(10,20),pady=(20,10))
        self.choice = None
        self.bind("<Escape>",self.cancel)
        self.action_display.place_cursor()
        center(self,600,200)
        self.deiconify()
        self.wait_window()
    def button1(self,event=None):
        self.choice = 1
        self.destroy()
    def button2(self,event=None):
        self.choice = 2
        self.destroy()
    def cancel(self,event=None):
        self.destroy()


class askconfirm(tk.Toplevel):
    def __init__(self,C,action,confirm_text="Confirm",cancel_text="Cancel",bgcolor="green",fgcolor="white", theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format = "gif", data = top_left_icon))
        self.title("Confirm Action - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.action_label = label(self, text = "Action:", font = EF)
        self.action_label.config(background=bgcolor,foreground=fgcolor)
        self.action_label.grid(row=0,column=0,sticky="nswe",pady=(20,5),padx=20)
        self.action_display = display_text(parent = self, text = action, theme = theme)
        self.action_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.action_display.config(height = 75)
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(10,20))
        self.button_frame.grid_columnconfigure(0, weight = True, uniform = "x")
        self.button_frame.grid_columnconfigure(1, weight = True, uniform = "x")
        self.button_frame.grid_rowconfigure(0, weight = True)
        self.confirm_button = button(self.button_frame,text=confirm_text,style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row = 0, column = 0, sticky = "nswe", padx = (0, 20))
        self.cancel_button = button(self.button_frame, text=cancel_text,style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe")
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.boolean = False
        self.action_display.place_cursor()
        center(self, 530, 155)
        self.deiconify()
        self.wait_window()
    def confirm(self,event=None):
        self.boolean = True
        self.destroy()
    def cancel(self,event=None):
        self.destroy()


class error_frame(tk.Frame):
    def __init__(self,parent,msg, theme = "dark"):
        tk.Frame.__init__(self,parent,height=150, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.C = parent
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.errorlabel = label(self,text="Error\nmessage:",font=EF, theme = theme)
        self.errorlabel.config(background="red",foreground="white")
        self.errorlabel.grid(row=0,column=0,sticky="nswe",padx=20)
        self.error_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.error_display.my_entry.config(font=ERR_ASK_FNT)
        self.error_display.set_my_value(msg)
        self.error_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.confirm_button = button(self,text="Okay",style="TF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(10,20))
    def confirm(self,event=None):
        self.destroy()

        
class treeview(ttk.Treeview):
    def __init__(self,parent,selectmode,style):
        ttk.Treeview.__init__(self,
                              parent,
                              selectmode=selectmode,
                              style=style)
        self.new_xview = tk.XView.xview
    def xview(self,*args):
        if "units" in args:
            if args[1] == "1":
                return self.new_xview(self,"scroll","27","units")
            elif args[1] == "-1":
                return self.new_xview(self,"scroll","-27","units")
        else:
            return self.new_xview(self,*args)


class working_text(tk.Text):
    def __init__(self,parent,wrap,font=("Calibri",12), theme = "dark", use_entry_bg = True, override_bg = None):
        tk.Text.__init__(self,
                         parent,
                         wrap=wrap,
                         font=font,
                         spacing1=5,spacing2=5)
        self.config(bg = theme_entry_bg(theme) if use_entry_bg else theme_bg(theme),
                    fg = theme_entry_fg(theme) if use_entry_bg else theme_fg(theme),
                    insertbackground = theme_entry_fg(theme) if use_entry_bg else theme_fg(theme))
        if override_bg is not None:
            self.config(bg = override_bg)
        self.rc_popup_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Select all",accelerator="Ctrl+A",command=self.select_all,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Cut",accelerator="Ctrl+X",command=self.cut,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Copy",accelerator="Ctrl+C",command=self.copy,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Paste",accelerator="Ctrl+V",command=self.paste,**menu_kwargs)
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
    def rc(self,event):
        self.focus_set()
        self.rc_popup_menu.tk_popup(event.x_root,event.y_root)
    def select_all(self,event=None):
        self.event_generate("<Control-a>")
        return "break"
    def cut(self,event=None):
        self.event_generate("<Control-x>")
        return "break"
    def copy(self,event=None):
        self.event_generate("<Control-c>")
        return "break"
    def paste(self,event=None):
        self.event_generate("<Control-v>")
        return "break"


class display_text(tk.Frame):
    def __init__(self, parent, text = "", theme = "dark"):
        tk.Frame.__init__(self, parent, bg = theme_bg(theme))
        self.C = parent
        self.grid_rowconfigure(0, weight = 1)
        self.grid_columnconfigure(0, weight = 1)
        self.textbox = working_text(self, wrap = "word", theme = theme, use_entry_bg = False)
        self.textbox.grid_propagate(False)
        self.grid_propagate(False)
        self.yscrollb = scrollbar(self,
                                  self.textbox.yview,
                                  "vertical",
                                  self.textbox)
        self.textbox.delete(1.0, "end")
        self.textbox.insert(1.0, text)
        self.textbox.config(state = "disabled", relief = "flat")
        self.textbox.grid(row = 0, column = 0, sticky = "nswe")
        self.yscrollb.grid(row = 0, column = 1, sticky = "ns")
        
    def place_cursor(self, index = None):
        if not index:
            self.textbox.focus_set()
            
    def get_my_value(self):
        return self.textbox.get("1.0", "end")
    
    def set_my_value(self, value):
        self.textbox.config(state = "normal")
        self.textbox.delete(1.0, "end")
        self.textbox.insert(1.0, value)
        self.textbox.config(state = "readonly")
        
    def change_my_state(self, new_state):
        self.current_state = new_state
        self.textbox.config(state = self.current_state)
        
    def change_my_width(self, new_width):
        self.textbox.config(width = new_width)
        
    def change_my_height(self, new_height):
        self.textbox.config(height = new_height)


class wrapped_text_with_find_and_yscroll(tk.Frame):
    def __init__(self,parent,text,current_state,height=None, theme = "dark"):
        tk.Frame.__init__(self, parent, bg = theme_bg(theme))
        self.C = parent
        self.theme = theme
        self.grid_rowconfigure(1,weight=1)
        self.grid_columnconfigure(0,weight=1)
        self.current_state = current_state
        self.word = ""
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        self.find_frame = frame(self, theme = theme)
        self.find_frame.grid(row=0,column=0,columnspan=2,sticky="nswe")
        self.save_text_button = button(self.find_frame,text="Save as",
                                       style="Std.TButton",
                                       command=self.save_text)
        self.save_text_button.pack(side="left",fill="x")
        self.find_button = button(self.find_frame,text="Find: ",
                                  style="Std.TButton",
                                  command=self.find)
        self.find_button.pack(side="left",fill="x")
        self.find_window = normal_entry(self.find_frame,font=BF, theme = theme)
        self.find_window.bind("<Return>",self.find)
        self.find_window.pack(side="left",fill="x",expand=True)
        self.find_reset_button = button(self.find_frame,text="X",style="Std.TButton",command=self.find_reset)
        self.find_reset_button.pack(side="left",fill="x")
        self.find_results_label = label(self.find_frame,"0/0",BF, theme = theme)
        self.find_results_label.pack(side="left",fill="x")
        self.find_up_button = button(self.find_frame,text="▲",style="Std.TButton",command=self.find_up)
        self.find_up_button.pack(side="left",fill="x")
        self.find_down_button = button(self.find_frame,text="▼",style="Std.TButton",command=self.find_down)
        self.find_down_button.pack(side="left",fill="x")
        self.textbox = working_text(self,wrap="word", theme = theme)
        if height:
            self.textbox.config(height=height)
        self.yscrollb = scrollbar(self,
                                  self.textbox.yview,
                                  "vertical",
                                  self.textbox)
        self.textbox.delete(1.0,"end")
        self.textbox.insert(1.0,text)
        self.textbox.config(state=self.current_state)
        self.textbox.grid(row=1,column=0,sticky="nswe")
        self.yscrollb.grid(row=1,column=1,sticky="ns")
        
    def place_cursor(self,index=None):
        if not index:
            self.textbox.focus_set()
            
    def get_my_value(self):
        return self.textbox.get("1.0","end")
    
    def set_my_value(self,value):
        self.textbox.config(state="normal")
        self.textbox.delete(1.0,"end")
        self.textbox.insert(1.0,value)
        self.textbox.config(state=self.current_state)
        
    def change_my_state(self,new_state):
        self.current_state = new_state
        self.textbox.config(state=self.current_state)
        
    def change_my_width(self,new_width):
        self.textbox.config(width=new_width)
        
    def change_my_height(self,new_height):
        self.textbox.config(height=new_height)
        
    def save_text(self):
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save text on popup window",
                                               filetypes=[('Text File','.txt'),('CSV File','.csv')],
                                               defaultextension=".txt",
                                               confirmoverwrite=True)
        if not newfile:
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv",".txt")):
            errorpopup = error(self.C,"Can only save .csv/.txt files", theme = self.theme)
            return
        self.save_text_button.change_text("Saving...")
        try:
            with open(newfile,"w") as fh:
                fh.write(self.textbox.get("1.0","end")) #remove last newline? [:-2]
        except:
            errorpopup = error(self.C,"Error saving file", theme = self.theme)
            self.save_text_button.change_text("Save as")
            return
        self.save_text_button.change_text("Save as")
        
    def find(self,event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        self.addon = "+" + str(len(self.word)) + "c"
        start = "1.0"
        while start:
            start = self.textbox.search(self.word,index=start,nocase=1,stopindex="end")
            if start:
                end = start + self.addon
                self.find_results.append(start)
                self.textbox.tag_add("i",start,end)
                start = end
        if self.find_results:
            self.textbox.tag_config("i",background="Yellow")
            self.find_results_label.config(text="1/"+str(len(self.find_results)))
            self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
            self.textbox.tag_config("c",background="Orange")
            self.textbox.see(self.find_results[self.results_number])
            
    def find_up(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_down(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_reset(self,newfind=False):
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        if newfind == False:
            self.find_window.delete(0,"end")
        for tag in self.textbox.tag_names():
            self.textbox.tag_delete(tag)
        self.find_results_label.config(text="0/0")


class scrollbar(ttk.Scrollbar):
    def __init__(self,parent,command,orient,widget):
        ttk.Scrollbar.__init__(self,
                               parent,
                               command=command,
                               orient=orient)
        self.orient = orient
        self.widget = widget
        if self.orient == "vertical":
            self.widget.configure(yscrollcommand=self.set)
        elif self.orient == "horizontal":
            self.widget.configure(xscrollcommand=self.set)


class readonly_entry(tk.Entry):
    def __init__(self, parent, font, width_ = None, theme = "dark"):
        tk.Entry.__init__(self, parent, font = font, state = "readonly",
                          background = theme_entry_bg(theme),
                          foreground = theme_entry_fg(theme),
                          disabledbackground = theme_entry_dbg(theme),
                          disabledforeground = theme_entry_dfg(theme),
                          insertbackground = theme_entry_cursor(theme),
                          readonlybackground = theme_entry_dbg(theme))
        if width_:
            self.config(width=width_)
        self.rc_popup_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Select all",accelerator="Ctrl+A",command=self.select_all,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Cut",accelerator="Ctrl+X",command=self.cut,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Copy",accelerator="Ctrl+C",command=self.copy,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Paste",accelerator="Ctrl+V",command=self.paste,**menu_kwargs)
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def rc(self,event):
        self.focus_set()
        self.rc_popup_menu.tk_popup(event.x_root,event.y_root)
        
    def select_all(self,event=None):
        self.event_generate("<Control-a>")
        return "break"
    
    def cut(self,event=None):
        self.event_generate("<Control-x>")
        return "break"
    
    def copy(self,event=None):
        self.event_generate("<Control-c>")
        return "break"
    
    def paste(self,event=None):
        self.event_generate("<Control-v>")
        return "break"
    
    def set_my_value(self,newvalue):
        self.config(state="normal")
        self.delete(0,"end")
        self.insert(0,str(newvalue))
        self.config(state="readonly")

    def change_theme(self, theme = "dark"):
        self.config(background = theme_entry_bg(theme),
                      foreground = theme_entry_fg(theme),
                      disabledbackground = theme_entry_dbg(theme),
                      disabledforeground = theme_entry_dfg(theme),
                      insertbackground = theme_entry_cursor(theme),
                      readonlybackground = theme_entry_dbg(theme))


class normal_entry(tk.Entry):
    def __init__(self,parent,font,width_=None,relief="sunken",
                 border=1,textvariable=None, theme = "dark"):
        tk.Entry.__init__(self,parent,font=font,relief=relief,
                          border=border,textvariable=textvariable,
                          background = theme_entry_bg(theme),
                          foreground = theme_entry_fg(theme),
                          disabledbackground = theme_entry_dbg(theme),
                          disabledforeground = theme_entry_dfg(theme),
                          insertbackground = theme_entry_cursor(theme),
                          readonlybackground = theme_entry_dbg(theme)
                          )
        if width_:
            self.config(width=width_)
        if textvariable:
            self.config(textvariable=textvariable)
        self.rc_popup_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Select all",accelerator="Ctrl+A",command=self.select_all,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Cut",accelerator="Ctrl+X",command=self.cut,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Copy",accelerator="Ctrl+C",command=self.copy,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Paste",accelerator="Ctrl+V",command=self.paste,**menu_kwargs)
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def rc(self,event):
        self.focus_set()
        self.rc_popup_menu.tk_popup(event.x_root,event.y_root)
        
    def select_all(self,event=None):
        self.event_generate("<Control-a>")
        return "break"
    
    def cut(self,event=None):
        self.event_generate("<Control-x>")
        return "break"
    
    def copy(self,event=None):
        self.event_generate("<Control-c>")
        return "break"
    
    def paste(self,event=None):
        self.event_generate("<Control-v>")
        return "break"
    
    def set_my_value(self,newvalue):
        self.delete(0,"end")
        self.insert(0,str(newvalue))
        
    def enable_me(self):
        self.config(state="normal")
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def disable_me(self):
        self.config(state="disabled")
        self.unbind("<1>")
        self.unbind(get_platform_rc_binding())
        

class readonly_entry_with_scrollbar(tk.Frame):
    def __init__(self,parent,font=EF, theme = "dark"):
        tk.Frame.__init__(self,parent, bg = theme_bg(theme)) 
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.my_entry = readonly_entry(self, font = font, theme = theme)
        self.my_entry.grid(row=0,column=0,sticky="nswe")
        self.my_scrollbar = scrollbar(self,self.my_entry.xview,
                                      "horizontal",
                                      self.my_entry)
        self.my_scrollbar.grid(row=1,column=0,sticky="ew")
        
    def change_my_state(self,state,event=None):
        self.my_entry.config(state=state)
        
    def place_cursor(self,event=None):
        self.my_entry.focus_set()
        
    def get_my_value(self,event=None):
        return self.my_entry.get()
    
    def set_my_value(self,val,event=None):
        self.my_entry.set_my_value(val)

    def change_text(self, text = ""):
        self.my_entry.set_my_value(text)

    def change_theme(self, theme = "dark"):
        self.config(bg = theme_bg(theme))
        self.my_entry.change_theme(theme)


class entry_with_scrollbar(tk.Frame):
    def __init__(self, parent, theme = "dark"):
        tk.Frame.__init__(self, parent, bg = theme_bg(theme))
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.my_entry = normal_entry(self, font = EF, theme = theme)
        self.my_entry.grid(row=0,column=0,sticky="nswe")
        self.my_scrollbar = scrollbar(self,self.my_entry.xview,
                                      "horizontal",
                                      self.my_entry)
        self.my_scrollbar.grid(row=1,column=0,sticky="ew")
        
    def change_my_state(self,state,event=None):
        self.my_entry.config(state=state)
        
    def place_cursor(self,event=None):
        self.my_entry.focus_set()
        
    def get_my_value(self,event=None):
        return self.my_entry.get()
    
    def set_my_value(self,val,event=None):
        self.my_entry.set_my_value(val)


class ez_dropdown(ttk.Combobox):
    def __init__(self,parent,font,width_=None):
        self.displayed = tk.StringVar()
        ttk.Combobox.__init__(self,parent,
                              font=font,
                              state="readonly",
                              textvariable=self.displayed)
        if width_:
            self.config(width=width_)
            
    def get_my_value(self,event=None):
        return self.displayed.get()
    
    def set_my_value(self,value,event=None):
        self.displayed.set(value)


class license_key_entry_popup(tk.Toplevel):
    def __init__(self, C, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False, False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title(f"Enter license key - Contact {contact_email} for more information")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(2,weight=1)

        self.label2 = label(self, text = "License key:", font = EF, theme = theme)
        self.label2.config(anchor = "e")
        self.label2.grid(row=1,column=0,sticky="nswe",pady=(0,17),padx=(20,10))
        
        self.display2 = entry_with_scrollbar(self, theme = theme)
        self.display2.set_my_value("")
        self.display2.grid(row=1,column=1,sticky="nswe",pady=(18,20),padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 2, column = 0, columnspan = 2, sticky = "nswe")
        self.confirm_button = button(self.bf,text="Confirm",style="EF.Std.TButton",
                                     command = self.confirm)
        self.confirm_button.grid(row=0, column = 0, sticky = "nswe",
                                 padx = 60,
                                 pady = 20)
        self.cancel_button = button(self.bf, text = "Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe",
                                padx = 60,
                                pady = 20)
        self.bind("<Return>",self.confirm)
        self.display2.bind("<Return>", self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.display2.place_cursor()
        center(self,670,141)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        cset = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789")
        self.license_key = "".join([c for c in "".join(self.display2.get_my_value().upper().split("-")) if c in cset])
        self.result = True
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()
        

class save_new_version_presave_popup(tk.Toplevel):
    def __init__(self, C, file_location, theme = "dark"):
        tk.Toplevel.__init__(self, C, width = "1", height = "1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm", "iconphoto", self._w,tk.PhotoImage(format = "gif", data = top_left_icon))
        self.title("Save new version - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.file_loc_label = label(self, text = "Your new version\nwill be saved here:", font = EF, theme = theme)
        self.file_loc_label.grid(row=0,column=0,sticky="nswe",padx=(20,10))
        self.file_loc_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.file_loc_display.set_my_value(file_location)
        self.file_loc_display.grid(row=0,column=1,sticky="nswe",pady=(20,20),padx=(0,20))
        self.choose_loc_button = button(self,text="Choose where to save a new version",style="EF.Std.TButton",command=self.choose_loc)
        self.choose_loc_button.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(0,10))
        self.confirm_button = button(self,text="Confirm save a new version here",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=2,column=0,columnspan=2,sticky="nswe",padx=20,pady=(5,20))
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.file_loc_display.place_cursor()
        center(self,550,170)
        self.deiconify()
        self.wait_window()
        
    def choose_loc(self,event=None):
        folder = os.path.normpath(filedialog.askdirectory(parent=self,title="Select a folder to save new version in"))
        if folder == ".":
            return
        self.file_loc_display.set_my_value(folder)
        
    def confirm(self,event=None):
        self.result = os.path.normpath(self.file_loc_display.get_my_value())
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class save_new_version_postsave_popup(tk.Toplevel):
    def __init__(self, C, file_location, filename, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1",background = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Success! New version saved - Click the Okay/X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.file_loc_label = label(self,text="Your new version\nwas saved here:",font=EF, theme = theme)
        self.file_loc_label.grid(row=0,column=0,sticky="nswe",padx=(20,10))
        self.file_loc_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.file_loc_display.set_my_value(file_location)
        self.file_loc_display.grid(row=0,column=1,sticky="nswe",pady=(20,20),padx=(0,20))
        self.file_name_label = label(self,text="This is the\n filename:",font=EF, theme = theme)
        self.file_name_label.grid(row=1,column=0,sticky="nswe",padx=(20,10))
        self.file_name_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.file_name_display.set_my_value(filename)
        self.file_name_display.grid(row=1,column=1,sticky="nswe",pady=(0,20),padx=(0,20))
        self.confirm_button = button(self,text="Okay",style="EF.Std.TButton",command=self.cancel)
        self.confirm_button.grid(row=2,column=0,columnspan=2,sticky="nswe",padx=20,pady=(0,20))
        self.bind("<Return>",self.cancel)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.file_name_display.place_cursor()
        center(self, 550, 185)
        self.deiconify()
        self.wait_window()
        
    def cancel(self,event=None):
        self.destroy()


class save_new_version_error_popup(tk.Toplevel):
    def __init__(self, C, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1",background= theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Error locating folder - Click the Okay/X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(0,weight=1)
        self.info_label = label(self,text="There was an error locating the saving folder. \n - Would you like to choose where to save a new version?",font=EF, theme = theme)
        self.info_label.grid(row=0,column=0,sticky="nswe",padx=20,pady=20)
        self.confirm_button = button(self,text="Choose where to save a new version",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(0,20))
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        center(self, 550, 130)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = True
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class sort_sheet_popup(tk.Toplevel):
    def __init__(self,C,headers,theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1",background=theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Sort sheet - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_columnconfigure(0,weight=1)
        self.sort_decision = {'type': None,
                              'col': None,
                              'order': None}
        self.sort_by_col_button = button(self,style="EF.Std.TButton",
                                         text="Sort by column",
                                         command=self.sort_by_col)
        self.sort_icon = tk.PhotoImage(format="gif",data=sort_icon)
        self.sort_by_col_button.config(image=self.sort_icon,compound="left")
        self.sort_by_col_button.grid(row=0,column=1, sticky="nswe",pady=(15,5),padx=70)
        self.order_label = label(self,text="Order:",font=EF, theme = theme)
        self.order_label.grid(row=1,column=0,sticky="nswe",pady=5,padx=(50,10))
        self.order_dropdown = ez_dropdown(self,EF)
        self.order_dropdown['values'] = ["ASCENDING","DESCENDING"]
        self.order_dropdown.set_my_value("ASCENDING")
        self.order_dropdown.grid(row=1,column=1,sticky="nswe",pady=5,padx=(0,70))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",pady=5,padx=(50,10))
        self.col_dropdown = ez_dropdown(self,EF)
        self.col_dropdown['values'] = headers
        self.col_dropdown.set_my_value(headers[0])
        self.col_dropdown.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,70))
        self.divider = frame(self)
        self.divider.config(bg = theme_fg(theme))
        self.divider.config(height = 5)
        self.divider.grid(row = 3, column = 0, columnspan = 2, padx = 20, pady = (10, 15), sticky = "ew")
        self.sort_by_tree_button = button(self,style="EF.Std.TButton",
                                          text="Sort by tree walk",
                                          command=self.sort_by_tree)
        self.tree_icon = tk.PhotoImage(format="gif",data=tree_icon)
        self.sort_by_tree_button.config(image=self.tree_icon,compound="left")
        self.sort_by_tree_button.grid(row=4,column=1, sticky="nswe",pady=(20,20),padx=70)
        self.bind("<Escape>",self.go_back)
        self.order_dropdown.bind("<<ComboboxSelected>>",lambda event: self.focus_set())
        self.col_dropdown.bind("<<ComboboxSelected>>",lambda event: self.focus_set())
        center(self, 550, 215)
        self.deiconify()
        self.wait_window()
        
    def sort_by_col(self,event=None):
        self.sort_decision = {'type': "by column",
                              'col': self.col_dropdown.get_my_value(),
                              'order': self.order_dropdown.get_my_value()}
        self.destroy()
        
    def sort_by_tree(self,event=None):
        self.sort_decision['type'] = "by tree"
        self.destroy()
        
    def go_back(self,event=None):
        self.destroy()


class numerical_entry_with_scrollbar(tk.Frame):
    def __init__(self, parent, theme = "dark"):
        tk.Frame.__init__(self, parent, background = theme_bg(theme))
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.my_entry = numerical_normal_entry(self, font = EF, theme = theme)
        self.my_entry.grid(row=0,column=0,sticky="nswe")
        self.my_scrollbar = scrollbar(self,self.my_entry.xview,
                                      "horizontal",
                                      self.my_entry)
        self.my_scrollbar.grid(row=1,column=0,sticky="ew")
        
    def change_my_state(self,state,event=None):
        self.my_entry.config(state=state)
        
    def place_cursor(self,event=None):
        self.my_entry.focus_set()
        
    def get_my_value(self,event=None):
        return self.my_entry.get()
    
    def set_my_value(self,val,event=None):
        self.my_entry.set_my_value(val)


class numerical_normal_entry(tk.Entry):
    def __init__(self, parent, font, width_ = None, theme = "dark"):
        tk.Entry.__init__(self,
                          parent,
                          font = font,
                          background = theme_entry_bg(theme),
                          foreground = theme_entry_fg(theme),
                          disabledbackground = theme_entry_dbg(theme),
                          disabledforeground = theme_entry_dfg(theme),
                          insertbackground = theme_entry_cursor(theme),
                          readonlybackground = theme_entry_dbg(theme))
        if width_:
            self.config(width=width_)
        self.allowed_chars = {"0","1","2","3","4","5","6","7","8","9","."}
        self.sv = tk.StringVar()
        self.config(textvariable=self.sv)
        self.sv.trace("w", lambda name, index, mode, sv=self.sv: self.validate_(self.sv))
        self.rc_popup_menu = tk.Menu(self,tearoff=0,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Select all",accelerator="Ctrl+A",command=self.select_all,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Cut",accelerator="Ctrl+X",command=self.cut,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Copy",accelerator="Ctrl+C",command=self.copy,**menu_kwargs)
        self.rc_popup_menu.add_command(label="Paste",accelerator="Ctrl+V",command=self.paste,**menu_kwargs)
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def validate_(self,sv):
        x = self.sv.get()
        dotidx = [i for i,c in enumerate(x) if c == "."]
        if dotidx:
            if len(dotidx) > 1:
                x = x[:dotidx[1]]
        if x.startswith("."):
            x = x[1:]
        if x.startswith("-"):
            self.sv.set("-" + "".join([c for c in x if c in self.allowed_chars]))
        else:
            self.sv.set("".join([c for c in x if c in self.allowed_chars]))
            
    def rc(self,event):
        self.focus_set()
        self.rc_popup_menu.tk_popup(event.x_root,event.y_root)
        
    def select_all(self,event=None):
        self.event_generate("<Control-a>")
        return "break"
    
    def cut(self,event=None):
        self.event_generate("<Control-x>")
        return "break"
    
    def copy(self,event=None):
        self.event_generate("<Control-c>")
        return "break"
    
    def paste(self,event=None):
        self.event_generate("<Control-v>")
        return "break"
    
    def set_my_value(self,newvalue):
        self.delete(0,"end")
        self.insert(0,str(newvalue))
        
    def enable_me(self):
        self.config(state="normal")
        self.bind("<1>",lambda event: self.focus_set())
        self.bind(get_platform_rc_binding(), self.rc)
        
    def disable_me(self):
        self.config(state="disabled")
        self.unbind("<1>")
        self.unbind(get_platform_rc_binding())


class date_entry(tk.Frame):
    def __init__(self,parent,DATE_FORM, theme = "dark"):
        tk.Frame.__init__(self,
                          parent,
                          relief="flat",
                          bg = theme_bg(theme),
                          highlightbackground = theme_fg(theme),
                          highlightthickness = 2,
                          border=2)
        self.C = parent

        self.allowed_chars = {"0","1","2","3","4","5","6","7","8","9"}
        self.sv_1 = tk.StringVar()
        self.sv_2 = tk.StringVar()
        self.sv_3 = tk.StringVar()
        self.entry_1 = normal_entry(self,font=("Calibri",30,"bold"),width_=4,relief="flat",border=0,
                                    textvariable=self.sv_1, theme = theme)
        self.sep = "/" if "/" in DATE_FORM else "-"
        self.label_1 = tk.Label(self,font=("Calibri",30,"bold"),text=self.sep,
                                background=theme_entry_bg(theme),relief="flat")
        
        self.entry_2 = normal_entry(self,font=("Calibri",30,"bold"),width_=2,relief="flat",border=0,
                                    textvariable=self.sv_2, theme = theme)
        
        self.label_2 = tk.Label(self,font=("Calibri",30,"bold"),text=self.sep,
                                background=theme_entry_bg(theme),relief="flat")
        
        self.entry_3 = normal_entry(self,font=("Calibri",30,"bold"),width_=2,relief="flat",border=0,
                                    textvariable=self.sv_3, theme = theme)
        
        self.sv_1.trace("w", lambda name, index, mode, sv=self.sv_1: self.validate_1(self.sv_1))
        self.sv_2.trace("w", lambda name, index, mode, sv=self.sv_2: self.validate_2(self.sv_2))
        self.sv_3.trace("w", lambda name, index, mode, sv=self.sv_3: self.validate_3(self.sv_3))
        self.entry_1.bind("<BackSpace>",self.e1_back)
        self.entry_2.bind("<BackSpace>",self.e2_back)
        self.entry_3.bind("<BackSpace>",self.e3_back)

        if DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
            self.entry_3.pack(side="left")
            self.label_2.pack(side="left")
            self.entry_2.pack(side="left")
            self.label_1.pack(side="left")
            self.entry_1.pack(side="left")
            self.entries = [self.entry_3,self.entry_2,self.entry_1]
        elif DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
            self.entry_1.pack(side="left")
            self.label_1.pack(side="left")
            self.entry_2.pack(side="left")
            self.label_2.pack(side="left")
            self.entry_3.pack(side="left")
            self.entries = [self.entry_1,self.entry_2,self.entry_3]
        elif DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
            self.entry_2.pack(side="left")
            self.label_2.pack(side="left")
            self.entry_3.pack(side="left")
            self.label_1.pack(side="left")
            self.entry_1.pack(side="left")
            self.entries = [self.entry_2,self.entry_3,self.entry_1]

        self.DATE_FORM = DATE_FORM
        
    def e1_back(self,event):
        x = self.sv_1.get()
        if self.DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
            if not x:
                self.entry_2.icursor(2)
                self.entry_2.focus_set()
                return "break"
            if len(x) == 1:
                x = ""
            else:
                x = x[:-1]
            self.sv_1.set(x)
        elif self.DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
            if len(x) == 1:
                x = ""
            else:
                x = x[:-1]
            self.sv_1.set(x)
        elif self.DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
            if not x:
                self.entry_3.icursor(2)
                self.entry_3.focus_set()
                return "break"
            if len(x) == 1:
                x = ""
            else:
                x = x[:-1]
            self.sv_1.set(x)
        return "break"

    def e2_back(self,event):
        x = self.sv_2.get()
        if self.DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
            if not x:
                self.entry_3.icursor(2)
                self.entry_3.focus_set()
                return "break"
            elif len(x) == 1:
                x = ""
            else:
                x = x[:-1]
            self.sv_2.set(x)
        elif self.DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
            if not x:
                self.entry_1.icursor(4)
                self.entry_1.focus_set()
                return "break"
            elif len(x) == 1:
                x = ""
            else:
                x = x[:-1]
            self.sv_2.set(x)
        elif self.DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
            if len(x) == 1:
                x = ""
            else:
                x = x[:-1]
            self.sv_2.set(x)
        return "break"
            
    def e3_back(self,event):
        x = self.sv_3.get()
        if self.DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
            if len(x) == 1:
                x = ""
            else:
                x = x[:-1]
            self.sv_3.set(x)
        elif self.DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
            if not x:
                self.entry_2.icursor(2)
                self.entry_2.focus_set()
                return "break"
            if len(x) == 1:
                x = ""
            else:
                x = x[:-1]
            self.sv_3.set(x)
        elif self.DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
            if not x:
                self.entry_2.icursor(2)
                self.entry_2.focus_set()
                return "break"
            if len(x) == 1:
                x = ""
            else:
                x = x[:-1]
            self.sv_3.set(x)
        return "break"

    def validate_1(self,sv):
        year = []
        for i,c in enumerate(self.sv_1.get()):
            if c in self.allowed_chars:
                year.append(c)
            if i > 3:
                break
        year = "".join(year)
        if len(year) > 4:
            year = year[:4]
        self.entry_1.set_my_value(year)
        if len(year) == 4:
            if self.DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
                pass
            if self.DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
                self.entry_2.focus_set()
            elif self.DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
                pass
        
    def validate_2(self,sv):
        month = []
        for i,c in enumerate(self.sv_2.get()):
            if c in self.allowed_chars:
                month.append(c)
            if i > 1:
                break
        move_cursor = False
        if len(month) > 2:
            month = month[:2]
        if not month:
            self.entry_2.set_my_value("")
            return
        if len(month) == 1 and int(month[0]) > 1:
            month = ["0",month[0]]
            move_cursor = True
        elif len(month) == 1 and int(month[0]) <= 1:
            return
        e0 = int(month[0])
        e1 = int(month[1])
        if e0 > 1:
            month[0] = "1"
            int(month[0])
        if e1 > 2 and e0 > 0:
            month[0] = "0"
            month[1] = str(e1)
        self.entry_2.set_my_value("".join(month))
        if move_cursor:
            self.entry_2.icursor(2)
        if len(month) == 2:
            if self.DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
                self.entry_1.focus_set()
            if self.DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
                self.entry_3.focus_set()
            elif self.DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
                self.entry_3.focus_set()
            
    def validate_3(self,sv):
        day = []
        for i,c in enumerate(self.sv_3.get()):
            if c in self.allowed_chars:
                day.append(c)
            if i > 1:
                break
        move_cursor = False
        if len(day) > 2:
            day = day[:2]
        if not day:
            self.entry_3.set_my_value("")
            return
        if len(day) == 1 and int(day[0]) > 3:
            day = ["0",day[0]]
            move_cursor = True
        elif len(day) == 1 and int(day[0]) <= 3:
            return
        e0 = int(day[0])
        e1 = int(day[1])
        if e0 > 3:
            day[0] = "1"
            int(day[0])
        if e0 >= 3 and e1 > 1:
            day[0] = "0"
            day[1] = str(e1)
        self.entry_3.set_my_value("".join(day))
        if move_cursor:
            self.entry_3.icursor(2)
        if len(day) == 2:
            if self.DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
                self.entry_2.focus_set()
            if self.DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
                pass
            elif self.DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
                self.entry_1.focus_set()
        
    def get_my_value(self):
        return self.sep.join([e.get() for e in self.entries])

    def set_my_value(self,date):
        date = re.split('|'.join(map(re.escape, ("/", "-"))), date)
        if self.DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
            try:
                self.sv_1.set(date[2])
            except:
                pass
            try:
                self.sv_2.set(date[1])
            except:
                pass
            try:
                self.sv_3.set(date[0])
            except:
                pass
        elif self.DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                self.sv_1.set(date[0])
            except:
                pass
            try:
                self.sv_2.set(date[1])
            except:
                pass
            try:
                self.sv_3.set(date[2])
            except:
                pass
        elif self.DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
            try:
                self.sv_1.set(date[2])
            except:
                pass
            try:
                self.sv_2.set(date[0])
            except:
                pass
            try:
                self.sv_3.set(date[1])
            except:
                pass
        
    def place_cursor(self,index=0):
        self.entries[index].focus()


class edit_detail_date_popup(tk.Toplevel):
    def __init__(self,C,ID,column,current_detail,DATE_FORM,validation_values=[],set_value=None, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1")
        self.withdraw()
        self.resizable(False, False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Change date detail - Click the X button or press escape to go back")
        self.C = C
        self.config(bg = theme_bg(theme))
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_label = label(self,text="ID:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",padx=20)
        self.col_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe")
        
        if validation_values:
            self.validation_dropdown = ez_dropdown(self,font=EF)
            self.validation_dropdown['values'] = validation_values
            if set_value is not None:
                self.validation_dropdown.set_my_value(set_value)
            else:
                self.validation_dropdown.set_my_value(validation_values[0])
            self.validation_dropdown.grid(row=3,column=0,columnspan=2,sticky="nswe",padx=20,pady=10)
            self.validation_dropdown.bind("<<ComboboxSelected>>", lambda focus: self.focus_set())
            width_ = 600
            height_ = 225
            self.bind("<Return>",self.confirm_validation)
        else:
            self.entries_frame = frame(self, theme = theme)
            self.entries_frame.grid_columnconfigure(3,weight=1)
            self.entries_frame.grid(row=3,column=0,columnspan=2,sticky="nswe",pady=10)
            if DATE_FORM in ("%d/%m/%Y", "%d-%m-%Y"):
                self.date_label = label(self.entries_frame,text="Set date DD/MM/YYYY:",font=EF, theme = theme)
            elif DATE_FORM in ("%Y/%m/%d", "%Y-%m-%d"):
                self.date_label = label(self.entries_frame,text="Set date YYYY/MM/DD:",font=EF, theme = theme)
            elif DATE_FORM in ("%m/%d/%Y", "%m-%d-%Y"):
                self.date_label = label(self.entries_frame,text="Set date MM/DD/YYYY:",font=EF, theme = theme)
            self.date_label.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=10)
            self.date_entry_widget = date_entry(self.entries_frame,DATE_FORM, theme = theme)
            self.date_entry_widget.grid(row=0,column=1,sticky="nswe",padx=(0,30),pady=10)
            self.numerical_label = label(self.entries_frame,text="OR set Number:",font=EF, theme = theme)
            self.numerical_label.grid(row=0,column=2,sticky="nswe",padx=(0,10),pady=10)
            self.numerical_entry_widget = numerical_entry_with_scrollbar(self.entries_frame, theme = theme)
            self.numerical_entry_widget.grid(row=0,column=3,sticky="nswe",padx=(0,20),pady=15)
            if "/" in current_detail or "-" in current_detail:
                self.date_entry_widget.set_my_value(current_detail)
            else:
                self.numerical_entry_widget.set_my_value(current_detail)
            self.numerical_entry_widget.my_entry.bind("<Return>",self.confirm_normal)
            self.date_entry_widget.entry_1.bind("<Return>",self.confirm_normal)
            self.date_entry_widget.entry_2.bind("<Return>",self.confirm_normal)
            self.date_entry_widget.entry_3.bind("<Return>",self.confirm_normal)
            self.date_entry_widget.place_cursor()
            width_ = 850
            height_ = 280

        self.confirm_button = button(self.bf,text="Save",style="EF.Std.TButton",
                                     command=self.confirm_validation if validation_values else self.confirm_normal)
        self.confirm_button.grid(row=0, column = 0, sticky = "nswe",
                                 padx = 70 if validation_values else 100,
                                 pady = 20)
        self.cancel_button = button(self.bf, text = "Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe",
                                padx = 70 if validation_values else 100,
                                pady = 20)
        
        self.result = False
        center(self,width_,height_)
        self.deiconify()
        self.bind("<Escape>",self.cancel)
        self.wait_window()
        
    def confirm_normal(self,event=None):
        self.result = True
        x1 = self.date_entry_widget.get_my_value()
        x2 = self.numerical_entry_widget.get_my_value()
        if not all(c in ("/", "-") for c in x1):
            self.saved_string = x1
        elif x2:
            self.saved_string = x2
        else:
            self.saved_string = ""
        self.destroy()
        
    def confirm_validation(self,event=None):
        self.result = True
        self.saved_string = self.validation_dropdown.get_my_value()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class edit_detail_numerical_popup(tk.Toplevel):
    def __init__(self,C,ID,column,current_detail,validation_values=[],set_value=None, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Change numerical detail - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_label = label(self,text="ID:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",padx=20)
        self.col_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe")

        if validation_values:
            self.validation_dropdown = ez_dropdown(self,font=EF)
            self.validation_dropdown['values'] = validation_values
            if set_value is not None:
                self.validation_dropdown.set_my_value(set_value)
            else:
                self.validation_dropdown.set_my_value(validation_values[0])
            self.validation_dropdown.grid(row=3,column=0,columnspan=2,sticky="nswe",padx=20,pady=10)
            self.validation_dropdown.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
            width_ = 600
            height_ = 225
            self.bind("<Return>",self.confirm_validation)
        else:
            self.entry_widget = numerical_entry_with_scrollbar(self, theme = theme)
            self.entry_widget.set_my_value(current_detail)
            self.entry_widget.grid(row=3,column=0,columnspan=2,sticky="nswe",padx=20,pady=10)
            self.entry_widget.my_entry.bind("<Return>",self.confirm_normal)
            width_ = 600
            height_ = 240
            self.entry_widget.place_cursor()
            
        self.confirm_button = button(self.bf,text="Save",style="EF.Std.TButton",
                                     command=self.confirm_validation if validation_values else self.confirm_normal)
        self.confirm_button.grid(row=0, column = 0, sticky = "nswe", padx = 70, pady = 20)
        self.cancel_button = button(self.bf, text = "Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = 70, pady = 20)
        
        self.result = False
        center(self,width_,height_)
        self.deiconify()
        self.bind("<Escape>",self.cancel)
        self.wait_window()
        
    def confirm_normal(self,event=None):
        self.result = True
        self.saved_string = self.entry_widget.get_my_value()
        self.destroy()
        
    def confirm_validation(self,event=None):
        self.result = True
        self.saved_string = self.validation_dropdown.get_my_value()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()
        

class edit_detail_text_popup(tk.Toplevel):
    def __init__(self,C,ID,column,current_detail,validation_values=[],set_value=None, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1")
        self.withdraw()
        self.resizable(False, False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Edit cell - Click the X button or press escape to go back")
        self.C = C
        self.config(bg = theme_bg(theme))
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_label = label(self,text="ID:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",padx=20)
        self.col_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe")
        
        if validation_values:
            self.validation_dropdown = ez_dropdown(self,font=EF)
            self.validation_dropdown['values'] = validation_values
            if set_value is not None:
                self.validation_dropdown.set_my_value(set_value)
            else:
                self.validation_dropdown.set_my_value(validation_values[0])
            self.validation_dropdown.grid(row=3,column=0,columnspan=2,sticky="nswe",padx=20,pady=10)
            self.validation_dropdown.bind("<<ComboboxSelected>>",lambda focus: self.focus_set())
            width_ = 620
            height_ = 225
            self.confirm_button = button(self.bf,text="Save",style="EF.Std.TButton",command=self.confirm_validation)
            self.bind("<Return>",self.confirm_validation)
        else:
            self.grid_rowconfigure(3,weight=1)
            self.text_widget = wrapped_text_with_find_and_yscroll(self,current_detail,"normal",15, theme = theme)
            self.text_widget.grid(row=3,column=0,sticky="nswe",columnspan=2)
            self.text_widget.place_cursor()
            width_ = 800
            height_ = 595
            self.confirm_button = button(self.bf,text="Save",style="EF.Std.TButton",command=self.confirm_normal)
        self.confirm_button.grid(row=0,column=0,sticky="nswe", padx = 75, pady = 20)
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = 75, pady = 20)
        center(self,width_,height_)
        self.result = False
        self.deiconify()
        self.bind("<Escape>",self.cancel)
        self.wait_window()
        
    def confirm_normal(self,event=None):
        self.result = True
        self.saved_string = self.text_widget.get_my_value().rstrip()
        self.destroy()
        
    def confirm_validation(self,event=None):
        self.result = True
        self.saved_string = self.validation_dropdown.get_my_value()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class view_column_text_popup(tk.Toplevel):
    def __init__(self,C,ID,column,text, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1")
        self.withdraw()
        self.resizable(False, False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("View text - Click the X button or press escape to go back")
        self.C = C
        self.config(bg = theme_bg(theme))
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_label = label(self,text="ID:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.col_label = label(self,text="Column:",font=EF, theme = theme)
        self.col_label.grid(row=2,column=0,sticky="nswe",padx=20)
        self.col_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.col_display.set_my_value(column)
        self.col_display.grid(row=2,column=1,sticky="nswe",pady=5,padx=(0,20))
        self.text_widget = wrapped_text_with_find_and_yscroll(self,text,"disabled",15, theme = theme)
        self.text_widget.grid(row=3,column=0,sticky="nswe",columnspan=2)
        self.cancel_button = button(self, text = "Close", style = "EF.Std.TButton", command = self.cancel)
        self.cancel_button.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe", padx = 220, pady = (25, 20))
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.text_widget.place_cursor()
        center(self,850,545)
        self.deiconify()
        self.wait_window()
        
    def cancel(self,event=None):
        self.destroy()


class add_top_id_popup(tk.Toplevel):
    def __init__(self, C, ss_selection, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Add top ID - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.id_name_label = label(self,text="ID name:",font=EF, theme = theme)
        self.id_name_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_name_display = entry_with_scrollbar(self, theme = theme)
        self.id_name_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        if self.C.tv_label_col != self.C.ic:
            self.id_tv_label = label(self,text="ID Treeview Label:",font=EF, theme = theme)
            self.id_tv_label.grid(row=1,column=0,sticky="nswe",padx=20)
            self.id_tv_display = entry_with_scrollbar(self, theme = theme)
            self.id_tv_display.grid(row=1,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.enter_ss_sel_button = button(self,text="Enter current sheet selection",style="EF.Std.TButton",command=self.enter_ss_sel)
        self.enter_ss_sel_button.grid(row=2,column=1,sticky="nswe",padx= (0, 20), pady=(10,5))
        if not ss_selection:
            self.enter_ss_sel_button.config(state="disabled")
        else:
            self.ss_sel = ss_selection
        
        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 3, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Add",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (75, 50), pady = (30, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 75), pady = (30, 20))
        
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.id_name_display.place_cursor()
        center(self,600,186)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = "".join(self.id_name_display.get_my_value().strip().split())
        try:
            self.id_label = self.id_tv_display.get_my_value().strip().split("\n")[0]
        except:
            pass
        self.destroy()
        
    def enter_ss_sel(self,event=None):
        self.id_name_display.set_my_value(self.ss_sel)
        if self.C.tv_label_col != self.C.ic:
            detail = self.C.sheet[self.C.rns[self.ss_sel.lower()]][self.C.tv_label_col]
            ni = detail.find("\n")
            if ni == -1:
                self.id_tv_display.set_my_value(detail)
            else:
                self.id_tv_display.set_my_value(detail[:ni])
                
    def cancel(self,event=None):
        self.destroy()


class add_child_or_sibling_id_popup(tk.Toplevel):
    def __init__(self, C, chld_or_sib, desired_parent, ss_selection, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", background = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        if chld_or_sib == "sibling":
            self.title("Add sibling ID - Click the X button or press escape to go back")
        elif chld_or_sib == "child":
            self.title("Add child ID - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.parent_label = label(self,text="Desired\nparent:",font=EF, theme = theme)
        self.parent_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.parent_display = readonly_entry_with_scrollbar(self, theme = theme)
        if desired_parent:
            self.parent_display.set_my_value(desired_parent)
        else:
            self.parent_display.set_my_value("** No parent - Top ID **")
        self.parent_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.id_name_label = label(self,text="ID name:",font=EF, theme = theme)
        self.id_name_label.grid(row=1,column=0,sticky="nswe",padx=20)
        self.id_name_display = entry_with_scrollbar(self, theme = theme)
        self.id_name_display.grid(row=1,column=1,sticky="nswe",pady=(5,10),padx=(0,20))
        if self.C.tv_label_col != self.C.ic:
            self.id_tv_label = label(self,text="ID Treeview Label:",font=EF, theme = theme)
            self.id_tv_label.grid(row=2,column=0,sticky="nswe",padx=20)
            self.id_tv_display = entry_with_scrollbar(self, theme = theme)
            self.id_tv_display.grid(row=2,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.enter_ss_sel_button = button(self,text="Enter current sheet selection",style="EF.Std.TButton",command=self.enter_ss_sel)
        self.enter_ss_sel_button.grid(row=3,column=1,sticky="nswe", padx= (0, 20), pady=(10,5))
        if not ss_selection:
            self.enter_ss_sel_button.config(state="disabled")
        else:
            self.ss_sel = ss_selection
        
        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 4, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Add",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (75, 50), pady = (30, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 75), pady = (30, 20))
        
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.id_name_display.place_cursor()
        center(self,600,237)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = "".join(self.id_name_display.get_my_value().strip().split())
        try:
            self.id_label = self.id_tv_display.get_my_value().strip().split("\n")[0]
        except:
            pass
        self.destroy()
        
    def enter_ss_sel(self,event=None):
        self.id_name_display.set_my_value(self.ss_sel)
        if self.C.tv_label_col != self.C.ic:
            detail = self.C.sheet[self.C.rns[self.ss_sel.lower()]][self.C.tv_label_col]
            ni = detail.find("\n")
            if ni == -1:
                self.id_tv_display.set_my_value(detail)
            else:
                self.id_tv_display.set_my_value(detail[:ni])
                
    def cancel(self,event=None):
        self.destroy()


class rename_id_popup(tk.Toplevel):
    def __init__(self, C, ID, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Rename ID - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(2,weight=1)
        self.id_label = label(self,text="Current ID\nname:",font=EF, theme = theme)
        self.id_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.id_display = readonly_entry_with_scrollbar(self, theme = theme)
        self.id_display.set_my_value(ID)
        self.id_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.new_name_label = label(self,text="New ID\nname:",font=EF, theme = theme)
        self.new_name_label.grid(row=1,column=0,sticky="nswe",padx=20)
        self.new_name_display = entry_with_scrollbar(self, theme = theme)
        self.new_name_display.grid(row=1,column=1,sticky="nswe",pady=5,padx=(0,20))

        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 2, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Rename",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (75, 50), pady = (20, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 75), pady = (20, 20))
        
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.new_name_display.place_cursor()
        center(self,600,185)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = "".join(self.new_name_display.get_my_value().strip().split())
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class rename_column_frame(tk.Frame):
    def __init__(self, C, current_col_name, type_of_col, theme = "dark"):
        tk.Frame.__init__(self, C, height = 190, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.C = C
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(2,weight=1)
        self.col_label = label(self,text="Current column\nname:",font=EF, theme = theme)
        self.col_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.col_display = readonly_entry_with_scrollbar(self, theme = theme)
        if type_of_col == "detail":
            self.col_display.set_my_value("Detail column: " + current_col_name)
        elif type_of_col == "hierarchy":
            self.col_display.set_my_value("Parent column: " + current_col_name)
        elif type_of_col == "ID":
            self.col_display.set_my_value("ID column: " + current_col_name)
        self.col_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.new_name_label = label(self,text="New column\nname:",font=EF, theme = theme)
        self.new_name_label.grid(row=1,column=0,sticky="nswe",padx=20)
        self.new_name_display = entry_with_scrollbar(self, theme = theme)
        self.new_name_display.grid(row=1,column=1,sticky="nswe",pady=5,padx=(0,20))
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text="Confirm",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(10,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(10,20))
        self.result = False
        self.new_name_display.place_cursor()
        
    def confirm(self,event=None):
        self.result = "".join(self.new_name_display.get_my_value().strip().split())
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class add_hierarchy_frame(tk.Frame):
    def __init__(self, C, theme = "dark"):
        tk.Frame.__init__(self, C, height = 150, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.C = C
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.hier_name_label = label(self,text="New hierarchy\nname:",font=EF, theme = theme)
        self.hier_name_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.hier_name_display = entry_with_scrollbar(self, theme = theme)
        self.hier_name_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=1,column=0,columnspan=2,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text="Add Hierarchy Column",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(10,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(10,20))
        self.result = False
        self.hier_name_display.place_cursor()
        
    def confirm(self,event=None):
        self.result = "".join(self.hier_name_display.get_my_value().strip().split())
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()
        

class add_detail_column_frame(tk.Frame):
    def __init__(self,C, theme = "dark"):
        tk.Frame.__init__(self, C, height = 150, bg = theme_bg(theme))
        self.grid_propagate(False)
        self.grid_columnconfigure(2,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.type_display = ez_dropdown(self,EF)
        self.type_display['values'] = ("Text Detail","Numerical Detail","Date Detail")
        self.type_display.set_my_value("Text Detail")
        self.type_display.grid(row=0,column=0,sticky="nswe",padx=(20,0),pady=(20,5))
        self.type_display.bind("<<ComboboxSelected>>",lambda focus: self.detail_name_display.place_cursor())
        self.detail_name_label = label(self,text="New detail\ncolumn name:",font=EF, theme = theme)
        self.detail_name_label.grid(row=0,column=1,sticky="nswe",padx=20,pady=(20,5))
        self.detail_name_display = entry_with_scrollbar(self, theme = theme)
        self.detail_name_display.grid(row=0,column=2,sticky="nswe",pady=(20,5),padx=(0,20))
        self.button_frame = frame(self, theme = theme)
        self.button_frame.grid(row=1,column=0,columnspan=3,sticky="nswe")
        self.button_frame.grid_columnconfigure(0,weight=1,uniform="x")
        self.button_frame.grid_columnconfigure(1,weight=1,uniform="x")
        self.button_frame.grid_rowconfigure(0,weight=1)
        self.confirm_button = button(self.button_frame,text="Add Detail Column",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx=(20,10),pady=(10,20))
        self.cancel_button = button(self.button_frame,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row=0,column=1,sticky="nswe",padx=(10,20),pady=(10,20))
        self.result = False
        self.type_ = "Text Detail"
        self.formula = ""
        self.detail_name_display.place_cursor()
        
    def confirm(self,event=None):
        self.result = "".join(self.detail_name_display.get_my_value().strip().split())
        self.type_ = self.type_display.get()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class enter_sheet_name_popup(tk.Toplevel):
    def __init__(self, C, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Enter desired sheet name - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.sheet_name_label = label(self,text="New sheet\nname:",font=EF, theme = theme)
        self.sheet_name_label.grid(row=0,column=0,sticky="nswe",padx=20)
        self.sheet_entry = entry_with_scrollbar(self, theme = theme)
        self.sheet_entry.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        
        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 1, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Confirm",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (75, 50), pady = (20, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 75), pady = (20, 20))
        
        self.bind("<Return>",self.confirm)
        self.bind("<Escape>",self.cancel)
        self.result = False
        self.sheet_entry.place_cursor()
        center(self,600,137)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.result = self.sheet_entry.get_my_value()
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class error(tk.Toplevel):
    def __init__(self, C, msg, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("*** Error *** - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.errorlabel = label(self,text="Error\nmessage:",font=EF, theme = theme)
        self.errorlabel.config(background="red",foreground="white")
        self.errorlabel.grid(row=0,column=0,sticky="nswe",pady=(20,5),padx=20)
        self.error_display = display_text(parent = self, text = msg, theme = theme)
        self.error_display.grid(row=0,column=1,sticky="nswe",pady=(20,5),padx=(0,20))
        self.error_display.config(height = 75)
        self.confirm_button = button(self,text="Okay",style="EF.Std.TButton",command=self.cancel)
        self.confirm_button.grid(row=1,column=0,columnspan=2,sticky="nswe",padx=20,pady=(10,20))
        self.bind("<Return>",self.cancel)
        self.bind("<Escape>",self.cancel)
        center(self,600,180)
        self.deiconify()
        self.wait_window()
        
    def cancel(self,event=None):
        self.destroy()


class frame(tk.Frame):
    def __init__(self,parent,background="white",highlightbackground="white",highlightthickness=0,theme = "dark"):
        tk.Frame.__init__(self,parent,
                          background=theme_bg(theme),
                          highlightbackground=highlightbackground,
                          highlightthickness=highlightthickness)
        

class button(ttk.Button):
    def __init__(self,parent,style="Std.TButton",text="",command=None,state="normal",underline=-1):
        ttk.Button.__init__(self,
                            parent,
                            style=style,
                            text=text,
                            command=command,
                            state=state,
                            underline=underline)
    def change_text(self,text):
        self.config(text=text)
        self.update_idletasks()


class StatusBar(tk.Label):
    def __init__(self, parent, text, theme = "dark"):
        tk.Label.__init__(self,
                          parent,
                          text = text,
                          font = ("Calibri",11,"normal"),
                          background = theme_bg(theme),
                          foreground = theme_status_fg(theme),
                          anchor = "w")
        self.text = text
    def change_text(self,text=""):
        self.config(text=text)
        self.text = text
        self.update_idletasks()


class label(tk.Label):
    def __init__(self,parent,text,font,theme = "dark", anchor = "center"):
        tk.Label.__init__(self,parent,
                          text=text,font=font,
                          background=theme_bg(theme),
                          foreground=theme_fg(theme),
                          anchor = anchor)
    def change_text(self,text):
        self.config(text=text)
        self.update_idletasks()


class displabel(tk.Label):
    def __init__(self,parent,text,font, theme = "dark"):
        tk.Label.__init__(self,parent,
                          background=theme_bg(theme),
                          text=text,font=font)
        self.config(anchor="w")
    def change_text(self,text):
        self.config(text=text)
        self.update_idletasks()


class treeview_id_finder(tk.Toplevel):
    def __init__(self, C, hiers, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("ID is in multiple hierarchies, choose which hierarchy to go to - Click the X button or press escape to go back")
        self.C = C
        self.GO = False
        self.selected = 0
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1,weight=1)
        self.dd_1 = ez_dropdown(self,EF)
        self.dd_1['values'] = hiers
        self.dd_1.set_my_value(hiers[0])
        self.dd_1.grid(row=0,column=0,sticky="nswe",columnspan=2,padx=20,pady=(20,5))
        
        self.bf = frame(self, theme = theme)
        self.bf.grid_columnconfigure(0, weight = 1, uniform = "x")
        self.bf.grid_columnconfigure(1, weight = 1, uniform = "x")
        self.bf.grid(row = 1, column = 0, columnspan = 2, sticky = "nswe")
        
        self.confirm_button = button(self.bf,text="Go",style="EF.Std.TButton",command=self.confirm)
        self.confirm_button.grid(row=0,column=0,sticky="nswe",padx = (90, 50), pady = (20, 20))
        self.cancel_button = button(self.bf,text="Cancel",style="EF.Std.TButton",command=self.cancel)
        self.cancel_button.grid(row = 0, column = 1, sticky = "nswe", padx = (50, 90), pady = (20, 20))
        
        self.bind("<Escape>",self.cancel)
        center(self,700,120)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        self.selected = self.dd_1.displayed.get()
        self.GO = True
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class ss_settings_chooser(tk.Toplevel):
    def __init__(self, C, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.resizable(False,False)
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Choose sheet settings - Click the X button or press escape to go back")
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.changed = False
        self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1,weight=1)
        self.settings_frame = frame(self, theme = theme)
        self.settings_frame.grid(row=0,column=0,sticky="nswe",columnspan=2,padx=20,pady=(20,5))
        
        self.dd_2_label = label(self.settings_frame,"Main table align: ",BF, theme = theme)
        self.dd_2_label.grid(row=0,column=0,sticky="nswe")
        self.dd_2 = ez_dropdown(self.settings_frame,EF)
        self.dd_2['values'] = ("Left","Center")
        x = self.C.sheetdisplay.align()
        if x == "w":
            self.dd_2.set_my_value("Left")
        elif x == "center":
            self.dd_2.set_my_value("Center")
        self.dd_2.grid(row=0,column=1,sticky="nswe")
        
        self.dd_3_label = label(self.settings_frame,"Row index align: ",BF, theme = theme)
        self.dd_3_label.grid(row=1,column=0,sticky="nswe")
        self.dd_3 = ez_dropdown(self.settings_frame,EF)
        self.dd_3['values'] = ("Left","Center")
        x = self.C.sheetdisplay.row_index_align()
        if x == "w":
            self.dd_3.set_my_value("Left")
        elif x == "center":
            self.dd_3.set_my_value("Center")
        self.dd_3.grid(row=1,column=1,sticky="nswe")
        
        self.dd_4_label = label(self.settings_frame,"Headers align: ",BF, theme = theme)
        self.dd_4_label.grid(row=2,column=0,sticky="nswe")
        self.dd_4 = ez_dropdown(self.settings_frame,EF)
        self.dd_4['values'] = ("Left","Center")
        x = self.C.sheetdisplay.header_align()
        if x == "w":
            self.dd_4.set_my_value("Left")
        elif x == "center":
            self.dd_4.set_my_value("Center")
        self.dd_4.grid(row=2,column=1,sticky="nswe")

        self.dd_5_label = label(self.settings_frame,"Theme: ",BF, theme = theme)
        self.dd_5_label.grid(row=3,column=0,sticky="nswe")
        self.dd_5 = ez_dropdown(self.settings_frame,EF)
        self.dd_5['values'] = ("Light","Dark")
        if self.C.sheetdisplay.MT.table_background == "white":
            self.dd_5.set_my_value("Light")
        else:
            self.dd_5.set_my_value("Dark")
        self.dd_5.grid(row=3,column=1,sticky="nswe")
        
        self.confirm_button = button(self,text="Confirm",
                                     style="EF.Std.TButton",
                                     command=self.confirm)
        self.confirm_button.grid(row=1,column=0,sticky="nswe",padx=20,pady=(15,20))
        self.cancel_button = button(self,text="Cancel",
                                    style="EF.Std.TButton",
                                    command=self.cancel)
        self.cancel_button.grid(row=1,column=1,sticky="nswe",padx=20,pady=(15,20))
        self.bind("<Escape>",self.cancel)
        center(self,500,185)
        self.deiconify()
        self.wait_window()
        
    def confirm(self,event=None):
        x = self.dd_2.displayed.get()
        if x == "Left":
            self.C.sheetdisplay.align("w",redraw=False)
        elif x == "Center":
            self.C.sheetdisplay.align("center",redraw=False)

        x = self.dd_3.displayed.get()
        if x == "Left":
            self.C.sheetdisplay.row_index_align("w",redraw=False)
        elif x == "Center":
            self.C.sheetdisplay.row_index_align("center",redraw=False)

        x = self.dd_4.displayed.get()
        if x == "Left":
            self.C.sheetdisplay.header_align("w",redraw=False)
        elif x == "Center":
            self.C.sheetdisplay.header_align("center",redraw=False)
        self.C.change_theme(self.dd_5.displayed.get().lower())
        self.changed = True
        self.destroy()
        
    def cancel(self,event=None):
        self.destroy()


class textpopup(tk.Toplevel):
    def __init__(self,C,text,width_=700,height_=650, theme = "dark", use_entry_bg = False):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.theme = theme
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.word = ""
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        self.grid_columnconfigure(0,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.find_frame = frame(self, theme = theme)
        self.find_frame.grid(row=0,column=0,columnspan=2,sticky="nswe")
        self.find_icon = tk.PhotoImage(format="gif",data=find_icon)
        self.search_button = button(self.find_frame,
                                    text=" Find:",
                                    command=self.find)
        self.search_button.config(image=self.find_icon,compound="left")
        self.search_button.pack(side="left",fill="x")
        self.find_window = normal_entry(self.find_frame,font=BF, theme = theme)
        self.find_window.bind("<Return>",self.find)
        self.find_window.pack(side="left",fill="x",expand=True)
        self.find_reset_button = button(self.find_frame,text="X",command=self.find_reset)
        self.find_reset_button.pack(side="left",fill="x")
        self.find_results_label = label(self.find_frame,"0/0",BF, theme = theme)
        self.find_results_label.pack(side="left",fill="x")
        self.find_up_button = button(self.find_frame,text="▲",command=self.find_up)
        self.find_up_button.pack(side="left",fill="x")
        self.find_down_button = button(self.find_frame,text="▼",command=self.find_down)
        self.find_down_button.pack(side="left",fill="x")
        self.textbox = working_text(self,
                                    wrap="none",
                                    theme = theme,
                                    use_entry_bg = use_entry_bg,
                                    override_bg = theme_treeview_bg(theme) if theme != "dark" else None)
        self.yscrollb = scrollbar(self,
                                  self.textbox.yview,
                                  "vertical",
                                  self.textbox)
        self.xscrollb = scrollbar(self,
                                  self.textbox.xview,
                                  "horizontal",
                                  self.textbox)
        self.textbox.delete(1.0,"end")
        self.textbox.insert(1.0,text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=1,column=0,sticky="nswe")
        self.yscrollb.grid(row=1,column=1,sticky="nswe")
        self.xscrollb.grid(row=2,column=0,columnspan=2,sticky="nswe")
        self.buttonframe = frame(self, theme = theme)
        self.buttonframe.grid(row=3,column=0,columnspan=2,sticky="nswe")
        self.save_text_button = button(self.buttonframe,text="Save as",
                                       style="EF.Std.TButton",
                                       command=lambda: self.save_text(text))
        self.save_text_button.pack(side="left",fill="x",padx=(20,40),pady=20)
        self.save_text_button.config(width=24)
        self.cancel_button = button(self.buttonframe,text="Close popup",
                                    style="EF.Std.TButton",
                                    command=self.cancel)
        self.cancel_button.pack(side="right",fill="x",padx=(40,20),pady=20)
        self.cancel_button.config(width=24)
        self.bind("<Escape>",self.cancel)
        center(self,width_,height_)
        self.deiconify()
        self.wait_window()
        
    def save_text(self,text):
        newfile = filedialog.asksaveasfilename(parent=self,
                                               title="Save text on popup window",
                                               filetypes=[('Text File','.txt'),('CSV File','.csv')],
                                               defaultextension=".txt",
                                               confirmoverwrite=True)
        if not newfile:
            return
        newfile = os.path.normpath(newfile)
        if not newfile.lower().endswith((".csv",".txt")):
            errorpopup = error(self,"Can only save .csv/.txt files", theme = self.theme)
            self.grab_set()
            return
        try:
            with open(newfile,"w") as fh:
                for line in text:
                    fh.write(line)
        except:
            errorpopup = error(self,"Error saving file", theme = self.theme)
            self.grab_set()
            return
        
    def find(self,event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        self.addon = "+" + str(len(self.word)) + "c"
        start = "1.0"
        while start:
            start = self.textbox.search(self.word,index=start,nocase=1,stopindex="end")
            if start:
                end = start + self.addon
                self.find_results.append(start)
                self.textbox.tag_add("i",start,end)
                start = end
        if self.find_results:
            self.textbox.tag_config("i",background="Yellow")
            self.find_results_label.config(text="1/"+str(len(self.find_results)))
            self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
            self.textbox.tag_config("c",background="Orange")
            self.textbox.see(self.find_results[self.results_number])
            
    def find_up(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_down(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_reset(self,newfind=False):
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        if newfind == False:
            self.find_window.delete(0,"end")
        for tag in self.textbox.tag_names():
            self.textbox.tag_delete(tag)
        self.find_results_label.config(text="0/0")
        
    def cancel(self,event=None):
        self.destroy()


class terms_popup(tk.Toplevel):
    def __init__(self, C, text, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1", bg = theme_bg(theme))
        self.withdraw()
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.title("Tree Surgeon© - EULA - ")
        self.resizable(False,False)
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.has_agreed = False
        self.grid_columnconfigure(0,weight=1,uniform="y")
        self.grid_columnconfigure(1,weight=1,uniform="y")
        self.grid_rowconfigure(0,weight=1)
        self.textbox = working_text(self,
                                    wrap="word",
                                    theme = theme,
                                    use_entry_bg = False,
                                    override_bg = theme_treeview_bg(theme) if theme != "dark" else None)
        self.textbox.config(height=18)
        self.yscrollb = scrollbar(self,
                                  self.textbox.yview,
                                  "vertical",
                                  self.textbox)
        self.textbox.delete(1.0,"end")
        self.textbox.insert(1.0,text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=0,column=0,columnspan=2,padx=(20,0),pady=20,sticky="nswe")
        self.yscrollb.grid(row=0,column=2,padx=(0,20),pady=20,sticky="nswe")
        self.agree_button = button(self,text="Agree",
                                   underline = 0,
                                   style="EF.Std.TButton",
                                   command=self.agree)
        self.agree_button.grid(row=1,column=0,sticky="nswe",padx=(20,40),pady=20)
        self.disagree_button = button(self,text="Disagree",
                                      underline = 0,
                                      style="EF.Std.TButton",
                                      command=self.disagree)
        self.disagree_button.grid(row=1,column=1,sticky="nswe",padx=(40,20),pady=20)
        self.bind("<Escape>", self.disagree)
        self.bind("<A>", self.agree)
        self.bind("<a>", self.agree)
        self.bind("<D>", self.disagree)
        self.bind("<d>", self.disagree)
        center(self,700,650)
        self.deiconify()
        self.wait_window()
        
    def agree(self, event = None):
        self.has_agreed = True
        self.destroy()
        
    def disagree(self, event = None):
        self.destroy()


class helppopup(tk.Toplevel):
    def __init__(self, C, text, theme = "dark"):
        tk.Toplevel.__init__(self,C,width="1",height="1",bg = theme_bg(theme))
        self.withdraw()
        self.title(" Tree Surgeon© - Help")
        self.tk.call("wm","iconphoto",self._w,tk.PhotoImage(format="gif",data=top_left_icon))
        self.C = C
        self.wm_transient(self.C)
        self.focus_force()
        self.grab_set()
        self.word = ""
        self.findpos_start = "1.0"
        self.findpos_end = "1.0"
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        self.grid_columnconfigure(1,weight=1)
        self.grid_rowconfigure(1,weight=1)
        self.find_frame = frame(self, theme = theme)
        self.find_frame.grid(row=0,column=0,columnspan=3,sticky="nswe")
        self.find_icon = tk.PhotoImage(format="gif",data=find_icon)
        self.search_button = button(self.find_frame,
                                    text=" Find:",
                                    command=self.find)
        self.search_button.config(image=self.find_icon,compound="left")
        self.search_button.pack(side="left",fill="x")
        self.find_window = normal_entry(self.find_frame,font=BF, theme = theme)
        self.find_window.bind("<Return>",self.find)
        self.find_window.pack(side="left",fill="x",expand=True)
        self.find_reset_button = button(self.find_frame,text="X",command=self.find_reset)
        self.find_reset_button.pack(side="left",fill="x")
        self.find_results_label = label(self.find_frame,"0/0",BF, theme = theme)
        self.find_results_label.pack(side="left",fill="x")
        self.find_up_button = button(self.find_frame,text="▲",command=self.find_up)
        self.find_up_button.pack(side="left",fill="x")
        self.find_down_button = button(self.find_frame,text="▼",command=self.find_down)
        self.find_down_button.pack(side="left",fill="x")
        self.buttonframe = frame(self, theme = theme)
        self.buttonframe.grid(row=1,column=0,rowspan=2,sticky="nswe")
        
        self.basics = button(self.buttonframe,text="Program\nBasics",
                             style="EF.Std.TButton",
                             command=lambda:self.scrollto("programbasics"))
        self.basics.pack(side="top",padx=2,pady=2,fill="x")
        
        self.tutorials = button(self.buttonframe,text="Tips &\nTutorials",
                             style="EF.Std.TButton",
                             command=lambda:self.scrollto("tutorials"))
        self.tutorials.pack(side="top",padx=2,pady=2,fill="x")
        
        self.xlsx_files = button(self.buttonframe,text=".XLSX\nFiles",
                                 style="EF.Std.TButton",
                                 command=lambda:self.scrollto("xlsxfiles"))
        self.xlsx_files.pack(side="top",padx=2,pady=2,fill="x")
        
        self.treecomparehelp = button(self.buttonframe,text="Tree\nCompare",
                                      style="EF.Std.TButton",
                                      command=lambda:self.scrollto("treecompare"))
        self.treecomparehelp.pack(side="top",padx=2,pady=2,fill="x")
        
        self.menubar = button(self.buttonframe,text="Treeview\nMenu",
                              style="EF.Std.TButton",
                              command=lambda:self.scrollto("treeviewmenu"))
        self.menubar.pack(side="top",padx=2,pady=2,fill="x")
        
        self.managecolumns = button(self.buttonframe,text="Manage\nColumns",
                                    style="EF.Std.TButton",
                                    command=lambda:self.scrollto("managecolumns"))
        self.managecolumns.pack(side="top",padx=2,pady=2,fill="x")
        
        self.buttons = button(self.buttonframe,text="Treeview\nButtons",
                              style="EF.Std.TButton",
                              command=lambda:self.scrollto("treeviewbuttons"))
        self.buttons.pack(side="top",padx=2,pady=2,fill="x")
        
        self.functions = button(self.buttonframe,text="Treeview\nFunctions",
                                style="EF.Std.TButton",
                                command=lambda:self.scrollto("treeviewfunctions"))
        self.functions.pack(side="top",padx=2,pady=2,fill="x")
        
        self.tsrgnfiles = button(self.buttonframe,text=".JSON FILES",
                                style="EF.Std.TButton",
                                command=lambda:self.scrollto("tsrgnfiles"))
        self.tsrgnfiles.pack(side="top",padx=2,pady=2,fill="x")
        
        self.api = button(self.buttonframe,text="Using the\nAPI",
                          style="EF.Std.TButton",
                          command=lambda:self.scrollto("api"))
        self.api.pack(side="top",padx=2,pady=2,fill="x")
        
        self.textbox = working_text(self,
                                    font = ("Calibri",12),
                                    wrap="word",
                                    theme = theme,
                                    use_entry_bg = False,
                                    override_bg = theme_treeview_bg(theme) if theme != "dark" else None)
        self.yscrollb = scrollbar(self,
                                  self.textbox.yview,
                                  "vertical",
                                  self.textbox)
        self.textbox.delete(1.0,"end")
        self.textbox.insert(1.0,text)
        self.textbox.config(state="disabled")
        self.textbox.grid(row=1,column=1,sticky="nswe")
        self.yscrollb.grid(row=1,column=2,sticky="nswe")
        self.textbox.focus_set()
        self.bind("<Escape>",self.cancel)
        center(self,975,650)
        self.deiconify()
        self.wait_window()
        
    def scrollto(self,option):
        if option == "programbasics":
            self.textbox.see(self.textbox.search("    PROGRAM BASICS    ","1.0").split(".")[0] + ".0")
        elif option == "xlsxfiles":
            self.textbox.see(self.textbox.search("    XLSX FILES    ","1.0").split(".")[0] + ".0")
        elif option == "tutorials":
            self.textbox.see(self.textbox.search("    HELPFUL TIPS AND TUTORIALS    ","1.0").split(".")[0] + ".0")
        elif option == "treecompare":
            self.textbox.see(self.textbox.search("    TREE COMPARE    ","1.0").split(".")[0] + ".0")
        elif option == "treeviewmenu":
            self.textbox.see(self.textbox.search("    TREEVIEW MENUBAR    ","1.0").split(".")[0] + ".0")
        elif option == "managecolumns":
            self.textbox.see(self.textbox.search("    MANAGE COLUMNS    ","1.0").split(".")[0] + ".0")
        elif option == "treeviewbuttons":
            self.textbox.see(self.textbox.search("    TREEVIEW BUTTONS    ","1.0").split(".")[0] + ".0")
        elif option == "treeviewfunctions":
            self.textbox.see(self.textbox.search("    TREEVIEW FUNCTIONS    ","1.0").split(".")[0] + ".0")
        elif option == "tsrgnfiles":
            self.textbox.see(self.textbox.search("    JSON FILES    ","1.0").split(".")[0] + ".0")
        elif option == "api":
            self.textbox.see(self.textbox.search("    USING THE API    ","1.0").split(".")[0] + ".0")
            
    def find(self,event=None):
        self.find_reset(True)
        self.word = self.find_window.get()
        if not self.word:
            return
        self.addon = "+" + str(len(self.word)) + "c"
        start = "1.0"
        while start:
            start = self.textbox.search(self.word,index=start,nocase=1,stopindex="end")
            if start:
                end = start + self.addon
                self.find_results.append(start)
                self.textbox.tag_add("i",start,end)
                start = end
        if self.find_results:
            self.textbox.tag_config("i",background="Yellow")
            self.find_results_label.config(text="1/"+str(len(self.find_results)))
            self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
            self.textbox.tag_config("c",background="Orange")
            self.textbox.see(self.find_results[self.results_number])
            
    def find_up(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == 0:
            self.results_number = len(self.find_results) - 1
        else:
            self.results_number -= 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_down(self,event=None):
        if not self.find_results or len(self.find_results) == 1:
            return
        self.textbox.tag_remove("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        if self.results_number == len(self.find_results) - 1:
            self.results_number = 0
        else:
            self.results_number += 1
        self.find_results_label.config(text=str(self.results_number+1)+"/"+str(len(self.find_results)))
        self.textbox.tag_add("c",self.find_results[self.results_number],self.find_results[self.results_number]+self.addon)
        self.textbox.tag_config("c",background="Orange")
        self.textbox.see(self.find_results[self.results_number])
        
    def find_reset(self,newfind=False):
        self.find_results = []
        self.results_number = 0
        self.addon = ""
        if newfind == False:
            self.find_window.delete(0,"end")
        for tag in self.textbox.tag_names():
            self.textbox.tag_delete(tag)
        self.find_results_label.config(text="0/0")
        
    def cancel(self,event=None):
        self.destroy()



